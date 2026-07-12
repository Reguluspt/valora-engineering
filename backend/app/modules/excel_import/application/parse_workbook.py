import re
import os
import struct
import zipfile
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter

from app.modules.excel_import.domain import (
    DEFAULT_LIMITS, ACCEPTED_EXTENSIONS,
    REQUIRED_ZIP_PARTS, FORBIDDEN_ZIP_PARTS, EXTERNAL_LINK_PATHS, COLUMN_ALIASES,
)

class ParseError(Exception):
    def __init__(self, status, error_code, detail, limit_category=None):
        self.status = status
        self.error_code = error_code
        self.detail = detail
        self.limit_category = limit_category
        super().__init__(detail)

def _fail(status, error_code, detail, limit_category=None):
    raise ParseError(status, error_code, detail, limit_category)

_UNSAFE_CHARS = {"\x00"}
_WINDOWS_DRIVE = re.compile(r"^[a-zA-Z]:")

def _is_unsafe_zip_path(path):
    if any(c in path for c in _UNSAFE_CHARS):
        return True
    path = path.replace("\\", "/")
    if path.startswith("/"):
        return True
    if _WINDOWS_DRIVE.match(path):
        return True
    if path.startswith("//"):
        return True
    if any(p == ".." for p in path.split("/")):
        return True
    return False

def sanitize_filename(fn):
    fn = fn.replace("\\", "/")
    fn = os.path.basename(fn)
    fn = "".join(c for c in fn if c.isprintable() and c not in "\x00\x01\x02\x03\x04\x05\x06\x07\x08\x0b\x0c\x0e\x0f")
    return fn or "import.xlsx"

def _validate_zip(file_obj, limits):
    try:
        zf = zipfile.ZipFile(file_obj)
    except (zipfile.BadZipFile, zipfile.LargeZipFile, struct.error):
        _fail(400, "invalid_xlsx", "Tệp Excel không hợp lệ hoặc không thể đọc được.")
    with zf:
        infos = zf.infolist()
        if len(infos) > limits.max_zip_entries:
            _fail(413, "zip_entry_limit", "Số lượng thành phần trong tệp ZIP vượt quá giới hạn cho phép.", limit_category="zip_entries")
        total = sum(i.file_size for i in infos)
        if total > limits.max_uncompressed_zip_bytes:
            _fail(413, "zip_expansion_limit", "Kích thước giải nén tệp ZIP vượt quá giới hạn cho phép.", limit_category="zip_size")
        names = {i.filename for i in infos}
        for part in REQUIRED_ZIP_PARTS:
            if part not in names:
                _fail(400, "invalid_xlsx", "Thiếu thành phần cấu trúc XLSX bắt buộc.")
        for info in infos:
            fn = info.filename
            if _is_unsafe_zip_path(fn):
                _fail(400, "unsafe_zip_path", "Đường dẫn tệp trong lưu trữ ZIP không an toàn.")
            if info.flag_bits & 0x1:
                _fail(400, "encrypted_archive", "Tệp mã hóa không được hỗ trợ.")
            if fn in FORBIDDEN_ZIP_PARTS:
                _fail(400, "macro_not_allowed", "Tệp Excel chứa macro hoặc mã VBA không được phép.")
            if any(fn.startswith(p) for p in EXTERNAL_LINK_PATHS):
                _fail(400, "external_link_not_allowed", "Tệp Excel chứa liên kết ngoài không được phép.")
    file_obj.seek(0)

def _copy_spool(file, limits):
    s = tempfile.SpooledTemporaryFile(max_size=min(limits.max_upload_bytes, 1048576))
    total = 0
    while True:
        chunk = file.file.read(limits.read_chunk_size)
        if not chunk:
            break
        total += len(chunk)
        if total > limits.max_upload_bytes:
            s.close()
            _fail(413, "upload_too_large", "Kích thước tệp tải lên vượt quá giới hạn cho phép.", limit_category="file_size")
        s.write(chunk)
    s.seek(0)
    return s

def get_request_size(request):
    if request is None:
        return None
    val = getattr(request, "headers", {}).get("content-length")
    if val is None:
        return None
    try:
        return int(val)
    except Exception:
        return -1

def enforce_request_limit(size, limits):
    if size is not None and size < 0:
        _fail(400, "request_too_large", "Kích thước yêu cầu HTTP không hợp lệ.", limit_category="request_size")
    if size is not None and size > limits.max_request_bytes:
        _fail(413, "request_too_large", "Kích thước yêu cầu HTTP vượt quá giới hạn cho phép.", limit_category="request_size")

def parse_workbook_lazy(file, source_sheet_name, limits=None):
    limits = limits or DEFAULT_LIMITS
    fn = file.filename or "import.xlsx"
    ext = fn[fn.rfind("."):] if "." in fn else ""
    if ext.lower() not in ACCEPTED_EXTENSIONS:
        _fail(400, "unsupported_extension", "Định dạng tệp không được hỗ trợ. Chỉ chấp nhận tệp .xlsx.")
    enforce_request_limit(file.size, limits)
    spool = _copy_spool(file, limits)
    wb = None
    try:
        _validate_zip(spool, limits)
        wb = openpyxl.load_workbook(spool, data_only=True, read_only=True, keep_links=False, keep_vba=False)
        return _LazyWorkbook(spool, wb, source_sheet_name, limits)
    except Exception:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        try:
            spool.close()
        except Exception:
            pass
        raise

class _LazyWorkbook:
    def __init__(self, spool, wb, source_sheet_name, limits):
        self._spool = spool
        self._wb = wb
        self._limits = limits
        self._closed = False
        self._yielded = 0
        self._sheet_name = _resolve_sheet(wb, source_sheet_name)
        ws = wb[self._sheet_name]
        self._gen = enumerate(ws.iter_rows(values_only=True))
        self._headers, self._header_idx = self._find_headers()
        self._mapping = _map_columns(self._headers)

    @property
    def resolved_sheet(self):
        return self._sheet_name

    @property
    def column_count(self):
        return len(self._headers)

    def _find_headers(self):
        for r_idx, row in self._gen:
            if r_idx >= self._limits.max_header_search_rows:
                _fail(400, "header_not_found", "Không tìm thấy dòng tiêu đề hợp lệ trong phạm vi cho phép.")
            if any(c is not None for c in row):
                hdrs = []
                row_char_sum = 0
                for c in row:
                    if c is not None:
                        c_str = str(c)
                        if len(c_str) > self._limits.max_cell_chars:
                            _fail(400, "cell_length_limit", "Ô tiêu đề vượt quá độ dài ký tự cho phép.")
                        row_char_sum += len(c_str)
                        hdrs.append(c_str)
                    else:
                        hdrs.append("")

                if row_char_sum > self._limits.max_row_chars:
                    _fail(400, "row_length_limit", "Độ dài dòng tiêu đề vượt quá giới hạn ký tự cho phép.")

                if len(hdrs) > self._limits.max_columns:
                    _fail(400, "column_limit", "Số lượng cột tiêu đề vượt quá giới hạn cho phép.", limit_category="columns")
                return hdrs, r_idx
        _fail(400, "header_not_found", "Tệp trống hoặc không chứa tiêu đề hợp lệ.")

    def __iter__(self):
        return self

    def __next__(self):
        if self._closed:
            raise StopIteration
        while True:
            try:
                r_idx, row = next(self._gen)
            except StopIteration:
                self.close()
                raise
            if r_idx >= self._limits.max_physical_rows:
                _fail(413, "physical_row_limit", "Số lượng dòng vật lý vượt quá giới hạn cho phép.", limit_category="rows")
            if not any(c is not None for c in row):
                continue
            if self._yielded >= self._limits.max_data_rows:
                _fail(413, "data_row_limit", f"Số lượng dòng dữ liệu vượt quá giới hạn {self._limits.max_data_rows} dòng.", limit_category="rows")
            if len(row) > self._limits.max_columns:
                _fail(400, "column_limit", "Số lượng cột dữ liệu vượt quá giới hạn cho phép.", limit_category="columns")
            rl = 0
            for c in row:
                if c is not None:
                    s = str(c)
                    rl += len(s)
                    if len(s) > self._limits.max_cell_chars:
                        _fail(400, "cell_length_limit", "Ô dữ liệu vượt quá độ dài ký tự cho phép.")
            if rl > self._limits.max_row_chars:
                _fail(400, "row_length_limit", "Độ dài dòng dữ liệu vượt quá giới hạn ký tự cho phép.")
            cells = _build_cells(self._headers, row)
            mp = {}
            props = {}
            for tk, ci in self._mapping.items():
                if ci < len(row):
                    v = str(row[ci]) if row[ci] is not None else ""
                    mp[tk] = v
                    props[tk] = v
            self._yielded += 1
            return {"source_row_number": r_idx + 1, "raw_cells": cells, "mapped_values": mp,
                    "proposed_asset_name": props.get("proposed_asset_name"),
                    "proposed_description": props.get("proposed_description"),
                    "proposed_quantity": props.get("proposed_quantity"),
                    "proposed_unit": props.get("proposed_unit"),
                    "proposed_raw_price": props.get("proposed_raw_price"),
                    "proposed_currency": props.get("proposed_currency"),
                    "proposed_appraised_unit_price": props.get("proposed_appraised_unit_price")}

    def close(self):
        if not self._closed:
            self._closed = True
            try:
                self._wb.close()
            except Exception:
                pass
            try:
                self._spool.close()
            except Exception:
                pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

def _resolve_sheet(wb, requested):
    if requested:
        if requested in wb.sheetnames:
            return requested
        _fail(400, "sheet_not_found", f"Trang tính '{requested}' không tồn tại trong tệp.")
    if not wb.sheetnames:
        _fail(400, "invalid_xlsx", "Không tìm thấy trang tính hợp lệ.")
    return wb.sheetnames[0]

def _normalize(h):
    if not h:
        return ""
    return re.sub(r"[\s\-]+", "_", str(h).strip().lower())

def _map_columns(headers):
    m = {}
    for i, h in enumerate(headers):
        n = _normalize(h)
        if not n:
            continue
        for tk, aliases in COLUMN_ALIASES.items():
            if tk in m:
                continue
            if n in aliases:
                m[tk] = i
                break
    return m

def _build_cells(headers, row):
    cells = []
    actual = max(len(headers), len(row))
    for ci in range(actual):
        cells.append({"column_index": ci + 1, "column_letter": get_column_letter(ci + 1),
                      "header": str(headers[ci]) if ci < len(headers) and headers[ci] else "",
                      "value": str(row[ci]) if ci < len(row) and row[ci] is not None else ""})
    return cells
