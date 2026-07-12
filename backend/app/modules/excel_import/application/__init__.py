import re
import io
import struct
import zipfile
import tempfile

import openpyxl
from openpyxl.utils import get_column_letter
from fastapi import HTTPException, UploadFile

from app.modules.excel_import.domain import (
    DEFAULT_LIMITS,
    ExcelImportLimits,
    ACCEPTED_EXTENSIONS,
    REQUIRED_ZIP_PARTS,
    FORBIDDEN_ZIP_PARTS,
    EXTERNAL_LINK_PATHS,
    COLUMN_ALIASES,
)


def _safe_413(msg: str) -> HTTPException:
    return HTTPException(status_code=413, detail=msg)


def _safe_400(msg: str) -> HTTPException:
    return HTTPException(status_code=400, detail=msg)


def parse_uploaded_workbook(
    file: UploadFile,
    source_sheet_name: str | None,
    limits: ExcelImportLimits | None = None,
) -> tuple[list[dict], list[str], list[dict], int, str]:
    limits = limits or DEFAULT_LIMITS

    # 1. Validate extension
    filename = file.filename or "import.xlsx"
    ext_idx = filename.rfind(".")
    ext = filename[ext_idx:] if ext_idx >= 0 else ""
    if ext.lower() not in ACCEPTED_EXTENSIONS:
        raise _safe_400("Định dạng tệp không được hỗ trợ. Vui lòng tải lên tệp .xlsx")
    sanitized_filename = filename.split("/")[-1].split("\\")[-1]

    # 2. Bounded file read into spooled temp
    spool = _copy_upload_with_limit(file, limits)
    try:
        file_bytes = spool.read()
    finally:
        spool.close()

    # 3. ZIP/XLSX validation
    _validate_zip_container(file_bytes, limits)

    # 4. openpyxl load
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
            read_only=True,
            keep_links=False,
            keep_vba=False,
        )
    except Exception:
        raise _safe_400("Không thể đọc tệp Excel. Vui lòng kiểm tra lại cấu trúc tệp.")

    try:
        # 5. Sheet selection
        sheet_name = _resolve_sheet(wb, source_sheet_name)

        # 6. Headers
        ws = wb[sheet_name]
        headers, header_row_idx = _find_headers(ws, limits)

        # 7. Column mapping
        mapping = _map_columns(headers)

        # 8. Streaming parse
        staging_rows, raw_cells_list, parsed_count = _parse_data_rows(
            ws, headers, header_row_idx, mapping, limits
        )
    finally:
        wb.close()

    return staging_rows, headers, raw_cells_list, parsed_count, sanitized_filename


def _copy_upload_with_limit(file: UploadFile, limits: ExcelImportLimits) -> tempfile.SpooledTemporaryFile:
    spool = tempfile.SpooledTemporaryFile(max_size=limits.max_upload_bytes)
    total = 0
    while True:
        chunk = file.file.read(limits.read_chunk_size)
        if not chunk:
            break
        total += len(chunk)
        if total > limits.max_upload_bytes:
            spool.close()
            raise _safe_413("Tệp tải lên vượt quá kích thước cho phép. Vui lòng giảm kích thước tệp.")
        spool.write(chunk)
    spool.seek(0)
    return spool


def _validate_zip_container(data: bytes, limits: ExcelImportLimits) -> None:
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except (zipfile.BadZipFile, zipfile.LargeZipFile, struct.error):
        raise _safe_400("Tệp Excel không hợp lệ. Tệp không phải là định dạng XLSX hợp lệ.")

    with zf:
        infos = zf.infolist()
        if len(infos) > limits.max_zip_entries:
            raise _safe_413("Tệp Excel có quá nhiều thành phần bên trong. Vui lòng giảm kích thước tệp.")

        total_uncompressed = sum(info.file_size for info in infos)
        if total_uncompressed > limits.max_uncompressed_zip_bytes:
            raise _safe_413("Kích thước giải nén của tệp Excel vượt quá giới hạn cho phép.")

        names = {info.filename for info in infos}
        for part in REQUIRED_ZIP_PARTS:
            if part not in names:
                raise _safe_400("Tệp Excel không hợp lệ. Thiếu thành phần cấu trúc XLSX bắt buộc.")

        for info in infos:
            fn = info.filename
            if fn.startswith("/") or ".." in fn:
                raise _safe_400("Tệp Excel chứa đường dẫn không an toàn.")
            if info.flag_bits & 0x1:
                raise _safe_400("Tệp Excel được mã hóa không được hỗ trợ. Vui lòng giải mã tệp trước khi tải lên.")
            if fn in FORBIDDEN_ZIP_PARTS:
                raise _safe_400("Tệp Excel chứa macro VBA không được hỗ trợ.")
            if any(fn.startswith(prefix) for prefix in EXTERNAL_LINK_PATHS):
                raise _safe_400("Tệp Excel chứa liên kết ngoài không được hỗ trợ.")


def _resolve_sheet(wb: openpyxl.Workbook, requested: str | None) -> str:
    if requested:
        if requested in wb.sheetnames:
            return requested
        raise _safe_400(f"Trang tính '{requested}' không tồn tại trong tệp Excel. Vui lòng kiểm tra tên trang tính.")
    if not wb.sheetnames:
        raise _safe_400("Tệp Excel không chứa trang tính nào.")
    return wb.sheetnames[0]


def _find_headers(ws: openpyxl.worksheet.worksheet.Worksheet, limits: ExcelImportLimits):
    header_row_idx = None
    headers = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx >= limits.max_header_search_rows:
            raise _safe_400("Không tìm thấy dòng tiêu đề trong tệp Excel. Vui lòng kiểm tra cấu trúc tệp.")
        if any(cell is not None for cell in row):
            header_row_idx = r_idx
            headers = [str(cell) if cell is not None else "" for cell in row]
            if len(headers) > limits.max_columns:
                raise _safe_400(f"Tệp Excel có quá nhiều cột. Số cột tối đa là {limits.max_columns}.")
            return headers, header_row_idx
    raise _safe_400("Tệp Excel trống hoặc không chứa dòng tiêu đề.")


def _normalize_header(h: str) -> str:
    if not h:
        return ""
    return re.sub(r"[\s\-]+", "_", str(h).strip().lower())


def _map_columns(headers: list[str]) -> dict:
    mapping: dict[str, int] = {}
    for i, h in enumerate(headers):
        normalized = _normalize_header(h)
        if not normalized:
            continue
        for target_key, alias_list in COLUMN_ALIASES.items():
            if target_key in mapping:
                continue
            if normalized in alias_list:
                mapping[target_key] = i
                break
    return mapping


def _build_raw_cells(
    headers: list[str],
    row: tuple,
    row_idx: int,
) -> list[dict]:
    cells = []
    actual_cols = max(len(headers), len(row))
    for col_idx in range(actual_cols):
        col_letter = get_column_letter(col_idx + 1)
        header = headers[col_idx] if col_idx < len(headers) else ""
        cell_val = row[col_idx] if col_idx < len(row) else None
        val_str = str(cell_val) if cell_val is not None else ""
        cells.append({
            "column_index": col_idx + 1,
            "column_letter": col_letter,
            "header": str(header) if header else "",
            "value": val_str,
        })
    return cells


def _parse_data_rows(
    ws,
    headers: list[str],
    header_row_idx: int,
    mapping: dict,
    limits: ExcelImportLimits,
) -> tuple[list[dict], list[list[dict]], int]:
    staging_rows = []
    raw_cells_list = []
    data_count = 0

    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_row_idx:
            continue
        if r_idx >= limits.max_physical_rows:
            raise _safe_413(f"Tệp Excel có quá nhiều dòng. Giới hạn tối đa là {limits.max_physical_rows} dòng.")

        if not any(cell is not None for cell in row):
            continue

        if data_count >= limits.max_data_rows:
            raise _safe_413(
                f"Tệp Excel vượt quá {limits.max_data_rows} dòng dữ liệu. "
                f"Toàn bộ tệp bị từ chối, không nhập dữ liệu nào."
            )

        if len(row) > limits.max_columns:
            raise _safe_400(f"Tệp Excel có dòng vượt quá {limits.max_columns} cột. Vui lòng giảm số cột.")

        row_str_len = 0
        for cell in row:
            if cell is not None:
                s = str(cell)
                row_str_len += len(s)
                if len(s) > limits.max_cell_chars:
                    raise _safe_400("Tệp Excel chứa ô dữ liệu quá dài. Vui lòng rút gọn nội dung ô.")
        if row_str_len > limits.max_row_chars:
            raise _safe_400("Tệp Excel chứa dòng dữ liệu quá lớn. Vui lòng rút gọn nội dung dòng.")

        raw_cells = _build_raw_cells(headers, row, r_idx)
        raw_cells_list.append(raw_cells)

        mapped_values = {}
        proposed: dict[str, str] = {}
        for target_key, col_idx in mapping.items():
            if col_idx < len(row):
                cell_val = row[col_idx]
                val_str = str(cell_val) if cell_val is not None else ""
                mapped_values[target_key] = val_str
                proposed[target_key] = val_str

        staging_rows.append({
            "source_row_number": r_idx + 1,
            "raw_cells": raw_cells,
            "mapped_values": mapped_values,
            "proposed_asset_name": proposed.get("proposed_asset_name"),
            "proposed_description": proposed.get("proposed_description"),
            "proposed_quantity": proposed.get("proposed_quantity"),
            "proposed_unit": proposed.get("proposed_unit"),
            "proposed_raw_price": proposed.get("proposed_raw_price"),
            "proposed_currency": proposed.get("proposed_currency"),
            "proposed_appraised_unit_price": proposed.get("proposed_appraised_unit_price"),
        })
        data_count += 1

    return staging_rows, raw_cells_list, data_count
