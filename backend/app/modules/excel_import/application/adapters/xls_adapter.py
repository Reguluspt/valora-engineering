"""Value-only .xls adapter: OLE/BIFF presence reject + xlrd values + full bounds."""
from __future__ import annotations

from openpyxl.utils import get_column_letter

from app.modules.excel_import.application.adapters.xls_safety import assert_xls_safety
from app.modules.excel_import.domain.source_artifact import (
    DEFAULT_SOURCE_LIMITS,
    SourceArtifactLimits,
    XLS_OLE_SIGNATURE,
)
from app.modules.excel_import.domain.workbook_adapter import (
    AdapterInspectionResult,
    CellValue,
    MergedRegion,
    SheetSummary,
    WorkbookFormat,
    fail_adapter,
)


class XlsWorkbookAdapter:
    """
    xlrd-based BIFF reader with fail-closed presence detection.

    Merged regions require formatting_info=True (no incorrect global BIFF fallback).
    """

    name = "xls-xlrd"
    version = "s13-pr-002-v3"
    format = WorkbookFormat.XLS

    def __init__(self, limits: SourceArtifactLimits | None = None):
        self._limits = limits or DEFAULT_SOURCE_LIMITS
        self._book = None

    def close(self) -> None:
        if self._book is not None:
            try:
                self._book.release_resources()
            except Exception:
                pass
            self._book = None

    def _assert_signature(self, path: str) -> None:
        with open(path, "rb") as f:
            sig = f.read(8)
        if not sig.startswith(XLS_OLE_SIGNATURE):
            fail_adapter(400, "signature_mismatch", "Chữ ký tệp không khớp định dạng .xls.")

    def _xlrd_open(self, path: str, *, formatting_info: bool):
        import xlrd

        try:
            return xlrd.open_workbook(
                path,
                formatting_info=formatting_info,
                on_demand=True,
                ragged_rows=False,
            )
        except xlrd.XLRDError as exc:
            msg = str(exc).lower()
            if "password" in msg or "encrypt" in msg or "file_pass" in msg:
                fail_adapter(400, "encrypted_workbook", "Tệp mã hóa không được hỗ trợ.")
            if formatting_info:
                fail_adapter(
                    400,
                    "invalid_xls",
                    "Không thể đọc metadata gộp ô an toàn từ tệp .xls.",
                )
            fail_adapter(400, "invalid_xls", "Tệp Excel .xls không hợp lệ hoặc không thể đọc được.")
        except Exception:
            if formatting_info:
                fail_adapter(
                    400,
                    "invalid_xls",
                    "Không thể đọc metadata gộp ô an toàn từ tệp .xls.",
                )
            fail_adapter(400, "invalid_xls", "Tệp Excel .xls không hợp lệ hoặc không thể đọc được.")

    def _merged_for_sheet(self, sh) -> tuple[MergedRegion, ...]:
        """Sheet-local merges from xlrd only (0-based half-open → 1-based inclusive)."""
        limits = self._limits
        regions: list[MergedRegion] = []
        raw = list(getattr(sh, "merged_cells", None) or [])
        for item in raw:
            if len(item) != 4:
                fail_adapter(400, "invalid_xls", "Tệp Excel .xls không hợp lệ hoặc không thể đọc được.")
            rlo, rhi, clo, chi = item
            # xlrd merged_cells: rlo/clo inclusive, rhi/chi exclusive (0-based)
            if rlo < 0 or clo < 0 or rhi <= rlo or chi <= clo:
                fail_adapter(400, "invalid_xls", "Tệp Excel .xls không hợp lệ hoặc không thể đọc được.")
            regions.append(
                MergedRegion(
                    min_row=rlo + 1,
                    min_col=clo + 1,
                    max_row=rhi,  # exclusive end → inclusive 1-based max is rhi
                    max_col=chi,
                )
            )
            if len(regions) > limits.max_merged_regions_per_sheet:
                fail_adapter(
                    413,
                    "merged_region_limit",
                    "Số vùng gộp ô vượt quá giới hạn cho phép.",
                    "merged",
                )
        return tuple(regions)

    def _cell_value(self, book, cell, r_1based: int, c_1based: int) -> CellValue:
        import xlrd

        val = cell.value
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                val = xlrd.xldate_as_datetime(cell.value, book.datemode).isoformat()
                ctype = "datetime"
            except Exception:
                ctype = "number"
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            val = bool(cell.value)
            ctype = "boolean"
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            ctype = "number"
        elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            val = None
            ctype = "empty"
        elif cell.ctype == xlrd.XL_CELL_ERROR:
            ctype = "error"
        elif cell.ctype == xlrd.XL_CELL_TEXT:
            ctype = "string"
        else:
            ctype = "other"

        if isinstance(val, str) and len(val) > self._limits.max_cell_chars:
            fail_adapter(
                400,
                "cell_length_limit",
                "Ô dữ liệu vượt quá độ dài ký tự cho phép.",
                "cell",
            )
        return CellValue(
            row=r_1based,
            column=c_1based,
            coordinate=f"{get_column_letter(c_1based)}{r_1based}",
            value=val,
            cell_type=ctype,
        )

    def _exhaust_sheet_bounds(self, book, sh) -> tuple[int, int, int]:
        limits = self._limits
        if sh.ncols > limits.max_columns:
            fail_adapter(
                413,
                "column_limit",
                "Số lượng cột vượt quá giới hạn cho phép.",
                "columns",
            )
        if sh.nrows > limits.max_physical_rows:
            fail_adapter(
                413,
                "physical_row_limit",
                "Số lượng dòng vật lý vượt quá giới hạn cho phép.",
                "rows",
            )
        cell_count = 0
        max_col_seen = 0
        width = sh.ncols
        for r_idx in range(sh.nrows):
            row_chars = 0
            for c_idx in range(width):
                cell = sh.cell(r_idx, c_idx)
                cell_count += 1
                if cell_count > limits.max_total_cells:
                    fail_adapter(
                        413,
                        "total_cell_limit",
                        "Tổng số ô vượt quá giới hạn cho phép.",
                        "cells",
                    )
                cv = self._cell_value(book, cell, r_idx + 1, c_idx + 1)
                if isinstance(cv.value, str):
                    row_chars += len(cv.value)
                if c_idx + 1 > max_col_seen:
                    max_col_seen = c_idx + 1
            if row_chars > limits.max_row_chars:
                fail_adapter(
                    413,
                    "row_char_limit",
                    "Tổng độ dài ký tự trên một dòng vượt quá giới hạn cho phép.",
                    "row_chars",
                )
        return sh.nrows, max(max_col_seen, width), cell_count

    def inspect(self, path: str) -> AdapterInspectionResult:
        self._assert_signature(path)
        assert_xls_safety(path)
        book = None
        try:
            # Merges require formatting_info — fail closed if unavailable
            book = self._xlrd_open(path, formatting_info=True)
            names = tuple(book.sheet_names())
            if len(names) > self._limits.max_sheets:
                fail_adapter(
                    413,
                    "sheet_limit",
                    "Số lượng sheet vượt quá giới hạn cho phép.",
                    "sheets",
                )
            sheets: list[SheetSummary] = []
            total_cells = 0
            total_merged = 0
            for name in names:
                sh = book.sheet_by_name(name)
                max_row, max_col, cells = self._exhaust_sheet_bounds(book, sh)
                total_cells += cells
                if total_cells > self._limits.max_total_cells:
                    fail_adapter(
                        413,
                        "total_cell_limit",
                        "Tổng số ô vượt quá giới hạn cho phép.",
                        "cells",
                    )
                merged = self._merged_for_sheet(sh)
                total_merged += len(merged)
                if total_merged > self._limits.max_merged_regions:
                    fail_adapter(
                        413,
                        "merged_region_limit",
                        "Số vùng gộp ô vượt quá giới hạn cho phép.",
                        "merged",
                    )
                sheets.append(
                    SheetSummary(
                        name=name,
                        max_row=max_row,
                        max_column=max_col,
                        merged_regions=merged,
                    )
                )
            return AdapterInspectionResult(
                format=WorkbookFormat.XLS,
                adapter_name=self.name,
                adapter_version=self.version,
                sheet_names=names,
                sheets=tuple(sheets),
                safe_metadata={
                    "sheet_count": len(names),
                    "limit_version": self._limits.limit_version,
                    "library": "xlrd+olefile",
                    "total_cells_inspected": total_cells,
                    "total_merged_regions": total_merged,
                    "merged_source": "xlrd_formatting_info",
                },
            )
        finally:
            if book is not None:
                try:
                    book.release_resources()
                except Exception:
                    pass

    def iter_rows(self, path: str, sheet_name: str | None = None):
        self._assert_signature(path)
        assert_xls_safety(path)
        # Value iteration does not need formatting_info
        book = self._xlrd_open(path, formatting_info=False)
        self._book = book
        try:
            if sheet_name and sheet_name in book.sheet_names():
                sh = book.sheet_by_name(sheet_name)
            else:
                sh = book.sheet_by_index(0)
            width = sh.ncols
            if width > self._limits.max_columns:
                fail_adapter(
                    413,
                    "column_limit",
                    "Số lượng cột vượt quá giới hạn cho phép.",
                    "columns",
                )
            for r_idx in range(sh.nrows):
                if r_idx + 1 > self._limits.max_physical_rows:
                    fail_adapter(
                        413,
                        "physical_row_limit",
                        "Số lượng dòng vật lý vượt quá giới hạn cho phép.",
                        "rows",
                    )
                row_chars = 0
                cells: list[CellValue] = []
                for c_idx in range(width):
                    cv = self._cell_value(book, sh.cell(r_idx, c_idx), r_idx + 1, c_idx + 1)
                    if isinstance(cv.value, str):
                        row_chars += len(cv.value)
                    cells.append(cv)
                if row_chars > self._limits.max_row_chars:
                    fail_adapter(
                        413,
                        "row_char_limit",
                        "Tổng độ dài ký tự trên một dòng vượt quá giới hạn cho phép.",
                        "row_chars",
                    )
                yield tuple(cells)
        finally:
            self.close()
