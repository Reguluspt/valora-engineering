"""Value-only .xlsx adapter: ZIP safety + openpyxl + full bounds + OOXML merges."""
from __future__ import annotations

import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.modules.excel_import.application.adapters.xlsx_merge import (
    extract_merged_regions_from_xlsx,
)
from app.modules.excel_import.application.parse_workbook import (
    _is_unsafe_zip_path,
)
from app.modules.excel_import.domain import (
    EXTERNAL_LINK_PATHS,
    FORBIDDEN_ZIP_PARTS,
    REQUIRED_ZIP_PARTS,
)
from app.modules.excel_import.domain.source_artifact import (
    DEFAULT_SOURCE_LIMITS,
    SourceArtifactLimits,
    XLSX_ZIP_SIGNATURE,
)
from app.modules.excel_import.domain.workbook_adapter import (
    AdapterInspectionResult,
    CellValue,
    SheetSummary,
    WorkbookFormat,
    fail_adapter,
)


class XlsxWorkbookAdapter:
    name = "xlsx-openpyxl"
    version = "s13-pr-002-v2"
    format = WorkbookFormat.XLSX

    def __init__(self, limits: SourceArtifactLimits | None = None):
        self._limits = limits or DEFAULT_SOURCE_LIMITS
        self._wb = None

    def close(self) -> None:
        if self._wb is not None:
            try:
                self._wb.close()
            except Exception:
                pass
            self._wb = None

    def _validate_zip(self, path: str) -> None:
        try:
            zf = zipfile.ZipFile(path)
        except (zipfile.BadZipFile, zipfile.LargeZipFile, OSError):
            fail_adapter(400, "invalid_xlsx", "Tệp Excel không hợp lệ hoặc không thể đọc được.")
        with zf:
            infos = zf.infolist()
            if len(infos) > self._limits.max_zip_entries:
                fail_adapter(
                    413,
                    "zip_entry_limit",
                    "Số lượng thành phần trong tệp ZIP vượt quá giới hạn cho phép.",
                    "zip_entries",
                )
            total = sum(i.file_size for i in infos)
            if total > self._limits.max_uncompressed_zip_bytes:
                fail_adapter(
                    413,
                    "zip_expansion_limit",
                    "Kích thước giải nén tệp ZIP vượt quá giới hạn cho phép.",
                    "zip_size",
                )
            names = {i.filename for i in infos}
            for part in REQUIRED_ZIP_PARTS:
                if part not in names:
                    fail_adapter(400, "invalid_xlsx", "Thiếu thành phần cấu trúc XLSX bắt buộc.")
            for info in infos:
                fn = info.filename
                if _is_unsafe_zip_path(fn):
                    fail_adapter(
                        400,
                        "unsafe_zip_path",
                        "Đường dẫn tệp trong lưu trữ ZIP không an toàn.",
                    )
                if info.flag_bits & 0x1:
                    fail_adapter(400, "encrypted_archive", "Tệp mã hóa không được hỗ trợ.")
                if fn in FORBIDDEN_ZIP_PARTS:
                    fail_adapter(
                        400,
                        "macro_not_allowed",
                        "Tệp Excel chứa macro hoặc mã VBA không được phép.",
                    )
                # Also reject case variants / nested vba
                lower = fn.lower()
                if "vbaproject" in lower.replace("\\", "/"):
                    fail_adapter(
                        400,
                        "macro_not_allowed",
                        "Tệp Excel chứa macro hoặc mã VBA không được phép.",
                    )
                if any(fn.startswith(p) for p in EXTERNAL_LINK_PATHS):
                    fail_adapter(
                        400,
                        "external_link_not_allowed",
                        "Tệp Excel chứa liên kết ngoài không được phép.",
                    )
                if "externallinks" in lower.replace("\\", "/"):
                    fail_adapter(
                        400,
                        "external_link_not_allowed",
                        "Tệp Excel chứa liên kết ngoài không được phép.",
                    )

    def _pad_row(self, row, width: int) -> list:
        vals = list(row)
        if width and len(vals) < width:
            vals.extend([None] * (width - len(vals)))
        return vals

    def _exhaust_sheet(self, ws) -> tuple[int, int, int]:
        """Iterate real rows/cells; return max_row, max_col, cell_count."""
        limits = self._limits
        # Prefer declared width so trailing blanks keep column positions
        declared_width = int(ws.max_column or 0)
        if declared_width > limits.max_columns:
            fail_adapter(
                413,
                "column_limit",
                "Số lượng cột vượt quá giới hạn cho phép.",
                "columns",
            )
        max_row = 0
        max_col = 0
        cell_count = 0
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx > limits.max_physical_rows:
                fail_adapter(
                    413,
                    "physical_row_limit",
                    "Số lượng dòng vật lý vượt quá giới hạn cho phép.",
                    "rows",
                )
            max_row = r_idx
            row_chars = 0
            width = max(declared_width, len(row))
            if width > limits.max_columns:
                fail_adapter(
                    413,
                    "column_limit",
                    "Số lượng cột vượt quá giới hạn cho phép.",
                    "columns",
                )
            vals = self._pad_row(row, width)
            for c_idx, val in enumerate(vals, start=1):
                cell_count += 1
                if cell_count > limits.max_total_cells:
                    fail_adapter(
                        413,
                        "total_cell_limit",
                        "Tổng số ô vượt quá giới hạn cho phép.",
                        "cells",
                    )
                if c_idx > max_col:
                    max_col = c_idx
                if isinstance(val, str):
                    if len(val) > limits.max_cell_chars:
                        fail_adapter(
                            400,
                            "cell_length_limit",
                            "Ô dữ liệu vượt quá độ dài ký tự cho phép.",
                            "cell",
                        )
                    row_chars += len(val)
            if row_chars > limits.max_row_chars:
                fail_adapter(
                    413,
                    "row_char_limit",
                    "Tổng độ dài ký tự trên một dòng vượt quá giới hạn cho phép.",
                    "row_chars",
                )
        return max_row, max(max_col, declared_width), cell_count

    def inspect(self, path: str) -> AdapterInspectionResult:
        with open(path, "rb") as f:
            sig = f.read(2)
        if not sig.startswith(XLSX_ZIP_SIGNATURE):
            fail_adapter(400, "signature_mismatch", "Chữ ký tệp không khớp định dạng .xlsx.")
        self._validate_zip(path)

        merges_by_sheet = extract_merged_regions_from_xlsx(
            path,
            max_merged_total=self._limits.max_merged_regions,
            max_merged_per_sheet=self._limits.max_merged_regions_per_sheet,
        )

        wb = None
        try:
            wb = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_links=False,
                keep_vba=False,
            )
            names = tuple(wb.sheetnames)
            if len(names) > self._limits.max_sheets:
                fail_adapter(
                    413,
                    "sheet_limit",
                    "Số lượng sheet vượt quá giới hạn cho phép.",
                    "sheets",
                )
            sheets: list[SheetSummary] = []
            total_cells = 0
            total_merged = 0
            for name in names:
                ws = wb[name]
                max_row, max_col, cells = self._exhaust_sheet(ws)
                total_cells += cells
                if total_cells > self._limits.max_total_cells:
                    fail_adapter(
                        413,
                        "total_cell_limit",
                        "Tổng số ô vượt quá giới hạn cho phép.",
                        "cells",
                    )
                merged = merges_by_sheet.get(name, ())
                total_merged += len(merged)
                if total_merged > self._limits.max_merged_regions:
                    fail_adapter(
                        413,
                        "merged_region_limit",
                        "Số vùng gộp ô vượt quá giới hạn cho phép.",
                        "merged",
                    )
                sheets.append(
                    SheetSummary(
                        name=name,
                        max_row=max_row,
                        max_column=max_col,
                        merged_regions=merged,
                    )
                )
            return AdapterInspectionResult(
                format=WorkbookFormat.XLSX,
                adapter_name=self.name,
                adapter_version=self.version,
                sheet_names=names,
                sheets=tuple(sheets),
                safe_metadata={
                    "sheet_count": len(names),
                    "limit_version": self._limits.limit_version,
                    "total_cells_inspected": total_cells,
                    "total_merged_regions": total_merged,
                },
            )
        except Exception as exc:
            from app.modules.excel_import.domain.workbook_adapter import AdapterError

            if isinstance(exc, AdapterError):
                raise
            fail_adapter(400, "invalid_xlsx", "Tệp Excel không hợp lệ hoặc không thể đọc được.")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def iter_rows(self, path: str, sheet_name: str | None = None):
        self._validate_zip(path)
        wb = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
        self._wb = wb
        try:
            name = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
            ws = wb[name]
            declared_width = int(ws.max_column or 0)
            if declared_width > self._limits.max_columns:
                fail_adapter(
                    413,
                    "column_limit",
                    "Số lượng cột vượt quá giới hạn cho phép.",
                    "columns",
                )
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if r_idx > self._limits.max_physical_rows:
                    fail_adapter(
                        413,
                        "physical_row_limit",
                        "Số lượng dòng vật lý vượt quá giới hạn cho phép.",
                        "rows",
                    )
                cells: list[CellValue] = []
                row_chars = 0
                width = max(declared_width, len(row))
                if width > self._limits.max_columns:
                    fail_adapter(
                        413,
                        "column_limit",
                        "Số lượng cột vượt quá giới hạn cho phép.",
                        "columns",
                    )
                vals = self._pad_row(row, width)
                for c_idx, val in enumerate(vals, start=1):
                    if isinstance(val, str) and len(val) > self._limits.max_cell_chars:
                        fail_adapter(
                            400,
                            "cell_length_limit",
                            "Ô dữ liệu vượt quá độ dài ký tự cho phép.",
                            "cell",
                        )
                    if isinstance(val, str):
                        row_chars += len(val)
                    cells.append(
                        CellValue(
                            row=r_idx,
                            column=c_idx,
                            coordinate=f"{get_column_letter(c_idx)}{r_idx}",
                            value=val,
                            cell_type=_cell_type(val),
                        )
                    )
                if row_chars > self._limits.max_row_chars:
                    fail_adapter(
                        413,
                        "row_char_limit",
                        "Tổng độ dài ký tự trên một dòng vượt quá giới hạn cho phép.",
                        "row_chars",
                    )
                yield tuple(cells)
        finally:
            self.close()


def _cell_type(val) -> str:
    if val is None:
        return "empty"
    if isinstance(val, bool):
        return "boolean"
    if isinstance(val, (int, float)):
        return "number"
    if isinstance(val, str):
        return "string"
    return "other"
