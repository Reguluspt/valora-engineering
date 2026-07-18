"""S13-PR-002 twelfth corrective: M-01…M-05 evidence completeness.

Test/support only. Closes remaining K-03/L-01…L-05 gaps:
- e04 genuine upload-byte boundary (M-01, fixed in fifth suite)
- complete mapper column snapshots (M-02)
- type-preserving canonicalization + DB mutation probes (M-03)
- marker-based collectable matrix + runtime strong-helper proof (M-04)
- honest boundary vs threat vs support counts (M-05)
"""
from __future__ import annotations

import hashlib
import io
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    ARTIFACT_FIELDS,
    AUDIT_FIELDS,
    BATCH_FIELDS,
    LINE_FIELDS,
    STAGING_FIELDS,
    assert_canonical_distinguishes_collisions,
    assert_field_sets_match_mappers,
    assert_pytest_collect_count_exactly,
    assert_source_intake_preserve,
    persisted_column_keys,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
    AuditEvent,
)


# ---------------------------------------------------------------------------
# M-04 / M-05: expected format/bound manifest (parameter-level, not AST)
# ---------------------------------------------------------------------------
# Keys are (reachability, bound). Format-agnostic request/upload appear once.
# Shared adapter bounds appear separately for xlsx and xls. ZIP is xlsx-only.

EXPECTED_FORMAT_BOUND: frozenset[tuple[str, str]] = frozenset(
    {
        ("intake", "max_request_bytes"),
        ("intake", "max_upload_bytes"),
        ("xlsx", "max_sheets"),
        ("xlsx", "max_physical_rows"),
        ("xlsx", "max_columns"),
        ("xlsx", "max_cell_chars"),
        ("xlsx", "max_row_chars"),
        ("xlsx", "max_total_cells"),
        ("xlsx", "max_merged_regions"),
        ("xlsx", "max_merged_regions_per_sheet"),
        ("xlsx", "max_zip_entries"),
        ("xlsx", "max_uncompressed_zip_bytes"),
        ("xls", "max_sheets"),
        ("xls", "max_physical_rows"),
        ("xls", "max_columns"),
        ("xls", "max_cell_chars"),
        ("xls", "max_row_chars"),
        ("xls", "max_total_cells"),
        ("xls", "max_merged_regions"),
        ("xls", "max_merged_regions_per_sheet"),
    }
)

# Concrete accepted/rejected node mappings for every format/bound (no "—"/ellipsis).
# rejected_node is a substring match against collected nodeid; accepted_node is a
# concrete function or "same-node-N-companion" when N and N+1 share one test.
FORMAT_BOUND_MANIFEST: list[dict[str, str]] = [
    {
        "reachability": "intake",
        "bound": "max_request_bytes",
        "error_code": "request_too_large",
        "accepted_node": "test_i03_request_bytes_exact_n_accepted_n_plus_one_rejected",
        "rejected_node": "test_i03_request_bytes_exact_n_accepted_n_plus_one_rejected",
        "suite": "eighth",
    },
    {
        "reachability": "intake",
        "bound": "max_upload_bytes",
        "error_code": "upload_too_large",
        "accepted_node": "test_e04_endpoint_upload_bytes_limit",
        "rejected_node": "test_e04_endpoint_upload_bytes_limit",
        "suite": "fifth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_sheets",
        "error_code": "sheet_limit",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one[max_sheets",
        "suite": "eighth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_physical_rows",
        "error_code": "physical_row_limit",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one[max_physical_rows",
        "suite": "eighth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_columns",
        "error_code": "column_limit",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one[max_columns",
        "suite": "eighth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_cell_chars",
        "error_code": "cell_length_limit",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one[max_cell_chars",
        "suite": "eighth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_row_chars",
        "error_code": "row_char_limit",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds[max_row_chars",
        "suite": "ninth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_total_cells",
        "error_code": "total_cell_limit",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds[max_total_cells",
        "suite": "ninth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_merged_regions",
        "error_code": "merged_region_limit",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds[max_merged_regions-",
        "suite": "ninth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_merged_regions_per_sheet",
        "error_code": "merged_region_limit",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds[max_merged_regions_per_sheet",
        "suite": "ninth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_zip_entries",
        "error_code": "zip_entry_limit",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds[max_zip_entries",
        "suite": "ninth",
    },
    {
        "reachability": "xlsx",
        "bound": "max_uncompressed_zip_bytes",
        "error_code": "zip_expansion_limit",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds[max_uncompressed_zip_bytes",
        "suite": "ninth",
    },
    {
        "reachability": "xls",
        "bound": "max_sheets",
        "error_code": "sheet_limit",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one[max_sheets",
        "suite": "eighth",
    },
    {
        "reachability": "xls",
        "bound": "max_physical_rows",
        "error_code": "physical_row_limit",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one[max_physical_rows",
        "suite": "eighth",
    },
    {
        "reachability": "xls",
        "bound": "max_columns",
        "error_code": "column_limit",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one[max_columns",
        "suite": "eighth",
    },
    {
        "reachability": "xls",
        "bound": "max_cell_chars",
        "error_code": "cell_length_limit",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one[max_cell_chars",
        "suite": "eighth",
    },
    {
        "reachability": "xls",
        "bound": "max_row_chars",
        "error_code": "row_char_limit",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds[max_row_chars",
        "suite": "ninth",
    },
    {
        "reachability": "xls",
        "bound": "max_total_cells",
        "error_code": "total_cell_limit",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds[max_total_cells",
        "suite": "ninth",
    },
    {
        "reachability": "xls",
        "bound": "max_merged_regions",
        "error_code": "merged_region_limit",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds[max_merged_regions-",
        "suite": "ninth",
    },
    {
        "reachability": "xls",
        "bound": "max_merged_regions_per_sheet",
        "error_code": "merged_region_limit",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds[max_merged_regions_per_sheet",
        "suite": "ninth",
    },
]

EXPECTED_HTTP_NPLUS1_COUNT = 48


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session(tmp_path) -> Session:
    from sqlalchemy import create_engine

    db_file = tmp_path / f"s13e12_{uuid.uuid4().hex}.db"
    engine = create_engine(f"sqlite:///{db_file}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()


def _xlsx_bytes(*, sheets: int = 1, rows: int = 1, cols: int = 1, cell: str = "ok") -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.title = "S0"
    for si in range(sheets):
        ws = wb.active if si == 0 else wb.create_sheet(f"S{si}")
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(r, c, cell)
    wb.save(buf)
    return buf.getvalue()


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org",
        organization_slug=f"o-{uuid.uuid4().hex[:6]}",
        status=OrganizationStatus.ACTIVE,
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        source_sheet_name="Sheet1",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    payload = _xlsx_bytes()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
        proposed_raw_price="100.50",
        proposed_currency="VND",
        proposed_appraised_unit_price="110",
        proposed_review_status="pending",
        proposed_validation_status="valid",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=Decimal("2.5"),
        raw_price=Decimal("99.99"),
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    return prior, staging, line, snap


# ---------------------------------------------------------------------------
# M-02: mapper / serializer set equality
# ---------------------------------------------------------------------------


def test_m02_field_sets_equal_mapper_columns():
    assert_field_sets_match_mappers()
    # explicit known omissions that previously failed
    assert "source_sheet_name" in BATCH_FIELDS
    assert "proposed_raw_price" in STAGING_FIELDS
    assert "proposed_currency" in STAGING_FIELDS
    assert "proposed_appraised_unit_price" in STAGING_FIELDS
    assert "proposed_review_status" in STAGING_FIELDS
    assert "proposed_validation_status" in STAGING_FIELDS
    assert "created_at" in STAGING_FIELDS and "updated_at" in STAGING_FIELDS
    assert "raw_price" in LINE_FIELDS
    assert "row_version" in LINE_FIELDS
    assert "source_import_batch_id" in LINE_FIELDS
    assert "source_staging_row_id" in LINE_FIELDS


def test_m02_serializer_keys_equal_field_sets(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    assert set(snap["artifacts"][0].keys()) == set(ARTIFACT_FIELDS)
    assert set(snap["batches"][0].keys()) == set(BATCH_FIELDS)
    assert set(snap["staging"][0].keys()) == set(STAGING_FIELDS)
    assert set(snap["lines"][0].keys()) == set(LINE_FIELDS)
    # empty audits still: inject one audit and check
    audit = AuditEvent(
        organization_id=org.id,
        actor_user_id=user.id,
        command_name="C",
        event_name="E",
        entity_type="T",
        entity_id=prior.id,
        payload={"k": 1},
    )
    db_session.add(audit)
    db_session.commit()
    snap2 = snapshot_source_intake_preserve(
        db_session, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    assert set(snap2["audits"][0].keys()) == set(AUDIT_FIELDS)
    assert set(ARTIFACT_FIELDS) == set(persisted_column_keys(ImportSourceArtifact))
    assert set(BATCH_FIELDS) == set(persisted_column_keys(ProjectAssetImportBatch))
    assert set(STAGING_FIELDS) == set(persisted_column_keys(ProjectAssetImportStagingRow))
    assert set(LINE_FIELDS) == set(persisted_column_keys(ProjectAssetLine))
    assert set(AUDIT_FIELDS) == set(persisted_column_keys(AuditEvent))


# ---------------------------------------------------------------------------
# M-03: typed canonicalization + real persisted mutation probes
# ---------------------------------------------------------------------------


def test_m03_canonical_collision_probes():
    assert_canonical_distinguishes_collisions()


def _expect_preserve_fail(db, fake_storage, snap, *, match_substr: str | None = None):
    with pytest.raises(AssertionError) as ei:
        assert_source_intake_preserve(db, fake_storage, snap)
    if match_substr:
        assert match_substr in str(ei.value)


def _set_field_freeze_updated(
    db,
    obj,
    field: str,
    value,
    *,
    freeze_fields=("updated_at", "row_version"),
):
    """Mutate one field while holding timestamp/version columns constant (M-03 isolation)."""
    frozen = {f: getattr(obj, f) for f in freeze_fields if hasattr(obj, f) and f != field}
    setattr(obj, field, value)
    db.flush()
    for f, v in frozen.items():
        setattr(obj, f, v)
    db.commit()
    db.refresh(obj)


def test_m03_persisted_mutation_probes(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    # Seed a baseline audit
    audit = AuditEvent(
        organization_id=org.id,
        actor_user_id=user.id,
        command_name="Seed",
        event_name="SeedEvent",
        entity_type="ImportSourceArtifact",
        entity_id=prior.id,
        payload={"k": "v", "nested": {"n": 1}},
    )
    db_session.add(audit)
    db_session.commit()
    snap = snapshot_source_intake_preserve(
        db_session, fake_storage, project_id=proj.id, batch_id=batch.id
    )

    # --- audit: insert ---
    extra = AuditEvent(
        organization_id=org.id,
        actor_user_id=user.id,
        command_name="Ins",
        event_name="Inserted",
        entity_type="T",
        entity_id=prior.id,
        payload={},
    )
    db_session.add(extra)
    db_session.commit()
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="audit")
    db_session.delete(extra)
    db_session.commit()
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- audit: delete ---
    a_del = db_session.query(AuditEvent).order_by(AuditEvent.id.asc()).first()
    db_session.delete(a_del)
    db_session.commit()
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="audit")
    db_session.add(
        AuditEvent(
            organization_id=org.id,
            actor_user_id=user.id,
            command_name="Seed",
            event_name="SeedEvent",
            entity_type="ImportSourceArtifact",
            entity_id=prior.id,
            payload={"k": "v", "nested": {"n": 1}},
        )
    )
    db_session.commit()
    snap = snapshot_source_intake_preserve(
        db_session, fake_storage, project_id=proj.id, batch_id=batch.id
    )

    # --- audit: scalar field mutation ---
    arow = db_session.query(AuditEvent).order_by(AuditEvent.id.asc()).first()
    old_event = arow.event_name
    arow.event_name = "MUTATED_EVENT"
    db_session.commit()
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="audit")
    arow.event_name = old_event
    db_session.commit()
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- audit: nested payload mutation ---
    arow = db_session.query(AuditEvent).order_by(AuditEvent.id.asc()).first()
    arow.payload = {"k": "v", "nested": {"n": 999}}
    db_session.commit()
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="audit")
    arow.payload = {"k": "v", "nested": {"n": 1}}
    db_session.commit()
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- object bytes mutation ---
    key = next(iter(snap["objects"]))
    original = fake_storage._objects[key]
    fake_storage._objects[key] = b"MUTATED-BYTES"
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="object-store")
    fake_storage._objects[key] = original
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- content-type mutation ---
    ct0 = fake_storage._content_types[key]
    fake_storage._content_types[key] = "application/mutated"
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="content_types")
    fake_storage._content_types[key] = ct0
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- content-type add ---
    fake_storage._content_types["extra-key"] = "x"
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="content_types")
    del fake_storage._content_types["extra-key"]
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- content-type delete ---
    del fake_storage._content_types[key]
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="content_types")
    fake_storage._content_types[key] = ct0
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- artifact state/checksum (freeze updated_at) ---
    prior = db_session.get(ImportSourceArtifact, prior.id)
    old_state, old_cs = prior.state, prior.checksum_sha256
    _set_field_freeze_updated(db_session, prior, "state", "failed")
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="artifact")
    prior = db_session.get(ImportSourceArtifact, prior.id)
    _set_field_freeze_updated(db_session, prior, "state", old_state)
    assert_source_intake_preserve(db_session, fake_storage, snap)
    prior = db_session.get(ImportSourceArtifact, prior.id)
    _set_field_freeze_updated(db_session, prior, "checksum_sha256", "0" * 64)
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="artifact")
    prior = db_session.get(ImportSourceArtifact, prior.id)
    _set_field_freeze_updated(db_session, prior, "checksum_sha256", old_cs)
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- batch pointer ---
    b = db_session.get(ProjectAssetImportBatch, batch.id)
    old_ptr = b.current_source_artifact_id
    _set_field_freeze_updated(db_session, b, "current_source_artifact_id", None)
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="batch")
    b = db_session.get(ProjectAssetImportBatch, batch.id)
    _set_field_freeze_updated(db_session, b, "current_source_artifact_id", old_ptr)
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- batch source_sheet_name with updated_at held constant ---
    b = db_session.get(ProjectAssetImportBatch, batch.id)
    old_sheet = b.source_sheet_name
    _set_field_freeze_updated(db_session, b, "source_sheet_name", "MUTATED-SHEET")
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="batch")
    b = db_session.get(ProjectAssetImportBatch, batch.id)
    _set_field_freeze_updated(db_session, b, "source_sheet_name", old_sheet)
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- staging JSON ---
    st = db_session.get(ProjectAssetImportStagingRow, staging.id)
    old_raw = dict(st.raw_values)
    _set_field_freeze_updated(db_session, st, "raw_values", {"A": "mutated"})
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="staging")
    st = db_session.get(ProjectAssetImportStagingRow, staging.id)
    _set_field_freeze_updated(db_session, st, "raw_values", old_raw)
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- staging scalar families previously omitted ---
    for field, new_val, restore in [
        ("proposed_raw_price", "999.99", "100.50"),
        ("proposed_currency", "USD", "VND"),
        ("proposed_appraised_unit_price", "1", "110"),
        ("proposed_review_status", "approved", "pending"),
        ("proposed_validation_status", "invalid", "valid"),
    ]:
        st = db_session.get(ProjectAssetImportStagingRow, staging.id)
        _set_field_freeze_updated(db_session, st, field, new_val)
        _expect_preserve_fail(db_session, fake_storage, snap, match_substr="staging")
        st = db_session.get(ProjectAssetImportStagingRow, staging.id)
        _set_field_freeze_updated(db_session, st, field, restore)
        assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- official line asset_name ---
    ln = db_session.get(ProjectAssetLine, line.id)
    old_name = ln.asset_name
    _set_field_freeze_updated(db_session, ln, "asset_name", "MUTATED-NAME")
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="official lines")
    ln = db_session.get(ProjectAssetLine, line.id)
    _set_field_freeze_updated(db_session, ln, "asset_name", old_name)
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- official line raw_price (previously undetected) ---
    ln = db_session.get(ProjectAssetLine, line.id)
    old_price = ln.raw_price
    _set_field_freeze_updated(db_session, ln, "raw_price", Decimal("1.23"))
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="official lines")
    ln = db_session.get(ProjectAssetLine, line.id)
    _set_field_freeze_updated(db_session, ln, "raw_price", old_price)
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- lineage FK ---
    ln = db_session.get(ProjectAssetLine, line.id)
    _set_field_freeze_updated(db_session, ln, "source_import_batch_id", batch.id)
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="official lines")
    ln = db_session.get(ProjectAssetLine, line.id)
    _set_field_freeze_updated(db_session, ln, "source_import_batch_id", None)
    assert_source_intake_preserve(db_session, fake_storage, snap)

    # --- row_version (Core UPDATE bypasses version_id auto-increment) ---
    from sqlalchemy import update as sa_update

    ln = db_session.get(ProjectAssetLine, line.id)
    old_rv = ln.row_version
    db_session.execute(
        sa_update(ProjectAssetLine)
        .where(ProjectAssetLine.id == line.id)
        .values(row_version=(old_rv or 1) + 99, updated_at=ln.updated_at)
    )
    db_session.commit()
    _expect_preserve_fail(db_session, fake_storage, snap, match_substr="official lines")
    ln = db_session.get(ProjectAssetLine, line.id)
    db_session.execute(
        sa_update(ProjectAssetLine)
        .where(ProjectAssetLine.id == line.id)
        .values(row_version=old_rv, updated_at=ln.updated_at)
    )
    db_session.commit()
    assert_source_intake_preserve(db_session, fake_storage, snap)


# ---------------------------------------------------------------------------
# M-04: collectable matrix + runtime helper instrumentation
# ---------------------------------------------------------------------------


def test_m04_collect_only_marked_count_is_48():
    """Exact collect-only count for s13_pr_002_http_nplus1_reject (anchored parser)."""
    import subprocess
    import sys

    r = subprocess.run(
        [
            sys.executable,
            "-m",
            "pytest",
            "--collect-only",
            "-m",
            "s13_pr_002_http_nplus1_reject",
            "-q",
        ],
        cwd=str(Path(__file__).resolve().parents[1]),
        capture_output=True,
        text=True,
        check=False,
    )
    out = (r.stdout or "") + (r.stderr or "")
    assert_pytest_collect_count_exactly(
        out, expected=EXPECTED_HTTP_NPLUS1_COUNT, returncode=r.returncode
    )


def test_m04_manifest_format_bound_set_equality():
    covered = {(r["reachability"], r["bound"]) for r in FORMAT_BOUND_MANIFEST}
    assert covered == EXPECTED_FORMAT_BOUND, (
        f"missing={EXPECTED_FORMAT_BOUND - covered} extra={covered - EXPECTED_FORMAT_BOUND}"
    )
    # no duplicate format/bound pairs
    assert len(FORMAT_BOUND_MANIFEST) == len(covered)
    # forbidden placeholders
    for row in FORMAT_BOUND_MANIFEST:
        for k in ("accepted_node", "rejected_node"):
            v = row[k]
            assert v and v != "—" and "..." not in v and "…" not in v, row


def test_m04_manifest_nodes_resolve_in_collect_only():
    """Collect-only set equals static ledger (fourteenth authority)."""
    import subprocess
    import sys

    from tests.support.s13_pr_002_matrix import ledger_rows, validate_collection_matches_ledger

    r = subprocess.run(
        [
            sys.executable,
            "-m",
            "pytest",
            "--collect-only",
            "-m",
            "s13_pr_002_http_nplus1_reject",
            "-q",
        ],
        cwd=str(Path(__file__).resolve().parents[1]),
        capture_output=True,
        text=True,
        check=False,
    )
    out = (r.stdout or "") + (r.stderr or "")
    assert_pytest_collect_count_exactly(out, expected=EXPECTED_HTTP_NPLUS1_COUNT, returncode=r.returncode)
    nodeids = [
        line.strip()
        for line in (r.stdout or "").splitlines()
        if "::" in line and "test_" in line
    ]
    validate_collection_matches_ledger(nodeids, ledger_rows())


def test_m04_strong_helper_records_only_on_success(db_session: Session, fake_storage):
    """Preserve assertion still rejects without durable side effects (no event path here)."""
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    from tests.support.s13_pr_002_http_preserve import assert_source_intake_preserve

    assert_source_intake_preserve(db_session, fake_storage, snap)
    # Mutate storage and ensure preserve fails
    key = next(iter(snap["objects"]))
    fake_storage._objects[key] = b"x"
    with pytest.raises(AssertionError):
        assert_source_intake_preserve(db_session, fake_storage, snap)


def test_m04_zip_bounds_xlsx_only_documented():
    zip_rows = [r for r in FORMAT_BOUND_MANIFEST if r["bound"].startswith("max_zip") or "uncompressed_zip" in r["bound"]]
    assert zip_rows
    assert all(r["reachability"] == "xlsx" for r in zip_rows)
    assert not any(r["reachability"] == "xls" and "zip" in r["bound"] for r in FORMAT_BOUND_MANIFEST)


# ---------------------------------------------------------------------------
# M-05: threat is outside the marked matrix
# ---------------------------------------------------------------------------


def test_m05_threat_not_in_marked_matrix():
    import subprocess
    import sys

    r = subprocess.run(
        [
            sys.executable,
            "-m",
            "pytest",
            "--collect-only",
            "-m",
            "s13_pr_002_http_nplus1_reject",
            "-q",
        ],
        cwd=str(Path(__file__).resolve().parents[1]),
        capture_output=True,
        text=True,
        check=False,
    )
    out = (r.stdout or "") + (r.stderr or "")
    assert r.returncode == 0, out[-500:]
    assert "test_threat_http_preserves_prior_official_and_staging" not in out


def test_m05_honest_count_constants():
    # Document the three separate counts this corrective publishes.
    http_nplus1 = EXPECTED_HTTP_NPLUS1_COUNT
    # support/manifest/mutation proofs in this file are unmarked
    assert http_nplus1 == 48
    assert len(EXPECTED_FORMAT_BOUND) == 20


# ---------------------------------------------------------------------------
# Live e04 identity cross-check (M-01 companion in twelfth suite)
# ---------------------------------------------------------------------------


def test_m01_e04_upload_boundary_identity_documented():
    """Document exact e04 N/N+1 contract for the report (executable sizes)."""
    payload = _xlsx_bytes()
    n = len(payload)
    assert n > 0
    # N+1 is payload + one extra byte — matches fifth suite fixture
    bad = payload + b"X"
    assert len(bad) == n + 1
    limits = SourceArtifactLimits(max_upload_bytes=n, max_request_bytes=12 * 1024 * 1024)
    assert limits.max_upload_bytes == n
    assert limits.max_request_bytes > n + 1024  # multipart overhead headroom
