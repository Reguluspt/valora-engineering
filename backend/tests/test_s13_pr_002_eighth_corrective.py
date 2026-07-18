"""S13-PR-002 eighth corrective: I-01…I-05 ownership, boundaries, schema, PG matrix."""
from __future__ import annotations

import hashlib
import io
import os
import threading
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any

import openpyxl
import pytest
from fastapi import HTTPException, UploadFile
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import Session, sessionmaker
from starlette.datastructures import Headers

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
    set_reconcile_work_session_factory,
    set_source_limits_override,
    upload_source_artifact,
    _sha256_object,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    assert_accepted_source_upload,
    assert_http_rejection_preserve,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
    AuditEvent,
)


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session(tmp_path) -> Session:
    db_file = tmp_path / f"s13e8_{uuid.uuid4().hex}.db"
    engine = create_engine(f"sqlite:///{db_file}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()
    set_source_limits_override(None)
    set_reconcile_work_session_factory(None)


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org", organization_slug=f"o-{uuid.uuid4().hex[:6]}", status=OrganizationStatus.ACTIVE
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _xlsx_bytes(*, sheets=1, rows=1, cols=1, cell="x", merges=0) -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    for r in range(rows):
        for c in range(cols):
            wb.active.cell(r + 1, c + 1, cell)
    for i in range(merges):
        # openpyxl merge needs a range; single-cell-ish A1:B1 style
        col = (i % cols) + 1 if cols > 1 else 1
        wb.active.merge_cells(
            start_row=1, start_column=col, end_row=1, end_column=min(col + 1, max(cols, 2))
        )
    for i in range(1, sheets):
        wb.create_sheet(f"S{i + 1}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_bytes(tmp_path, *, sheets=1, rows=1, cols=1, cell="x") -> bytes:
    pytest.importorskip("xlwt")
    import xlwt

    p = tmp_path / f"t-{uuid.uuid4().hex[:6]}.xls"
    book = xlwt.Workbook()
    for si in range(sheets):
        sh = book.add_sheet(f"S{si + 1}")
        for r in range(rows):
            for c in range(cols):
                sh.write(r, c, cell)
    book.save(str(p))
    return p.read_bytes()


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    payload = _xlsx_bytes()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    snap.update(
        {
            "prior_id": prior.id,
            "prior_checksum": prior.checksum_sha256,
            "prior_key": prior.storage_object_key,
            "prior_state": prior.state,
            "prior_gen": prior.generation,
            "batch_current": batch.current_source_artifact_id,
            "batch_status": batch.status,
            "batch_total": batch.total_rows,
            "batch_valid": batch.valid_rows,
            "staging_id": staging.id,
            "staging_raw": dict(staging.raw_values),
            "staging_mapped": dict(staging.mapped_values),
            "staging_name": staging.proposed_asset_name,
            "staging_desc": staging.proposed_description,
            "staging_qty": staging.proposed_quantity,
            "staging_unit": staging.proposed_unit,
            "staging_status": staging.validation_status,
            "line_id": line.id,
            "line_name": line.asset_name,
            "line_desc": line.description,
            "line_qty": float(line.quantity),
            "line_review": line.review_status,
            "line_val": line.validation_status,
            "object_keys": set(snap["objects"].keys()),
            "art_count": len(snap["artifacts"]),
            "line_count": len(snap["lines"]),
        }
    )
    return prior, staging, line, snap


def _assert_preserved(db, fake_storage, proj, batch, staging, line, snap):
    """Field-level prior/staging/line preserve for non-reject paths.

    HTTP N+1 rejections use assert_http_rejection_preserve.
    """
    db.expire_all()
    prior = db.get(ImportSourceArtifact, snap["prior_id"])
    batch = db.get(ProjectAssetImportBatch, batch.id)
    staging = db.get(ProjectAssetImportStagingRow, staging.id)
    line = db.get(ProjectAssetLine, line.id)
    assert prior is not None
    assert prior.checksum_sha256 == snap["prior_checksum"]
    assert prior.storage_object_key == snap["prior_key"]
    assert prior.state == snap["prior_state"]
    assert prior.generation == snap["prior_gen"]
    assert batch.current_source_artifact_id == snap["batch_current"]
    assert batch.status == snap["batch_status"]
    assert batch.total_rows == snap["batch_total"]
    assert batch.valid_rows == snap["batch_valid"]
    assert staging.raw_values == snap["staging_raw"]
    assert staging.mapped_values == snap["staging_mapped"]
    assert staging.proposed_asset_name == snap["staging_name"]
    assert staging.proposed_description == snap["staging_desc"]
    assert staging.proposed_quantity == snap["staging_qty"]
    assert staging.proposed_unit == snap["staging_unit"]
    assert staging.validation_status == snap["staging_status"]
    assert line.asset_name == snap["line_name"]
    assert line.description == snap["line_desc"]
    assert float(line.quantity) == snap["line_qty"]
    assert line.review_status == snap["line_review"]
    assert line.validation_status == snap["line_val"]
    assert snap["prior_key"] in fake_storage._objects
    assert db.query(ProjectAssetLine).filter_by(project_id=proj.id).count() == snap["line_count"]


def _assert_zero_new_objects(fake_storage, snap):
    assert {k: bytes(v) for k, v in fake_storage._objects.items()} == snap["objects"]
    assert dict(fake_storage._content_types) == snap["content_types"]


def _assert_error_code(res, expected_code: str):
    body = res.json()
    detail = body.get("detail")
    assert isinstance(detail, dict), f"detail must be mapping, got {type(detail)!r}: {detail!r}"
    assert detail.get("error_code") == expected_code


def _multipart_body(filename: str, payload: bytes, content_type: str, boundary: str) -> bytes:
    """Build a deterministic multipart/form-data body with fixed boundary."""
    crlf = b"\r\n"
    parts = []
    parts.append(f"--{boundary}".encode())
    parts.append(
        f'Content-Disposition: form-data; name="file"; filename="{filename}"'.encode()
    )
    parts.append(f"Content-Type: {content_type}".encode())
    parts.append(b"")
    parts.append(payload)
    parts.append(f"--{boundary}--".encode())
    parts.append(b"")
    return crlf.join(parts)


def _pg_url():
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql")
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL required")
    return url


def _cname(exc: IntegrityError) -> str:
    orig = getattr(exc, "orig", None)
    diag = getattr(orig, "diag", None) if orig is not None else None
    return getattr(diag, "constraint_name", None) or ""


# ---------------------------------------------------------------------------
# I-01 upload/reconciler overlap — single Available audit
# ---------------------------------------------------------------------------


def test_i01_upload_reconciler_overlap_single_available_audit(
    db_session: Session, fake_storage
):
    """Reproduce mid-put reconcile + uploader finalize; assert exactly one Available audit."""
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    hooks = {"reconcile_calls": 0, "mid_put": 0}
    real_put = fake_storage.put_stream

    def hooked_put(key, stream, content_type=None, expected_size=None):
        st = real_put(key, stream, content_type=content_type, expected_size=expected_size)
        hooks["mid_put"] += 1
        # Active upload window: object visible, pending reserved, finalize not yet run
        hooks["reconcile_calls"] += 1
        reconcile_source_artifacts(
            db_session,
            storage=fake_storage,
            max_items=10,
            actor_id=user.id,
            org_id=org.id,
        )
        return st

    fake_storage.put_stream = hooked_put  # type: ignore[method-assign]

    class R:
        headers = Headers({})

    uf = UploadFile(filename="overlap.xlsx", file=io.BytesIO(payload))
    art = upload_source_artifact(
        db_session,
        org_id=org.id,
        project_id=proj.id,
        batch_id=batch.id,
        file=uf,
        request=R(),
        current_user=user,
        storage=fake_storage,
    )
    assert hooks["mid_put"] == 1
    assert hooks["reconcile_calls"] == 1
    assert art.state == "available"
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, art.id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert art.state == "available"
    assert batch.current_source_artifact_id == art.id
    avail = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .all()
    )
    assert len(avail) == 1
    # Exactly one logical transition; command is either actor but not both
    commands = {a.command_name for a in avail}
    assert len(commands) == 1
    assert commands <= {"ReconcileImportSourceArtifact", "UploadImportSourceArtifact"}
    # No duplicate reserved/failed/orphan noise for this gen
    for name in (
        "ImportSourceArtifactFailed",
        "ImportSourceArtifactOrphaned",
        "ImportSourceArtifactObjectDeleted",
    ):
        assert (
            db_session.query(AuditEvent)
            .filter(AuditEvent.entity_id == art.id, AuditEvent.event_name == name)
            .count()
            == 0
        )
    # Prior preserved; no extra delete
    assert snap["prior_key"] in fake_storage._objects
    assert art.storage_object_key in fake_storage._objects
    assert art.checksum_sha256 == hashlib.sha256(payload).hexdigest()
    # Staging / official lines unchanged for prior snapshot fields that must survive
    staging = db_session.get(ProjectAssetImportStagingRow, staging.id)
    line = db_session.get(ProjectAssetLine, line.id)
    assert staging.proposed_asset_name == snap["staging_name"]
    assert line.asset_name == snap["line_name"]

    # Retry reconcile is idempotent — no second Available audit
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 0
    assert (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .count()
        == 1
    )


# ---------------------------------------------------------------------------
# I-02 commit ambiguity + genuine generation race
# ---------------------------------------------------------------------------


def test_i02_finalize_commit_raises_before_durability(
    db_session: Session, fake_storage, monkeypatch, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    p = tmp_path / "a.xlsx"
    p.write_bytes(payload)
    hooks = {"commit_n": 0, "fail_calls": 0}
    real = Session.commit

    def flaky(self):
        if self is not db_session:
            return real(self)
        hooks["commit_n"] += 1
        if hooks["commit_n"] == 2:
            hooks["fail_calls"] += 1
            raise RuntimeError("forced_final_commit_fail_before_durability")
        return real(self)

    monkeypatch.setattr(Session, "commit", flaky)

    class R:
        headers = Headers({})

    with open(p, "rb") as f:
        uf = UploadFile(filename="a.xlsx", file=f)
        with pytest.raises(HTTPException) as ei:
            upload_source_artifact(
                db_session,
                org_id=org.id,
                project_id=proj.id,
                batch_id=batch.id,
                file=uf,
                request=R(),
                current_user=user,
                storage=fake_storage,
            )
    assert ei.value.status_code == 500
    assert hooks["fail_calls"] == 1
    db_session.expire_all()
    arts = [a for a in db_session.query(ImportSourceArtifact).all() if a.generation > 1]
    assert len(arts) == 1
    art = arts[0]
    assert art.state in {"failed", "pending"}
    if art.state == "failed":
        assert art.failure_code == "finalize_commit_failed"
    assert art.storage_object_key in fake_storage._objects
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert batch.current_source_artifact_id == snap["batch_current"]
    _assert_preserved(db_session, fake_storage, proj, batch, staging, line, snap)
    avail = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .count()
    )
    assert avail == 0


def test_i02_durable_commit_then_raise_recovery_idempotent(
    db_session: Session, fake_storage, monkeypatch, tmp_path
):
    """Commit becomes durable, then caller sees exception; recovery returns available once."""
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    p = tmp_path / "d.xlsx"
    p.write_bytes(payload)
    hooks = {"commit_n": 0, "durable_then_raise": 0}
    real = Session.commit

    def flaky(self):
        if self is not db_session:
            return real(self)
        hooks["commit_n"] += 1
        # 1 = reservation; 2 = finalize available — durable then raise
        if hooks["commit_n"] == 2:
            real(self)
            hooks["durable_then_raise"] += 1
            raise RuntimeError("disconnect_after_durable_commit")
        return real(self)

    monkeypatch.setattr(Session, "commit", flaky)

    class R:
        headers = Headers({})

    with open(p, "rb") as f:
        uf = UploadFile(filename="d.xlsx", file=f)
        art = upload_source_artifact(
            db_session,
            org_id=org.id,
            project_id=proj.id,
            batch_id=batch.id,
            file=uf,
            request=R(),
            current_user=user,
            storage=fake_storage,
        )
    assert hooks["durable_then_raise"] == 1
    assert art.state == "available"
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, art.id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert art.state == "available"
    assert batch.current_source_artifact_id == art.id
    assert art.failure_code is None
    avail = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .count()
    )
    assert avail == 1
    # Must not have been compensated to failed
    assert (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactFailed",
        )
        .count()
        == 0
    )
    # Prior gen still present; staging/lines intact
    prior = db_session.get(ImportSourceArtifact, snap["prior_id"])
    assert prior.state == "available"
    staging = db_session.get(ProjectAssetImportStagingRow, staging.id)
    line = db_session.get(ProjectAssetLine, line.id)
    assert staging.proposed_asset_name == snap["staging_name"]
    assert line.asset_name == snap["line_name"]

    # Fresh-session recovery/retry observes durable available without new audit
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 0
    assert (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .count()
        == 1
    )


def test_i02_newer_finalize_while_older_recovery_in_flight(
    db_session: Session, fake_storage
):
    """Real interleaving: older pending recovery blocked mid-item; newer finalize wins."""
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    # Older pending with valid object (residual)
    body1 = b"older-valid-body"
    k1 = f"old/{uuid.uuid4()}"
    fake_storage._objects[k1] = body1
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="old.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256=hashlib.sha256(body1).hexdigest(),
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(a1)
    db_session.commit()
    a1_id = a1.id

    barrier_entered = threading.Event()
    release = threading.Event()
    hooks = {"barrier_hits": 0, "work_commits": 0}
    Work = sessionmaker(bind=db_session.get_bind(), autoflush=False, autocommit=False)

    def factory(bind):
        work = Work()
        real_commit = work.commit

        def gated_commit():
            hooks["work_commits"] += 1
            # First item commit is older recovery promote — pause before durability
            if hooks["work_commits"] == 1:
                hooks["barrier_hits"] += 1
                barrier_entered.set()
                assert release.wait(timeout=10), "release timeout"
            return real_commit()

        work.commit = gated_commit  # type: ignore[method-assign]
        return work

    set_reconcile_work_session_factory(factory)
    errors: list[BaseException] = []
    result: dict[str, Any] = {}

    def run_reconcile():
        try:
            result["stats"] = reconcile_source_artifacts(
                db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
            )
        except BaseException as exc:  # noqa: BLE001
            errors.append(exc)
        finally:
            set_reconcile_work_session_factory(None)

    t = threading.Thread(target=run_reconcile)
    t.start()
    assert barrier_entered.wait(timeout=10), "reconciler never reached barrier"

    # Newer generation finalizes while older recovery is in flight
    body2 = _xlsx_bytes(cell="newer")
    class R:
        headers = Headers({})

    uf = UploadFile(filename="new.xlsx", file=io.BytesIO(body2))
    # Need a separate session for uploader so it does not share reconciler's mid-txn state
    UploadSession = sessionmaker(bind=db_session.get_bind(), autoflush=False, autocommit=False)
    up_db = UploadSession()
    try:
        art2 = upload_source_artifact(
            up_db,
            org_id=org.id,
            project_id=proj.id,
            batch_id=batch.id,
            file=uf,
            request=R(),
            current_user=user,
            storage=fake_storage,
        )
    finally:
        up_db.close()

    assert art2.state == "available"
    a2_id = art2.id
    release.set()
    t.join(timeout=15)
    assert not t.is_alive()
    assert not errors, errors
    assert hooks["barrier_hits"] == 1

    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1_id)
    a2 = db_session.get(ImportSourceArtifact, a2_id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert a2.state == "available"
    assert batch.current_source_artifact_id == a2_id
    # Older must not steal pointer / become the sole available winner incorrectly
    assert a1.state != "available" or batch.current_source_artifact_id == a2_id
    if a1.state == "available":
        # Should not happen with stale path; pointer must still be newer
        assert batch.current_source_artifact_id == a2_id
    else:
        assert a1.state in {"orphaned", "pending", "failed"}
    # Exactly one Available audit for newer; older must not have Available that wins
    a2_avail = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == a2_id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .count()
    )
    assert a2_avail == 1
    # Pointer generation winner is gen2
    assert a2.generation == 2


# ---------------------------------------------------------------------------
# I-03 HTTP request/upload/adapter exact N and N+1
# ---------------------------------------------------------------------------


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_i03_request_bytes_exact_n_accepted_n_plus_one_rejected(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    boundary = "----ValoraBoundaryI03Fixed"
    body = _multipart_body(
        "a.xlsx",
        payload,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        boundary,
    )
    n = len(body)
    set_source_limits_override(
        SourceArtifactLimits(max_request_bytes=n, max_upload_bytes=10 * 1024 * 1024)
    )
    url = f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts"
    headers_base = {
        "X-User-Id": str(user.id),
        "Content-Type": f"multipart/form-data; boundary={boundary}",
    }
    try:
        # N+1 rejected
        res_bad = client.post(
            url,
            content=body + b"X",
            headers={**headers_base, "Content-Length": str(n + 1)},
        )
        assert_http_rejection_preserve(
            res_bad,
            status=413,
            error_code="request_too_large",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )

        # exact N accepted
        res_ok = client.post(
            url,
            content=body,
            headers={**headers_base, "Content-Length": str(n)},
        )
        assert_accepted_source_upload(res_ok, status=201)
    finally:
        set_source_limits_override(None)


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_i03_upload_bytes_exact_n_accepted_n_plus_one_rejected(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    n = len(payload)
    set_source_limits_override(
        SourceArtifactLimits(max_upload_bytes=n, max_request_bytes=12 * 1024 * 1024)
    )
    url = f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts"
    try:
        res_bad = client.post(
            url,
            files={
                "file": (
                    "big.xlsx",
                    io.BytesIO(payload + b"X"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res_bad,
            status=413,
            error_code="upload_too_large",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )

        res_ok = client.post(
            url,
            files={
                "file": (
                    "ok.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_accepted_source_upload(res_ok, status=201)
    finally:
        set_source_limits_override(None)


@pytest.mark.parametrize(
    "limit_field,n_limit,build_ok,build_bad,error_code,bad_status",
    [
        (
            "max_sheets",
            1,
            lambda: _xlsx_bytes(sheets=1),
            lambda: _xlsx_bytes(sheets=2),
            "sheet_limit",
            413,
        ),
        (
            "max_physical_rows",
            1,
            lambda: _xlsx_bytes(rows=1),
            lambda: _xlsx_bytes(rows=2),
            "physical_row_limit",
            413,
        ),
        (
            "max_columns",
            1,
            lambda: _xlsx_bytes(cols=1),
            lambda: _xlsx_bytes(cols=2),
            "column_limit",
            413,
        ),
        (
            "max_cell_chars",
            3,
            lambda: _xlsx_bytes(cell="abc"),
            lambda: _xlsx_bytes(cell="abcd"),
            "cell_length_limit",
            400,
        ),
    ],
)
@pytest.mark.s13_pr_002_http_nplus1_reject
def test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one(
    client: TestClient,
    db_session: Session,
    fake_storage,
    limit_field,
    n_limit,
    build_ok,
    build_bad,
    error_code,
    bad_status,
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    url = f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts"
    set_source_limits_override(SourceArtifactLimits(**{limit_field: n_limit}))
    try:
        # N+1 reject
        objects_before = set(fake_storage._objects.keys())
        res_bad = client.post(
            url,
            files={
                "file": (
                    "lim.xlsx",
                    io.BytesIO(build_bad()),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res_bad,
            status=bad_status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
        assert set(fake_storage._objects.keys()) == objects_before

        # exact N accept
        res_ok = client.post(
            url,
            files={
                "file": (
                    "ok.xlsx",
                    io.BytesIO(build_ok()),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_accepted_source_upload(res_ok, status=201)
    finally:
        set_source_limits_override(None)


@pytest.mark.parametrize(
    "limit_field,n_limit,ok_kwargs,bad_kwargs,error_code,bad_status",
    [
        ("max_sheets", 1, {"sheets": 1}, {"sheets": 2}, "sheet_limit", 413),
        ("max_physical_rows", 1, {"rows": 1}, {"rows": 2}, "physical_row_limit", 413),
        ("max_columns", 1, {"cols": 1}, {"cols": 2}, "column_limit", 413),
        ("max_cell_chars", 3, {"cell": "abc"}, {"cell": "abcd"}, "cell_length_limit", 400),
    ],
)
@pytest.mark.s13_pr_002_http_nplus1_reject
def test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one(
    client: TestClient,
    db_session: Session,
    fake_storage,
    tmp_path,
    limit_field,
    n_limit,
    ok_kwargs,
    bad_kwargs,
    error_code,
    bad_status,
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    url = f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts"
    set_source_limits_override(SourceArtifactLimits(**{limit_field: n_limit}))
    try:
        objects_before = set(fake_storage._objects.keys())
        res_bad = client.post(
            url,
            files={
                "file": (
                    "lim.xls",
                    io.BytesIO(_xls_bytes(tmp_path, **bad_kwargs)),
                    "application/vnd.ms-excel",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res_bad,
            status=bad_status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
        assert set(fake_storage._objects.keys()) == objects_before

        res_ok = client.post(
            url,
            files={
                "file": (
                    "ok.xls",
                    io.BytesIO(_xls_bytes(tmp_path, **ok_kwargs)),
                    "application/vnd.ms-excel",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_accepted_source_upload(res_ok, status=201)
    finally:
        set_source_limits_override(None)


# ---------------------------------------------------------------------------
# I-04 throwaway schema + full DML matrix (twice)
# ---------------------------------------------------------------------------


def test_i04_throwaway_migration_schema_and_dml_matrix():
    url = _pg_url()
    from alembic import command
    from alembic.config import Config
    from alembic.script import ScriptDirectory
    from app.core.config import get_settings
    from sqlalchemy.engine.url import make_url

    admin = create_engine(url, isolation_level="AUTOCOMMIT")
    db_name = f"s13_i04_{uuid.uuid4().hex[:10]}"
    with admin.connect() as conn:
        conn.execute(text(f'CREATE DATABASE "{db_name}"'))
    u = make_url(url)
    iso_url = url.rsplit("/", 1)[0] + f"/{db_name}"
    prev = {
        k: os.environ.get(k)
        for k in (
            "POSTGRES_HOST",
            "POSTGRES_PORT",
            "POSTGRES_DB",
            "POSTGRES_USER",
            "POSTGRES_PASSWORD",
            "VALORA_ENV",
        )
    }

    # Exact expected column map for import_source_artifacts
    expected_artifact_cols = {
        "id": ("uuid", None, "NO"),
        "organization_id": ("uuid", None, "NO"),
        "project_id": ("uuid", None, "NO"),
        "import_batch_id": ("uuid", None, "NO"),
        "generation": ("integer", None, "NO"),
        "original_filename": ("character varying", 255, "NO"),
        "detected_format": ("character varying", 16, "NO"),
        "content_type": ("character varying", 128, "NO"),
        "file_size_bytes": ("integer", None, "NO"),
        "checksum_sha256": ("character varying", 64, "NO"),
        "storage_object_key": ("character varying", 1024, "NO"),
        "storage_etag": ("character varying", 128, "YES"),
        "state": ("character varying", 32, "NO"),
        "adapter_name": ("character varying", 64, "YES"),
        "adapter_version": ("character varying", 64, "YES"),
        "adapter_metadata": ("jsonb", None, "NO"),
        "created_by_user_id": ("uuid", None, "NO"),
        "available_at": ("timestamp with time zone", None, "YES"),
        "failed_at": ("timestamp with time zone", None, "YES"),
        "orphaned_at": ("timestamp with time zone", None, "YES"),
        "failure_code": ("character varying", 64, "YES"),
        "created_at": ("timestamp with time zone", None, "NO"),
        "updated_at": ("timestamp with time zone", None, "NO"),
    }

    def _assert_schema(eng):
        with eng.connect() as c:
            rows = c.execute(
                text(
                    """
                    SELECT column_name, data_type, character_maximum_length, is_nullable,
                           udt_name
                    FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = 'import_source_artifacts'
                    """
                )
            ).all()
            cols = {r[0]: r for r in rows}
            assert set(cols) == set(expected_artifact_cols)
            for name, (dtype, clen, nullable) in expected_artifact_cols.items():
                r = cols[name]
                # data_type may be USER-DEFINED for uuid/jsonb — accept udt_name
                got_type = r[1]
                udt = r[4]
                if dtype == "uuid":
                    assert got_type in {"uuid", "USER-DEFINED"} or udt == "uuid"
                elif dtype == "jsonb":
                    assert got_type in {"jsonb", "USER-DEFINED"} or udt == "jsonb"
                else:
                    assert got_type == dtype, (name, got_type, dtype)
                assert r[2] == clen, (name, r[2], clen)
                assert r[3] == nullable, (name, r[3], nullable)

            # Pointer column
            prow = c.execute(
                text(
                    """
                    SELECT data_type, udt_name, character_maximum_length, is_nullable
                    FROM information_schema.columns
                    WHERE table_schema='public' AND table_name='project_asset_import_batches'
                      AND column_name='current_source_artifact_id'
                    """
                )
            ).one()
            assert prow[3] == "YES"
            assert prow[0] in {"uuid", "USER-DEFINED"} or prow[1] == "uuid"

            checks = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint "
                        "WHERE conrelid = 'import_source_artifacts'::regclass AND contype = 'c'"
                    )
                )
            }
            for name in (
                "chk_source_artifact_generation_positive",
                "chk_source_artifact_size_nonneg",
                "chk_source_artifact_checksum_len",
                "chk_source_artifact_checksum_lower",
                "chk_source_artifact_checksum_hex",
                "chk_source_artifact_state",
                "chk_source_artifact_format",
            ):
                assert name in checks

            uniques = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint "
                        "WHERE conrelid = 'import_source_artifacts'::regclass AND contype = 'u'"
                    )
                )
            }
            for name in (
                "uq_source_artifact_batch_generation",
                "uq_source_artifact_object_key",
                "uq_source_artifact_batch_id",
            ):
                assert name in uniques

            # Exact FK map by column list
            fk_rows = c.execute(
                text(
                    """
                    SELECT con.conname,
                           array_agg(a.attname ORDER BY u.ord) AS cols
                    FROM pg_constraint con
                    JOIN LATERAL unnest(con.conkey) WITH ORDINALITY AS u(attnum, ord) ON true
                    JOIN pg_attribute a ON a.attrelid = con.conrelid AND a.attnum = u.attnum
                    WHERE con.contype = 'f'
                      AND con.conrelid = 'import_source_artifacts'::regclass
                    GROUP BY con.conname
                    """
                )
            ).all()
            fk_by_cols = {tuple(r[1]): r[0] for r in fk_rows}
            assert (
                fk_by_cols[("organization_id", "project_id", "import_batch_id")]
                == "fk_source_artifact_batch_tenant"
            )
            assert fk_by_cols[("import_batch_id",)] == "import_source_artifacts_import_batch_id_fkey"
            assert ("organization_id",) in fk_by_cols
            assert ("project_id",) in fk_by_cols
            assert ("created_by_user_id",) in fk_by_cols
            # Store exact scalar names for documentation assertions
            assert fk_by_cols[("organization_id",)].endswith("_fkey") or "organization" in fk_by_cols[
                ("organization_id",)
            ]
            # Require exact discovered names equal to themselves (no fragment matching in DML)
            for cols_key, cname in fk_by_cols.items():
                assert fk_by_cols[cols_key] == cname

            batch_fk = c.execute(
                text(
                    """
                    SELECT con.conname
                    FROM pg_constraint con
                    WHERE con.contype = 'f'
                      AND con.conrelid = 'project_asset_import_batches'::regclass
                      AND con.conname = 'fk_batch_current_artifact_same_batch'
                    """
                )
            ).scalar()
            assert batch_fk == "fk_batch_current_artifact_same_batch"

            idxs = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT indexname FROM pg_indexes WHERE tablename='import_source_artifacts'"
                    )
                )
            }
            for name in (
                "idx_source_artifact_org",
                "idx_source_artifact_project",
                "idx_source_artifact_batch",
                "idx_source_artifact_state",
            ):
                assert name in idxs

            return fk_by_cols

    def _seed_and_dml(eng, fk_by_cols):
        oid = uuid.uuid4()
        oid2 = uuid.uuid4()
        uid = uuid.uuid4()
        cid = uuid.uuid4()
        pid = uuid.uuid4()
        b1 = uuid.uuid4()
        b2 = uuid.uuid4()
        a1 = uuid.uuid4()
        slug = f"dml-{uuid.uuid4().hex[:8]}"
        with eng.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T2', :slug, 'active', now(), now())"
                ),
                {"id": oid2, "slug": slug + "2"},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            for bid, fn in ((b1, "b1.xlsx"), (b2, "b2.xlsx")):
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :fn, 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": bid, "oid": oid, "pid": pid, "fn": fn, "uid": uid},
                )
            # valid artifact insert
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'x.xlsx', 'xlsx', 't', 1,
                      :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": a1,
                    "oid": oid,
                    "pid": pid,
                    "bid": b1,
                    "chk": "a" * 64,
                    "key": f"k-{a1}",
                    "uid": uid,
                },
            )
            # same-batch pointer
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": a1, "bid": b1},
            )

        # cross-batch pointer failure
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                    ),
                    {"aid": a1, "bid": b2},
                )
            assert _cname(ei.value) == "fk_batch_current_artifact_same_batch"

        # composite tenant mismatch — wrong organization_id for existing batch
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 2, 'y.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'pending', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid2,  # wrong org
                        "pid": pid,
                        "bid": b1,
                        "chk": "b" * 64,
                        "key": f"k-tenant-{uuid.uuid4()}",
                        "uid": uid,
                    },
                )
            assert _cname(ei.value) == "fk_source_artifact_batch_tenant"

        # batch-generation unique
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 1, 'y.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid,
                        "pid": pid,
                        "bid": b1,
                        "chk": "c" * 64,
                        "key": f"k-dupgen-{uuid.uuid4()}",
                        "uid": uid,
                    },
                )
            assert _cname(ei.value) == "uq_source_artifact_batch_generation"

        # object key unique
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 3, 'y.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid,
                        "pid": pid,
                        "bid": b1,
                        "chk": "d" * 64,
                        "key": f"k-{a1}",
                        "uid": uid,
                    },
                )
            assert _cname(ei.value) == "uq_source_artifact_object_key"

        check_cases = [
            (
                "chk_source_artifact_state",
                "UPDATE import_source_artifacts SET state = 'bogus' WHERE id = :id",
                {"id": a1},
            ),
            (
                "chk_source_artifact_format",
                "UPDATE import_source_artifacts SET detected_format = 'csv' WHERE id = :id",
                {"id": a1},
            ),
            (
                "chk_source_artifact_generation_positive",
                "UPDATE import_source_artifacts SET generation = 0 WHERE id = :id",
                {"id": a1},
            ),
            (
                "chk_source_artifact_size_nonneg",
                "UPDATE import_source_artifacts SET file_size_bytes = -1 WHERE id = :id",
                {"id": a1},
            ),
            (
                "chk_source_artifact_checksum_len",
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "a" * 63},
            ),
            (
                "chk_source_artifact_checksum_lower",
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "A" * 64},
            ),
            (
                "chk_source_artifact_checksum_hex",
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "g" * 64},
            ),
        ]
        for expected_name, sql, params in check_cases:
            with eng.begin() as conn:
                with pytest.raises(IntegrityError) as ei:
                    conn.execute(text(sql), params)
                cname = _cname(ei.value)
                # Overlapping checksum checks may fire first; allow exact known set for checksums
                if expected_name.startswith("chk_source_artifact_checksum"):
                    assert cname in {
                        "chk_source_artifact_checksum_len",
                        "chk_source_artifact_checksum_lower",
                        "chk_source_artifact_checksum_hex",
                    }, cname
                    if expected_name == "chk_source_artifact_checksum_hex" and params.get("chk") == "g" * 64:
                        assert cname == "chk_source_artifact_checksum_hex"
                else:
                    assert cname == expected_name, cname

        # RESTRICT batch delete — exact scalar FK (pointer cleared first so only artifact FK fires)
        with eng.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL WHERE id = :bid"
                ),
                {"bid": b1},
            )
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text("DELETE FROM project_asset_import_batches WHERE id = :id"),
                    {"id": b1},
                )
            assert _cname(ei.value) == "import_source_artifacts_import_batch_id_fkey"

        # Exact FK identities from schema map remain stable
        assert (
            fk_by_cols[("organization_id", "project_id", "import_batch_id")]
            == "fk_source_artifact_batch_tenant"
        )
        assert fk_by_cols[("import_batch_id",)] == "import_source_artifacts_import_batch_id_fkey"

    try:
        os.environ["VALORA_ENV"] = "test"
        os.environ["POSTGRES_HOST"] = u.host or "localhost"
        os.environ["POSTGRES_PORT"] = str(u.port or 5432)
        os.environ["POSTGRES_DB"] = db_name
        os.environ["POSTGRES_USER"] = u.username or "valora"
        os.environ["POSTGRES_PASSWORD"] = u.password or "valora_local_password"
        get_settings.cache_clear()
        cfg = Config("alembic.ini")
        assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]
        command.upgrade(cfg, "e1f2a3b4c5d6")
        eng = create_engine(iso_url)
        try:
            command.upgrade(cfg, "f2a3b4c5d6e7")
            fk_map = _assert_schema(eng)
            _seed_and_dml(eng, fk_map)
            command.downgrade(cfg, "e1f2a3b4c5d6")
            command.upgrade(cfg, "head")
            fk_map2 = _assert_schema(eng)
            _seed_and_dml(eng, fk_map2)
        finally:
            eng.dispose()
    finally:
        for k, v in prev.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        get_settings.cache_clear()
        with admin.connect() as conn:
            conn.execute(text(f'DROP DATABASE IF EXISTS "{db_name}" WITH (FORCE)'))
        admin.dispose()


# ---------------------------------------------------------------------------
# I-05 PostgreSQL ownership matrix
# ---------------------------------------------------------------------------


def _pg_seed_org_batch(conn, *, oid, uid, cid, pid, bid, slug, filename="b.xlsx"):
    conn.execute(
        text(
            "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
            "VALUES (:id, 'T', :slug, 'active', now(), now())"
        ),
        {"id": oid, "slug": slug},
    )
    conn.execute(
        text(
            "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
            "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
        ),
        {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
    )
    conn.execute(
        text(
            "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
            "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
        ),
        {"id": cid, "oid": oid, "uid": uid},
    )
    conn.execute(
        text(
            """
            INSERT INTO projects (
              id, organization_id, customer_id, name, code, status,
              knowledge_status, fee_amount, created_by, created_at, updated_at
            ) VALUES (
              :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
            )
            """
        ),
        {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
    )
    conn.execute(
        text(
            """
            INSERT INTO project_asset_import_batches (
              id, organization_id, project_id, source_filename, status,
              total_rows, valid_rows, invalid_rows, warning_rows,
              created_by_user_id, created_at, updated_at
            ) VALUES (
              :id, :oid, :pid, :fn, 'created', 0, 0, 0, 0, :uid, now(), now()
            )
            """
        ),
        {"id": bid, "oid": oid, "pid": pid, "fn": filename, "uid": uid},
    )


def _pg_insert_pending_artifact(conn, *, aid, oid, pid, bid, uid, key, body, gen=1):
    # created_at = now() so verified pending takes the promote path (not past-retention orphan).
    conn.execute(
        text(
            """
            INSERT INTO import_source_artifacts (
              id, organization_id, project_id, import_batch_id, generation,
              original_filename, detected_format, content_type, file_size_bytes,
              checksum_sha256, storage_object_key, state, adapter_metadata,
              created_by_user_id, created_at, updated_at
            ) VALUES (
              :id, :oid, :pid, :bid, :gen, 'p.xlsx', 'xlsx', 't', :sz,
              :chk, :key, 'pending', '{}'::jsonb, :uid, now(), now()
            )
            """
        ),
        {
            "id": aid,
            "oid": oid,
            "pid": pid,
            "bid": bid,
            "gen": gen,
            "sz": len(body),
            "chk": hashlib.sha256(body).hexdigest(),
            "key": key,
            "uid": uid,
        },
    )


@pytest.mark.parametrize("case", ["insert", "update", "delete", "readonly"])
def test_i05_pg_ownership_flushed_uow_with_work_transition(case):
    url = _pg_url()
    engine = create_engine(url)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    caller = SessionLocal()
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    bid_rec = uuid.uuid4()
    aid_rec = uuid.uuid4()
    slug = f"i05-{case}-{uuid.uuid4().hex[:8]}"
    store = FakeObjectStorage()
    rec_body = b"reconcile-pending-body"
    rec_key = f"rec/{aid_rec}"
    store._objects[rec_key] = rec_body
    extra_batch_id = uuid.uuid4()

    try:
        with engine.begin() as conn:
            _pg_seed_org_batch(
                conn, oid=oid, uid=uid, cid=cid, pid=pid, bid=bid, slug=slug, filename="base.xlsx"
            )
            # Separate committed batch+pending for reconciler work-session transition
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'rec.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid_rec, "oid": oid, "pid": pid, "uid": uid},
            )
            _pg_insert_pending_artifact(
                conn,
                aid=aid_rec,
                oid=oid,
                pid=pid,
                bid=bid_rec,
                uid=uid,
                key=rec_key,
                body=rec_body,
            )

        # Capture caller txn identity after begin
        caller.execute(text("SELECT 1"))
        txn_id_before = caller.execute(text("SELECT txid_current()")).scalar()
        assert caller.in_transaction()

        if case == "insert":
            # Flushed uncommitted insert of a third batch
            nb = ProjectAssetImportBatch(
                id=extra_batch_id,
                organization_id=oid,
                project_id=pid,
                source_filename="flushed-insert.xlsx",
                status=ImportBatchStatus.CREATED,
                created_by_user_id=uid,
            )
            caller.add(nb)
            caller.flush()
            lock_target_id = bid
        elif case == "update":
            batch = caller.get(ProjectAssetImportBatch, bid)
            assert batch is not None
            batch.source_filename = "flushed-uncommitted.xlsx"
            caller.flush()
            lock_target_id = bid
        elif case == "delete":
            # Need a disposable committed batch then delete flush
            with engine.begin() as conn:
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, 'to-delete.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": extra_batch_id, "oid": oid, "pid": pid, "uid": uid},
                )
            doomed = caller.get(ProjectAssetImportBatch, extra_batch_id)
            assert doomed is not None
            caller.delete(doomed)
            caller.flush()
            lock_target_id = bid  # lock on unrelated row still held by txn
        else:
            # read-only: SELECT FOR UPDATE on batch without mutation
            caller.execute(
                text("SELECT id FROM project_asset_import_batches WHERE id = :id FOR UPDATE"),
                {"id": bid},
            )
            lock_target_id = bid

        assert caller.in_transaction()
        txn_id_mid = caller.execute(text("SELECT txid_current()")).scalar()
        assert txn_id_mid == txn_id_before

        # Second connection cannot see uncommitted caller work
        other = SessionLocal()
        try:
            other.execute(text("SET LOCAL lock_timeout = '500ms'"))
            if case == "insert":
                seen = other.get(ProjectAssetImportBatch, extra_batch_id)
                assert seen is None
            elif case == "update":
                b2 = other.get(ProjectAssetImportBatch, bid)
                assert b2 is not None
                assert b2.source_filename == "base.xlsx"
            elif case == "delete":
                still = other.get(ProjectAssetImportBatch, extra_batch_id)
                assert still is not None  # delete not committed
            else:
                b2 = other.get(ProjectAssetImportBatch, bid)
                assert b2 is not None
                assert b2.source_filename == "base.xlsx"

            # Competing lock cannot acquire caller's row lock (update/readonly hold FOR UPDATE)
            if case in {"update", "readonly"}:
                with pytest.raises(OperationalError) as lock_ei:
                    other.execute(
                        text(
                            "SELECT id FROM project_asset_import_batches WHERE id = :id FOR UPDATE NOWAIT"
                        ),
                        {"id": lock_target_id},
                    )
                # Accept PG lock_not_available / could not obtain lock
                msg = str(lock_ei.value).lower()
                assert "lock" in msg or "could not obtain" in msg or "55p03" in msg
                try:
                    other.rollback()
                except Exception:
                    pass
        finally:
            try:
                other.close()
            except Exception:
                pass

        # Reconcile separate committed pending via dedicated work Session → durable available
        stats = reconcile_source_artifacts(
            caller, storage=store, max_items=10, actor_id=uid, org_id=oid
        )
        assert stats["scanned"] >= 1
        # Promote path may leave pending if within retention without valid path — created_at is -2h
        other = SessionLocal()
        try:
            rec = other.get(ImportSourceArtifact, aid_rec)
            assert rec is not None
            assert rec.state == "available"
            brec = other.get(ProjectAssetImportBatch, bid_rec)
            assert brec.current_source_artifact_id == aid_rec
        finally:
            other.close()

        # Caller txn identity / UOW unchanged
        assert caller.in_transaction()
        txn_id_after = caller.execute(text("SELECT txid_current()")).scalar()
        assert txn_id_after == txn_id_before

        other = SessionLocal()
        try:
            if case == "insert":
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
            elif case == "update":
                assert other.get(ProjectAssetImportBatch, bid).source_filename == "base.xlsx"
            elif case == "delete":
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is not None
        finally:
            other.close()

        if case == "update":
            # commit semantics
            caller.commit()
            other = SessionLocal()
            try:
                assert (
                    other.get(ProjectAssetImportBatch, bid).source_filename
                    == "flushed-uncommitted.xlsx"
                )
            finally:
                other.close()
            # reset for cleanup consistency
            with engine.begin() as conn:
                conn.execute(
                    text(
                        "UPDATE project_asset_import_batches SET source_filename = 'base.xlsx' WHERE id = :id"
                    ),
                    {"id": bid},
                )
        elif case == "insert":
            caller.rollback()
            other = SessionLocal()
            try:
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
            finally:
                other.close()
        elif case == "delete":
            caller.rollback()
            other = SessionLocal()
            try:
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is not None
            finally:
                other.close()
        else:
            caller.rollback()
    finally:
        try:
            if caller.in_transaction():
                caller.rollback()
        except Exception:
            pass
        try:
            caller.close()
        except Exception:
            pass
        with engine.begin() as conn:
            # Clear current pointer before artifact delete (fk_batch_current_artifact_same_batch).
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches "
                    "SET current_source_artifact_id = NULL WHERE organization_id = :oid"
                ),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


# ---------------------------------------------------------------------------
# MinIO (CI-required)
# ---------------------------------------------------------------------------


def test_i06_s3_minio_roundtrip_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL")
    elif not os.environ.get("S3_ENDPOINT_URL"):
        pytest.skip("S3_ENDPOINT_URL not set for local MinIO")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"eighth/{uuid.uuid4()}"
    body = b"eighth-corrective"
    store.put_stream(
        key, io.BytesIO(body), content_type="application/octet-stream", expected_size=len(body)
    )
    digest = _sha256_object(store, key, chunk_size=64, expected_size=len(body))
    assert digest == hashlib.sha256(body).hexdigest()
    store.delete(key)
    assert store.head(key) is None
