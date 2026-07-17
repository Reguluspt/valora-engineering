"""S13-PR-002 fourth corrective: F-01..F-09 / D-01..D-10 proofs."""
from __future__ import annotations

import hashlib
import io
import os
import uuid
from datetime import datetime, timedelta, timezone

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.adapters.xlsx_adapter import XlsxWorkbookAdapter
from app.modules.excel_import.application.adapters.xls_adapter import XlsWorkbookAdapter
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
    _sha256_object,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.domain.workbook_adapter import AdapterError
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    ObjectStorageError,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    assert_http_rejection_preserve,
    assert_source_intake_preserve,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
)
from tests.fixtures.s13_pr_002.ole_builder import write_threat_xls


@pytest.fixture
def db_session() -> Session:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org", organization_slug=f"o-{uuid.uuid4().hex[:6]}", status=OrganizationStatus.ACTIVE
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    """Prior available + staging + official line with field snapshot."""
    data = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.append(["h"])
    wb.save(data)
    payload = data.getvalue()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    snap.update(
        {
            "prior_id": prior.id,
            "prior_checksum": prior.checksum_sha256,
            "prior_key": prior.storage_object_key,
            "prior_state": prior.state,
            "prior_gen": prior.generation,
            "batch_current": batch.current_source_artifact_id,
            "batch_status": batch.status,
            "batch_total": batch.total_rows,
            "batch_valid": batch.valid_rows,
            "staging_id": staging.id,
            "staging_raw": dict(staging.raw_values),
            "staging_mapped": dict(staging.mapped_values),
            "staging_name": staging.proposed_asset_name,
            "staging_desc": staging.proposed_description,
            "staging_qty": staging.proposed_quantity,
            "staging_unit": staging.proposed_unit,
            "staging_status": staging.validation_status,
            "line_id": line.id,
            "line_name": line.asset_name,
            "line_desc": line.description,
            "line_qty": float(line.quantity),
            "line_review": line.review_status,
            "line_val": line.validation_status,
            "art_count": len(snap["artifacts"]),
            "line_count": len(snap["lines"]),
        }
    )
    assert snap["line_count"] >= 1
    return prior, staging, line, snap


def _assert_preserved(db, fake_storage, org, user, proj, batch, staging, line, snap):
    """HTTP reject paths (threat / cell limit) use full L-02/L-03 contract."""
    if "artifacts" in snap and "audits" in snap and "content_types" in snap:
        assert_source_intake_preserve(db, fake_storage, snap)
        return
    db.refresh(batch)
    db.refresh(staging)
    db.refresh(line)
    prior = db.get(ImportSourceArtifact, snap["prior_id"])
    assert prior is not None
    assert prior.checksum_sha256 == snap["prior_checksum"]
    assert prior.storage_object_key == snap["prior_key"]
    assert prior.state == snap["prior_state"]
    assert prior.generation == snap["prior_gen"]
    assert batch.current_source_artifact_id == snap["batch_current"]
    assert batch.status == snap["batch_status"]
    assert batch.total_rows == snap["batch_total"]
    assert batch.valid_rows == snap["batch_valid"]
    assert staging.id == snap["staging_id"]
    assert staging.raw_values == snap["staging_raw"]
    assert staging.mapped_values == snap["staging_mapped"]
    assert staging.proposed_asset_name == snap["staging_name"]
    assert staging.proposed_description == snap["staging_desc"]
    assert staging.proposed_quantity == snap["staging_qty"]
    assert staging.proposed_unit == snap["staging_unit"]
    assert staging.validation_status == snap["staging_status"]
    assert line.id == snap["line_id"]
    assert line.asset_name == snap["line_name"]
    assert line.description == snap["line_desc"]
    assert float(line.quantity) == snap["line_qty"]
    assert line.review_status == snap["line_review"]
    assert line.validation_status == snap["line_val"]
    assert fake_storage._objects == snap["objects"]
    assert db.query(ImportSourceArtifact).count() == snap["art_count"]
    assert db.query(ProjectAssetLine).filter_by(project_id=proj.id).count() == snap["line_count"]


# --- D-01 short-read ---


def test_short_read_not_checksum_mismatch(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    payload = b"0123456789abcdef"
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = payload
    fake_storage.truncate_open_to = 4  # clean EOF short-read
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    with pytest.raises(ObjectStorageError) as ei:
        _sha256_object(fake_storage, key, chunk_size=64, expected_size=len(payload))
    assert ei.value.code == "short_read"
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art)
    assert art.state == "pending"
    assert art.failure_code is None
    assert stats["errors"] >= 1
    assert stats["marked_failed"] == 0


def test_true_checksum_mismatch_after_full_stream(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    payload = b"full-body-content"
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = payload
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(payload),
        checksum_sha256="a" * 64,
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art)
    assert art.state == "failed"
    assert art.failure_code == "checksum_mismatch"
    assert stats["marked_failed"] >= 1


# --- D-02 transaction closure ---


def test_reconciler_empty_run_uses_dedicated_session(db_session: Session, fake_storage):
    """Reconciler owns a dedicated session; caller session is not rolled back."""
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    pending = ProjectAssetLine(
        project_id=proj.id,
        asset_name="CallerKeep",
        description="x",
        quantity=1,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db_session.add(pending)
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
    )
    assert stats["scanned"] == 0
    assert pending in db_session.new


def test_reconciler_skip_paths_preserve_caller(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = b"x"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="b" * 64,
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
    )
    db_session.add(art)
    db_session.commit()
    pending = ProjectAssetLine(
        project_id=proj.id,
        asset_name="SkipKeep",
        description="x",
        quantity=1,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db_session.add(pending)
    reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
    )
    assert pending in db_session.new


# --- D-03 / D-04 threat matrix with official lines ---


THREATS = [
    "filepass",
    "addin_supbook",
    "external_supbook",
    "dcon",
    "dconname",
    "dconref",
    "macro_boundsheet",
    "vba_boundsheet",
    "macro_name",
    "binary_name",
    "externname",
    "truncated_biff",
]


@pytest.mark.parametrize("threat", THREATS)
def test_threat_http_preserves_prior_official_and_staging(
    client: TestClient, db_session: Session, fake_storage, tmp_path, threat
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    path = tmp_path / f"{threat}.xls"
    write_threat_xls(path, threat)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": (f"{threat}.xls", io.BytesIO(path.read_bytes()), "application/vnd.ms-excel")},
        headers={"X-User-Id": str(user.id)},
    )
    assert res.status_code == 400, res.text
    _assert_preserved(db_session, fake_storage, org, user, proj, batch, staging, line, snap)


# --- D-05 cached formula values ---


def _xlsx_with_formula_cache(path, *, a1: int = 10, formula: str = "A1*2", cached: float = 20.0):
    """Write formula cell and inject known OOXML cached <v> so data_only reads it.

    openpyxl builds differ: empty <v></v>, missing <v>, or style attrs on <c>.
    """
    import re
    import zipfile

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = a1
    ws["B1"] = f"={formula}"
    wb.save(path)
    cached_s = f"{cached:g}"
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if "worksheets/" in item.filename and item.filename.endswith(".xml"):
                s = data.decode("utf-8")
                m = re.search(r'<c r="B1"[^>]*>.*?</c>', s, flags=re.DOTALL)
                if m is None:
                    m = re.search(r'<c[^>]*r="B1"[^>]*>.*?</c>', s, flags=re.DOTALL)
                assert m is not None, f"B1 cell missing in {item.filename}: {s[:500]!r}"
                old_cell = m.group(0)
                fm = re.search(r'<f[^>]*>.*?</f>', old_cell, flags=re.DOTALL)
                f_xml = fm.group(0) if fm else f"<f>{formula}</f>"
                new_cell = f'<c r="B1">{f_xml}<v>{cached_s}</v></c>'
                s2 = s[: m.start()] + new_cell + s[m.end() :]
                assert f"<v>{cached_s}</v>" in s2, f"inject failed; old cell={old_cell!r}"
                data = s2.encode("utf-8")
            zout.writestr(item, data)
    path.write_bytes(buf.getvalue())


def test_xlsx_cached_formula_value(tmp_path):
    """Exact cached result returned; formula text never returned; no execution."""
    p = tmp_path / "f.xlsx"
    _xlsx_with_formula_cache(p, a1=10, formula="A1*2", cached=20.0)
    rows = list(XlsxWorkbookAdapter().iter_rows(str(p)))
    vals = [c.value for r in rows for c in r]
    assert 20 in vals or 20.0 in vals
    assert all(not (isinstance(v, str) and str(v).startswith("=")) for v in vals)
    # Prove formula text is still present in non-data_only mode (cache, not execution)
    from openpyxl import load_workbook

    wb_f = load_workbook(p, data_only=False)
    assert wb_f.active["B1"].value == "=A1*2"
    wb_f.close()
    # Wrong cache would not equal a re-executed 10*2 if we had executed; we assert exact 20
    assert rows[0][1].value == 20


def test_xls_formula_value_not_formula_text(tmp_path):
    """xlrd 2.x does not execute formulas; formula string is not returned as cell value."""
    pytest.importorskip("xlwt")
    import xlwt

    book = xlwt.Workbook()
    sh = book.add_sheet("S")
    sh.write(0, 0, 2)
    sh.write(0, 1, xlwt.Formula("A1*3"))
    p = tmp_path / "f.xls"
    book.save(str(p))
    rows = list(XlsWorkbookAdapter().iter_rows(str(p)))
    v = rows[0][1].value
    assert not (isinstance(v, str) and v.startswith("="))
    assert v != "A1*3"
    assert v != "=A1*3"
    # xlrd2 typically surfaces formula cells as empty text — still non-executing
    assert v in (None, "", 0, 0.0) or isinstance(v, (int, float))


# --- D-01 extras: missing / oversize ---


def test_pending_object_missing_marks_failed(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    key = f"missing/{uuid.uuid4()}"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=3,
        checksum_sha256="a" * 64,
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art)
    assert art.state == "failed"
    assert art.failure_code == "pending_object_missing"
    assert stats["marked_failed"] >= 1


def test_object_too_large_is_infra_not_checksum(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    payload = b"0123456789"
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = payload
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=4,  # expected smaller than actual stream
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    with pytest.raises(ObjectStorageError) as ei:
        _sha256_object(fake_storage, key, chunk_size=64, expected_size=4)
    assert ei.value.code == "object_too_large"
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art)
    assert art.state == "pending"
    assert art.failure_code is None
    assert stats["marked_failed"] == 0
    assert stats["errors"] >= 1


# --- D-06 sample boundaries exact/max+1 ---


def test_xlsx_sheets_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    wb.create_sheet("S2")
    wb.save(p)  # 2 sheets
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_sheets=2)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_sheets=1)).inspect(str(p))
    assert ei.value.error_code == "sheet_limit"


def test_xlsx_rows_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "r.xlsx"
    wb = openpyxl.Workbook()
    for i in range(3):
        wb.active.append([i])
    wb.save(p)
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_physical_rows=3)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_physical_rows=2)).inspect(str(p))
    assert ei.value.error_code == "physical_row_limit"


def test_xlsx_columns_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "c.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append([1, 2, 3])
    wb.save(p)
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_columns=3)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_columns=2)).inspect(str(p))
    assert ei.value.error_code == "column_limit"


def test_xlsx_row_chars_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "rc.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["aa", "bb"])  # 4 chars
    wb.save(p)
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_row_chars=4)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_row_chars=3)).inspect(str(p))
    assert ei.value.error_code == "row_char_limit"


def test_xlsx_merge_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook()
    wb.active.merge_cells("A1:B1")
    wb.active["A1"] = "x"
    wb.save(p)
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_merged_regions=1)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_merged_regions=0)).inspect(str(p))
    assert ei.value.error_code == "merged_region_limit"


def test_xlsx_total_cells_exact_and_max_plus_one_iter(tmp_path):
    p = tmp_path / "tc.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append([1, 2])
    wb.active.append([3, 4])  # 4 cells
    wb.save(p)
    limits_ok = SourceArtifactLimits(max_total_cells=4)
    list(XlsxWorkbookAdapter(limits=limits_ok).iter_rows(str(p)))
    with pytest.raises(AdapterError) as ei:
        list(XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=3)).iter_rows(str(p)))
    assert ei.value.error_code == "total_cell_limit"


def test_xls_total_cells_exact_and_max_plus_one_iter(tmp_path):
    pytest.importorskip("xlwt")
    import xlwt

    p = tmp_path / "tc.xls"
    book = xlwt.Workbook()
    sh = book.add_sheet("S")
    sh.write(0, 0, 1)
    sh.write(0, 1, 2)
    sh.write(1, 0, 3)
    sh.write(1, 1, 4)
    book.save(str(p))
    list(XlsWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=4)).iter_rows(str(p)))
    with pytest.raises(AdapterError) as ei:
        list(XlsWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=3)).iter_rows(str(p)))
    assert ei.value.error_code == "total_cell_limit"


def test_xlsx_cell_chars_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "cc.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["abcd"])  # 4 chars
    wb.save(p)
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_cell_chars=4)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_cell_chars=3)).inspect(str(p))
    assert ei.value.error_code == "cell_length_limit"


def test_endpoint_cell_limit_no_reservation(
    client: TestClient, db_session: Session, fake_storage, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    # Cannot inject adapter limits via API — use default max 10000 with huge cell
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Z" * 10001
    buf = io.BytesIO()
    wb.save(buf)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "huge.xlsx",
                io.BytesIO(buf.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert_http_rejection_preserve(
        res,
        status=400,
        error_code="cell_length_limit",
        db=db_session,
        fake_storage=fake_storage,
        snap=snap,
    )


# --- D-07 retention boundary ---


def test_orphan_retention_boundary(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    now = datetime.now(timezone.utc)
    key_old = f"k-old/{uuid.uuid4()}"
    key_new = f"k-new/{uuid.uuid4()}"
    fake_storage._objects[key_old] = b"x"
    fake_storage._objects[key_new] = b"y"
    art_old = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="old.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="c" * 64,
        storage_object_key=key_old,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=now - timedelta(days=2),
        orphaned_at=now - timedelta(days=1),  # well past 3600s retention
    )
    art_new = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="new.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="d" * 64,
        storage_object_key=key_new,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=now,
        orphaned_at=now,  # not past retention
    )
    db_session.add_all([art_old, art_new])
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert key_old not in fake_storage._objects
    assert key_new in fake_storage._objects
    assert stats["deleted_objects"] == 1


def test_referenced_current_source_never_deleted(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    now = datetime.now(timezone.utc)
    key = f"k-cur/{uuid.uuid4()}"
    fake_storage._objects[key] = b"z"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="cur.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="e" * 64,
        storage_object_key=key,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=now - timedelta(days=3),
        orphaned_at=now - timedelta(days=2),
    )
    db_session.add(art)
    db_session.flush()
    batch.current_source_artifact_id = art.id
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert key in fake_storage._objects
    assert stats["deleted_objects"] == 0


def test_multi_item_later_error_keeps_earlier_failed(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    # item1: true checksum mismatch → durable failed
    key1 = f"k1/{uuid.uuid4()}"
    body1 = b"body-one"
    fake_storage._objects[key1] = body1
    art1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256="f" * 64,
        storage_object_key=key1,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=3),
    )
    # item2: short-read infrastructure error
    key2 = f"k2/{uuid.uuid4()}"
    body2 = b"body-two-xx"
    fake_storage._objects[key2] = body2
    fake_storage.truncate_open_to = 2
    art2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="b.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body2),
        checksum_sha256=hashlib.sha256(body2).hexdigest(),
        storage_object_key=key2,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add_all([art1, art2])
    db_session.commit()
    # Process art1 first without truncation, then enable truncation for art2
    fake_storage.truncate_open_to = None
    # Use a side-effect: after first open_stream of full body, set truncate
    original_open = fake_storage.open_stream
    calls = {"n": 0}

    def open_once(key):
        calls["n"] += 1
        if calls["n"] >= 2:
            fake_storage.truncate_open_to = 2
        return original_open(key)

    fake_storage.open_stream = open_once  # type: ignore[method-assign]
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art1)
    db_session.refresh(art2)
    assert art1.state == "failed"
    assert art1.failure_code == "checksum_mismatch"
    assert art2.state == "pending"
    assert stats["marked_failed"] >= 1
    assert stats["errors"] >= 1


# --- D-09 migration round-trip (PG) ---


def test_pg_migration_roundtrip_s13():
    """S13 revision round-trip on the CI PostgreSQL database.

    Uses the shared TEST_DATABASE_URL (already migrated to head by CI smoke).
    Avoids CREATE DATABASE (not required) and keeps Alembic on the same URL
    as POSTGRES_* / get_settings().database_url.

    Sequence: head present → downgrade parent → table gone → upgrade f2a3 →
    columns present → upgrade head → single head.
    """
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql")
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL required for migration round-trip")

    from alembic import command
    from alembic.config import Config
    from alembic.script import ScriptDirectory
    from app.core.config import get_settings
    from sqlalchemy.engine.url import make_url

    u = make_url(url)
    prev = {
        k: os.environ.get(k)
        for k in (
            "POSTGRES_HOST",
            "POSTGRES_PORT",
            "POSTGRES_DB",
            "POSTGRES_USER",
            "POSTGRES_PASSWORD",
            "VALORA_ENV",
        )
    }
    os.environ["VALORA_ENV"] = "test"
    os.environ["POSTGRES_HOST"] = u.host or "localhost"
    os.environ["POSTGRES_PORT"] = str(u.port or 5432)
    os.environ["POSTGRES_DB"] = u.database or "valora"
    os.environ["POSTGRES_USER"] = u.username or "valora"
    os.environ["POSTGRES_PASSWORD"] = u.password or "valora_local_password"
    get_settings.cache_clear()
    try:
        settings_url = get_settings().database_url
        # Driver may differ slightly; host/db must match the live test DB
        assert get_settings().postgres_db == (u.database or "valora")
        assert get_settings().postgres_host == (u.host or "localhost")

        cfg = Config("alembic.ini")
        script = ScriptDirectory.from_config(cfg)
        assert script.get_heads() == ["f2a3b4c5d6e7"]

        eng = create_engine(url)
        try:
            # Normalize to head first (CI smoke already did this)
            command.upgrade(cfg, "head")
            with eng.connect() as c:
                r = c.execute(
                    text("SELECT to_regclass('public.import_source_artifacts')")
                ).scalar()
                assert r is not None

            command.downgrade(cfg, "e1f2a3b4c5d6")
            with eng.connect() as c:
                r = c.execute(
                    text("SELECT to_regclass('public.import_source_artifacts')")
                ).scalar()
                assert r is None
                # Prior schema still valid
                r2 = c.execute(
                    text("SELECT to_regclass('public.project_asset_import_batches')")
                ).scalar()
                assert r2 is not None

            command.upgrade(cfg, "f2a3b4c5d6e7")
            with eng.connect() as c:
                r = c.execute(
                    text("SELECT to_regclass('public.import_source_artifacts')")
                ).scalar()
                assert r is not None
                cols = c.execute(
                    text(
                        "SELECT column_name FROM information_schema.columns "
                        "WHERE table_name='import_source_artifacts'"
                    )
                ).fetchall()
                names = {x[0] for x in cols}
                assert "checksum_sha256" in names
                assert "storage_object_key" in names
                assert "state" in names

            command.upgrade(cfg, "head")
            with eng.connect() as c:
                r = c.execute(
                    text("SELECT to_regclass('public.import_source_artifacts')")
                ).scalar()
                assert r is not None
            assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]
            # silence unused if settings_url only for debug
            assert "postgresql" in settings_url
        finally:
            eng.dispose()
    finally:
        for k, v in prev.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        get_settings.cache_clear()


# --- MinIO CI ---


def test_s3_minio_roundtrip_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL")
    elif not os.environ.get("S3_ENDPOINT_URL"):
        pytest.skip("S3_ENDPOINT_URL not set for local MinIO")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"fourth/{uuid.uuid4()}"
    body = b"fourth-corrective"
    store.put_stream(key, io.BytesIO(body), content_type="application/octet-stream", expected_size=len(body))
    digest = _sha256_object(store, key, chunk_size=64, expected_size=len(body))
    assert digest == hashlib.sha256(body).hexdigest()
    store.delete(key)
    assert store.head(key) is None
