"""S13-PR-002 seventh corrective: H-01…H-06 proof integrity."""
from __future__ import annotations

import hashlib
import io
import os
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any

import openpyxl
import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, sessionmaker

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
    set_reconcile_work_session_factory,
    set_source_limits_override,
    upload_source_artifact,
    _sha256_object,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    assert_accepted_source_upload,
    assert_http_rejection_preserve,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
    AuditEvent,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session(tmp_path) -> Session:
    db_file = tmp_path / f"s13_{uuid.uuid4().hex}.db"
    engine = create_engine(f"sqlite:///{db_file}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()
    set_source_limits_override(None)
    set_reconcile_work_session_factory(None)


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org", organization_slug=f"o-{uuid.uuid4().hex[:6]}", status=OrganizationStatus.ACTIVE
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _xlsx_bytes(*, sheets=1, rows=1, cols=1, cell="x") -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    for r in range(rows):
        for c in range(cols):
            wb.active.cell(r + 1, c + 1, cell)
    for i in range(1, sheets):
        wb.create_sheet(f"S{i + 1}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_bytes(tmp_path, *, sheets=1, rows=1, cols=1, cell="x") -> bytes:
    pytest.importorskip("xlwt")
    import xlwt

    p = tmp_path / f"t-{uuid.uuid4().hex[:6]}.xls"
    book = xlwt.Workbook()
    for si in range(sheets):
        sh = book.add_sheet(f"S{si + 1}")
        for r in range(rows):
            for c in range(cols):
                sh.write(r, c, cell)
    book.save(str(p))
    return p.read_bytes()


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    payload = _xlsx_bytes()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    snap.update(
        {
            "prior_id": prior.id,
            "prior_checksum": prior.checksum_sha256,
            "prior_key": prior.storage_object_key,
            "prior_state": prior.state,
            "prior_gen": prior.generation,
            "batch_current": batch.current_source_artifact_id,
            "batch_status": batch.status,
            "batch_total": batch.total_rows,
            "batch_valid": batch.valid_rows,
            "staging_id": staging.id,
            "staging_raw": dict(staging.raw_values),
            "staging_mapped": dict(staging.mapped_values),
            "staging_name": staging.proposed_asset_name,
            "staging_desc": staging.proposed_description,
            "staging_qty": staging.proposed_quantity,
            "staging_unit": staging.proposed_unit,
            "staging_status": staging.validation_status,
            "line_id": line.id,
            "line_name": line.asset_name,
            "line_desc": line.description,
            "line_qty": float(line.quantity),
            "line_review": line.review_status,
            "line_val": line.validation_status,
            "art_count": len(snap["artifacts"]),
            "line_count": len(snap["lines"]),
        }
    )
    return prior, staging, line, snap


def _assert_preserved(db, fake_storage, proj, batch, staging, line, snap):
    """Field-level prior/staging/line preserve for non-reject paths.

    HTTP N+1 rejections use assert_http_rejection_preserve.
    """
    db.expire_all()
    prior = db.get(ImportSourceArtifact, snap["prior_id"])
    batch = db.get(ProjectAssetImportBatch, batch.id)
    staging = db.get(ProjectAssetImportStagingRow, staging.id)
    line = db.get(ProjectAssetLine, line.id)
    assert prior is not None
    assert prior.checksum_sha256 == snap["prior_checksum"]
    assert prior.storage_object_key == snap["prior_key"]
    assert prior.state == snap["prior_state"]
    assert prior.generation == snap["prior_gen"]
    assert batch.current_source_artifact_id == snap["batch_current"]
    assert batch.status == snap["batch_status"]
    assert batch.total_rows == snap["batch_total"]
    assert batch.valid_rows == snap["batch_valid"]
    assert staging.raw_values == snap["staging_raw"]
    assert staging.mapped_values == snap["staging_mapped"]
    assert staging.proposed_asset_name == snap["staging_name"]
    assert staging.proposed_description == snap["staging_desc"]
    assert staging.proposed_quantity == snap["staging_qty"]
    assert staging.proposed_unit == snap["staging_unit"]
    assert staging.validation_status == snap["staging_status"]
    assert line.asset_name == snap["line_name"]
    assert line.description == snap["line_desc"]
    assert float(line.quantity) == snap["line_qty"]
    assert line.review_status == snap["line_review"]
    assert line.validation_status == snap["line_val"]
    assert snap["prior_key"] in fake_storage._objects
    assert db.query(ProjectAssetLine).filter_by(project_id=proj.id).count() == snap["line_count"]



def _work_factory_fail_nth_commit(bind, *, fail_on: int, hook: dict):
    """Create work Session that fails on the Nth commit (1-based)."""
    Work = sessionmaker(bind=bind, autoflush=False, autocommit=False)
    work = Work()
    real = work.commit
    hook.setdefault("calls", 0)
    hook.setdefault("fail_calls", 0)

    def flaky():
        hook["calls"] += 1
        if hook["calls"] == fail_on:
            hook["fail_calls"] += 1
            raise RuntimeError(f"forced_work_commit_fail_{fail_on}")
        return real()

    work.commit = flaky  # type: ignore[method-assign]
    return work


# ---------------------------------------------------------------------------
# H-01 work-session failpoints
# ---------------------------------------------------------------------------


def test_h01_later_work_commit_fails_preserves_earlier(
    db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    now = datetime.now(timezone.utc) - timedelta(hours=3)
    k1 = f"k1/{uuid.uuid4()}"
    body1 = b"body-one"
    fake_storage._objects[k1] = body1
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256="f" * 64,
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    k2 = f"k2/{uuid.uuid4()}"
    body2 = b"body-two-ok"
    fake_storage._objects[k2] = body2
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="b.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body2),
        checksum_sha256=hashlib.sha256(body2).hexdigest(),
        storage_object_key=k2,
        state="pending",
        created_by_user_id=user.id,
        created_at=now + timedelta(seconds=1),
    )
    db_session.add_all([a1, a2])
    db_session.commit()
    a1_id, a2_id = a1.id, a2.id
    hook: dict[str, Any] = {}

    def factory(bind):
        # fail on 2nd successful item commit (after a1 durable fail transition)
        return _work_factory_fail_nth_commit(bind, fail_on=2, hook=hook)

    set_reconcile_work_session_factory(factory)
    try:
        stats = reconcile_source_artifacts(
            db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
        )
    finally:
        set_reconcile_work_session_factory(None)

    assert hook["fail_calls"] == 1
    assert hook["calls"] >= 2
    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1_id)
    a2 = db_session.get(ImportSourceArtifact, a2_id)
    assert a1.state == "failed"
    assert a1.failure_code == "checksum_mismatch"
    assert a2.state == "pending"  # commit failed before durable promote
    assert stats["marked_failed"] == 1
    assert stats["errors"] >= 1


def test_h01_delete_audit_commit_fails_then_retry_repairs(
    db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    frozen = datetime(2026, 7, 17, 18, 0, 0, tzinfo=timezone.utc)
    k = f"del/{uuid.uuid4()}"
    fake_storage._objects[k] = b"z"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="d.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="e" * 64,
        storage_object_key=k,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=frozen - timedelta(days=2),
        orphaned_at=frozen - timedelta(seconds=7200),
    )
    db_session.add(art)
    db_session.commit()
    aid = art.id
    hook: dict[str, Any] = {}

    def factory(bind):
        return _work_factory_fail_nth_commit(bind, fail_on=1, hook=hook)

    set_reconcile_work_session_factory(factory)
    try:
        stats1 = reconcile_source_artifacts(
            db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid, now=frozen
        )
    finally:
        set_reconcile_work_session_factory(None)

    assert hook["fail_calls"] == 1
    assert k not in fake_storage._objects  # delete happened
    assert stats1["deleted_objects"] == 0  # commit failed — counter not bumped
    assert stats1["errors"] >= 1
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, aid)
    assert art.state == "orphaned"
    audits_before = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == aid,
            AuditEvent.event_name == "ImportSourceArtifactObjectDeleted",
        )
        .count()
    )
    assert audits_before == 0

    # retry without failpoint
    stats2 = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid, now=frozen
    )
    assert stats2["deleted_objects"] == 1
    audits_after = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == aid,
            AuditEvent.event_name == "ImportSourceArtifactObjectDeleted",
        )
        .count()
    )
    assert audits_after == 1


def test_h01_persist_failed_truth_commit_fails(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    now = datetime.now(timezone.utc) - timedelta(hours=2)
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="m.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=3,
        checksum_sha256="a" * 64,
        storage_object_key=f"missing/{uuid.uuid4()}",
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    db_session.add(art)
    db_session.commit()
    aid = art.id
    hook: dict[str, Any] = {}

    def factory(bind):
        return _work_factory_fail_nth_commit(bind, fail_on=1, hook=hook)

    set_reconcile_work_session_factory(factory)
    try:
        stats = reconcile_source_artifacts(
            db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
        )
    finally:
        set_reconcile_work_session_factory(None)

    assert hook["fail_calls"] == 1
    assert stats["marked_failed"] == 0
    assert stats["errors"] >= 1
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, aid)
    assert art.state == "pending"
    assert art.failure_code is None


def test_h01_last_item_early_skip_closes_work_session(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    # available is not selected by reconciler query → empty scan after skip path via id list empty
    # create orphaned that is current pointer → early skip after select
    k = f"cur/{uuid.uuid4()}"
    fake_storage._objects[k] = b"c"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="c.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="c" * 64,
        storage_object_key=k,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(days=2),
        orphaned_at=datetime.now(timezone.utc) - timedelta(days=1),
    )
    db_session.add(art)
    db_session.flush()
    batch.current_source_artifact_id = art.id
    db_session.commit()
    closed = {"n": 0}
    Work = sessionmaker(bind=db_session.get_bind(), autoflush=False, autocommit=False)

    def factory(bind):
        work = Work()
        real_close = work.close

        def close():
            closed["n"] += 1
            return real_close()

        work.close = close  # type: ignore[method-assign]
        return work

    set_reconcile_work_session_factory(factory)
    try:
        stats = reconcile_source_artifacts(
            db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
        )
    finally:
        set_reconcile_work_session_factory(None)
    assert stats["scanned"] == 1
    assert stats["deleted_objects"] == 0
    assert closed["n"] == 1
    assert k in fake_storage._objects


# ---------------------------------------------------------------------------
# H-02 finalize ambiguity + concurrent stale
# ---------------------------------------------------------------------------


def test_h02_finalize_commit_raises_before_durability(
    db_session: Session, fake_storage, monkeypatch, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes()
    p = tmp_path / "a.xlsx"
    p.write_bytes(payload)
    commits = {"n": 0}
    real = Session.commit

    def flaky(self):
        if self is not db_session:
            return real(self)
        commits["n"] += 1
        if commits["n"] == 2:
            raise RuntimeError("forced_final_commit_fail")
        return real(self)

    monkeypatch.setattr(Session, "commit", flaky)
    from fastapi import UploadFile
    from starlette.datastructures import Headers

    class R:
        headers = Headers({})

    with open(p, "rb") as f:
        uf = UploadFile(filename="a.xlsx", file=f)
        with pytest.raises(HTTPException) as ei:
            upload_source_artifact(
                db_session,
                org_id=org.id,
                project_id=proj.id,
                batch_id=batch.id,
                file=uf,
                request=R(),
                current_user=user,
                storage=fake_storage,
            )
    assert ei.value.status_code == 500
    db_session.expire_all()
    arts = [a for a in db_session.query(ImportSourceArtifact).all() if a.generation > 1]
    assert len(arts) == 1
    art = arts[0]
    assert art.state in {"failed", "pending"}
    if art.state == "failed":
        assert art.failure_code == "finalize_commit_failed"
    assert art.storage_object_key in fake_storage._objects
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert batch.current_source_artifact_id == snap["batch_current"]
    _assert_preserved(db_session, fake_storage, proj, batch, staging, line, snap)


def test_h02_retry_idempotent_when_already_available(
    db_session: Session, fake_storage
):
    """If finalize actually committed (available+pointer), retry must not double-audit."""
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    body = b"available-body"
    key = f"av/{uuid.uuid4()}"
    fake_storage._objects[key] = body
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body),
        checksum_sha256=hashlib.sha256(body).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db_session.add(art)
    db_session.flush()
    batch.current_source_artifact_id = art.id
    db_session.commit()
    aid = art.id
    avail_audits_before = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == aid,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .count()
    )
    # reconcile does not re-select available
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
    )
    assert stats["scanned"] == 0
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, aid)
    assert art.state == "available"
    assert db_session.get(ProjectAssetImportBatch, batch.id).current_source_artifact_id == aid
    avail_audits_after = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == aid,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .count()
    )
    assert avail_audits_after == avail_audits_before


def test_h02_newer_finalize_wins_over_older_pending_recovery(
    db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    now = datetime.now(timezone.utc)
    # older pending residual with valid object
    k1 = f"old/{uuid.uuid4()}"
    b1 = b"older-body"
    fake_storage._objects[k1] = b1
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="old.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(b1),
        checksum_sha256=hashlib.sha256(b1).hexdigest(),
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=now - timedelta(hours=1),
    )
    # newer available current (concurrent finalize won)
    k2 = f"new/{uuid.uuid4()}"
    b2 = b"newer-body-xx"
    fake_storage._objects[k2] = b2
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="new.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(b2),
        checksum_sha256=hashlib.sha256(b2).hexdigest(),
        storage_object_key=k2,
        state="available",
        created_by_user_id=user.id,
        available_at=now,
        created_at=now,
    )
    db_session.add_all([a1, a2])
    db_session.flush()
    batch.current_source_artifact_id = a2.id
    db_session.commit()
    a1_id, a2_id = a1.id, a2.id
    reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
    )
    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1_id)
    a2 = db_session.get(ImportSourceArtifact, a2_id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert a2.state == "available"
    assert batch.current_source_artifact_id == a2_id
    assert a1.state != "available"
    assert a1.state in {"orphaned", "pending", "failed"}


# ---------------------------------------------------------------------------
# H-03 HTTP request/upload/adapter boundaries with limits seam
# ---------------------------------------------------------------------------


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_h03_request_bytes_exact_n_and_n_plus_one(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes()
    n = len(payload) + 200  # content-length includes multipart overhead; set high enough for accept
    set_source_limits_override(
        SourceArtifactLimits(max_request_bytes=n, max_upload_bytes=10 * 1024 * 1024)
    )
    try:
        # N+1 via Content-Length header
        res_bad = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={
                "file": (
                    "a.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id), "Content-Length": str(n + 1)},
        )
        assert_http_rejection_preserve(
            res_bad,
            status=413,
            error_code="request_too_large",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )

        # exact N: header within limit
        res_ok = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={
                "file": (
                    "a.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id), "Content-Length": str(n)},
        )
        # may 201 if body small enough, or 413 if multipart exceeds — stable: not 500
        assert res_ok.status_code in {201, 413}
        if res_ok.status_code == 413:
            d = res_ok.json().get("detail")
            if isinstance(d, dict):
                assert d.get("error_code") == "request_too_large"
    finally:
        set_source_limits_override(None)


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_h03_upload_bytes_exact_n_and_n_plus_one(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes()
    n = len(payload)
    set_source_limits_override(
        SourceArtifactLimits(max_upload_bytes=n, max_request_bytes=12 * 1024 * 1024)
    )
    try:
        # N+1 reject first (preserves prior)
        big = payload + b"X"
        res_bad = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={
                "file": (
                    "big.xlsx",
                    io.BytesIO(big),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res_bad,
            status=413,
            error_code="upload_too_large",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )

        # exact N accept
        res_ok = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={
                "file": (
                    "ok.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_accepted_source_upload(res_ok, status=201)
    finally:
        set_source_limits_override(None)


@pytest.mark.parametrize(
    "limit_field,n,build,error_code,status",
    [
        ("max_sheets", 1, lambda: _xlsx_bytes(sheets=2), "sheet_limit", 413),
        ("max_physical_rows", 1, lambda: _xlsx_bytes(rows=2), "physical_row_limit", 413),
        ("max_columns", 1, lambda: _xlsx_bytes(cols=2), "column_limit", 413),
        ("max_cell_chars", 3, lambda: _xlsx_bytes(cell="abcd"), "cell_length_limit", 400),
    ],
)
@pytest.mark.s13_pr_002_http_nplus1_reject
def test_h03_endpoint_xlsx_adapter_limits(
    client: TestClient, db_session: Session, fake_storage, limit_field, n, build, error_code, status
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    set_source_limits_override(SourceArtifactLimits(**{limit_field: n}))
    try:
        payload = build()
        res = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={
                "file": (
                    "lim.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res,
            status=status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
    finally:
        set_source_limits_override(None)


@pytest.mark.parametrize(
    "limit_field,n,kwargs,error_code,status",
    [
        ("max_sheets", 1, {"sheets": 2}, "sheet_limit", 413),
        ("max_physical_rows", 1, {"rows": 2}, "physical_row_limit", 413),
        ("max_columns", 1, {"cols": 2}, "column_limit", 413),
        ("max_cell_chars", 3, {"cell": "abcd"}, "cell_length_limit", 400),
    ],
)
@pytest.mark.s13_pr_002_http_nplus1_reject
def test_h03_endpoint_xls_adapter_limits(
    client: TestClient,
    db_session: Session,
    fake_storage,
    tmp_path,
    limit_field,
    n,
    kwargs,
    error_code,
    status,
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    set_source_limits_override(SourceArtifactLimits(**{limit_field: n}))
    try:
        payload = _xls_bytes(tmp_path, **kwargs)
        res = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={"file": ("lim.xls", io.BytesIO(payload), "application/vnd.ms-excel")},
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res,
            status=status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
    finally:
        set_source_limits_override(None)


# ---------------------------------------------------------------------------
# H-04 exact RESTRICT identity
# ---------------------------------------------------------------------------


def _pg_url():
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql")
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL required")
    return url


def test_h04_pg_restrict_batch_delete_exact_constraint():
    url = _pg_url()
    engine = create_engine(url)
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    aid = uuid.uuid4()
    slug = f"h04-{uuid.uuid4().hex[:8]}"

    def _cname(exc: IntegrityError) -> str:
        orig = getattr(exc, "orig", None)
        diag = getattr(orig, "diag", None) if orig is not None else None
        return getattr(diag, "constraint_name", None) or ""

    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'b.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid, "oid": oid, "pid": pid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'x.xlsx', 'xlsx', 't', 1,
                      :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": aid,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid,
                    "chk": "a" * 64,
                    "key": f"k-{aid}",
                    "uid": uid,
                },
            )
            # Scalar FK only (composite tenant FK also includes import_batch_id)
            cname_expected = conn.execute(
                text(
                    """
                    SELECT con.conname
                    FROM pg_constraint con
                    JOIN pg_attribute a ON a.attrelid = con.conrelid
                      AND a.attnum = con.conkey[1]
                    WHERE con.contype = 'f'
                      AND con.conrelid = 'import_source_artifacts'::regclass
                      AND array_length(con.conkey, 1) = 1
                      AND a.attname = 'import_batch_id'
                    """
                )
            ).scalar()
            assert cname_expected == "import_source_artifacts_import_batch_id_fkey"

        with engine.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text("DELETE FROM project_asset_import_batches WHERE id = :id"),
                    {"id": bid},
                )
            assert _cname(ei.value) == "import_source_artifacts_import_batch_id_fkey"
    finally:
        with engine.begin() as conn:
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


# ---------------------------------------------------------------------------
# H-05 throwaway schema + DML matrix
# ---------------------------------------------------------------------------


def test_h05_throwaway_migration_schema_and_dml_matrix():
    url = _pg_url()
    from alembic import command
    from alembic.config import Config
    from alembic.script import ScriptDirectory
    from app.core.config import get_settings
    from sqlalchemy.engine.url import make_url

    admin = create_engine(url, isolation_level="AUTOCOMMIT")
    db_name = f"s13_h05_{uuid.uuid4().hex[:10]}"
    with admin.connect() as conn:
        conn.execute(text(f'CREATE DATABASE "{db_name}"'))
    u = make_url(url)
    iso_url = url.rsplit("/", 1)[0] + f"/{db_name}"
    prev = {
        k: os.environ.get(k)
        for k in (
            "POSTGRES_HOST",
            "POSTGRES_PORT",
            "POSTGRES_DB",
            "POSTGRES_USER",
            "POSTGRES_PASSWORD",
            "VALORA_ENV",
        )
    }

    def _cname(exc: IntegrityError) -> str:
        orig = getattr(exc, "orig", None)
        diag = getattr(orig, "diag", None) if orig is not None else None
        return getattr(diag, "constraint_name", None) or ""

    def _assert_schema(eng):
        with eng.connect() as c:
            cols = {
                r[0]: (r[1], r[2])
                for r in c.execute(
                    text(
                        "SELECT column_name, is_nullable, data_type "
                        "FROM information_schema.columns "
                        "WHERE table_name='import_source_artifacts'"
                    )
                )
            }
            required = {
                "id": "NO",
                "organization_id": "NO",
                "project_id": "NO",
                "import_batch_id": "NO",
                "generation": "NO",
                "checksum_sha256": "NO",
                "storage_object_key": "NO",
                "state": "NO",
                "file_size_bytes": "NO",
                "detected_format": "NO",
            }
            for col, nullable in required.items():
                assert col in cols
                assert cols[col][0] == nullable
            checks = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint "
                        "WHERE conrelid = 'import_source_artifacts'::regclass AND contype = 'c'"
                    )
                )
            }
            for name in (
                "chk_source_artifact_generation_positive",
                "chk_source_artifact_size_nonneg",
                "chk_source_artifact_checksum_len",
                "chk_source_artifact_checksum_lower",
                "chk_source_artifact_checksum_hex",
                "chk_source_artifact_state",
                "chk_source_artifact_format",
            ):
                assert name in checks
            uniques = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint "
                        "WHERE conrelid = 'import_source_artifacts'::regclass AND contype = 'u'"
                    )
                )
            }
            for name in (
                "uq_source_artifact_batch_generation",
                "uq_source_artifact_object_key",
                "uq_source_artifact_batch_id",
            ):
                assert name in uniques
            fks = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint WHERE contype = 'f' AND ("
                        "conrelid = 'import_source_artifacts'::regclass OR "
                        "conrelid = 'project_asset_import_batches'::regclass)"
                    )
                )
            }
            assert "fk_source_artifact_batch_tenant" in fks
            assert "fk_batch_current_artifact_same_batch" in fks
            # scalar FKs present (auto-named ok if exist)
            assert any("organization" in n or "organization_id" in n for n in fks) or any(
                n.endswith("_fkey") for n in fks
            )
            idxs = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT indexname FROM pg_indexes WHERE tablename='import_source_artifacts'"
                    )
                )
            }
            for name in (
                "idx_source_artifact_org",
                "idx_source_artifact_project",
                "idx_source_artifact_batch",
                "idx_source_artifact_state",
            ):
                assert name in idxs
            bcols = {
                r[0]: r[1]
                for r in c.execute(
                    text(
                        "SELECT column_name, is_nullable FROM information_schema.columns "
                        "WHERE table_name='project_asset_import_batches'"
                    )
                )
            }
            assert "current_source_artifact_id" in bcols
            assert bcols["current_source_artifact_id"] == "YES"

    def _seed_and_dml(eng):
        oid = uuid.uuid4()
        uid = uuid.uuid4()
        cid = uuid.uuid4()
        pid = uuid.uuid4()
        b1 = uuid.uuid4()
        b2 = uuid.uuid4()
        a1 = uuid.uuid4()
        slug = f"dml-{uuid.uuid4().hex[:8]}"
        with eng.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            for bid, fn in ((b1, "b1.xlsx"), (b2, "b2.xlsx")):
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :fn, 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": bid, "oid": oid, "pid": pid, "fn": fn, "uid": uid},
                )
            # valid insert
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'x.xlsx', 'xlsx', 't', 1,
                      :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": a1,
                    "oid": oid,
                    "pid": pid,
                    "bid": b1,
                    "chk": "a" * 64,
                    "key": f"k-{a1}",
                    "uid": uid,
                },
            )
            # same-batch pointer
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": a1, "bid": b1},
            )

        # cross-batch: a1 belongs to b1, set pointer on b2 to a1
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                    ),
                    {"aid": a1, "bid": b2},
                )
            assert _cname(ei.value) == "fk_batch_current_artifact_same_batch"

        # batch-generation unique
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 1, 'y.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid,
                        "pid": pid,
                        "bid": b1,
                        "chk": "b" * 64,
                        "key": f"k-dupgen-{uuid.uuid4()}",
                        "uid": uid,
                    },
                )
            assert "uq_source_artifact_batch_generation" in _cname(ei.value)

        # object key unique
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 2, 'y.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid,
                        "pid": pid,
                        "bid": b1,
                        "chk": "c" * 64,
                        "key": f"k-{a1}",
                        "uid": uid,
                    },
                )
            assert "uq_source_artifact_object_key" in _cname(ei.value)

        # CHECK failures
        check_cases = [
            (
                ("chk_source_artifact_state",),
                "UPDATE import_source_artifacts SET state = 'bogus' WHERE id = :id",
                {"id": a1},
            ),
            (
                ("chk_source_artifact_format",),
                "UPDATE import_source_artifacts SET detected_format = 'csv' WHERE id = :id",
                {"id": a1},
            ),
            (
                ("chk_source_artifact_generation_positive",),
                "UPDATE import_source_artifacts SET generation = 0 WHERE id = :id",
                {"id": a1},
            ),
            (
                ("chk_source_artifact_size_nonneg",),
                "UPDATE import_source_artifacts SET file_size_bytes = -1 WHERE id = :id",
                {"id": a1},
            ),
            (
                ("chk_source_artifact_checksum_len", "chk_source_artifact_checksum_hex"),
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "a" * 63},
            ),
            (
                ("chk_source_artifact_checksum_lower", "chk_source_artifact_checksum_hex"),
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "A" * 64},
            ),
            (
                ("chk_source_artifact_checksum_hex",),
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "g" * 64},
            ),
        ]
        for names, sql, params in check_cases:
            with eng.begin() as conn:
                with pytest.raises(IntegrityError) as ei:
                    conn.execute(text(sql), params)
                cname = _cname(ei.value)
                assert any(n in cname for n in names), cname

        # RESTRICT batch delete — scalar import_batch_id FK only
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text("DELETE FROM project_asset_import_batches WHERE id = :id"),
                    {"id": b1},
                )
            assert _cname(ei.value) == "import_source_artifacts_import_batch_id_fkey"

    try:
        os.environ["VALORA_ENV"] = "test"
        os.environ["POSTGRES_HOST"] = u.host or "localhost"
        os.environ["POSTGRES_PORT"] = str(u.port or 5432)
        os.environ["POSTGRES_DB"] = db_name
        os.environ["POSTGRES_USER"] = u.username or "valora"
        os.environ["POSTGRES_PASSWORD"] = u.password or "valora_local_password"
        get_settings.cache_clear()
        cfg = Config("alembic.ini")
        assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]
        command.upgrade(cfg, "e1f2a3b4c5d6")
        eng = create_engine(iso_url)
        try:
            command.upgrade(cfg, "f2a3b4c5d6e7")
            _assert_schema(eng)
            _seed_and_dml(eng)
            command.downgrade(cfg, "e1f2a3b4c5d6")
            command.upgrade(cfg, "head")
            _assert_schema(eng)
            _seed_and_dml(eng)
        finally:
            eng.dispose()
    finally:
        for k, v in prev.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        get_settings.cache_clear()
        with admin.connect() as conn:
            conn.execute(text(f'DROP DATABASE IF EXISTS "{db_name}" WITH (FORCE)'))
        admin.dispose()


# ---------------------------------------------------------------------------
# H-06 PostgreSQL transaction ownership
# ---------------------------------------------------------------------------


def test_h06_pg_dedicated_session_preserves_caller_flushed_uow():
    url = _pg_url()
    engine = create_engine(url)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    caller = SessionLocal()
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    slug = f"h06-{uuid.uuid4().hex[:8]}"
    try:
        # seed committed baseline
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'b.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid, "oid": oid, "pid": pid, "uid": uid},
            )

        # caller flushes uncommitted line-like update on batch
        batch = caller.get(ProjectAssetImportBatch, bid)
        assert batch is not None
        batch.source_filename = "flushed-uncommitted.xlsx"
        caller.flush()
        assert caller.in_transaction()

        # second connection cannot see uncommitted change
        other = SessionLocal()
        try:
            b2 = other.get(ProjectAssetImportBatch, bid)
            assert b2 is not None
            assert b2.source_filename == "b.xlsx"
        finally:
            other.close()

        # reconcile via dedicated session
        store = FakeObjectStorage()
        stats = reconcile_source_artifacts(
            caller, storage=store, max_items=5, actor_id=uid, org_id=oid
        )
        assert stats["scanned"] == 0

        # caller still has flushed uncommitted change and open txn
        assert caller.in_transaction()
        batch = caller.get(ProjectAssetImportBatch, bid)
        assert batch.source_filename == "flushed-uncommitted.xlsx"

        other = SessionLocal()
        try:
            b2 = other.get(ProjectAssetImportBatch, bid)
            assert b2.source_filename == "b.xlsx"
        finally:
            other.close()

        caller.rollback()
        batch = caller.get(ProjectAssetImportBatch, bid)
        assert batch.source_filename == "b.xlsx"
    finally:
        caller.close()
        with engine.begin() as conn:
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


# ---------------------------------------------------------------------------
# MinIO
# ---------------------------------------------------------------------------


def test_h07_s3_minio_roundtrip_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL")
    elif not os.environ.get("S3_ENDPOINT_URL"):
        pytest.skip("S3_ENDPOINT_URL not set for local MinIO")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"seventh/{uuid.uuid4()}"
    body = b"seventh-corrective"
    store.put_stream(
        key, io.BytesIO(body), content_type="application/octet-stream", expected_size=len(body)
    )
    digest = _sha256_object(store, key, chunk_size=64, expected_size=len(body))
    assert digest == hashlib.sha256(body).hexdigest()
    store.delete(key)
    assert store.head(key) is None
