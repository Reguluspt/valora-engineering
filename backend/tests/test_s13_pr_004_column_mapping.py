"""Focused SQLite development proof for S13-PR-004 Column Mapping Memory."""
from __future__ import annotations

import hashlib
import io
import os
import uuid
from copy import deepcopy
from dataclasses import replace
from datetime import datetime, timezone

import openpyxl
import pytest
from fastapi import HTTPException
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.db import Base
import app.modules.excel_import.models  # noqa: F401
import app.modules.excel_import.application.column_mapping_service as mapping_service
from app.modules.excel_import.application.column_mapping_service import (
    confirm_column_mapping,
    materialize_confirmed_mapping_to_staging,
    propose_column_mapping,
    reject_column_mapping,
)
from app.modules.excel_import.domain.column_mapping import (
    ColumnMappingContractError,
    MappingField,
    SemanticRole,
    build_mapping_snapshot,
    build_template_fingerprint,
    canonical_sha256,
    column_letter,
    mapping_digest,
    normalize_header,
    profile_mapping_digest,
    project_asset_row,
    similar_template_remap,
    suggest_mapping,
    validate_mapping_snapshot,
)
from app.modules.excel_import.domain.workbook_adapter import AdapterError, CellValue
from app.modules.excel_import.domain.workbook_structure import canonical_payload_digest
from app.modules.excel_import.infrastructure.object_storage import FakeObjectStorage
from app.modules.excel_import.models import (
    ColumnMappingDecision,
    ColumnMappingField,
    ColumnMappingProfile,
    ColumnMappingProfileUsage,
    ImportSourceArtifact,
    WorkbookStructureSnapshot,
)
from app.modules.project_master_data.models import (
    AuditEvent,
    Customer,
    CustomerStatus,
    ImportBatchStatus,
    ImportRowValidationStatus,
    OrganizationProfile,
    OrganizationStatus,
    Project,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ProjectAssetLine,
    ProjectWorkflowStatus,
    User,
    UserStatus,
)


@pytest.fixture
def mapping_db() -> Session:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(engine)


def _xlsx_bytes(*, asset_rows: int = 3) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "PD-001"
    sheet["A1"] = "BẢNG KÊ TÀI SẢN"
    headers = ["STT", "TÊN VẬT TƯ", "Đặc điểm", "ĐVT", "KHỐI LƯỢNG", "GIÁ TĐ", None]
    for column, value in enumerate(headers, 1):
        sheet.cell(5, column, value)
    row_number = 6
    sheet.cell(row_number, 1, "PHẦN ĐIỆN")
    row_number += 1
    for index in range(1, asset_rows + 1):
        values = [index, f"Tài sản {index}", f"Mô tả {index}", "cái", index, index * 100, "E"]
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
        row_number += 1
    sheet.cell(row_number, 1, "Cộng phần điện")
    row_number += 1
    sheet.cell(row_number, 1, "TỔNG CỘNG")
    row_number += 1
    sheet.cell(row_number, 1, "Ghi chú: kiểm tra")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xls_bytes(*, asset_rows: int = 3) -> bytes:
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("PD-001")
    sheet.write(0, 0, "BẢNG KÊ TÀI SẢN")
    headers = ["STT", "TÊN VẬT TƯ", "Đặc điểm", "ĐVT", "KHỐI LƯỢNG", "GIÁ TĐ", None]
    for column, value in enumerate(headers):
        if value is not None:
            sheet.write(4, column, value)
    row_index = 5
    sheet.write(row_index, 0, "PHẦN ĐIỆN")
    row_index += 1
    for index in range(1, asset_rows + 1):
        values = [index, f"Tài sản {index}", f"Mô tả {index}", "cái", index, index * 100, "E"]
        for column, value in enumerate(values):
            sheet.write(row_index, column, value)
        row_index += 1
    sheet.write(row_index, 0, "Cộng phần điện")
    row_index += 1
    sheet.write(row_index, 0, "TỔNG CỘNG")
    row_index += 1
    sheet.write(row_index, 0, "Ghi chú: kiểm tra")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _candidate(*, max_row: int = 12, headers: list[str | None] | None = None) -> dict:
    labels = headers or [
        "STT",
        "TÊN VẬT TƯ",
        "Đặc điểm",
        "ĐVT",
        "KHỐI LƯỢNG",
        "GIÁ TĐ",
        None,
    ]
    return {
        "sheet_name": "PD-001",
        "header_start_row": 5,
        "header_end_row": 5,
        "data_start_row": 6,
        "candidate_table_bounds": {
            "min_row": 5,
            "max_row": max_row,
            "min_column": 1,
            "max_column": len(labels),
        },
        "boundary_reason": "sheet_end",
        "boundary_flags": [],
        "confidence": 0.9,
        "reasons": ["header_vocabulary"],
        "header_labels": labels,
    }


def _seed(
    db: Session,
    *,
    storage: FakeObjectStorage | None = None,
    asset_rows: int = 3,
    org: OrganizationProfile | None = None,
    user: User | None = None,
    customer: Customer | None = None,
    workbook_format: str = "xlsx",
) -> dict:
    if org is None:
        org = OrganizationProfile(
            legal_name="Mapping Org",
            organization_slug=f"mapping-{uuid.uuid4().hex[:8]}",
            status=OrganizationStatus.ACTIVE,
        )
        db.add(org)
        db.flush()
    if user is None:
        user = User(
            organization_id=org.id,
            email=f"mapping-{uuid.uuid4().hex[:8]}@example.com",
            full_name="Mapping Human",
            status=UserStatus.ACTIVE,
        )
        db.add(user)
        db.flush()
    if customer is None:
        customer = Customer(
            organization_id=org.id,
            legal_name="Mapping Customer",
            status=CustomerStatus.ACTIVE,
            created_by=user.id,
        )
        db.add(customer)
        db.flush()
    project = Project(
        organization_id=org.id,
        customer_id=customer.id,
        name="Mapping Project",
        code=f"MAP-{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(project)
    db.flush()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=project.id,
        source_filename="pd001.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.flush()

    if workbook_format == "xlsx":
        content = _xlsx_bytes(asset_rows=asset_rows)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        adapter_name = "xlsx-openpyxl"
        adapter_version = "s13-pr-002-v2"
    elif workbook_format == "xls":
        content = _xls_bytes(asset_rows=asset_rows)
        content_type = "application/vnd.ms-excel"
        adapter_name = "xls-xlrd"
        adapter_version = "s13-pr-002-v3"
    else:
        raise ValueError("unsupported test workbook format")
    checksum = hashlib.sha256(content).hexdigest()
    filename = f"pd001.{workbook_format}"
    key = f"mapping/{uuid.uuid4()}.{workbook_format}"
    artifact = ImportSourceArtifact(
        organization_id=org.id,
        project_id=project.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename=filename,
        detected_format=workbook_format,
        content_type=content_type,
        file_size_bytes=len(content),
        checksum_sha256=checksum,
        storage_object_key=key,
        state="available",
        adapter_name=adapter_name,
        adapter_version=adapter_version,
        adapter_metadata={},
        created_by_user_id=user.id,
    )
    db.add(artifact)
    db.flush()
    batch.current_source_artifact_id = artifact.id
    candidate = _candidate(max_row=asset_rows + 9)
    payload = {
        "rule_version": "s13-pr-003-v3",
        "rule_config": {},
        "disposition": "proposed",
        "disposition_reasons": [],
        "proposed_candidate_index": 0,
        "candidate_count": 1,
        "candidates": [candidate],
        "row_classification": {
            "sheet_name": "PD-001",
            "data_start_row": 6,
            "candidate_table_bounds": candidate["candidate_table_bounds"],
            "physical_rows_classified": asset_rows + 4,
            "counts": {},
            "preview": [],
            "preview_truncated": asset_rows > 200,
        },
        "source": {
            "source_artifact_id": str(artifact.id),
            "source_generation": 1,
            "source_checksum_sha256": checksum,
            "detected_format": workbook_format,
            "adapter_name": adapter_name,
            "adapter_version": adapter_version,
        },
        "drift_reference": None,
    }
    digest = canonical_payload_digest(payload)
    snapshot = WorkbookStructureSnapshot(
        organization_id=org.id,
        project_id=project.id,
        import_batch_id=batch.id,
        source_artifact_id=artifact.id,
        snapshot_version=1,
        source_checksum_sha256=checksum,
        rule_version="s13-pr-003-v3",
        adapter_name=adapter_name,
        adapter_version=adapter_version,
        disposition="proposed",
        candidate_count=1,
        structure_payload=payload,
        analysis_digest_sha256=digest,
        created_by_user_id=user.id,
    )
    db.add(snapshot)
    db.flush()
    db.add(
        AuditEvent(
            organization_id=org.id,
            actor_user_id=user.id,
            event_name="WorkbookStructureAnalyzed",
            entity_type="WorkbookStructureSnapshot",
            entity_id=snapshot.id,
            command_name="AnalyzeWorkbookStructure",
            payload={
                "import_batch_id": str(batch.id),
                "source_artifact_id": str(artifact.id),
                "source_generation": 1,
                "snapshot_version": 1,
                "rule_version": "s13-pr-003-v3",
                "disposition": "proposed",
                "candidate_count": 1,
                "analysis_digest_sha256": digest,
            },
        )
    )
    db.commit()
    if storage is not None:
        storage.put_stream(
            key,
            io.BytesIO(content),
            content_type=artifact.content_type,
            expected_size=len(content),
        )
    return {
        "org": org,
        "user": user,
        "customer": customer,
        "project": project,
        "batch": batch,
        "artifact": artifact,
        "snapshot": snapshot,
        "candidate": candidate,
    }


def _propose(db: Session, seeded: dict):
    return propose_column_mapping(
        db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        artifact_id=seeded["artifact"].id,
        snapshot_id=seeded["snapshot"].id,
        candidate_index=0,
        command_id=uuid.uuid4(),
    )


def _replace_candidate(db: Session, seeded: dict, headers: list[str | None]) -> None:
    snapshot = seeded["snapshot"]
    payload = deepcopy(snapshot.structure_payload)
    candidate = deepcopy(payload["candidates"][0])
    candidate["header_labels"] = headers
    candidate["candidate_table_bounds"]["max_column"] = len(headers)
    payload["candidates"][0] = candidate
    payload["row_classification"]["candidate_table_bounds"] = candidate[
        "candidate_table_bounds"
    ]
    digest = canonical_payload_digest(payload)
    snapshot.structure_payload = payload
    snapshot.analysis_digest_sha256 = digest
    event = db.query(AuditEvent).filter(
        AuditEvent.entity_id == snapshot.id,
        AuditEvent.event_name == "WorkbookStructureAnalyzed",
    ).one()
    event.payload = {**event.payload, "analysis_digest_sha256": digest}
    db.commit()
    seeded["candidate"] = candidate


def test_semantic_registry_is_accent_insensitive_and_blank_stays_ignore():
    suggestion = suggest_mapping(_candidate())
    roles = {field.source_column_index: field.semantic_role for field in suggestion.fields}
    assert roles[2] is SemanticRole.RAW_ASSET_NAME
    assert roles[5] is SemanticRole.QUANTITY
    assert roles[6] is SemanticRole.APPRAISER_PROPOSED_PRICE
    assert roles[7] is SemanticRole.IGNORE
    assert normalize_header("  KHỐI   LƯỢNG ") == "khoi luong"
    assert suggestion.review_required is True


def test_blank_header_can_be_human_confirmed_as_evidence_note():
    candidate = _candidate()
    fields = list(suggest_mapping(candidate).fields)
    fields[-1] = replace(fields[-1], semantic_role=SemanticRole.EVIDENCE_NOTE)
    snapshot = build_mapping_snapshot(
        source_artifact_id=str(uuid.uuid4()),
        source_generation=1,
        source_checksum_sha256="a" * 64,
        structure_snapshot_id=str(uuid.uuid4()),
        snapshot_version=1,
        structure_rule_version="s13-pr-003-v3",
        structure_digest_sha256="b" * 64,
        candidate_index=0,
        candidate=candidate,
        fields=fields,
    )
    assert validate_mapping_snapshot(snapshot)[-1].semantic_role is SemanticRole.EVIDENCE_NOTE


def test_mapping_cardinality_and_canonical_json_fail_closed():
    candidate = _candidate()
    fields = list(suggest_mapping(candidate).fields)
    fields[0] = replace(fields[0], semantic_role=SemanticRole.RAW_ASSET_NAME)
    with pytest.raises(ColumnMappingContractError) as exc:
        build_mapping_snapshot(
            source_artifact_id=str(uuid.uuid4()),
            source_generation=1,
            source_checksum_sha256="a" * 64,
            structure_snapshot_id=str(uuid.uuid4()),
            snapshot_version=1,
            structure_rule_version="s13-pr-003-v3",
            structure_digest_sha256="b" * 64,
            candidate_index=0,
            candidate=candidate,
            fields=fields,
        )
    assert exc.value.error_code == "mapping_role_cardinality_invalid"
    with pytest.raises(ColumnMappingContractError):
        canonical_sha256({"bad": float("nan")})


def test_digest_is_stable_and_profile_digest_ignores_generation_lineage():
    candidate = _candidate()
    fields = suggest_mapping(candidate).fields
    kwargs = dict(
        source_artifact_id=str(uuid.uuid4()),
        source_generation=1,
        source_checksum_sha256="a" * 64,
        structure_snapshot_id=str(uuid.uuid4()),
        snapshot_version=1,
        structure_rule_version="s13-pr-003-v3",
        structure_digest_sha256="b" * 64,
        candidate_index=0,
        candidate=candidate,
        fields=fields,
    )
    first = build_mapping_snapshot(**kwargs)
    second = build_mapping_snapshot(**dict(reversed(list(kwargs.items()))))
    assert mapping_digest(first) == mapping_digest(second)
    assert profile_mapping_digest(candidate, fields) == profile_mapping_digest(candidate, fields)


def test_similar_template_requires_unique_same_multiset_labels():
    profile_headers = ["STT", "Tên vật tư", "ĐVT"]
    fields = (
        MappingField(1, "A", "STT", SemanticRole.ROW_NUMBER),
        MappingField(2, "B", "Tên vật tư", SemanticRole.RAW_ASSET_NAME),
        MappingField(3, "C", "ĐVT", SemanticRole.UNIT),
    )
    result = similar_template_remap(
        current_headers=["ĐVT", "STT", "TÊN VẬT TƯ"],
        current_min_column=1,
        current_header_height=1,
        profile_fields=fields,
        profile_headers=profile_headers,
        profile_header_height=1,
    )
    assert result.qualifies is True
    assert [field.semantic_role for field in result.fields] == [
        SemanticRole.UNIT,
        SemanticRole.ROW_NUMBER,
        SemanticRole.RAW_ASSET_NAME,
    ]
    duplicate = similar_template_remap(
        current_headers=["STT", "STT", "Tên vật tư"],
        current_min_column=1,
        current_header_height=1,
        profile_fields=fields,
        profile_headers=profile_headers,
        profile_header_height=1,
    )
    assert duplicate.qualifies is False


def test_projection_preserves_positions_and_never_projects_appraised_price_to_raw_price():
    fields = suggest_mapping(_candidate()).fields
    row = tuple(
        CellValue(7, index, f"{column_letter(index)}7", value, "string")
        for index, value in enumerate([1, "Máy bơm", "Mô tả", "cái", 2, 900, "note"], 1)
    )
    projected = project_asset_row(row, fields)
    assert list(projected["raw_values"]) == [f"column_{index:04d}" for index in range(1, 8)]
    assert projected["proposed_asset_name"] == "Máy bơm"
    assert projected["proposed_quantity"] == "2"
    assert projected["proposed_raw_price"] is None
    assert projected["proposed_appraised_unit_price"] == "900"
    assert projected["proposed_currency"] is None


def test_proposal_is_append_only_audited_and_does_not_touch_staging(mapping_db: Session):
    seeded = _seed(mapping_db)
    result = _propose(mapping_db, seeded)
    assert result.review_required is True
    assert result.decision.decision_kind == "proposal"
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 0
    events = mapping_db.query(AuditEvent).filter(AuditEvent.event_name == "ColumnMappingProposed").all()
    assert len(events) == 1
    assert "original_header" not in events[0].payload


def test_proposal_idempotency_rejects_changed_input(mapping_db: Session):
    seeded = _seed(mapping_db)
    command_id = uuid.uuid4()
    first = propose_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        artifact_id=seeded["artifact"].id,
        snapshot_id=seeded["snapshot"].id,
        candidate_index=0,
        command_id=command_id,
    )
    retry = propose_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        artifact_id=seeded["artifact"].id,
        snapshot_id=seeded["snapshot"].id,
        candidate_index=0,
        command_id=command_id,
    )
    assert retry.decision.id == first.decision.id
    with pytest.raises(HTTPException) as exc:
        propose_column_mapping(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            artifact_id=seeded["artifact"].id,
            snapshot_id=seeded["snapshot"].id,
            candidate_index=1,
            command_id=command_id,
        )
    assert exc.value.detail["error_code"] == "idempotency_key_reused"


@pytest.mark.parametrize("memory_scope,expected_profiles", [("none", 0), ("customer", 1)])
def test_human_confirmation_persists_exact_decision_and_optional_memory(
    mapping_db: Session, memory_scope: str, expected_profiles: int
):
    seeded = _seed(mapping_db)
    proposal = _propose(mapping_db, seeded).decision
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        memory_scope=memory_scope,
        command_id=uuid.uuid4(),
    )
    assert confirmation.outcome == "accepted"
    assert confirmation.memory_scope == memory_scope
    assert mapping_db.query(ColumnMappingProfile).count() == expected_profiles
    assert mapping_db.query(ColumnMappingField).count() == (7 if expected_profiles else 0)
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 0


def test_rejection_cannot_materialize(mapping_db: Session):
    seeded = _seed(mapping_db)
    proposal = _propose(mapping_db, seeded).decision
    rejection = reject_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        command_id=uuid.uuid4(),
        reason_code="wrong_mapping",
    )
    with pytest.raises(HTTPException) as exc:
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=rejection.id,
            command_id=uuid.uuid4(),
            storage=FakeObjectStorage(),
        )
    assert exc.value.detail["error_code"] == "mapping_confirmation_required"
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 0


def test_full_source_materialization_exceeds_preview_and_filters_non_assets(mapping_db: Session):
    storage = FakeObjectStorage()
    seeded = _seed(mapping_db, storage=storage, asset_rows=230)
    proposal = _propose(mapping_db, seeded).decision
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        memory_scope="none",
        command_id=uuid.uuid4(),
    )
    usage = materialize_confirmed_mapping_to_staging(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        confirmation_decision_id=confirmation.id,
        command_id=uuid.uuid4(),
        storage=storage,
    )
    assert usage.materialized_asset_row_count == 230
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 230
    assert mapping_db.query(ProjectAssetLine).count() == 0
    batch = mapping_db.get(ProjectAssetImportBatch, seeded["batch"].id)
    assert batch.status == ImportBatchStatus.PARSED
    assert (batch.total_rows, batch.valid_rows, batch.invalid_rows, batch.warning_rows) == (230, 0, 0, 0)
    assert mapping_db.query(ColumnMappingProfileUsage).count() == 1
    assert mapping_db.query(AuditEvent).filter(
        AuditEvent.event_name == "ConfirmedMappingMaterialized"
    ).count() == 1


def test_cross_tenant_and_tampered_snapshot_fail_closed(mapping_db: Session):
    seeded = _seed(mapping_db)
    other_org = OrganizationProfile(
        legal_name="Other", organization_slug=f"other-{uuid.uuid4().hex[:8]}", status="active"
    )
    mapping_db.add(other_org)
    mapping_db.flush()
    other_user = User(
        organization_id=other_org.id,
        email=f"other-{uuid.uuid4().hex[:8]}@example.com",
        full_name="Other",
        status="active",
    )
    mapping_db.add(other_user)
    mapping_db.commit()
    with pytest.raises(HTTPException) as cross:
        propose_column_mapping(
            mapping_db,
            actor=other_user,
            org_id=other_org.id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            artifact_id=seeded["artifact"].id,
            snapshot_id=seeded["snapshot"].id,
            candidate_index=0,
            command_id=uuid.uuid4(),
        )
    assert cross.value.status_code == 404
    seeded["snapshot"].analysis_digest_sha256 = "f" * 64
    mapping_db.commit()
    with pytest.raises(HTTPException) as tampered:
        _propose(mapping_db, seeded)
    assert tampered.value.detail["error_code"] == "mapping_structure_integrity_failure"


def test_applied_batch_cannot_be_proposed_or_materialized(mapping_db: Session):
    seeded = _seed(mapping_db)
    seeded["batch"].status = ImportBatchStatus.APPLIED
    mapping_db.commit()
    with pytest.raises(HTTPException) as exc:
        _propose(mapping_db, seeded)
    assert exc.value.detail["error_code"] == "mapping_batch_already_applied"


def test_exact_customer_profile_prefills_and_reuses_without_new_version(mapping_db: Session):
    first = _seed(mapping_db)
    first_proposal = _propose(mapping_db, first).decision
    first_confirmation = confirm_column_mapping(
        mapping_db,
        actor=first["user"],
        org_id=first["org"].id,
        project_id=first["project"].id,
        batch_id=first["batch"].id,
        proposal_decision_id=first_proposal.id,
        mapping_snapshot=first_proposal.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    second = _seed(
        mapping_db,
        org=first["org"],
        user=first["user"],
        customer=first["customer"],
    )
    second_proposal = _propose(mapping_db, second)
    assert second_proposal.exact_profile_id == first_confirmation.profile_id
    assert second_proposal.similar_profile_ids == ()
    second_confirmation = confirm_column_mapping(
        mapping_db,
        actor=second["user"],
        org_id=second["org"].id,
        project_id=second["project"].id,
        batch_id=second["batch"].id,
        proposal_decision_id=second_proposal.decision.id,
        mapping_snapshot=second_proposal.decision.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    assert second_confirmation.profile_id == first_confirmation.profile_id
    assert mapping_db.query(ColumnMappingProfile).count() == 1


def test_similar_customer_candidate_is_review_only_and_beats_org_template(mapping_db: Session):
    first = _seed(mapping_db)
    proposal = _propose(mapping_db, first).decision
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=first["user"],
        org_id=first["org"].id,
        project_id=first["project"].id,
        batch_id=first["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    second = _seed(
        mapping_db,
        org=first["org"],
        user=first["user"],
        customer=first["customer"],
    )
    reordered = [
        "TÊN VẬT TƯ",
        "STT",
        "Đặc điểm",
        "ĐVT",
        "KHỐI LƯỢNG",
        "GIÁ TĐ",
        None,
    ]
    _replace_candidate(mapping_db, second, reordered)
    result = _propose(mapping_db, second)
    assert result.exact_profile_id is None
    assert result.similar_profile_ids == (confirmation.profile_id,)
    assert "profile_not_selected" in result.review_reasons
    assert result.decision.profile_id is None


def test_approved_organization_template_is_retrieval_only_and_requires_confirmation(
    mapping_db: Session,
):
    first = _seed(mapping_db)
    proposal = _propose(mapping_db, first).decision
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=first["user"],
        org_id=first["org"].id,
        project_id=first["project"].id,
        batch_id=first["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    profile = mapping_db.get(ColumnMappingProfile, confirmation.profile_id)
    user_id = first["user"].id
    profile.approved_by_user_id = user_id
    profile.approved_at = datetime.now(timezone.utc)
    profile.scope_type = "organization_template"
    profile.customer_id = None
    mapping_db.commit()

    second = _seed(
        mapping_db,
        org=first["org"],
        user=first["user"],
        customer=first["customer"],
    )
    result = _propose(mapping_db, second)
    assert result.organization_template_id == profile.id
    assert result.exact_profile_id is None
    assert result.decision.profile_id is None
    assert "approved_organization_template_prefill" in result.review_reasons
    assert not hasattr(mapping_service, "publish_organization_template")


def test_profiles_never_cross_customer_scope(mapping_db: Session):
    first = _seed(mapping_db)
    proposal = _propose(mapping_db, first).decision
    confirm_column_mapping(
        mapping_db,
        actor=first["user"],
        org_id=first["org"].id,
        project_id=first["project"].id,
        batch_id=first["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    other_customer = Customer(
        organization_id=first["org"].id,
        legal_name="Other customer",
        status="active",
        created_by=first["user"].id,
    )
    mapping_db.add(other_customer)
    mapping_db.commit()
    second = _seed(
        mapping_db,
        org=first["org"],
        user=first["user"],
        customer=other_customer,
    )
    result = _propose(mapping_db, second)
    assert result.exact_profile_id is None
    assert result.similar_profile_ids == ()


def test_profile_correction_versions_without_rewriting_prior_evidence(mapping_db: Session):
    seeded = _seed(mapping_db)
    first_proposal = _propose(mapping_db, seeded).decision
    first_confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=first_proposal.id,
        mapping_snapshot=first_proposal.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    original_profile_id = first_confirmation.profile_id
    original_fields = [
        (row.source_column_index, row.semantic_role)
        for row in mapping_db.query(ColumnMappingField)
        .filter(ColumnMappingField.profile_id == original_profile_id)
        .order_by(ColumnMappingField.source_column_index)
    ]
    second_proposal = _propose(mapping_db, seeded).decision
    corrected = deepcopy(second_proposal.mapping_snapshot)
    corrected["fields"][2]["semantic_role"] = "ignore"
    corrected["fields"][6]["semantic_role"] = "raw_description"
    corrected_digest = mapping_digest(corrected)
    assert corrected_digest != second_proposal.mapping_digest_sha256
    second_confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=second_proposal.id,
        mapping_snapshot=corrected,
        memory_scope="customer",
        supersedes_profile_id=original_profile_id,
        command_id=uuid.uuid4(),
    )
    old_profile = mapping_db.get(ColumnMappingProfile, original_profile_id)
    new_profile = mapping_db.get(ColumnMappingProfile, second_confirmation.profile_id)
    assert (old_profile.status, old_profile.profile_version) == ("superseded", 1)
    assert (new_profile.status, new_profile.profile_version) == ("active", 2)
    assert new_profile.supersedes_profile_id == old_profile.id
    assert [
        (row.source_column_index, row.semantic_role)
        for row in mapping_db.query(ColumnMappingField)
        .filter(ColumnMappingField.profile_id == original_profile_id)
        .order_by(ColumnMappingField.source_column_index)
    ] == original_fields
    assert mapping_db.query(ColumnMappingDecision).count() == 4


def test_materialization_spool_failure_rolls_back_prior_staging_and_audit(
    mapping_db: Session, monkeypatch
):
    storage = FakeObjectStorage()
    seeded = _seed(mapping_db, storage=storage, asset_rows=3)
    old = ProjectAssetImportStagingRow(
        organization_id=seeded["org"].id,
        project_id=seeded["project"].id,
        import_batch_id=seeded["batch"].id,
        source_row_number=999,
        raw_values={"old": True},
        mapped_values={},
        normalized_preview={},
        validation_status="valid",
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="Old",
    )
    mapping_db.add(old)
    seeded["batch"].status = ImportBatchStatus.READY_FOR_REVIEW
    seeded["batch"].total_rows = 1
    seeded["batch"].valid_rows = 1
    mapping_db.commit()
    proposal = _propose(mapping_db, seeded).decision
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        memory_scope="none",
        command_id=uuid.uuid4(),
    )
    original_reader = mapping_service._read_spool

    def failing_reader(path: str, *, expected_digest: str):
        rows = original_reader(path, expected_digest=expected_digest)
        yield next(rows)
        raise RuntimeError("injected spool read failure")

    monkeypatch.setattr(mapping_service, "_read_spool", failing_reader)
    with pytest.raises(RuntimeError, match="injected spool"):
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=confirmation.id,
            command_id=uuid.uuid4(),
            storage=storage,
        )
    rows = mapping_db.query(ProjectAssetImportStagingRow).all()
    assert len(rows) == 1 and rows[0].id == old.id and rows[0].raw_values == {"old": True}
    batch = mapping_db.get(ProjectAssetImportBatch, seeded["batch"].id)
    assert (batch.status, batch.total_rows, batch.valid_rows) == (
        ImportBatchStatus.READY_FOR_REVIEW,
        1,
        1,
    )
    assert mapping_db.query(ColumnMappingProfileUsage).count() == 0
    assert mapping_db.query(AuditEvent).filter(
        AuditEvent.event_name == "ConfirmedMappingMaterialized"
    ).count() == 0


def test_source_pointer_and_object_checksum_drift_fail_without_success_evidence(
    mapping_db: Session,
):
    storage = FakeObjectStorage()
    seeded = _seed(mapping_db, storage=storage)
    proposal = _propose(mapping_db, seeded).decision
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        memory_scope="none",
        command_id=uuid.uuid4(),
    )
    storage._objects[seeded["artifact"].storage_object_key] += b"tamper"
    with pytest.raises(HTTPException) as checksum:
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=confirmation.id,
            command_id=uuid.uuid4(),
            storage=storage,
        )
    assert checksum.value.detail["error_code"] == "mapping_materialization_stale"
    assert mapping_db.query(ColumnMappingProfileUsage).count() == 0
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 0


def test_xls_xlsx_materialization_parity_for_identical_value_only_content(mapping_db: Session):
    storage = FakeObjectStorage()
    outputs = []
    for workbook_format in ("xlsx", "xls"):
        seeded = _seed(
            mapping_db,
            storage=storage,
            asset_rows=4,
            workbook_format=workbook_format,
        )
        proposal = _propose(mapping_db, seeded).decision
        confirmation = confirm_column_mapping(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            proposal_decision_id=proposal.id,
            mapping_snapshot=proposal.mapping_snapshot,
            memory_scope="none",
            command_id=uuid.uuid4(),
        )
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=confirmation.id,
            command_id=uuid.uuid4(),
            storage=storage,
        )
        rows = (
            mapping_db.query(ProjectAssetImportStagingRow)
            .filter(ProjectAssetImportStagingRow.import_batch_id == seeded["batch"].id)
            .order_by(ProjectAssetImportStagingRow.source_row_number)
            .all()
        )
        outputs.append(
            [
                (
                    row.source_row_number,
                    row.raw_values,
                    row.mapped_values,
                    row.proposed_asset_name,
                    row.proposed_quantity,
                    row.proposed_appraised_unit_price,
                )
                for row in rows
            ]
        )
    assert outputs[0] == outputs[1]


def _attacker_reseal_mapping_snapshot(snapshot: dict) -> None:
    candidate = snapshot["candidate"]
    sealed_shape = {
        "sheet_name": candidate["sheet_name"],
        "header_start_row": candidate["header_start_row"],
        "header_end_row": candidate["header_end_row"],
        "data_start_row": candidate["data_start_row"],
        "candidate_table_bounds": {
            "min_row": candidate["min_row"],
            "max_row": candidate["max_row"],
            "min_column": candidate["min_column"],
            "max_column": candidate["max_column"],
        },
        "header_labels": [field["original_header"] for field in snapshot["fields"]],
    }
    snapshot["template_fingerprint_sha256"] = build_template_fingerprint(
        sealed_shape, rule_version=snapshot["structure"]["rule_version"]
    )


@pytest.mark.parametrize("tamper", ["sheet", "bounds", "headers"])
def test_confirmation_rebuilds_canonical_snapshot_and_rejects_candidate_tamper(
    mapping_db: Session, tamper: str
):
    seeded = _seed(mapping_db)
    proposal = _propose(mapping_db, seeded).decision
    attacked = deepcopy(proposal.mapping_snapshot)
    if tamper == "sheet":
        attacked["candidate"]["sheet_name"] = "ATTACKER-SHEET"
    elif tamper == "bounds":
        attacked["candidate"]["max_row"] += 1
    else:
        attacked["fields"][1]["original_header"] = "ATTACKER HEADER"
    _attacker_reseal_mapping_snapshot(attacked)

    with pytest.raises(HTTPException) as caught:
        confirm_column_mapping(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            proposal_decision_id=proposal.id,
            mapping_snapshot=attacked,
            command_id=uuid.uuid4(),
        )
    assert caught.value.status_code == 409
    assert caught.value.detail["error_code"] == "mapping_proposal_not_current"
    assert (
        mapping_db.query(ColumnMappingDecision)
        .filter(ColumnMappingDecision.decision_kind == "confirmation")
        .count()
        == 0
    )


def test_confirmation_persists_server_rebuilt_snapshot_not_caller_alias(mapping_db: Session):
    seeded = _seed(mapping_db)
    proposal = _propose(mapping_db, seeded).decision
    submitted = deepcopy(proposal.mapping_snapshot)
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=submitted,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    submitted["candidate"]["sheet_name"] = "MUTATED-AFTER-CALL"
    submitted["fields"][1]["original_header"] = "MUTATED-AFTER-CALL"
    mapping_db.expire_all()
    persisted = mapping_db.get(ColumnMappingDecision, confirmation.id)
    assert persisted.mapping_snapshot == proposal.mapping_snapshot
    assert persisted.mapping_digest_sha256 == mapping_digest(proposal.mapping_snapshot)
    profile = mapping_db.get(ColumnMappingProfile, persisted.profile_id)
    assert profile.sheet_name == seeded["candidate"]["sheet_name"]
    assert profile.mapping_digest_sha256 == profile_mapping_digest(
        seeded["candidate"], validate_mapping_snapshot(persisted.mapping_snapshot)
    )


class _MidIOProbeStorage(FakeObjectStorage):
    def __init__(self):
        super().__init__()
        self.on_open = None
        self.open_count = 0

    def open_stream(self, key: str):
        stream = super().open_stream(key)
        self.open_count += 1
        if self.on_open is not None:
            callback, self.on_open = self.on_open, None
            callback()
        return stream


def _confirmed_seed(db: Session, storage: FakeObjectStorage) -> tuple[dict, ColumnMappingDecision]:
    seeded = _seed(db, storage=storage)
    proposal = _propose(db, seeded).decision
    confirmation = confirm_column_mapping(
        db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        command_id=uuid.uuid4(),
    )
    return seeded, confirmation


@pytest.mark.parametrize(
    ("drift", "error_code"),
    [
        ("decision_digest", "mapping_materialization_stale"),
        ("decision_snapshot", "mapping_materialization_stale"),
        ("customer", "mapping_materialization_stale"),
        ("source_pointer", "mapping_materialization_stale"),
        ("structure_payload", "mapping_materialization_stale"),
        ("structure_seal", "mapping_materialization_stale"),
        ("applied_batch", "mapping_batch_already_applied"),
    ],
)
def test_materialization_requeries_frozen_primitives_after_object_io(
    mapping_db: Session, drift: str, error_code: str
):
    storage = _MidIOProbeStorage()
    seeded, confirmation = _confirmed_seed(mapping_db, storage)
    alternate = None
    if drift == "customer":
        alternate = Customer(
            organization_id=seeded["org"].id,
            legal_name="Drift Customer",
            status=CustomerStatus.ACTIVE,
            created_by=seeded["user"].id,
        )
        mapping_db.add(alternate)
        mapping_db.commit()

    def inject_drift() -> None:
        if drift == "decision_digest":
            confirmation.mapping_digest_sha256 = "f" * 64
        elif drift == "decision_snapshot":
            value = deepcopy(confirmation.mapping_snapshot)
            value["candidate"]["sheet_name"] = "DRIFTED"
            confirmation.mapping_snapshot = value
        elif drift == "customer":
            seeded["project"].customer_id = alternate.id
        elif drift == "source_pointer":
            seeded["batch"].current_source_artifact_id = None
        elif drift == "structure_payload":
            value = deepcopy(seeded["snapshot"].structure_payload)
            value["candidate_count"] = 2
            seeded["snapshot"].structure_payload = value
        elif drift == "structure_seal":
            seeded["snapshot"].analysis_digest_sha256 = "e" * 64
        else:
            seeded["batch"].status = ImportBatchStatus.APPLIED
        mapping_db.flush()

    storage.on_open = inject_drift
    with pytest.raises(HTTPException) as caught:
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=confirmation.id,
            command_id=uuid.uuid4(),
            storage=storage,
        )
    assert caught.value.detail["error_code"] == error_code
    assert mapping_db.query(ColumnMappingProfileUsage).count() == 0
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 0
    assert (
        mapping_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "ConfirmedMappingMaterialized")
        .count()
        == 0
    )
    mapping_db.rollback()


@pytest.mark.parametrize("inactive_target", ["actor", "organization"])
def test_active_actor_and_organization_are_reloaded_before_idempotency(
    mapping_db: Session, inactive_target: str
):
    seeded = _seed(mapping_db)
    command_id = uuid.uuid4()
    propose_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        artifact_id=seeded["artifact"].id,
        snapshot_id=seeded["snapshot"].id,
        candidate_index=0,
        command_id=command_id,
    )
    if inactive_target == "actor":
        seeded["user"].status = UserStatus.INACTIVE
    else:
        seeded["org"].status = OrganizationStatus.INACTIVE
    mapping_db.commit()
    with pytest.raises(HTTPException) as caught:
        propose_column_mapping(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            artifact_id=seeded["artifact"].id,
            snapshot_id=seeded["snapshot"].id,
            candidate_index=0,
            command_id=command_id,
        )
    assert caught.value.status_code == 404


def test_canonical_idempotency_binds_actor_operation_and_all_command_inputs(
    mapping_db: Session
):
    storage = FakeObjectStorage()
    seeded = _seed(mapping_db, storage=storage)
    second_actor = User(
        organization_id=seeded["org"].id,
        email=f"second-{uuid.uuid4().hex[:8]}@example.com",
        full_name="Second Actor",
        status=UserStatus.ACTIVE,
    )
    mapping_db.add(second_actor)
    mapping_db.commit()

    proposal_command = uuid.uuid4()
    proposal = propose_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        artifact_id=seeded["artifact"].id,
        snapshot_id=seeded["snapshot"].id,
        candidate_index=0,
        command_id=proposal_command,
    ).decision
    with pytest.raises(HTTPException) as actor_reuse:
        propose_column_mapping(
            mapping_db,
            actor=second_actor,
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            artifact_id=seeded["artifact"].id,
            snapshot_id=seeded["snapshot"].id,
            candidate_index=0,
            command_id=proposal_command,
        )
    assert actor_reuse.value.detail["error_code"] == "idempotency_key_reused"

    confirmation_command = uuid.uuid4()
    confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        command_id=confirmation_command,
    )
    same_confirmation = confirm_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=deepcopy(proposal.mapping_snapshot),
        command_id=confirmation_command,
    )
    assert same_confirmation.id == confirmation.id
    with pytest.raises(HTTPException) as scope_reuse:
        confirm_column_mapping(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            proposal_decision_id=proposal.id,
            mapping_snapshot=proposal.mapping_snapshot,
            memory_scope="customer",
            command_id=confirmation_command,
        )
    assert scope_reuse.value.detail["error_code"] == "idempotency_key_reused"

    usage_command = uuid.uuid4()
    usage = materialize_confirmed_mapping_to_staging(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        confirmation_decision_id=confirmation.id,
        command_id=usage_command,
        storage=storage,
    )
    same_usage = materialize_confirmed_mapping_to_staging(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        confirmation_decision_id=confirmation.id,
        command_id=usage_command,
        storage=storage,
    )
    assert same_usage.id == usage.id
    with pytest.raises(HTTPException) as usage_actor_reuse:
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=second_actor,
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=confirmation.id,
            command_id=usage_command,
            storage=storage,
        )
    assert usage_actor_reuse.value.detail["error_code"] == "idempotency_key_reused"


def test_rejection_idempotency_binds_actor_project_batch_proposal_and_reason(mapping_db: Session):
    seeded = _seed(mapping_db)
    proposal = _propose(mapping_db, seeded).decision
    command_id = uuid.uuid4()
    rejection = reject_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        reason_code="header_wrong",
        reason_text="human checked",
        command_id=command_id,
    )
    retry = reject_column_mapping(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        proposal_decision_id=proposal.id,
        reason_code="header_wrong",
        reason_text="human checked",
        command_id=command_id,
    )
    assert retry.id == rejection.id
    with pytest.raises(HTTPException) as changed_reason:
        reject_column_mapping(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            proposal_decision_id=proposal.id,
            reason_code="other",
            reason_text="human checked",
            command_id=command_id,
        )
    assert changed_reason.value.detail["error_code"] == "idempotency_key_reused"


def test_unrelated_profile_cannot_be_superseded_and_no_authority_starts_family_v1(
    mapping_db: Session
):
    first = _seed(mapping_db)
    first_proposal = _propose(mapping_db, first).decision
    first_confirmation = confirm_column_mapping(
        mapping_db,
        actor=first["user"],
        org_id=first["org"].id,
        project_id=first["project"].id,
        batch_id=first["batch"].id,
        proposal_decision_id=first_proposal.id,
        mapping_snapshot=first_proposal.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    unrelated_profile = mapping_db.get(ColumnMappingProfile, first_confirmation.profile_id)

    second = _seed(
        mapping_db,
        org=first["org"],
        user=first["user"],
        customer=first["customer"],
    )
    _replace_candidate(
        mapping_db,
        second,
        ["STT", "TÊN VẬT TƯ", "SERIAL", "MODEL", "NĂM", "XUẤT XỨ", "NOTE X"],
    )
    proposal = _propose(mapping_db, second)
    assert proposal.exact_profile_id is None
    assert proposal.similar_profile_ids == ()
    with pytest.raises(HTTPException) as unrelated:
        confirm_column_mapping(
            mapping_db,
            actor=second["user"],
            org_id=second["org"].id,
            project_id=second["project"].id,
            batch_id=second["batch"].id,
            proposal_decision_id=proposal.decision.id,
            mapping_snapshot=proposal.decision.mapping_snapshot,
            memory_scope="customer",
            supersedes_profile_id=unrelated_profile.id,
            command_id=uuid.uuid4(),
        )
    assert unrelated.value.detail["error_code"] == "mapping_profile_stale"
    created = confirm_column_mapping(
        mapping_db,
        actor=second["user"],
        org_id=second["org"].id,
        project_id=second["project"].id,
        batch_id=second["batch"].id,
        proposal_decision_id=proposal.decision.id,
        mapping_snapshot=proposal.decision.mapping_snapshot,
        memory_scope="customer",
        command_id=uuid.uuid4(),
    )
    profile = mapping_db.get(ColumnMappingProfile, created.profile_id)
    assert profile.profile_version == 1
    assert profile.profile_family_id == profile.id
    assert unrelated_profile.status == "active"


def test_materialization_exact_audit_privacy_cardinality_and_no_apply_side_effect(
    mapping_db: Session
):
    storage = FakeObjectStorage()
    seeded, confirmation = _confirmed_seed(mapping_db, storage)
    asset_lines_before = mapping_db.query(ProjectAssetLine).count()
    usage = materialize_confirmed_mapping_to_staging(
        mapping_db,
        actor=seeded["user"],
        org_id=seeded["org"].id,
        project_id=seeded["project"].id,
        batch_id=seeded["batch"].id,
        confirmation_decision_id=confirmation.id,
        command_id=uuid.uuid4(),
        storage=storage,
    )
    assert mapping_db.query(ProjectAssetLine).count() == asset_lines_before
    assert seeded["batch"].status == ImportBatchStatus.PARSED
    rows = mapping_db.query(ProjectAssetImportStagingRow).all()
    assert len(rows) == usage.materialized_asset_row_count == 3
    assert {row.validation_status for row in rows} == {ImportRowValidationStatus.PENDING}
    event = (
        mapping_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "ConfirmedMappingMaterialized")
        .one()
    )
    assert set(event.payload) == {
        "organization_id",
        "project_id",
        "batch_id",
        "source_artifact_id",
        "structure_snapshot_id",
        "decision_id",
        "profile_id",
        "usage_id",
        "mapping_contract_version",
        "materialization_contract_version",
        "template_fingerprint_sha256",
        "mapping_digest_sha256",
        "source_generation",
        "materialized_asset_row_count",
    }
    serialized = str(event.payload)
    assert "TÊN VẬT TƯ" not in serialized
    assert "Tài sản 1" not in serialized


@pytest.mark.parametrize("failure", ["adapter", "read", "close"])
def test_adapter_read_close_failures_leave_no_success_evidence(
    mapping_db: Session, monkeypatch: pytest.MonkeyPatch, failure: str
):
    storage = FakeObjectStorage()
    seeded, confirmation = _confirmed_seed(mapping_db, storage)
    if failure == "adapter":
        def fail_adapter(*args, **kwargs):
            raise AdapterError(422, "probe", "adapter probe")

        monkeypatch.setattr(mapping_service, "detect_format_and_adapter", fail_adapter)
    elif failure == "read":
        storage.fail_open_stream = True
    else:
        original_detect = mapping_service.detect_format_and_adapter

        def close_failing_adapter(*args, **kwargs):
            detected, adapter = original_detect(*args, **kwargs)
            original_close = adapter.close

            def fail_close():
                original_close()
                raise RuntimeError("close probe")

            adapter.close = fail_close
            return detected, adapter

        monkeypatch.setattr(
            mapping_service, "detect_format_and_adapter", close_failing_adapter
        )
    with pytest.raises((HTTPException, RuntimeError)):
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=confirmation.id,
            command_id=uuid.uuid4(),
            storage=storage,
        )
    assert mapping_db.query(ColumnMappingProfileUsage).count() == 0
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 0
    assert (
        mapping_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "ConfirmedMappingMaterialized")
        .count()
        == 0
    )


@pytest.mark.parametrize("failure", ["flush", "savepoint_release", "outer_commit"])
def test_transaction_failure_injection_rolls_back_and_cleans_spool(
    mapping_db: Session, monkeypatch: pytest.MonkeyPatch, failure: str
):
    storage = FakeObjectStorage()
    seeded, confirmation = _confirmed_seed(mapping_db, storage)
    spool_paths: list[str] = []
    original_write_spool = mapping_service._write_spool

    def recording_write_spool(rows):
        result = original_write_spool(rows)
        spool_paths.append(result[0])
        return result

    monkeypatch.setattr(mapping_service, "_write_spool", recording_write_spool)
    if failure == "flush":
        original_flush = mapping_db.flush
        original_begin_nested = mapping_db.begin_nested
        flush_calls = 0
        savepoint_started = False
        flush_failed = False

        def begin_nested_for_flush_probe():
            nonlocal savepoint_started
            transaction = original_begin_nested()
            savepoint_started = True
            return transaction

        def fail_flush(*args, **kwargs):
            nonlocal flush_calls, flush_failed
            flush_calls += 1
            if savepoint_started and not flush_failed:
                flush_failed = True
                raise RuntimeError("flush probe")
            return original_flush(*args, **kwargs)

        monkeypatch.setattr(mapping_db, "begin_nested", begin_nested_for_flush_probe)
        monkeypatch.setattr(mapping_db, "flush", fail_flush)
    elif failure == "savepoint_release":
        original_begin_nested = mapping_db.begin_nested

        class _SavepointProbe:
            def __init__(self, transaction):
                self.transaction = transaction

            def commit(self):
                raise RuntimeError("savepoint release probe")

            def rollback(self):
                return self.transaction.rollback()

        def begin_nested_probe():
            return _SavepointProbe(original_begin_nested())

        monkeypatch.setattr(mapping_db, "begin_nested", begin_nested_probe)
    else:
        def fail_commit():
            raise RuntimeError("outer commit probe")

        monkeypatch.setattr(mapping_db, "commit", fail_commit)

    with pytest.raises(RuntimeError, match="probe"):
        materialize_confirmed_mapping_to_staging(
            mapping_db,
            actor=seeded["user"],
            org_id=seeded["org"].id,
            project_id=seeded["project"].id,
            batch_id=seeded["batch"].id,
            confirmation_decision_id=confirmation.id,
            command_id=uuid.uuid4(),
            storage=storage,
        )
    assert spool_paths and all(not os.path.exists(path) for path in spool_paths)
    assert mapping_db.query(ColumnMappingProfileUsage).count() == 0
    assert mapping_db.query(ProjectAssetImportStagingRow).count() == 0
    assert (
        mapping_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "ConfirmedMappingMaterialized")
        .count()
        == 0
    )
    mapping_db.expire_all()
    batch = mapping_db.get(ProjectAssetImportBatch, seeded["batch"].id)
    assert batch.status == ImportBatchStatus.CREATED


def test_proposal_confirmation_rejection_audits_have_exact_private_cardinalities(
    mapping_db: Session
):
    confirmed_seed = _seed(mapping_db)
    proposal = _propose(mapping_db, confirmed_seed).decision
    confirm_column_mapping(
        mapping_db,
        actor=confirmed_seed["user"],
        org_id=confirmed_seed["org"].id,
        project_id=confirmed_seed["project"].id,
        batch_id=confirmed_seed["batch"].id,
        proposal_decision_id=proposal.id,
        mapping_snapshot=proposal.mapping_snapshot,
        command_id=uuid.uuid4(),
    )
    rejected_seed = _seed(
        mapping_db,
        org=confirmed_seed["org"],
        user=confirmed_seed["user"],
        customer=confirmed_seed["customer"],
    )
    rejected_proposal = _propose(mapping_db, rejected_seed).decision
    reject_column_mapping(
        mapping_db,
        actor=rejected_seed["user"],
        org_id=rejected_seed["org"].id,
        project_id=rejected_seed["project"].id,
        batch_id=rejected_seed["batch"].id,
        proposal_decision_id=rejected_proposal.id,
        command_id=uuid.uuid4(),
        reason_code="human_rejected",
        reason_text="private explanation",
    )
    expected_keys = {
        "ColumnMappingProposed": {
            "organization_id",
            "project_id",
            "batch_id",
            "source_artifact_id",
            "structure_snapshot_id",
            "decision_id",
            "mapping_contract_version",
            "rule_version",
            "source_generation",
            "template_fingerprint_sha256",
            "mapping_digest_sha256",
            "outcome",
            "role_counts",
        },
        "ColumnMappingConfirmed": {
            "organization_id",
            "project_id",
            "batch_id",
            "source_artifact_id",
            "structure_snapshot_id",
            "decision_id",
            "profile_id",
            "profile_version",
            "mapping_contract_version",
            "template_fingerprint_sha256",
            "mapping_digest_sha256",
            "source_generation",
            "outcome",
            "role_counts",
        },
        "ColumnMappingRejected": {
            "organization_id",
            "project_id",
            "batch_id",
            "source_artifact_id",
            "structure_snapshot_id",
            "decision_id",
            "mapping_contract_version",
            "template_fingerprint_sha256",
            "mapping_digest_sha256",
            "source_generation",
            "outcome",
        },
    }
    for event_name, keys in expected_keys.items():
        events = mapping_db.query(AuditEvent).filter(AuditEvent.event_name == event_name).all()
        assert events
        for event in events:
            assert set(event.payload) == keys
            serialized = str(event.payload)
            assert "TÊN VẬT TƯ" not in serialized
            assert "private explanation" not in serialized
            if "role_counts" in event.payload:
                assert sum(event.payload["role_counts"].values()) == 7


def test_composite_tenant_lineage_constraints_are_present_in_model_metadata():
    expected = {
        "users": {"uq_s13_pr004_user_tenant_id"},
        "customers": {"uq_s13_pr004_customer_tenant_id"},
        "projects": {"uq_s13_pr004_project_tenant_customer_id"},
        "import_source_artifacts": {"fk_source_artifact_creator_tenant"},
        "workbook_structure_snapshots": {
            "uq_workbook_structure_tenant_source_id",
            "fk_workbook_structure_creator_tenant",
        },
        "column_mapping_profiles": {
            "fk_mapping_profile_customer_tenant",
            "fk_mapping_profile_source_customer_tenant",
            "fk_mapping_profile_source_project_tenant",
            "fk_mapping_profile_source_batch_tenant",
            "fk_mapping_profile_source_artifact_tenant",
            "fk_mapping_profile_structure_tenant",
            "fk_mapping_profile_confirmer_tenant",
        },
        "column_mapping_decisions": {
            "uq_mapping_decision_tenant_lineage_id",
            "fk_mapping_decision_customer_tenant",
            "fk_mapping_decision_project_customer_tenant",
            "fk_mapping_decision_artifact_tenant",
            "fk_mapping_decision_structure_tenant",
            "fk_mapping_decision_proposal_lineage",
            "fk_mapping_decision_actor_tenant",
        },
        "column_mapping_profile_usages": {
            "fk_mapping_usage_customer_tenant",
            "fk_mapping_usage_project_customer_tenant",
            "fk_mapping_usage_artifact_tenant",
            "fk_mapping_usage_structure_tenant",
            "fk_mapping_usage_confirmation_lineage",
            "fk_mapping_usage_creator_tenant",
        },
    }
    for table_name, constraint_names in expected.items():
        actual = {
            constraint.name for constraint in Base.metadata.tables[table_name].constraints
        }
        assert constraint_names <= actual
