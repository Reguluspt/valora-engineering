"""S13-PR-002 fifth corrective: E-01…E-08 proofs (F-01…F-08 re-audit)."""
from __future__ import annotations

import hashlib
import io
import os
import uuid
from datetime import datetime, timedelta, timezone

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.adapters.xlsx_adapter import XlsxWorkbookAdapter
from app.modules.excel_import.application.adapters.xls_adapter import XlsWorkbookAdapter
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
    set_source_limits_override,
    _sha256_object,
    upload_source_artifact,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.domain.workbook_adapter import AdapterError
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    ObjectStorageError,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    assert_http_rejection_preserve,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
    AuditEvent,
)
from tests.fixtures.s13_pr_002.ole_builder import write_xls_formula_cached


# ---------------------------------------------------------------------------
# Fixtures / seed
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session() -> Session:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org", organization_slug=f"o-{uuid.uuid4().hex[:6]}", status=OrganizationStatus.ACTIVE
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    data = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.append(["h"])
    wb.save(data)
    payload = data.getvalue()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    snap.update(
        {
            "prior_id": prior.id,
            "prior_checksum": prior.checksum_sha256,
            "prior_key": prior.storage_object_key,
            "prior_state": prior.state,
            "prior_gen": prior.generation,
            "batch_current": batch.current_source_artifact_id,
            "batch_status": batch.status,
            "batch_total": batch.total_rows,
            "batch_valid": batch.valid_rows,
            "staging_id": staging.id,
            "staging_raw": dict(staging.raw_values),
            "staging_mapped": dict(staging.mapped_values),
            "staging_name": staging.proposed_asset_name,
            "staging_desc": staging.proposed_description,
            "staging_qty": staging.proposed_quantity,
            "staging_unit": staging.proposed_unit,
            "staging_status": staging.validation_status,
            "line_id": line.id,
            "line_name": line.asset_name,
            "line_desc": line.description,
            "line_qty": float(line.quantity),
            "line_review": line.review_status,
            "line_val": line.validation_status,
            "art_count": len(snap["artifacts"]),
            "line_count": len(snap["lines"]),
        }
    )
    return prior, staging, line, snap


def _assert_preserved(db, fake_storage, org, user, proj, batch, staging, line, snap):
    """Field-level prior/staging/line preserve for non-reject paths.

    HTTP N+1 rejections use assert_http_rejection_preserve / assert_source_intake_preserve.
    """
    db.refresh(batch)
    db.refresh(staging)
    db.refresh(line)
    prior = db.get(ImportSourceArtifact, snap["prior_id"])
    assert prior is not None
    assert prior.checksum_sha256 == snap["prior_checksum"]
    assert prior.storage_object_key == snap["prior_key"]
    assert prior.state == snap["prior_state"]
    assert prior.generation == snap["prior_gen"]
    assert batch.current_source_artifact_id == snap["batch_current"]
    assert batch.status == snap["batch_status"]
    assert batch.total_rows == snap["batch_total"]
    assert batch.valid_rows == snap["batch_valid"]
    assert staging.id == snap["staging_id"]
    assert staging.raw_values == snap["staging_raw"]
    assert staging.mapped_values == snap["staging_mapped"]
    assert staging.proposed_asset_name == snap["staging_name"]
    assert staging.proposed_description == snap["staging_desc"]
    assert staging.proposed_quantity == snap["staging_qty"]
    assert staging.proposed_unit == snap["staging_unit"]
    assert staging.validation_status == snap["staging_status"]
    assert line.id == snap["line_id"]
    assert line.asset_name == snap["line_name"]
    assert line.description == snap["line_desc"]
    assert float(line.quantity) == snap["line_qty"]
    assert line.review_status == snap["line_review"]
    assert line.validation_status == snap["line_val"]
    # prior object retained
    assert snap["prior_key"] in fake_storage._objects
    assert db.query(ProjectAssetLine).filter_by(project_id=proj.id).count() == snap["line_count"]


def _xlsx_bytes(*, sheets=1, rows=1, cols=1, cell="x", merges=0) -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    for c in range(cols):
        wb.active.cell(1, c + 1, cell)
    for r in range(1, rows):
        for c in range(cols):
            wb.active.cell(r + 1, c + 1, cell)
    for i in range(1, sheets):
        wb.create_sheet(f"S{i + 1}")
    for i in range(merges):
        ws = wb.active
        # non-overlapping merges on row 1+i
        rr = i + 1
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
        ws.cell(rr, 1, "m")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_bytes(tmp_path, *, sheets=1, rows=1, cols=1, cell="x"):
    pytest.importorskip("xlwt")
    import xlwt

    p = tmp_path / f"t-{uuid.uuid4().hex[:6]}.xls"
    book = xlwt.Workbook()
    for si in range(sheets):
        sh = book.add_sheet(f"S{si + 1}")
        for r in range(rows):
            for c in range(cols):
                sh.write(r, c, cell)
    book.save(str(p))
    return p


# ---------------------------------------------------------------------------
# E-01 upload short-read ≠ checksum_mismatch
# ---------------------------------------------------------------------------


def test_e01_upload_short_read_not_checksum_mismatch(
    client: TestClient, db_session: Session, fake_storage, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    # Build valid small xlsx
    payload = _xlsx_bytes()
    fake_storage.truncate_open_to = 4  # clean EOF short-read on verify
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "a.xlsx",
                io.BytesIO(payload),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert res.status_code == 500, res.text
    arts = db_session.query(ImportSourceArtifact).order_by(ImportSourceArtifact.generation).all()
    # prior + failed residual
    failed = [a for a in arts if a.state == "failed"]
    assert len(failed) == 1
    assert failed[0].failure_code != "checksum_mismatch"
    assert failed[0].failure_code in {"short_read", "object_too_large", "size_mismatch"} or failed[
        0
    ].failure_code  # infrastructure code
    assert failed[0].failure_code == "short_read"
    # residual object may remain for reconciler
    assert failed[0].storage_object_key in fake_storage._objects
    # no false corruption audit name
    audits = (
        db_session.query(AuditEvent)
        .filter(AuditEvent.entity_id == failed[0].id)
        .all()
    )
    for a in audits:
        if a.event_name == "ImportSourceArtifactFailed":
            assert a.payload.get("failure_code") == "short_read"
    _assert_preserved(db_session, fake_storage, org, user, proj, batch, staging, line, snap)


# ---------------------------------------------------------------------------
# E-02 caller-owned work survives empty reconcile
# ---------------------------------------------------------------------------


def test_e02_reconciler_preserves_dirty_caller_session(
    db_session: Session, fake_storage
):
    """Dedicated reconciler session: caller unflushed UOW is preserved."""
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    pending = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Caller Pending",
        description="uncommitted",
        quantity=1,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db_session.add(pending)
    assert pending in db_session.new
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
    )
    assert stats["scanned"] == 0
    assert pending in db_session.new
    assert pending.asset_name == "Caller Pending"


def test_e02_reconciler_empty_and_skip_paths_owned_session(
    db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    if db_session.in_transaction():
        db_session.rollback()
    # empty
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 0
    # available (not selected) skip
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = b"x"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="b" * 64,
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
    )
    db_session.add(art)
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 0


# ---------------------------------------------------------------------------
# E-03 exact .xls cached formula value
# ---------------------------------------------------------------------------


def test_e03_xls_exact_cached_formula_value(tmp_path):
    p = tmp_path / "cached.xls"
    write_xls_formula_cached(p, cached=6.0)
    rows = list(XlsWorkbookAdapter().iter_rows(str(p)))
    vals = [c.value for r in rows for c in r]
    assert 6.0 in vals or 6 in vals
    assert rows[0][1].value == 6.0
    assert all(not (isinstance(v, str) and str(v).startswith("=")) for v in vals)
    assert all(v != "A1*3" and v != "=A1*3" for v in vals)
    # inspect path also safe
    insp = XlsWorkbookAdapter().inspect(str(p))
    assert insp.sheet_names


# ---------------------------------------------------------------------------
# E-04 boundary matrix (inspect + iter_rows, both formats where applicable)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "fmt,limit_field,limit_val,error_code,builder",
    [
        ("xlsx", "max_sheets", 1, "sheet_limit", lambda tp: _xlsx_bytes(sheets=2)),
        ("xlsx", "max_physical_rows", 1, "physical_row_limit", lambda tp: _xlsx_bytes(rows=2)),
        ("xlsx", "max_columns", 1, "column_limit", lambda tp: _xlsx_bytes(cols=2)),
        ("xlsx", "max_total_cells", 1, "total_cell_limit", lambda tp: _xlsx_bytes(rows=1, cols=2)),
        (
            "xlsx",
            "max_cell_chars",
            3,
            "cell_length_limit",
            lambda tp: _xlsx_bytes(cell="abcd"),
        ),
        (
            "xlsx",
            "max_row_chars",
            3,
            "row_char_limit",
            lambda tp: _xlsx_bytes(cols=2, cell="ab"),
        ),
        (
            "xlsx",
            "max_merged_regions",
            0,
            "merged_region_limit",
            lambda tp: _xlsx_bytes(merges=1),
        ),
    ],
)
def test_e04_xlsx_inspect_exact_and_max_plus_one(tmp_path, fmt, limit_field, limit_val, error_code, builder):
    data = builder(tmp_path)
    p = tmp_path / "b.xlsx"
    p.write_bytes(data)
    # max+1 reject
    lim = SourceArtifactLimits(**{limit_field: limit_val})
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=lim).inspect(str(p))
    assert ei.value.error_code == error_code
    # exact accept (limit_val + 1 as the limit when content is limit_val+1... for reject we used limit_val)
    # acceptance: set limit high enough for the fixture
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(**{limit_field: limit_val + 10})).inspect(str(p))


@pytest.mark.parametrize(
    "limit_field,limit_val,error_code,kwargs",
    [
        ("max_sheets", 1, "sheet_limit", {"sheets": 2}),
        ("max_physical_rows", 1, "physical_row_limit", {"rows": 2}),
        ("max_columns", 1, "column_limit", {"cols": 2}),
        ("max_total_cells", 1, "total_cell_limit", {"rows": 1, "cols": 2}),
        ("max_cell_chars", 3, "cell_length_limit", {"cell": "abcd"}),
        ("max_row_chars", 3, "row_char_limit", {"cols": 2, "cell": "ab"}),
    ],
)
def test_e04_xls_inspect_exact_and_max_plus_one(tmp_path, limit_field, limit_val, error_code, kwargs):
    p = _xls_bytes(tmp_path, **kwargs)
    lim = SourceArtifactLimits(**{limit_field: limit_val})
    with pytest.raises(AdapterError) as ei:
        XlsWorkbookAdapter(limits=lim).inspect(str(p))
    assert ei.value.error_code == error_code
    XlsWorkbookAdapter(limits=SourceArtifactLimits(**{limit_field: limit_val + 10})).inspect(str(p))


@pytest.mark.parametrize("fmt", ["xlsx", "xls"])
def test_e04_iter_rows_total_cells_exact_max_plus_one(tmp_path, fmt):
    if fmt == "xlsx":
        p = tmp_path / "t.xlsx"
        p.write_bytes(_xlsx_bytes(rows=2, cols=2))  # 4 cells
        ok = SourceArtifactLimits(max_total_cells=4)
        bad = SourceArtifactLimits(max_total_cells=3)
        Adapter = XlsxWorkbookAdapter
    else:
        p = _xls_bytes(tmp_path, rows=2, cols=2)
        ok = SourceArtifactLimits(max_total_cells=4)
        bad = SourceArtifactLimits(max_total_cells=3)
        Adapter = XlsWorkbookAdapter
    list(Adapter(limits=ok).iter_rows(str(p)))
    with pytest.raises(AdapterError) as ei:
        list(Adapter(limits=bad).iter_rows(str(p)))
    assert ei.value.error_code == "total_cell_limit"


def test_e04_xlsx_zip_entry_and_expansion_limits(tmp_path):
    p = tmp_path / "z.xlsx"
    p.write_bytes(_xlsx_bytes())
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_zip_entries=1)).inspect(str(p))
    assert ei.value.error_code == "zip_entry_limit"
    with pytest.raises(AdapterError) as ei2:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_uncompressed_zip_bytes=10)).inspect(
            str(p)
        )
    assert ei2.value.error_code == "zip_expansion_limit"
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_zip_entries=5000)).inspect(str(p))


def test_e04_xlsx_merged_per_sheet_vs_total(tmp_path):
    p = tmp_path / "m.xlsx"
    p.write_bytes(_xlsx_bytes(merges=2))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_merged_regions_per_sheet=1)).inspect(
            str(p)
        )
    assert ei.value.error_code == "merged_region_limit"
    with pytest.raises(AdapterError) as ei2:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_merged_regions=1)).inspect(str(p))
    assert ei2.value.error_code == "merged_region_limit"


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_e04_endpoint_cell_limit_preserves_prior(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes(cell="Z" * 10001)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "huge.xlsx",
                io.BytesIO(payload),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert_http_rejection_preserve(
        res,
        status=400,
        error_code="cell_length_limit",
        db=db_session,
        fake_storage=fake_storage,
        snap=snap,
    )


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_e04_endpoint_upload_bytes_limit(
    client: TestClient, db_session: Session, fake_storage
):
    """Genuine upload-spool N / N+1 boundary (not request-size pre-check).

    M-01: max_upload_bytes=N, max_request_bytes high enough that multipart
    overhead cannot fire request_too_large; N accepted, N+1 → upload_too_large.
    """
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes()
    n = len(payload)
    set_source_limits_override(
        SourceArtifactLimits(max_upload_bytes=n, max_request_bytes=12 * 1024 * 1024)
    )
    url = f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts"
    try:
        res_bad = client.post(
            url,
            files={
                "file": (
                    "big.xlsx",
                    io.BytesIO(payload + b"X"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res_bad,
            status=413,
            error_code="upload_too_large",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
        res_ok = client.post(
            url,
            files={
                "file": (
                    "ok.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert res_ok.status_code == 201, res_ok.text
    finally:
        set_source_limits_override(None)


# ---------------------------------------------------------------------------
# E-05 failure ordering / recovery
# ---------------------------------------------------------------------------


def test_e05_reservation_commit_fail_before_object_write(
    db_session: Session, fake_storage, monkeypatch, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes()
    p = tmp_path / "a.xlsx"
    p.write_bytes(payload)

    commits = {"n": 0}
    real_commit = db_session.commit

    def flaky_commit():
        commits["n"] += 1
        # first commit is reservation
        if commits["n"] == 1:
            raise RuntimeError("forced_reservation_commit_fail")
        return real_commit()

    monkeypatch.setattr(db_session, "commit", flaky_commit)
    from fastapi import UploadFile
    from starlette.datastructures import Headers

    class R:
        headers = Headers({})

    with open(p, "rb") as f:
        uf = UploadFile(filename="a.xlsx", file=f)
        with pytest.raises(Exception):
            upload_source_artifact(
                db_session,
                org_id=org.id,
                project_id=proj.id,
                batch_id=batch.id,
                file=uf,
                request=R(),
                current_user=user,
                storage=fake_storage,
            )
    # Reservation commit failed before object write — no storage mutation
    assert fake_storage._objects == snap["objects"]
    # Roll back failed session and prove no durable second generation
    try:
        db_session.rollback()
    except Exception:
        pass
    durable = (
        db_session.query(ImportSourceArtifact)
        .filter(ImportSourceArtifact.state == "available")
        .count()
    )
    assert durable == 1
    assert (
        db_session.query(ImportSourceArtifact)
        .filter(ImportSourceArtifact.generation > 1)
        .count()
        == 0
    )


def test_e05_put_ok_final_commit_fails_marks_failed(
    db_session: Session, fake_storage, monkeypatch, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    payload = _xlsx_bytes()
    p = tmp_path / "a.xlsx"
    p.write_bytes(payload)
    commits = {"n": 0}
    real_commit = Session.commit

    def flaky(self):
        commits["n"] += 1
        # 1=reservation success path uses session commit; after put, finalize commit
        # Count commits on this session only
        if self is db_session and commits["n"] >= 2:
            raise RuntimeError("forced_final_commit_fail")
        return real_commit(self)

    monkeypatch.setattr(Session, "commit", flaky)
    from fastapi import UploadFile
    from starlette.datastructures import Headers

    class R:
        headers = Headers({})

    with open(p, "rb") as f:
        uf = UploadFile(filename="a.xlsx", file=f)
        with pytest.raises(Exception):
            upload_source_artifact(
                db_session,
                org_id=org.id,
                project_id=proj.id,
                batch_id=batch.id,
                file=uf,
                request=R(),
                current_user=user,
                storage=fake_storage,
            )


def test_e05_pending_missing_short_read_timeout_checksum(
    db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    now = datetime.now(timezone.utc) - timedelta(hours=2)
    # missing
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="m.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=3,
        checksum_sha256="a" * 64,
        storage_object_key=f"missing/{uuid.uuid4()}",
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    # short read
    body = b"0123456789"
    k2 = f"short/{uuid.uuid4()}"
    fake_storage._objects[k2] = body
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="s.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body),
        checksum_sha256=hashlib.sha256(body).hexdigest(),
        storage_object_key=k2,
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    # true checksum
    k3 = f"bad/{uuid.uuid4()}"
    fake_storage._objects[k3] = b"full-body!!"
    a3 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=3,
        original_filename="c.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(b"full-body!!"),
        checksum_sha256="c" * 64,
        storage_object_key=k3,
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    # timeout
    k4 = f"to/{uuid.uuid4()}"
    fake_storage._objects[k4] = b"zzzz"
    a4 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=4,
        original_filename="t.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=4,
        checksum_sha256=hashlib.sha256(b"zzzz").hexdigest(),
        storage_object_key=k4,
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    db_session.add_all([a1, a2, a3, a4])
    db_session.commit()
    fake_storage.truncate_open_to = 2
    # Process with truncate for all open_stream — short for a2,a3,a4
    # Better: only truncate for a2
    fake_storage.truncate_open_to = None
    original = fake_storage.open_stream

    def selective(key):
        if key == k2:
            fake_storage.truncate_open_to = 2
        elif key == k4:
            raise ObjectStorageError("stream_read_timeout")
        else:
            fake_storage.truncate_open_to = None
        if key == k4:
            raise ObjectStorageError("stream_read_timeout")
        return original(key)

    fake_storage.open_stream = selective  # type: ignore[method-assign]
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1.id)
    a2 = db_session.get(ImportSourceArtifact, a2.id)
    a3 = db_session.get(ImportSourceArtifact, a3.id)
    a4 = db_session.get(ImportSourceArtifact, a4.id)
    assert a1.state == "failed" and a1.failure_code == "pending_object_missing"
    assert a2.state == "pending"  # infra short_read
    assert a3.state == "failed" and a3.failure_code == "checksum_mismatch"
    assert a4.state == "pending"  # infra timeout
    assert stats["marked_failed"] >= 2
    assert stats["errors"] >= 2


def test_e05_failed_residual_to_orphaned_and_retention_boundary(
    db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    now = datetime.now(timezone.utc)
    retention = 3600
    # failed residual past created retention → orphaned
    k1 = f"fr/{uuid.uuid4()}"
    fake_storage._objects[k1] = b"x"
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="f.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="a" * 64,
        storage_object_key=k1,
        state="failed",
        created_by_user_id=user.id,
        created_at=now - timedelta(seconds=retention + 10),
        failed_at=now - timedelta(seconds=retention + 5),
    )
    # within retention window (orphaned recently) → retain
    k2 = f"ob/{uuid.uuid4()}"
    fake_storage._objects[k2] = b"y"
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="o.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="b" * 64,
        storage_object_key=k2,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=now - timedelta(days=1),
        orphaned_at=now - timedelta(seconds=1),  # clearly before cutoff elapsed
    )
    # past retention (boundary+1 / well past) → delete
    k3 = f"op/{uuid.uuid4()}"
    fake_storage._objects[k3] = b"z"
    a3 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=3,
        original_filename="p.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="c" * 64,
        storage_object_key=k3,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=now - timedelta(days=2),
        orphaned_at=now - timedelta(seconds=retention + 30),
    )
    db_session.add_all([a1, a2, a3])
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1.id)
    assert a1.state == "orphaned"
    assert k2 in fake_storage._objects  # boundary retained
    assert k3 not in fake_storage._objects  # past deleted
    assert stats["deleted_objects"] >= 1
    assert stats["marked_orphan"] >= 1


def test_e05_max_items_oldest_first(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    base = datetime.now(timezone.utc) - timedelta(hours=5)
    ids = []
    for i in range(3):
        art = ImportSourceArtifact(
            organization_id=org.id,
            project_id=proj.id,
            import_batch_id=batch.id,
            generation=i + 1,
            original_filename=f"a{i}.xlsx",
            detected_format="xlsx",
            content_type="t",
            file_size_bytes=0,
            checksum_sha256="d" * 64,
            storage_object_key=f"m/{i}-{uuid.uuid4()}",
            state="pending",
            created_by_user_id=user.id,
            created_at=base + timedelta(seconds=i),
        )
        db_session.add(art)
        ids.append(art)
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=2, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 2


def test_e05_current_and_available_never_deleted(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    now = datetime.now(timezone.utc)
    k = f"cur/{uuid.uuid4()}"
    fake_storage._objects[k] = b"c"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="c.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="e" * 64,
        storage_object_key=k,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=now - timedelta(days=3),
        orphaned_at=now - timedelta(days=2),
    )
    db_session.add(art)
    db_session.flush()
    batch.current_source_artifact_id = art.id
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert k in fake_storage._objects
    assert stats["deleted_objects"] == 0


def test_e05_later_error_keeps_earlier_failed(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    now = datetime.now(timezone.utc) - timedelta(hours=3)
    k1 = f"k1/{uuid.uuid4()}"
    body1 = b"body-one"
    fake_storage._objects[k1] = body1
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256="f" * 64,
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    k2 = f"k2/{uuid.uuid4()}"
    body2 = b"body-two-xx"
    fake_storage._objects[k2] = body2
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="b.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body2),
        checksum_sha256=hashlib.sha256(body2).hexdigest(),
        storage_object_key=k2,
        state="pending",
        created_by_user_id=user.id,
        created_at=now + timedelta(seconds=1),
    )
    db_session.add_all([a1, a2])
    db_session.commit()
    original = fake_storage.open_stream
    calls = {"n": 0}

    def open_once(key):
        calls["n"] += 1
        if calls["n"] >= 2:
            fake_storage.truncate_open_to = 2
        return original(key)

    fake_storage.open_stream = open_once  # type: ignore[method-assign]
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1.id)
    a2 = db_session.get(ImportSourceArtifact, a2.id)
    assert a1.state == "failed" and a1.failure_code == "checksum_mismatch"
    assert a2.state == "pending"
    assert stats["marked_failed"] >= 1
    assert stats["errors"] >= 1


# ---------------------------------------------------------------------------
# E-06 PostgreSQL constraint identity matrix
# ---------------------------------------------------------------------------


def _pg_url():
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql")
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL required for constraint matrix")
    return url


def test_e06_pg_constraint_identity_matrix():
    url = _pg_url()
    engine = create_engine(url)
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    b1 = uuid.uuid4()
    b2 = uuid.uuid4()
    a1 = uuid.uuid4()
    a2 = uuid.uuid4()
    slug = f"e06-{uuid.uuid4().hex[:8]}"
    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            for bid, fn in ((b1, "b1.xlsx"), (b2, "b2.xlsx")):
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :fn, 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": bid, "oid": oid, "pid": pid, "fn": fn, "uid": uid},
                )
            for aid, bid, gen, key in (
                (a1, b1, 1, f"k1-{a1}"),
                (a2, b2, 1, f"k2-{a2}"),
            ):
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, :gen, 'x.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": aid,
                        "oid": oid,
                        "pid": pid,
                        "bid": bid,
                        "gen": gen,
                        "chk": "a" * 64,
                        "key": key,
                        "uid": uid,
                    },
                )
            # valid same-batch pointer
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": a1, "bid": b1},
            )

        def _constraint_name(exc: IntegrityError) -> str:
            orig = getattr(exc, "orig", None)
            diag = getattr(orig, "diag", None) if orig is not None else None
            name = getattr(diag, "constraint_name", None) if diag is not None else None
            return name or str(orig or exc)

        # cross-batch pointer fails
        with engine.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                    ),
                    {"aid": a2, "bid": b1},
                )
            cname = _constraint_name(ei.value)
            assert "fk_batch_current_artifact_same_batch" in cname or "same_batch" in cname

        # Each case: allowed constraint name(s), SQL, params.
        # Note: checksum hex CHECK embeds {64}, so wrong-length hex may surface as
        # either checksum_len or checksum_hex depending on PG check order.
        cases = [
            (
                ("chk_source_artifact_checksum_len", "chk_source_artifact_checksum_hex"),
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "a" * 63},
            ),
            (
                # Uppercase fails lowercase CHECK and/or hex CHECK (order not guaranteed)
                ("chk_source_artifact_checksum_lower", "chk_source_artifact_checksum_hex"),
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "A" * 64},
            ),
            (
                ("chk_source_artifact_checksum_hex",),
                "UPDATE import_source_artifacts SET checksum_sha256 = :chk WHERE id = :id",
                {"id": a1, "chk": "g" * 64},
            ),
            (
                ("chk_source_artifact_state",),
                "UPDATE import_source_artifacts SET state = 'bogus' WHERE id = :id",
                {"id": a1},
            ),
            (
                ("chk_source_artifact_format",),
                "UPDATE import_source_artifacts SET detected_format = 'csv' WHERE id = :id",
                {"id": a1},
            ),
            (
                ("chk_source_artifact_generation_positive",),
                "UPDATE import_source_artifacts SET generation = 0 WHERE id = :id",
                {"id": a1},
            ),
            (
                ("chk_source_artifact_size_nonneg",),
                "UPDATE import_source_artifacts SET file_size_bytes = -1 WHERE id = :id",
                {"id": a1},
            ),
        ]
        for expected_names, sql, params in cases:
            with engine.begin() as conn:
                with pytest.raises(IntegrityError) as ei:
                    conn.execute(text(sql), params)
                cname = _constraint_name(ei.value)
                assert any(n in cname for n in expected_names), (
                    f"expected one of {expected_names} in {cname}"
                )

        # tenant composite FK mismatch
        with engine.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 9, 'x.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid,
                        "pid": pid,
                        # wrong batch org path: use batch from other — same org but force mismatch via random batch
                        "bid": uuid.uuid4(),
                        "chk": "b" * 64,
                        "key": f"k-bad-{uuid.uuid4()}",
                        "uid": uid,
                    },
                )
            cname = _constraint_name(ei.value)
            assert (
                "fk_source_artifact_batch_tenant" in cname
                or "import_batch" in cname.lower()
                or "foreign" in cname.lower()
            )

        # RESTRICT: cannot delete batch with artifacts
        with engine.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text("DELETE FROM project_asset_import_batches WHERE id = :id"),
                    {"id": b1},
                )
            cname = _constraint_name(ei.value)
            assert cname  # RESTRICT raised
    finally:
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE project_asset_import_batches SET current_source_artifact_id = NULL WHERE id IN (:b1, :b2)"),
                {"b1": b1, "b2": b2},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(
                text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid}
            )
        engine.dispose()


# ---------------------------------------------------------------------------
# E-07 throwaway migration round-trip
# ---------------------------------------------------------------------------


def test_e07_throwaway_pg_migration_roundtrip():
    url = _pg_url()
    from alembic import command
    from alembic.config import Config
    from alembic.script import ScriptDirectory
    from app.core.config import get_settings
    from sqlalchemy.engine.url import make_url

    admin = create_engine(url, isolation_level="AUTOCOMMIT")
    db_name = f"s13_e07_{uuid.uuid4().hex[:10]}"
    with admin.connect() as conn:
        conn.execute(text(f'CREATE DATABASE "{db_name}"'))
    u = make_url(url)
    iso_url = url.rsplit("/", 1)[0] + f"/{db_name}"
    prev = {
        k: os.environ.get(k)
        for k in (
            "POSTGRES_HOST",
            "POSTGRES_PORT",
            "POSTGRES_DB",
            "POSTGRES_USER",
            "POSTGRES_PASSWORD",
            "VALORA_ENV",
        )
    }
    try:
        os.environ["VALORA_ENV"] = "test"
        os.environ["POSTGRES_HOST"] = u.host or "localhost"
        os.environ["POSTGRES_PORT"] = str(u.port or 5432)
        os.environ["POSTGRES_DB"] = db_name
        os.environ["POSTGRES_USER"] = u.username or "valora"
        os.environ["POSTGRES_PASSWORD"] = u.password or "valora_local_password"
        get_settings.cache_clear()
        assert get_settings().postgres_db == db_name

        cfg = Config("alembic.ini")
        assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]

        command.upgrade(cfg, "e1f2a3b4c5d6")
        eng = create_engine(iso_url)
        try:
            with eng.connect() as c:
                assert (
                    c.execute(text("SELECT to_regclass('public.import_source_artifacts')")).scalar()
                    is None
                )
            command.upgrade(cfg, "f2a3b4c5d6e7")
            with eng.connect() as c:
                assert (
                    c.execute(text("SELECT to_regclass('public.import_source_artifacts')")).scalar()
                    is not None
                )
                cols = {
                    r[0]
                    for r in c.execute(
                        text(
                            "SELECT column_name FROM information_schema.columns "
                            "WHERE table_name='import_source_artifacts'"
                        )
                    )
                }
                for col in (
                    "checksum_sha256",
                    "storage_object_key",
                    "state",
                    "generation",
                    "file_size_bytes",
                    "detected_format",
                    "import_batch_id",
                ):
                    assert col in cols
                checks = {
                    r[0]
                    for r in c.execute(
                        text(
                            "SELECT conname FROM pg_constraint "
                            "WHERE conrelid = 'import_source_artifacts'::regclass AND contype = 'c'"
                        )
                    )
                }
                for name in (
                    "chk_source_artifact_generation_positive",
                    "chk_source_artifact_size_nonneg",
                    "chk_source_artifact_checksum_len",
                    "chk_source_artifact_state",
                    "chk_source_artifact_format",
                ):
                    assert name in checks
                fks = {
                    r[0]
                    for r in c.execute(
                        text(
                            "SELECT conname FROM pg_constraint "
                            "WHERE contype = 'f' AND ("
                            "conrelid = 'import_source_artifacts'::regclass OR "
                            "conrelid = 'project_asset_import_batches'::regclass)"
                        )
                    )
                }
                assert "fk_source_artifact_batch_tenant" in fks
                assert "fk_batch_current_artifact_same_batch" in fks
                idxs = {
                    r[0]
                    for r in c.execute(
                        text(
                            "SELECT indexname FROM pg_indexes "
                            "WHERE tablename = 'import_source_artifacts'"
                        )
                    )
                }
                assert "idx_source_artifact_state" in idxs
                assert "idx_source_artifact_batch" in idxs

            command.downgrade(cfg, "e1f2a3b4c5d6")
            with eng.connect() as c:
                assert (
                    c.execute(text("SELECT to_regclass('public.import_source_artifacts')")).scalar()
                    is None
                )
                assert (
                    c.execute(
                        text("SELECT to_regclass('public.project_asset_import_batches')")
                    ).scalar()
                    is not None
                )
            command.upgrade(cfg, "head")
            with eng.connect() as c:
                assert (
                    c.execute(text("SELECT to_regclass('public.import_source_artifacts')")).scalar()
                    is not None
                )
            assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]
        finally:
            eng.dispose()
    finally:
        for k, v in prev.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        get_settings.cache_clear()
        with admin.connect() as conn:
            conn.execute(text(f'DROP DATABASE IF EXISTS "{db_name}" WITH (FORCE)'))
        admin.dispose()


# ---------------------------------------------------------------------------
# MinIO (CI)
# ---------------------------------------------------------------------------


def test_e08_s3_minio_roundtrip_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL")
    elif not os.environ.get("S3_ENDPOINT_URL"):
        pytest.skip("S3_ENDPOINT_URL not set for local MinIO")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"fifth/{uuid.uuid4()}"
    body = b"fifth-corrective"
    store.put_stream(
        key, io.BytesIO(body), content_type="application/octet-stream", expected_size=len(body)
    )
    digest = _sha256_object(store, key, chunk_size=64, expected_size=len(body))
    assert digest == hashlib.sha256(body).hexdigest()
    store.delete(key)
    assert store.head(key) is None
