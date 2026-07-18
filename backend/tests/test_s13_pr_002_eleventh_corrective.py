"""S13-PR-002 eleventh corrective: L-01…L-04 full HTTP rejection preservation matrix.

Evidence-completeness only — no production CAS changes.
Every retained HTTP N+1 rejection path must use the shared immutable snapshot
contract in tests.support.s13_pr_002_http_preserve.
"""

from __future__ import annotations

import ast
import hashlib
import io
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    ARTIFACT_FIELDS,
    AUDIT_FIELDS,
    BATCH_FIELDS,
    CaseInput,
    LINE_FIELDS,
    STAGING_FIELDS,
    assert_audit_snapshot_detects_mutations,
    assert_http_rejection_preserve,
    snapshot_source_intake_preserve,
    source_case_limits,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
)


# ---------------------------------------------------------------------------
# L-01 / L-04: literal coverage manifest (executable set equality)
# ---------------------------------------------------------------------------
# Every externally reachable bound that rejects at the HTTP source-artifact
# endpoint must appear here with at least one retained rejected node that uses
# assert_http_rejection_preserve (or assert_source_intake_preserve for the
# fifth upload-bytes soft-status path).

HTTP_REJECTION_COVERAGE_MANIFEST: list[dict[str, str]] = [
    # request / upload size
    {
        "bound": "max_request_bytes",
        "format": "multipart",
        "accepted_node": "test_i03_request_bytes_exact_n_accepted_n_plus_one_rejected",
        "rejected_node": "test_i03_request_bytes_exact_n_accepted_n_plus_one_rejected",
        "error_code": "request_too_large",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_request_bytes",
        "format": "multipart",
        "accepted_node": "test_h03_request_bytes_exact_n_and_n_plus_one",
        "rejected_node": "test_h03_request_bytes_exact_n_and_n_plus_one",
        "error_code": "request_too_large",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_upload_bytes",
        "format": "xlsx",
        "accepted_node": "test_i03_upload_bytes_exact_n_accepted_n_plus_one_rejected",
        "rejected_node": "test_i03_upload_bytes_exact_n_accepted_n_plus_one_rejected",
        "error_code": "upload_too_large",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_upload_bytes",
        "format": "xlsx",
        "accepted_node": "test_h03_upload_bytes_exact_n_and_n_plus_one",
        "rejected_node": "test_h03_upload_bytes_exact_n_and_n_plus_one",
        "error_code": "upload_too_large",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_upload_bytes",
        "format": "xlsx",
        "accepted_node": "test_k03_upload_too_large_full_snapshot",
        "rejected_node": "test_k03_upload_too_large_full_snapshot",
        "error_code": "upload_too_large",
        "helper": "assert_http_rejection_preserve",
        "suite": "tenth",
    },
    {
        "bound": "max_upload_bytes",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_e04_endpoint_upload_bytes_limit",
        "error_code": "status_varies_400_413_500",
        "helper": "assert_source_intake_preserve",
        "suite": "fifth",
    },
    # xlsx core adapter bounds (eighth + seventh + tenth representatives)
    {
        "bound": "max_sheets",
        "format": "xlsx",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one[max_sheets-1-...]",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "error_code": "sheet_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_physical_rows",
        "format": "xlsx",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "error_code": "physical_row_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_columns",
        "format": "xlsx",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "error_code": "column_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_cell_chars",
        "format": "xlsx",
        "accepted_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xlsx_adapter_exact_n_and_n_plus_one",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_cell_chars",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_e04_endpoint_cell_limit_preserves_prior",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "fifth",
    },
    {
        "bound": "max_cell_chars",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_g03_endpoint_cell_limit_stable_status",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "sixth",
    },
    {
        "bound": "max_cell_chars",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_endpoint_cell_limit_no_reservation",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "fourth",
    },
    {
        "bound": "max_sheets",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xlsx_adapter_limits",
        "error_code": "sheet_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_physical_rows",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xlsx_adapter_limits",
        "error_code": "physical_row_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_columns",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xlsx_adapter_limits",
        "error_code": "column_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_cell_chars",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xlsx_adapter_limits",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_sheets",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_k03_reject_preserves_objects_content_types_and_all_audits",
        "error_code": "sheet_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "tenth",
    },
    {
        "bound": "max_physical_rows",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_k03_xlsx_rejects_full_snapshot",
        "error_code": "physical_row_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "tenth",
    },
    {
        "bound": "max_columns",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_k03_xlsx_rejects_full_snapshot",
        "error_code": "column_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "tenth",
    },
    {
        "bound": "max_cell_chars",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_k03_xlsx_rejects_full_snapshot",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "tenth",
    },
    {
        "bound": "max_row_chars",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_k03_xlsx_rejects_full_snapshot",
        "error_code": "row_char_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "tenth",
    },
    {
        "bound": "max_total_cells",
        "format": "xlsx",
        "accepted_node": "—",
        "rejected_node": "test_k03_xlsx_rejects_full_snapshot",
        "error_code": "total_cell_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "tenth",
    },
    # xlsx extra bounds (ninth)
    {
        "bound": "max_zip_entries",
        "format": "xlsx",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "error_code": "zip_entry_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_uncompressed_zip_bytes",
        "format": "xlsx",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "error_code": "zip_expansion_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_row_chars",
        "format": "xlsx",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "error_code": "row_char_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_total_cells",
        "format": "xlsx",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "error_code": "total_cell_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_merged_regions",
        "format": "xlsx",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "error_code": "merged_region_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_merged_regions_per_sheet",
        "format": "xlsx",
        "accepted_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xlsx_extra_adapter_bounds",
        "error_code": "merged_region_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    # xls core (eighth + seventh)
    {
        "bound": "max_sheets",
        "format": "xls",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "error_code": "sheet_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_physical_rows",
        "format": "xls",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "error_code": "physical_row_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_columns",
        "format": "xls",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "error_code": "column_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_cell_chars",
        "format": "xls",
        "accepted_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "rejected_node": "test_i03_endpoint_xls_adapter_exact_n_and_n_plus_one",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "eighth",
    },
    {
        "bound": "max_sheets",
        "format": "xls",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xls_adapter_limits",
        "error_code": "sheet_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_physical_rows",
        "format": "xls",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xls_adapter_limits",
        "error_code": "physical_row_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_columns",
        "format": "xls",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xls_adapter_limits",
        "error_code": "column_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    {
        "bound": "max_cell_chars",
        "format": "xls",
        "accepted_node": "—",
        "rejected_node": "test_h03_endpoint_xls_adapter_limits",
        "error_code": "cell_length_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "seventh",
    },
    # xls extra (ninth) — zip bounds intentionally not reachable
    {
        "bound": "max_row_chars",
        "format": "xls",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "error_code": "row_char_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_total_cells",
        "format": "xls",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "error_code": "total_cell_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_merged_regions",
        "format": "xls",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "error_code": "merged_region_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
    {
        "bound": "max_merged_regions_per_sheet",
        "format": "xls",
        "accepted_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "rejected_node": "test_j03_endpoint_xls_extra_adapter_bounds",
        "error_code": "merged_region_limit",
        "helper": "assert_http_rejection_preserve",
        "suite": "ninth",
    },
]

# Externally reachable SourceArtifactLimits bounds that reject at HTTP intake.
EXPECTED_HTTP_REACHABLE_BOUNDS = frozenset(
    {
        "max_request_bytes",
        "max_upload_bytes",
        "max_sheets",
        "max_physical_rows",
        "max_columns",
        "max_cell_chars",
        "max_row_chars",
        "max_total_cells",
        "max_merged_regions",
        "max_merged_regions_per_sheet",
        "max_zip_entries",  # xlsx only
        "max_uncompressed_zip_bytes",  # xlsx only
    }
)

# Non-HTTP / non-reject bounds intentionally excluded from the HTTP matrix.
INTENTIONAL_NON_HTTP_LIMIT_FIELDS = frozenset(
    {
        "spool_max_size",
        "reconcilers_max_items",
        "read_chunk_size",
        "orphan_retention_seconds",
    }
)

REQUIRED_REJECT_NODES_WITH_FULL_HELPER = frozenset(
    row["rejected_node"] for row in HTTP_REJECTION_COVERAGE_MANIFEST
)

SUITE_FILES = {
    "fourth": "test_s13_pr_002_fourth_corrective.py",
    "fifth": "test_s13_pr_002_fifth_corrective.py",
    "sixth": "test_s13_pr_002_sixth_corrective.py",
    "seventh": "test_s13_pr_002_seventh_corrective.py",
    "eighth": "test_s13_pr_002_eighth_corrective.py",
    "ninth": "test_s13_pr_002_ninth_corrective.py",
    "tenth": "test_s13_pr_002_tenth_corrective.py",
}

FULL_HELPER_NAMES = frozenset(
    {
        "assert_http_rejection_preserve",
        "assert_source_intake_preserve",
        "_assert_reject_preserve",  # ninth wrapper → assert_http_rejection_preserve
    }
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session(tmp_path) -> Session:
    from sqlalchemy import create_engine

    db_file = tmp_path / f"s13e11_{uuid.uuid4().hex}.db"
    engine = create_engine(f"sqlite:///{db_file}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()


def _xlsx_bytes(*, sheets: int = 1, rows: int = 1, cols: int = 1, cell: str = "ok") -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.title = "S0"
    for si in range(sheets):
        ws = wb.active if si == 0 else wb.create_sheet(f"S{si}")
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(r, c, cell)
    wb.save(buf)
    return buf.getvalue()


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org",
        organization_slug=f"o-{uuid.uuid4().hex[:6]}",
        status=OrganizationStatus.ACTIVE,
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    payload = _xlsx_bytes()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(db, fake_storage, project_id=proj.id, batch_id=batch.id)
    return prior, staging, line, snap


def _tests_dir() -> Path:
    return Path(__file__).resolve().parent


def _function_uses_full_helper(func_node: ast.FunctionDef) -> bool:
    for node in ast.walk(func_node):
        if isinstance(node, ast.Name) and node.id in FULL_HELPER_NAMES:
            return True
        if isinstance(node, ast.Attribute) and node.attr in FULL_HELPER_NAMES:
            return True
    return False


def _load_suite_function_map() -> dict[str, set[str]]:
    """Map suite file basename → set of function names that call a full helper."""
    out: dict[str, set[str]] = {}
    for suite, fname in SUITE_FILES.items():
        path = _tests_dir() / fname
        tree = ast.parse(path.read_text(encoding="utf-8"))
        full_funcs: set[str] = set()
        for node in tree.body:
            if isinstance(node, ast.FunctionDef) and node.name.startswith("test_"):
                if _function_uses_full_helper(node):
                    full_funcs.add(node.name)
            # also allow wrapper helpers in module that tests call
        # detect wrappers that themselves call full helpers
        wrappers: set[str] = set()
        for node in tree.body:
            if isinstance(node, ast.FunctionDef) and not node.name.startswith("test_"):
                if _function_uses_full_helper(node):
                    wrappers.add(node.name)
        if wrappers:
            for node in tree.body:
                if isinstance(node, ast.FunctionDef) and node.name.startswith("test_"):
                    for n in ast.walk(node):
                        if isinstance(n, ast.Name) and n.id in wrappers:
                            full_funcs.add(node.name)
                        if isinstance(n, ast.Attribute) and n.attr in wrappers:
                            full_funcs.add(node.name)
        # fourth threat path uses _assert_preserved which routes to full when snap is full
        if suite == "fourth":
            for node in tree.body:
                if isinstance(node, ast.FunctionDef) and node.name == "_assert_preserved":
                    if "assert_source_intake_preserve" in ast.dump(node):
                        for tnode in tree.body:
                            if isinstance(tnode, ast.FunctionDef) and tnode.name.startswith(
                                "test_"
                            ):
                                for n in ast.walk(tnode):
                                    if isinstance(n, ast.Name) and n.id == "_assert_preserved":
                                        full_funcs.add(tnode.name)
        out[suite] = full_funcs
    return out


# ---------------------------------------------------------------------------
# L-02 self-test + schema contract
# ---------------------------------------------------------------------------


def test_l02_audit_snapshot_detects_insert_delete_payload_mutation():
    assert_audit_snapshot_detects_mutations()


def test_l02_snapshot_schema_fields_are_complete():
    # Exact field lists required by L-02 — detect accidental narrowing.
    assert "checksum_sha256" in ARTIFACT_FIELDS
    assert "storage_object_key" in ARTIFACT_FIELDS
    assert "current_source_artifact_id" in BATCH_FIELDS
    assert "raw_values" in STAGING_FIELDS
    assert "asset_name" in LINE_FIELDS
    assert "payload" in AUDIT_FIELDS
    assert "event_name" in AUDIT_FIELDS
    assert "correlation_id" in AUDIT_FIELDS


def test_l02_snapshot_deep_copy_isolation(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    assert set(snap.keys()) >= {
        "objects",
        "content_types",
        "artifacts",
        "batches",
        "staging",
        "lines",
        "audits",
    }
    # Mutating live storage after snapshot must not mutate the snap.
    key = next(iter(snap["objects"]))
    fake_storage._objects[key] = b"MUTATED"
    assert snap["objects"][key] != b"MUTATED"
    # content_types isolation
    fake_storage._content_types[key] = "mutated/type"
    assert snap["content_types"][key] != "mutated/type"
    # restore for cleanliness
    fake_storage._objects[key] = snap["objects"][key]
    fake_storage._content_types[key] = snap["content_types"][key]


# ---------------------------------------------------------------------------
# L-04: exhaustive bound set equality + helper presence
# ---------------------------------------------------------------------------


def test_l04_manifest_covers_every_http_reachable_bound():
    covered = {row["bound"] for row in HTTP_REJECTION_COVERAGE_MANIFEST}
    assert covered == EXPECTED_HTTP_REACHABLE_BOUNDS, (
        f"manifest missing={EXPECTED_HTTP_REACHABLE_BOUNDS - covered} "
        f"extra={covered - EXPECTED_HTTP_REACHABLE_BOUNDS}"
    )


def test_l04_limit_dataclass_fields_accounted():
    from dataclasses import fields

    field_names = {f.name for f in fields(SourceArtifactLimits)}
    # Every limit field is either HTTP-reachable (in matrix) or intentional non-HTTP.
    assert field_names == EXPECTED_HTTP_REACHABLE_BOUNDS | INTENTIONAL_NON_HTTP_LIMIT_FIELDS


def test_l04_manifest_xlsx_and_xls_both_cover_shared_adapter_bounds():
    shared = {
        "max_sheets",
        "max_physical_rows",
        "max_columns",
        "max_cell_chars",
        "max_row_chars",
        "max_total_cells",
        "max_merged_regions",
        "max_merged_regions_per_sheet",
    }
    xlsx = {
        row["bound"]
        for row in HTTP_REJECTION_COVERAGE_MANIFEST
        if row["format"] == "xlsx" and row["bound"] in shared
    }
    xls = {
        row["bound"]
        for row in HTTP_REJECTION_COVERAGE_MANIFEST
        if row["format"] == "xls" and row["bound"] in shared
    }
    assert xlsx == shared
    assert xls == shared


def test_l04_zip_bounds_xlsx_only_in_manifest():
    zip_rows = [
        row
        for row in HTTP_REJECTION_COVERAGE_MANIFEST
        if row["bound"] in {"max_zip_entries", "max_uncompressed_zip_bytes"}
    ]
    assert zip_rows
    assert all(row["format"] == "xlsx" for row in zip_rows)
    assert not any(
        row["format"] == "xls" and row["bound"].startswith("max_zip")
        for row in HTTP_REJECTION_COVERAGE_MANIFEST
    )


def test_l04_every_manifest_rejected_node_calls_full_helper():
    suite_map = _load_suite_function_map()
    missing: list[str] = []
    for row in HTTP_REJECTION_COVERAGE_MANIFEST:
        suite = row["suite"]
        node = row["rejected_node"]
        full = suite_map.get(suite, set())
        if node not in full:
            missing.append(f"{suite}::{node} helper={row['helper']}")
    assert not missing, "nodes missing full preserve helper:\n" + "\n".join(missing)


def test_l04_ninth_xlsx_extra_params_still_six():
    """Do not shrink the ninth xlsx extra-bound matrix."""
    path = _tests_dir() / SUITE_FILES["ninth"]
    src = path.read_text(encoding="utf-8")
    tree = ast.parse(src)
    for node in tree.body:
        if (
            isinstance(node, ast.FunctionDef)
            and node.name == "test_j03_endpoint_xlsx_extra_adapter_bounds"
        ):
            # decorator parametrize list length
            for d in node.decorator_list:
                if isinstance(d, ast.Call):
                    # args: names, values
                    if len(d.args) >= 2 and isinstance(d.args[1], ast.List):
                        assert len(d.args[1].elts) == 6, len(d.args[1].elts)
                        return
    raise AssertionError("could not locate ninth xlsx extra param list")


def test_l04_ninth_xls_extra_params_still_four():
    path = _tests_dir() / SUITE_FILES["ninth"]
    src = path.read_text(encoding="utf-8")
    tree = ast.parse(src)
    for node in tree.body:
        if (
            isinstance(node, ast.FunctionDef)
            and node.name == "test_j03_endpoint_xls_extra_adapter_bounds"
        ):
            for d in node.decorator_list:
                if isinstance(d, ast.Call) and len(d.args) >= 2 and isinstance(d.args[1], ast.List):
                    assert len(d.args[1].elts) == 4, len(d.args[1].elts)
                    return
    raise AssertionError("could not locate ninth xls extra param list")


# ---------------------------------------------------------------------------
# L-03 live branch identity: N+1 fires intended code, full preserve runs
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "case,n,build_bad,error_code,status",
    [
        (CaseInput("xlsx", "max_sheets"), 1, lambda: _xlsx_bytes(sheets=2), "sheet_limit", 413),
        (
            CaseInput("xlsx", "max_physical_rows"),
            1,
            lambda: _xlsx_bytes(rows=2),
            "physical_row_limit",
            413,
        ),
        (CaseInput("xlsx", "max_columns"), 1, lambda: _xlsx_bytes(cols=2), "column_limit", 413),
        (
            CaseInput("xlsx", "max_cell_chars"),
            3,
            lambda: _xlsx_bytes(cell="abcd"),
            "cell_length_limit",
            400,
        ),
        (
            CaseInput("xlsx", "max_row_chars"),
            4,
            lambda: _xlsx_bytes(cols=3, cell="ab"),
            "row_char_limit",
            413,
        ),
        (
            CaseInput("xlsx", "max_total_cells"),
            4,
            lambda: _xlsx_bytes(rows=2, cols=3),
            "total_cell_limit",
            413,
        ),
    ],
    ids=[
        "max_sheets",
        "max_physical_rows",
        "max_columns",
        "max_cell_chars",
        "max_row_chars",
        "max_total_cells",
    ],
)
@pytest.mark.s13_pr_002_http_nplus1_reject
def test_l03_endpoint_xlsx_n_plus_one_full_preserve(
    client: TestClient,
    db_session: Session,
    fake_storage,
    case,
    n,
    build_bad,
    error_code,
    status,
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = case.build_artifact(xlsx=build_bad)
    with source_case_limits(case, n):
        res = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={
                "file": (
                    "lim.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res,
            status=status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
        # prove pointer and prior identity untouched
        db_session.expire_all()
        b = db_session.get(ProjectAssetImportBatch, batch.id)
        assert b is not None
        assert b.current_source_artifact_id == prior.id
        assert db_session.query(ImportSourceArtifact).count() == 1


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_l03_upload_too_large_full_preserve(client: TestClient, db_session: Session, fake_storage):
    case = CaseInput(reachability="intake", bound="max_upload_bytes")
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    base = _xlsx_bytes()
    payload = case.build_artifact(intake=lambda: base + b"X")
    with source_case_limits(case, len(base), max_request_bytes=12 * 1024 * 1024):
        res = client.post(
            f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
            files={
                "file": (
                    "big.xlsx",
                    io.BytesIO(payload),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers={"X-User-Id": str(user.id)},
        )
        assert_http_rejection_preserve(
            res,
            status=413,
            error_code="upload_too_large",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )


def test_l01_inventory_required_nodes_nonempty():
    assert len(REQUIRED_REJECT_NODES_WITH_FULL_HELPER) >= 15
    assert len(HTTP_REJECTION_COVERAGE_MANIFEST) >= 30
