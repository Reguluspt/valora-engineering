"""S13-PR-002 tenth corrective: K-01…K-04 atomic pointer CAS and proof matrix."""
from __future__ import annotations

import hashlib
import io
import os
import threading
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any

import openpyxl
import pytest
from fastapi import UploadFile
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session, sessionmaker
from starlette.datastructures import Headers

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
    set_pointer_probe_hook,
    set_source_limits_override,
    upload_source_artifact,
    _sha256_object,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    assert_http_rejection_preserve,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
    AuditEvent,
)


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session(tmp_path) -> Session:
    db_file = tmp_path / f"s13e10_{uuid.uuid4().hex}.db"
    engine = create_engine(f"sqlite:///{db_file}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()
    set_source_limits_override(None)
    set_pointer_probe_hook(None)


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org",
        organization_slug=f"o-{uuid.uuid4().hex[:6]}",
        status=OrganizationStatus.ACTIVE,
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _xlsx_bytes(*, sheets=1, rows=1, cols=1, cell="x", merges=0) -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    for r in range(rows):
        for c in range(cols):
            wb.active.cell(r + 1, c + 1, cell)
    for i in range(merges):
        col = (i % max(cols, 1)) + 1
        end_col = min(col + 1, max(cols + 1, 2))
        wb.active.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
    for i in range(1, sheets):
        wb.create_sheet(f"S{i + 1}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    payload = _xlsx_bytes()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    snap.update(
        {
            "prior_id": prior.id,
            "prior_checksum": prior.checksum_sha256,
            "prior_key": prior.storage_object_key,
            "prior_state": prior.state,
            "prior_gen": prior.generation,
            "batch_current": batch.current_source_artifact_id,
            "batch_status": batch.status,
            "batch_total": batch.total_rows,
            "batch_valid": batch.valid_rows,
            "staging_id": staging.id,
            "staging_raw": dict(staging.raw_values),
            "staging_mapped": dict(staging.mapped_values),
            "staging_name": staging.proposed_asset_name,
            "staging_desc": staging.proposed_description,
            "staging_qty": staging.proposed_quantity,
            "staging_unit": staging.proposed_unit,
            "staging_status": staging.validation_status,
            "line_id": line.id,
            "line_name": line.asset_name,
            "line_desc": line.description,
            "line_qty": float(line.quantity),
            "line_review": line.review_status,
            "line_val": line.validation_status,
            "art_count": len(snap["artifacts"]),
            "line_count": len(snap["lines"]),
        }
    )
    return prior, staging, line, snap


def _assert_preserved_full(db, fake_storage, proj, batch, staging, line, snap):
    """Full L-02/L-03 contract for HTTP rejection nodes."""
    from tests.support.s13_pr_002_http_preserve import assert_source_intake_preserve

    assert_source_intake_preserve(db, fake_storage, snap)


def _assert_error_code(res, expected_code: str):
    detail = res.json().get("detail")
    assert isinstance(detail, dict), f"detail must be mapping: {detail!r}"
    assert detail.get("error_code") == expected_code


def _audit_count(db, entity_id, event_name) -> int:
    return (
        db.query(AuditEvent)
        .filter(AuditEvent.entity_id == entity_id, AuditEvent.event_name == event_name)
        .count()
    )


def _pg_url():
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql")
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL required")
    return url


def _pg_sqlstate(exc: BaseException) -> str:
    orig = getattr(exc, "orig", None)
    return getattr(orig, "sqlstate", None) or getattr(orig, "pgcode", None) or ""


# ---------------------------------------------------------------------------
# K-01 atomic CAS TOCTOU regression (post-observe / pre-durability)
# ---------------------------------------------------------------------------


def test_k01_post_probe_pre_commit_tocou_sqlite(db_session: Session, fake_storage):
    """
    At 6038e8b: observe old pointer → pause → gen2 finalize → older commit stole pointer.
    After CAS: gen2 remains current; gen1 orphaned with zero Available audits.
    """
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    body1 = b"older-valid-tocou"
    k1 = f"old/{uuid.uuid4()}"
    fake_storage._objects[k1] = body1
    fake_storage._content_types[k1] = "application/octet-stream"
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="old.xlsx",
        detected_format="xlsx",
        content_type="application/octet-stream",
        file_size_bytes=len(body1),
        checksum_sha256=hashlib.sha256(body1).hexdigest(),
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(minutes=5),
    )
    db_session.add(a1)
    db_session.commit()
    a1_id = a1.id

    barrier_entered = threading.Event()
    release = threading.Event()
    hooks = {"probe_hits": 0}
    errors: list[BaseException] = []
    result: dict[str, Any] = {}

    def probe_hook():
        hooks["probe_hits"] += 1
        if hooks["probe_hits"] == 1:
            barrier_entered.set()
            assert release.wait(timeout=15), "release timeout"

    set_pointer_probe_hook(probe_hook)

    def run_reconcile():
        try:
            result["stats"] = reconcile_source_artifacts(
                db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
            )
        except BaseException as exc:  # noqa: BLE001
            errors.append(exc)
        finally:
            set_pointer_probe_hook(None)

    t = threading.Thread(target=run_reconcile)
    t.start()
    assert barrier_entered.wait(timeout=15), "never reached post-observe CAS barrier"

    body2 = _xlsx_bytes(cell="newer")
    UploadSession = sessionmaker(bind=db_session.get_bind(), autoflush=False, autocommit=False)
    up_db = UploadSession()
    try:

        class R:
            headers = Headers({})

        art2 = upload_source_artifact(
            up_db,
            org_id=org.id,
            project_id=proj.id,
            batch_id=batch.id,
            file=UploadFile(filename="new.xlsx", file=io.BytesIO(body2)),
            request=R(),
            current_user=user,
            storage=fake_storage,
        )
    finally:
        up_db.close()

    assert art2.state == "available"
    a2_id = art2.id
    release.set()
    t.join(timeout=20)
    assert not t.is_alive()
    assert not errors, errors
    assert hooks["probe_hits"] >= 1
    stats = result["stats"]
    assert stats["scanned"] == 1
    assert stats["errors"] == 0

    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1_id)
    a2 = db_session.get(ImportSourceArtifact, a2_id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert a2.state == "available"
    assert batch.current_source_artifact_id == a2_id
    assert a1.state == "orphaned"
    assert a1.failure_code == "stale_generation"
    assert _audit_count(db_session, a1_id, "ImportSourceArtifactAvailable") == 0
    orphan = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == a1_id,
            AuditEvent.event_name == "ImportSourceArtifactOrphaned",
        )
        .all()
    )
    assert len(orphan) == 1
    reason = (orphan[0].payload or {}).get("reason", "")
    assert "stale" in reason
    avail2 = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == a2_id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .all()
    )
    assert len(avail2) == 1
    assert avail2[0].command_name == "UploadImportSourceArtifact"
    assert fake_storage._objects[k1] == body1
    assert fake_storage._content_types[k1] == "application/octet-stream"
    assert a2.storage_object_key in fake_storage._objects
    assert a2.storage_object_key in fake_storage._content_types


def test_k01_older_durable_first_then_gen2_wins(db_session: Session, fake_storage):
    """Older reconcile promotes first; gen2 finalize still becomes current."""
    org, user, proj, batch = _seed(db_session)
    body1 = b"older-first"
    k1 = f"old/{uuid.uuid4()}"
    fake_storage._objects[k1] = body1
    fake_storage._content_types[k1] = "t"
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="old.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256=hashlib.sha256(body1).hexdigest(),
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc),
    )
    db_session.add(a1)
    db_session.commit()
    a1_id = a1.id
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats["errors"] == 0
    db_session.expire_all()
    assert db_session.get(ImportSourceArtifact, a1_id).state == "available"
    assert (
        db_session.get(ProjectAssetImportBatch, batch.id).current_source_artifact_id == a1_id
    )

    class R:
        headers = Headers({})

    art2 = upload_source_artifact(
        db_session,
        org_id=org.id,
        project_id=proj.id,
        batch_id=batch.id,
        file=UploadFile(filename="new.xlsx", file=io.BytesIO(_xlsx_bytes(cell="g2"))),
        request=R(),
        current_user=user,
        storage=fake_storage,
    )
    assert art2.state == "available"
    db_session.expire_all()
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    a1 = db_session.get(ImportSourceArtifact, a1_id)
    assert batch.current_source_artifact_id == art2.id
    # gen1 may remain available (successful history) — pointer is gen2
    assert a1.state == "available"
    assert _audit_count(db_session, art2.id, "ImportSourceArtifactAvailable") == 1
    assert _audit_count(db_session, a1_id, "ImportSourceArtifactAvailable") == 1


# ---------------------------------------------------------------------------
# K-02 genuine concurrent SQLite + PostgreSQL races
# ---------------------------------------------------------------------------


def test_k02_sqlite_concurrent_gen2_first_older_loses(db_session: Session, fake_storage):
    """Same controlled concurrent timeline as K-01 (SQLite dialect proof)."""
    # Reuse the exact TOCTOU interleaving node assertions via shared body.
    test_k01_post_probe_pre_commit_tocou_sqlite(db_session, fake_storage)


def test_k02_pg_concurrent_gen2_first_real_finalize():
    """PostgreSQL: real concurrent finalize vs older promote CAS (no pre-seed gen2)."""
    url = _pg_url()
    engine = create_engine(url)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    a1 = uuid.uuid4()
    slug = f"k02-{uuid.uuid4().hex[:8]}"
    store = FakeObjectStorage()
    body1 = b"pg-old-pending"
    k1 = f"pg/{a1}"
    store._objects[k1] = body1
    store._content_types[k1] = "t"
    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'b.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid, "oid": oid, "pid": pid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'old.xlsx', 'xlsx', 't', :sz,
                      :chk, :key, 'pending', '{}'::jsonb, :uid, now() - interval '5 minutes', now()
                    )
                    """
                ),
                {
                    "id": a1,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid,
                    "sz": len(body1),
                    "chk": hashlib.sha256(body1).hexdigest(),
                    "key": k1,
                    "uid": uid,
                },
            )

        caller = SessionLocal()
        # Need ORM user for upload
        user = caller.get(User, uid)
        barrier = threading.Event()
        release = threading.Event()
        hooks = {"probe_hits": 0}
        errors: list[BaseException] = []
        result: dict[str, Any] = {}

        def probe_hook():
            hooks["probe_hits"] += 1
            if hooks["probe_hits"] == 1:
                barrier.set()
                assert release.wait(timeout=20)

        set_pointer_probe_hook(probe_hook)

        def run_reconcile():
            try:
                result["stats"] = reconcile_source_artifacts(
                    caller, storage=store, max_items=10, actor_id=uid, org_id=oid
                )
            except BaseException as exc:  # noqa: BLE001
                errors.append(exc)
            finally:
                set_pointer_probe_hook(None)

        t = threading.Thread(target=run_reconcile)
        t.start()
        assert barrier.wait(timeout=20), "PG reconcile never reached CAS barrier"

        up = SessionLocal()
        try:

            class R:
                headers = Headers({})

            art2 = upload_source_artifact(
                up,
                org_id=oid,
                project_id=pid,
                batch_id=bid,
                file=UploadFile(filename="new.xlsx", file=io.BytesIO(_xlsx_bytes(cell="pg2"))),
                request=R(),
                current_user=user,
                storage=store,
            )
        finally:
            up.close()

        assert art2.state == "available"
        a2_id = art2.id
        release.set()
        t.join(timeout=30)
        assert not t.is_alive()
        assert not errors, errors
        assert hooks["probe_hits"] >= 1
        assert result["stats"]["errors"] == 0
        assert result["stats"]["scanned"] == 1

        fresh = SessionLocal()
        try:
            old = fresh.get(ImportSourceArtifact, a1)
            new = fresh.get(ImportSourceArtifact, a2_id)
            batch = fresh.get(ProjectAssetImportBatch, bid)
            assert new.state == "available"
            assert batch.current_source_artifact_id == a2_id
            assert old.state == "orphaned"
            assert old.failure_code == "stale_generation"
            assert (
                fresh.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == a1,
                    AuditEvent.event_name == "ImportSourceArtifactAvailable",
                )
                .count()
                == 0
            )
            assert (
                fresh.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == a1,
                    AuditEvent.event_name == "ImportSourceArtifactOrphaned",
                )
                .count()
                == 1
            )
            avail2 = (
                fresh.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == a2_id,
                    AuditEvent.event_name == "ImportSourceArtifactAvailable",
                )
                .all()
            )
            assert len(avail2) == 1
            assert avail2[0].command_name == "UploadImportSourceArtifact"
        finally:
            fresh.close()
            caller.close()
    finally:
        set_pointer_probe_hook(None)
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL "
                    "WHERE organization_id = :oid"
                ),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


# ---------------------------------------------------------------------------
# K-03 HTTP rejection exact store + audit snapshots
# ---------------------------------------------------------------------------


def _post_source(client, proj, batch, user, filename, payload, content_type):
    return client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": (filename, io.BytesIO(payload), content_type)},
        headers={"X-User-Id": str(user.id)},
    )


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_k03_reject_preserves_objects_content_types_and_all_audits(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    set_source_limits_override(SourceArtifactLimits(max_sheets=1))
    try:
        res = _post_source(
            client,
            proj,
            batch,
            user,
            "lim.xlsx",
            _xlsx_bytes(sheets=2),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        assert_http_rejection_preserve(
            res,
            status=413,
            error_code="sheet_limit",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
    finally:
        set_source_limits_override(None)


@pytest.mark.parametrize(
    "limit_field,n,build_bad,error_code,status",
    [
        ("max_physical_rows", 1, lambda: _xlsx_bytes(rows=2), "physical_row_limit", 413),
        ("max_columns", 1, lambda: _xlsx_bytes(cols=2), "column_limit", 413),
        ("max_cell_chars", 3, lambda: _xlsx_bytes(cell="abcd"), "cell_length_limit", 400),
        ("max_row_chars", 4, lambda: _xlsx_bytes(cols=3, cell="ab"), "row_char_limit", 413),
        ("max_total_cells", 4, lambda: _xlsx_bytes(rows=2, cols=3), "total_cell_limit", 413),
    ],
)
@pytest.mark.s13_pr_002_http_nplus1_reject
def test_k03_xlsx_rejects_full_snapshot(
    client: TestClient,
    db_session: Session,
    fake_storage,
    limit_field,
    n,
    build_bad,
    error_code,
    status,
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    set_source_limits_override(SourceArtifactLimits(**{limit_field: n}))
    try:
        res = _post_source(
            client,
            proj,
            batch,
            user,
            "lim.xlsx",
            build_bad(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        assert_http_rejection_preserve(
            res,
            status=status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
    finally:
        set_source_limits_override(None)


@pytest.mark.s13_pr_002_http_nplus1_reject
def test_k03_upload_too_large_full_snapshot(client: TestClient, db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    set_source_limits_override(
        SourceArtifactLimits(max_upload_bytes=len(payload), max_request_bytes=12 * 1024 * 1024)
    )
    try:
        res = _post_source(
            client,
            proj,
            batch,
            user,
            "big.xlsx",
            payload + b"X",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        assert_http_rejection_preserve(
            res,
            status=413,
            error_code="upload_too_large",
            db=db_session,
            fake_storage=fake_storage,
            snap=snap,
        )
    finally:
        set_source_limits_override(None)


# ---------------------------------------------------------------------------
# K-04 SessionTransaction object identity
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("case", ["insert", "update", "delete", "readonly"])
def test_k04_pg_session_transaction_identity_and_locks(case):
    url = _pg_url()
    engine = create_engine(url)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    caller = SessionLocal()
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    bid_rec = uuid.uuid4()
    aid_rec = uuid.uuid4()
    extra_batch_id = uuid.uuid4()
    slug = f"k04-{case}-{uuid.uuid4().hex[:8]}"
    store = FakeObjectStorage()
    rec_body = b"reconcile-pending-body"
    rec_key = f"rec/{aid_rec}"
    store._objects[rec_key] = rec_body
    store._content_types[rec_key] = "t"

    def _assert_lock_unavailable(target_id: uuid.UUID):
        other = SessionLocal()
        try:
            other.execute(text("SET LOCAL lock_timeout = '200ms'"))
            with pytest.raises(OperationalError) as ei:
                other.execute(
                    text(
                        "SELECT id FROM project_asset_import_batches WHERE id = :id FOR UPDATE NOWAIT"
                    ),
                    {"id": target_id},
                )
            assert _pg_sqlstate(ei.value) == "55P03"
            try:
                other.rollback()
            except Exception:
                pass
        finally:
            other.close()

    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'base.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid, "oid": oid, "pid": pid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'rec.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid_rec, "oid": oid, "pid": pid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'p.xlsx', 'xlsx', 't', :sz,
                      :chk, :key, 'pending', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": aid_rec,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid_rec,
                    "sz": len(rec_body),
                    "chk": hashlib.sha256(rec_body).hexdigest(),
                    "key": rec_key,
                    "uid": uid,
                },
            )

        caller.execute(text("SELECT 1"))
        tx_before = caller.get_transaction()
        assert tx_before is not None
        txn_before = caller.execute(text("SELECT txid_current()")).scalar()

        if case == "insert":
            nb = ProjectAssetImportBatch(
                id=extra_batch_id,
                organization_id=oid,
                project_id=pid,
                source_filename="flushed-insert.xlsx",
                status=ImportBatchStatus.CREATED,
                created_by_user_id=uid,
            )
            caller.add(nb)
            caller.flush()
            lock_target = None
        elif case == "update":
            batch = caller.get(ProjectAssetImportBatch, bid)
            batch.source_filename = "flushed-uncommitted.xlsx"
            caller.flush()
            lock_target = bid
        elif case == "delete":
            with engine.begin() as conn:
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, 'to-delete.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": extra_batch_id, "oid": oid, "pid": pid, "uid": uid},
                )
            doomed = caller.get(ProjectAssetImportBatch, extra_batch_id)
            caller.delete(doomed)
            caller.flush()
            lock_target = extra_batch_id
        else:
            caller.execute(
                text("SELECT id FROM project_asset_import_batches WHERE id = :id FOR UPDATE"),
                {"id": bid},
            )
            lock_target = bid

        if case == "insert":
            other = SessionLocal()
            try:
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
            finally:
                other.close()
        else:
            _assert_lock_unavailable(lock_target)

        stats = reconcile_source_artifacts(
            caller, storage=store, max_items=10, actor_id=uid, org_id=oid
        )
        assert stats["scanned"] >= 1
        assert stats["errors"] == 0

        tx_after = caller.get_transaction()
        assert tx_after is tx_before
        assert caller.in_transaction()
        txn_after = caller.execute(text("SELECT txid_current()")).scalar()
        assert txn_after == txn_before

        fresh = SessionLocal()
        try:
            rec = fresh.get(ImportSourceArtifact, aid_rec)
            assert rec.state == "available"
            assert fresh.get(ProjectAssetImportBatch, bid_rec).current_source_artifact_id == aid_rec
            assert (
                fresh.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == aid_rec,
                    AuditEvent.event_name == "ImportSourceArtifactAvailable",
                )
                .count()
                == 1
            )
        finally:
            fresh.close()

        if case == "insert":
            other = SessionLocal()
            try:
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
            finally:
                other.close()
        else:
            _assert_lock_unavailable(lock_target)

        if case == "update":
            caller.commit()
            other = SessionLocal()
            try:
                assert (
                    other.get(ProjectAssetImportBatch, bid).source_filename
                    == "flushed-uncommitted.xlsx"
                )
            finally:
                other.close()
        else:
            caller.rollback()
            if case == "delete":
                other = SessionLocal()
                try:
                    assert other.get(ProjectAssetImportBatch, extra_batch_id) is not None
                finally:
                    other.close()
            if case == "insert":
                other = SessionLocal()
                try:
                    assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
                finally:
                    other.close()
    finally:
        try:
            if caller.in_transaction():
                caller.rollback()
        except Exception:
            pass
        try:
            caller.close()
        except Exception:
            pass
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL "
                    "WHERE organization_id = :oid"
                ),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


# ---------------------------------------------------------------------------
# MinIO
# ---------------------------------------------------------------------------


def test_k05_s3_minio_roundtrip_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL")
    elif not os.environ.get("S3_ENDPOINT_URL"):
        pytest.skip("S3_ENDPOINT_URL not set for local MinIO")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"tenth/{uuid.uuid4()}"
    body = b"tenth-corrective"
    store.put_stream(
        key, io.BytesIO(body), content_type="application/octet-stream", expected_size=len(body)
    )
    digest = _sha256_object(store, key, chunk_size=64, expected_size=len(body))
    assert digest == hashlib.sha256(body).hexdigest()
    store.delete(key)
    assert store.head(key) is None
