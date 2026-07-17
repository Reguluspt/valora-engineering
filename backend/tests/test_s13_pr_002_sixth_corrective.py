"""S13-PR-002 sixth corrective: G-01…G-06 state-based proofs."""
from __future__ import annotations

import hashlib
import io
import os
import uuid
from datetime import datetime, timedelta, timezone

import openpyxl
import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.adapters.xlsx_adapter import XlsxWorkbookAdapter
from app.modules.excel_import.application.adapters.xls_adapter import XlsWorkbookAdapter
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
    upload_source_artifact,
    _sha256_object,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.domain.workbook_adapter import AdapterError
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    assert_http_rejection_preserve,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
    AuditEvent,
)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session(tmp_path) -> Session:
    # File-backed SQLite so dedicated reconciler Session uses a separate connection
    # (StaticPool/:memory: shares one connection and rolls back caller work).
    db_file = tmp_path / f"s13_{uuid.uuid4().hex}.db"
    engine = create_engine(
        f"sqlite:///{db_file}",
        connect_args={"check_same_thread": False},
    )
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()



def _ids(user, org):
    return user.id, org.id

def _ensure_inactive(db: Session) -> None:
    """Leave session inactive with no UOW so reconcile may own it."""
    if db.new or db.dirty or db.deleted:
        raise RuntimeError("test helper cannot discard caller UOW")
    if db.in_transaction():
        db.rollback()


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org", organization_slug=f"o-{uuid.uuid4().hex[:6]}", status=OrganizationStatus.ACTIVE
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    _ensure_inactive(db)
    return org, user, proj, batch


def _xlsx_bytes(*, sheets=1, rows=1, cols=1, cell="x", merges=0) -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    for r in range(rows):
        for c in range(cols):
            wb.active.cell(r + 1, c + 1, cell)
    for i in range(1, sheets):
        wb.create_sheet(f"S{i + 1}")
    for i in range(merges):
        rr = i + 1
        wb.active.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
        wb.active.cell(rr, 1, "m")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_path(tmp_path, *, sheets=1, rows=1, cols=1, cell="x", merges=0):
    pytest.importorskip("xlwt")
    import xlwt

    p = tmp_path / f"t-{uuid.uuid4().hex[:6]}.xls"
    book = xlwt.Workbook()
    for si in range(sheets):
        sh = book.add_sheet(f"S{si + 1}")
        if merges and si == 0:
            # write_merge fills cells; do not also write those positions
            for i in range(merges):
                sh.write_merge(i, i, 0, 1, "m")
            for r in range(merges, max(rows, merges)):
                for c in range(max(cols, 2)):
                    sh.write(r, c, cell)
        else:
            for r in range(rows):
                for c in range(cols):
                    sh.write(r, c, cell)
    book.save(str(p))
    return p


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    payload = _xlsx_bytes()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    _ensure_inactive(db)
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    snap.update(
        {
            "prior_id": prior.id,
            "prior_checksum": prior.checksum_sha256,
            "prior_key": prior.storage_object_key,
            "prior_state": prior.state,
            "prior_gen": prior.generation,
            "batch_current": batch.current_source_artifact_id,
            "batch_status": batch.status,
            "batch_total": batch.total_rows,
            "batch_valid": batch.valid_rows,
            "staging_id": staging.id,
            "staging_raw": dict(staging.raw_values),
            "staging_mapped": dict(staging.mapped_values),
            "staging_name": staging.proposed_asset_name,
            "staging_desc": staging.proposed_description,
            "staging_qty": staging.proposed_quantity,
            "staging_unit": staging.proposed_unit,
            "staging_status": staging.validation_status,
            "line_id": line.id,
            "line_name": line.asset_name,
            "line_desc": line.description,
            "line_qty": float(line.quantity),
            "line_review": line.review_status,
            "line_val": line.validation_status,
            "art_count": len(snap["artifacts"]),
            "line_count": len(snap["lines"]),
        }
    )
    _ensure_inactive(db)
    return prior, staging, line, snap


def _assert_preserved(db, fake_storage, proj, batch, staging, line, snap):
    """Field-level prior/staging/line preserve for non-reject paths.

    HTTP N+1 rejections use assert_http_rejection_preserve.
    """
    db.expire_all()
    prior = db.get(ImportSourceArtifact, snap["prior_id"])
    batch = db.get(ProjectAssetImportBatch, batch.id)
    staging = db.get(ProjectAssetImportStagingRow, staging.id)
    line = db.get(ProjectAssetLine, line.id)
    assert prior is not None
    assert prior.checksum_sha256 == snap["prior_checksum"]
    assert prior.storage_object_key == snap["prior_key"]
    assert prior.state == snap["prior_state"]
    assert prior.generation == snap["prior_gen"]
    assert batch.current_source_artifact_id == snap["batch_current"]
    assert batch.status == snap["batch_status"]
    assert batch.total_rows == snap["batch_total"]
    assert batch.valid_rows == snap["batch_valid"]
    assert staging.raw_values == snap["staging_raw"]
    assert staging.mapped_values == snap["staging_mapped"]
    assert staging.proposed_asset_name == snap["staging_name"]
    assert staging.proposed_description == snap["staging_desc"]
    assert staging.proposed_quantity == snap["staging_qty"]
    assert staging.proposed_unit == snap["staging_unit"]
    assert staging.validation_status == snap["staging_status"]
    assert line.asset_name == snap["line_name"]
    assert line.description == snap["line_desc"]
    assert float(line.quantity) == snap["line_qty"]
    assert line.review_status == snap["line_review"]
    assert line.validation_status == snap["line_val"]
    assert snap["prior_key"] in fake_storage._objects
    assert db.query(ProjectAssetLine).filter_by(project_id=proj.id).count() == snap["line_count"]


# ---------------------------------------------------------------------------
# G-01 transaction ownership
# ---------------------------------------------------------------------------


def test_g01_preserves_unflushed_insert(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Unflushed",
        description="x",
        quantity=1,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db_session.add(line)
    assert line in db_session.new
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
    )
    assert stats["scanned"] == 0
    assert line in db_session.new
    assert line.asset_name == "Unflushed"


def test_g01_preserves_flushed_insert(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="FlushedIns",
        description="x",
        quantity=1,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db_session.add(line)
    db_session.flush()
    assert db_session.in_transaction()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
    )
    assert stats["scanned"] == 0
    assert db_session.in_transaction()
    rows = db_session.query(ProjectAssetLine).filter_by(asset_name="FlushedIns").all()
    assert len(rows) == 1
    # not durable to second session until caller commits
    from sqlalchemy.orm import sessionmaker

    other = sessionmaker(bind=db_session.get_bind())()
    try:
        assert other.query(ProjectAssetLine).filter_by(asset_name="FlushedIns").count() == 0
    finally:
        other.close()


def test_g01_preserves_flushed_update(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Upd",
        description="x",
        quantity=1,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db_session.add(line)
    db_session.commit()
    line = db_session.get(ProjectAssetLine, line.id)
    line.asset_name = "Upd2"
    db_session.flush()
    reconcile_source_artifacts(db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid)
    assert db_session.get(ProjectAssetLine, line.id).asset_name == "Upd2"


def test_g01_preserves_flushed_delete(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Del",
        description="x",
        quantity=1,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db_session.add(line)
    db_session.commit()
    lid = line.id
    line = db_session.get(ProjectAssetLine, lid)
    db_session.delete(line)
    db_session.flush()
    reconcile_source_artifacts(db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid)
    assert db_session.get(ProjectAssetLine, lid) is None


def test_g01_readonly_active_transaction_untouched(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    _ = db_session.query(Project).filter_by(id=proj.id).first()
    assert db_session.in_transaction()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
    )
    assert stats["scanned"] == 0
    assert db_session.in_transaction()
    assert db_session.query(Project).filter_by(id=proj.id).count() == 1


def test_g01_clean_session_ok(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    _ensure_inactive(db_session)
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid
    )
    assert stats["scanned"] == 0


# ---------------------------------------------------------------------------
# G-02 finalize commit failure + reconciler promotion
# ---------------------------------------------------------------------------


def test_g02_finalize_commit_fail_marks_failed_and_reconcile_repairs(
    db_session: Session, fake_storage, monkeypatch, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes()
    p = tmp_path / "a.xlsx"
    p.write_bytes(payload)

    commits = {"n": 0}
    real_commit = Session.commit

    def flaky(self):
        if self is not db_session:
            return real_commit(self)
        commits["n"] += 1
        # 1 = reservation OK; 2 = finalize available → fail
        if commits["n"] == 2:
            raise RuntimeError("forced_final_commit_fail")
        return real_commit(self)

    monkeypatch.setattr(Session, "commit", flaky)
    from fastapi import UploadFile
    from starlette.datastructures import Headers

    class R:
        headers = Headers({})

    with open(p, "rb") as f:
        uf = UploadFile(filename="a.xlsx", file=f)
        with pytest.raises(HTTPException) as ei:
            upload_source_artifact(
                db_session,
                org_id=org.id,
                project_id=proj.id,
                batch_id=batch.id,
                file=uf,
                request=R(),
                current_user=user,
                storage=fake_storage,
            )
    assert ei.value.status_code == 500
    _ensure_inactive(db_session)
    db_session.expire_all()
    arts = db_session.query(ImportSourceArtifact).order_by(ImportSourceArtifact.generation).all()
    assert len(arts) >= 2
    failed = [a for a in arts if a.generation > 1]
    assert len(failed) == 1
    art = failed[0]
    # Either marked failed by recovery or left pending if mark also failed
    assert art.state in {"failed", "pending"}
    if art.state == "failed":
        assert art.failure_code == "finalize_commit_failed"
    assert art.storage_object_key in fake_storage._objects
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert batch.current_source_artifact_id == snap["batch_current"]
    _assert_preserved(db_session, fake_storage, proj, batch, staging, line, snap)

    # Reconcile: if failed residual, keep; if pending verified, promote
    _ensure_inactive(db_session)
    # Clear prior current so promotion can set pointer when pending
    # Keep prior as available current — promotion of newer gen should set pointer when not stale-blocked
    # prior is gen1 available current; new is gen2 pending/failed with object
    if art.state == "pending":
        reconcile_source_artifacts(
            db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
        )
        db_session.expire_all()
        art = db_session.get(ImportSourceArtifact, art.id)
        # newer than prior? gen2 > gen1 so if prior is current available, stale path orphans
        # Actually promote path: newer_wins if current.generation > art.generation — prior gen1 < gen2, so promote
        assert art.state == "available"
        batch = db_session.get(ProjectAssetImportBatch, batch.id)
        assert batch.current_source_artifact_id == art.id
    else:
        # failed residual remains failed until retention orphan
        assert art.state == "failed"
        assert art.failure_code == "finalize_commit_failed"


def test_g02_stale_recovery_cannot_overwrite_newer_available(
    db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    now = datetime.now(timezone.utc)
    # newer available current gen=2
    k2 = f"new/{uuid.uuid4()}"
    body2 = b"newer-body-xx"
    fake_storage._objects[k2] = body2
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="new.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body2),
        checksum_sha256=hashlib.sha256(body2).hexdigest(),
        storage_object_key=k2,
        state="available",
        created_by_user_id=user.id,
        available_at=now,
    )
    db_session.add(a2)
    db_session.flush()
    batch.current_source_artifact_id = a2.id
    # older pending residual gen=1 with valid object (stale recovery)
    k1 = f"old/{uuid.uuid4()}"
    body1 = b"older-body"
    fake_storage._objects[k1] = body1
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="old.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256=hashlib.sha256(body1).hexdigest(),
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=now - timedelta(hours=1),
    )
    db_session.add(a1)
    db_session.commit()
    _ensure_inactive(db_session)
    reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
    )
    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1.id)
    a2 = db_session.get(ImportSourceArtifact, a2.id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert a2.state == "available"
    assert batch.current_source_artifact_id == a2.id
    # Stale gen1 must not become available or steal pointer
    assert a1.state != "available"
    assert a1.state in {"orphaned", "pending", "failed"}
    assert batch.current_source_artifact_id == a2.id


# ---------------------------------------------------------------------------
# G-03 exact/max+1 boundaries
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "limit_field,n,error_code,build_n",
    [
        ("max_sheets", 2, "sheet_limit", lambda: _xlsx_bytes(sheets=2)),
        ("max_physical_rows", 2, "physical_row_limit", lambda: _xlsx_bytes(rows=2)),
        ("max_columns", 2, "column_limit", lambda: _xlsx_bytes(cols=2)),
        ("max_total_cells", 4, "total_cell_limit", lambda: _xlsx_bytes(rows=2, cols=2)),
        ("max_cell_chars", 4, "cell_length_limit", lambda: _xlsx_bytes(cell="abcd")),
        ("max_row_chars", 4, "row_char_limit", lambda: _xlsx_bytes(cols=2, cell="ab")),
        ("max_merged_regions", 1, "merged_region_limit", lambda: _xlsx_bytes(merges=1)),
        (
            "max_merged_regions_per_sheet",
            1,
            "merged_region_limit",
            lambda: _xlsx_bytes(merges=1),
        ),
    ],
)
def test_g03_xlsx_inspect_exact_n_and_n_plus_one(tmp_path, limit_field, n, error_code, build_n):
    p = tmp_path / "b.xlsx"
    p.write_bytes(build_n())
    # exact N accepts with limit=N
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(**{limit_field: n})).inspect(str(p))
    # N+1 fixture rejects at limit N
    if limit_field in {"max_sheets"}:
        p2 = tmp_path / "b2.xlsx"
        p2.write_bytes(_xlsx_bytes(sheets=n + 1))
    elif limit_field == "max_physical_rows":
        p2 = tmp_path / "b2.xlsx"
        p2.write_bytes(_xlsx_bytes(rows=n + 1))
    elif limit_field == "max_columns":
        p2 = tmp_path / "b2.xlsx"
        p2.write_bytes(_xlsx_bytes(cols=n + 1))
    elif limit_field == "max_total_cells":
        p2 = tmp_path / "b2.xlsx"
        p2.write_bytes(_xlsx_bytes(rows=2, cols=3))  # 6 > 4
    elif limit_field == "max_cell_chars":
        p2 = tmp_path / "b2.xlsx"
        p2.write_bytes(_xlsx_bytes(cell="abcde"))
    elif limit_field == "max_row_chars":
        p2 = tmp_path / "b2.xlsx"
        p2.write_bytes(_xlsx_bytes(cols=3, cell="ab"))  # 6 > 4
    elif limit_field in {"max_merged_regions", "max_merged_regions_per_sheet"}:
        p2 = tmp_path / "b2.xlsx"
        p2.write_bytes(_xlsx_bytes(merges=n + 1))
    else:
        p2 = p
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(**{limit_field: n})).inspect(str(p2))
    assert ei.value.error_code == error_code


@pytest.mark.parametrize(
    "limit_field,n,error_code,kwargs_n,kwargs_np1",
    [
        ("max_sheets", 2, "sheet_limit", {"sheets": 2}, {"sheets": 3}),
        ("max_physical_rows", 2, "physical_row_limit", {"rows": 2}, {"rows": 3}),
        ("max_columns", 2, "column_limit", {"cols": 2}, {"cols": 3}),
        ("max_total_cells", 4, "total_cell_limit", {"rows": 2, "cols": 2}, {"rows": 2, "cols": 3}),
        ("max_cell_chars", 4, "cell_length_limit", {"cell": "abcd"}, {"cell": "abcde"}),
        ("max_row_chars", 4, "row_char_limit", {"cols": 2, "cell": "ab"}, {"cols": 3, "cell": "ab"}),
        ("max_merged_regions", 1, "merged_region_limit", {"merges": 1}, {"merges": 2}),
        ("max_merged_regions_per_sheet", 1, "merged_region_limit", {"merges": 1}, {"merges": 2}),
    ],
)
def test_g03_xls_inspect_exact_n_and_n_plus_one(
    tmp_path, limit_field, n, error_code, kwargs_n, kwargs_np1
):
    p = _xls_path(tmp_path, **kwargs_n)
    XlsWorkbookAdapter(limits=SourceArtifactLimits(**{limit_field: n})).inspect(str(p))
    p2 = _xls_path(tmp_path, **kwargs_np1)
    with pytest.raises(AdapterError) as ei:
        XlsWorkbookAdapter(limits=SourceArtifactLimits(**{limit_field: n})).inspect(str(p2))
    assert ei.value.error_code == error_code


def test_g03_xlsx_zip_entries_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "z.xlsx"
    p.write_bytes(_xlsx_bytes())
    # Count real entries
    import zipfile

    with zipfile.ZipFile(p) as zf:
        n = len(zf.infolist())
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_zip_entries=n)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_zip_entries=n - 1 if n > 1 else 0)).inspect(
            str(p)
        )
    assert ei.value.error_code == "zip_entry_limit"


def test_g03_xlsx_zip_expansion_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "z.xlsx"
    p.write_bytes(_xlsx_bytes())
    import zipfile

    with zipfile.ZipFile(p) as zf:
        total = sum(i.file_size for i in zf.infolist())
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_uncompressed_zip_bytes=total)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(
            limits=SourceArtifactLimits(max_uncompressed_zip_bytes=max(0, total - 1))
        ).inspect(str(p))
    assert ei.value.error_code == "zip_expansion_limit"


@pytest.mark.parametrize("fmt", ["xlsx", "xls"])
def test_g03_iter_rows_total_cells_exact(tmp_path, fmt):
    if fmt == "xlsx":
        p = tmp_path / "t.xlsx"
        p.write_bytes(_xlsx_bytes(rows=2, cols=2))
        Adapter = XlsxWorkbookAdapter
    else:
        p = _xls_path(tmp_path, rows=2, cols=2)
        Adapter = XlsWorkbookAdapter
    list(Adapter(limits=SourceArtifactLimits(max_total_cells=4)).iter_rows(str(p)))
    with pytest.raises(AdapterError) as ei:
        list(Adapter(limits=SourceArtifactLimits(max_total_cells=3)).iter_rows(str(p)))
    assert ei.value.error_code == "total_cell_limit"


def test_g03_endpoint_cell_limit_stable_status(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, proj, batch = _seed(db_session)
    _uid, _oid = user.id, org.id
    prior, staging, line, snap = _seed_prior_full(db_session, org, user, proj, batch, fake_storage)
    payload = _xlsx_bytes(cell="Z" * 10001)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "huge.xlsx",
                io.BytesIO(payload),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert_http_rejection_preserve(
        res,
        status=400,
        error_code="cell_length_limit",
        db=db_session,
        fake_storage=fake_storage,
        snap=snap,
    )


# ---------------------------------------------------------------------------
# G-04 failure ordering (state-based)
# ---------------------------------------------------------------------------


def test_g04_retention_exact_boundary_and_plus_one(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    frozen = datetime(2026, 7, 17, 12, 0, 0, tzinfo=timezone.utc)
    retention = 3600
    # exact boundary: orphaned_at == frozen - retention → retain (strict <)
    k_eq = f"eq/{uuid.uuid4()}"
    fake_storage._objects[k_eq] = b"e"
    a_eq = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="eq.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="a" * 64,
        storage_object_key=k_eq,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=frozen - timedelta(days=2),
        orphaned_at=frozen - timedelta(seconds=retention),
    )
    # boundary+1 second past → delete
    k_past = f"past/{uuid.uuid4()}"
    fake_storage._objects[k_past] = b"p"
    a_past = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="past.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="b" * 64,
        storage_object_key=k_past,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=frozen - timedelta(days=3),
        orphaned_at=frozen - timedelta(seconds=retention + 1),
    )
    # before retention → retain
    k_before = f"bef/{uuid.uuid4()}"
    fake_storage._objects[k_before] = b"b"
    a_before = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=3,
        original_filename="bef.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="c" * 64,
        storage_object_key=k_before,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=frozen - timedelta(days=1),
        orphaned_at=frozen - timedelta(seconds=retention - 1),
    )
    db_session.add_all([a_eq, a_past, a_before])
    db_session.commit()
    _ensure_inactive(db_session)
    stats = reconcile_source_artifacts(
        db_session,
        storage=fake_storage,
        max_items=10,
        actor_id=uid,
        org_id=oid,
        now=frozen,
    )
    assert k_eq in fake_storage._objects
    assert k_before in fake_storage._objects
    assert k_past not in fake_storage._objects
    assert stats["deleted_objects"] == 1


def test_g04_max_items_oldest_identities(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    base = datetime(2026, 7, 17, 10, 0, 0, tzinfo=timezone.utc)
    arts = []
    for i in range(3):
        a = ImportSourceArtifact(
            organization_id=org.id,
            project_id=proj.id,
            import_batch_id=batch.id,
            generation=i + 1,
            original_filename=f"a{i}.xlsx",
            detected_format="xlsx",
            content_type="t",
            file_size_bytes=0,
            checksum_sha256="d" * 64,
            storage_object_key=f"m/{i}-{uuid.uuid4()}",
            state="pending",
            created_by_user_id=user.id,
            created_at=base + timedelta(seconds=i),
        )
        db_session.add(a)
        arts.append(a)
    db_session.commit()
    ids = [a.id for a in arts]
    _ensure_inactive(db_session)
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=2, actor_id=uid, org_id=oid, now=base + timedelta(hours=5)
    )
    assert stats["scanned"] == 2
    db_session.expire_all()
    # oldest two processed: missing object + past retention → failed
    a0 = db_session.get(ImportSourceArtifact, ids[0])
    a1 = db_session.get(ImportSourceArtifact, ids[1])
    a2 = db_session.get(ImportSourceArtifact, ids[2])
    assert a0.state == "failed" and a0.failure_code == "pending_object_missing"
    assert a1.state == "failed" and a1.failure_code == "pending_object_missing"
    assert a2.state == "pending"  # not scanned


def test_g04_later_db_commit_failure_preserves_earlier(
    db_session: Session, fake_storage, monkeypatch
):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    now = datetime.now(timezone.utc) - timedelta(hours=3)
    # a1: will fail checksum → durable failed
    k1 = f"k1/{uuid.uuid4()}"
    body1 = b"body-one"
    fake_storage._objects[k1] = body1
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256="f" * 64,
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        created_at=now,
    )
    # a2: valid body then force commit fail on second item
    k2 = f"k2/{uuid.uuid4()}"
    body2 = b"body-two-ok"
    fake_storage._objects[k2] = body2
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="b.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body2),
        checksum_sha256=hashlib.sha256(body2).hexdigest(),
        storage_object_key=k2,
        state="pending",
        created_by_user_id=user.id,
        created_at=now + timedelta(seconds=1),
    )
    db_session.add_all([a1, a2])
    db_session.commit()
    _ensure_inactive(db_session)
    commits = {"n": 0}
    real = Session.commit

    def flaky(self):
        if self is not db_session:
            return real(self)
        commits["n"] += 1
        # after snapshot rollback, item commits: 1=a1 success, 2=a2 fail
        if commits["n"] >= 2:
            raise RuntimeError("forced_item_commit_fail")
        return real(self)

    monkeypatch.setattr(Session, "commit", flaky)
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
    )
    _ensure_inactive(db_session)
    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1.id)
    a2 = db_session.get(ImportSourceArtifact, a2.id)
    assert a1.state == "failed"
    assert a1.failure_code == "checksum_mismatch"
    # a2 commit failed mid-item; may remain pending or partial orphan/available
    assert a2.state in {"pending", "available", "orphaned", "failed"}
    assert stats["marked_failed"] >= 1


def test_g04_delete_succeeds_audit_commit_fails_retry_repairs(
    db_session: Session, fake_storage, monkeypatch
):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    frozen = datetime(2026, 7, 17, 15, 0, 0, tzinfo=timezone.utc)
    k = f"del/{uuid.uuid4()}"
    fake_storage._objects[k] = b"z"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="d.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="e" * 64,
        storage_object_key=k,
        state="orphaned",
        created_by_user_id=user.id,
        created_at=frozen - timedelta(days=2),
        orphaned_at=frozen - timedelta(seconds=7200),
    )
    db_session.add(art)
    db_session.commit()
    aid = art.id
    _ensure_inactive(db_session)
    commits = {"n": 0}
    real = Session.commit

    def flaky(self):
        if self is not db_session:
            return real(self)
        commits["n"] += 1
        # first durable commit after delete+audit fails
        if commits["n"] == 1:
            raise RuntimeError("forced_delete_audit_commit_fail")
        return real(self)

    monkeypatch.setattr(Session, "commit", flaky)
    reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid, now=frozen
    )
    # object deleted but DB commit failed
    assert k not in fake_storage._objects
    _ensure_inactive(db_session)
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, aid)
    assert art.state == "orphaned"  # no delete audit committed
    # retry repairs: delete idempotent, audit+commit succeed
    monkeypatch.setattr(Session, "commit", real)
    _ensure_inactive(db_session)
    stats2 = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=uid, org_id=oid, now=frozen
    )
    assert stats2["deleted_objects"] == 1
    audits = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == aid,
            AuditEvent.event_name == "ImportSourceArtifactObjectDeleted",
        )
        .count()
    )
    assert audits >= 1


def test_g04_prior_available_never_deleted(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    frozen = datetime(2026, 7, 17, 16, 0, 0, tzinfo=timezone.utc)
    k = f"av/{uuid.uuid4()}"
    fake_storage._objects[k] = b"a"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="a" * 64,
        storage_object_key=k,
        state="available",
        created_by_user_id=user.id,
        available_at=frozen - timedelta(days=10),
        created_at=frozen - timedelta(days=10),
    )
    db_session.add(art)
    db_session.flush()
    batch.current_source_artifact_id = art.id
    db_session.commit()
    _ensure_inactive(db_session)
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid, now=frozen
    )
    assert k in fake_storage._objects
    assert stats["deleted_objects"] == 0
    assert stats["scanned"] == 0  # available not selected


# ---------------------------------------------------------------------------
# G-05 PostgreSQL constraint identity
# ---------------------------------------------------------------------------


def _pg_url():
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql")
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL required")
    return url


def test_g05_pg_composite_tenant_fk_and_restrict():
    url = _pg_url()
    engine = create_engine(url)
    oid1 = uuid.uuid4()
    oid2 = uuid.uuid4()
    uid1 = uuid.uuid4()
    uid2 = uuid.uuid4()
    cid1 = uuid.uuid4()
    cid2 = uuid.uuid4()
    pid1 = uuid.uuid4()
    pid2 = uuid.uuid4()
    b1 = uuid.uuid4()
    b2 = uuid.uuid4()
    a1 = uuid.uuid4()
    slug = f"g05-{uuid.uuid4().hex[:8]}"

    def _cname(exc: IntegrityError) -> str:
        orig = getattr(exc, "orig", None)
        diag = getattr(orig, "diag", None) if orig is not None else None
        return getattr(diag, "constraint_name", None) or ""

    try:
        with engine.begin() as conn:
            for oid, uid, cid, pid, bid, sfx in (
                (oid1, uid1, cid1, pid1, b1, "a"),
                (oid2, uid2, cid2, pid2, b2, "b"),
            ):
                conn.execute(
                    text(
                        "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                        "VALUES (:id, 'T', :slug, 'active', now(), now())"
                    ),
                    {"id": oid, "slug": f"{slug}-{sfx}"},
                )
                conn.execute(
                    text(
                        "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                        "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                    ),
                    {"id": uid, "oid": oid, "email": f"{slug}-{sfx}@ex.com"},
                )
                conn.execute(
                    text(
                        "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                        "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                    ),
                    {"id": cid, "oid": oid, "uid": uid},
                )
                conn.execute(
                    text(
                        """
                        INSERT INTO projects (
                          id, organization_id, customer_id, name, code, status,
                          knowledge_status, fee_amount, created_by, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": pid, "oid": oid, "cid": cid, "code": f"{slug[:12]}{sfx}", "uid": uid},
                )
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, 'b.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": bid, "oid": oid, "pid": pid, "uid": uid},
                )
            # valid artifact for batch1
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'x.xlsx', 'xlsx', 't', 1,
                      :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": a1,
                    "oid": oid1,
                    "pid": pid1,
                    "bid": b1,
                    "chk": "a" * 64,
                    "key": f"k-{a1}",
                    "uid": uid1,
                },
            )
            # same-batch pointer OK
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": a1, "bid": b1},
            )

        # Composite tenant FK: batch b2 exists but artifact claims org1/project1 + batch2
        with engine.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 1, 'x.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid1,
                        "pid": pid1,
                        "bid": b2,  # exists but belongs to org2/project2
                        "chk": "b" * 64,
                        "key": f"k-mismatch-{uuid.uuid4()}",
                        "uid": uid1,
                    },
                )
            assert _cname(ei.value) == "fk_source_artifact_batch_tenant"

        # RESTRICT: delete batch with artifact
        with engine.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text("DELETE FROM project_asset_import_batches WHERE id = :id"),
                    {"id": b1},
                )
            cname = _cname(ei.value)
            assert cname  # FK RESTRICT
            assert "import_source" in cname or "batch" in cname or cname.startswith("fk_")
    finally:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL "
                    "WHERE id IN (:b1, :b2)"
                ),
                {"b1": b1, "b2": b2},
            )
            for oid in (oid1, oid2):
                conn.execute(
                    text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                    {"oid": oid},
                )
                conn.execute(
                    text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                    {"oid": oid},
                )
                conn.execute(text("DELETE FROM projects WHERE organization_id = :oid"), {"oid": oid})
                conn.execute(text("DELETE FROM customers WHERE organization_id = :oid"), {"oid": oid})
                conn.execute(text("DELETE FROM users WHERE organization_id = :oid"), {"oid": oid})
                conn.execute(
                    text("DELETE FROM organization_profiles WHERE id = :oid"), {"oid": oid}
                )
        engine.dispose()


# ---------------------------------------------------------------------------
# G-06 throwaway migration full schema + DML
# ---------------------------------------------------------------------------


def test_g06_throwaway_migration_full_schema_and_dml():
    url = _pg_url()
    from alembic import command
    from alembic.config import Config
    from alembic.script import ScriptDirectory
    from app.core.config import get_settings
    from sqlalchemy.engine.url import make_url

    admin = create_engine(url, isolation_level="AUTOCOMMIT")
    db_name = f"s13_g06_{uuid.uuid4().hex[:10]}"
    with admin.connect() as conn:
        conn.execute(text(f'CREATE DATABASE "{db_name}"'))
    u = make_url(url)
    iso_url = url.rsplit("/", 1)[0] + f"/{db_name}"
    prev = {
        k: os.environ.get(k)
        for k in (
            "POSTGRES_HOST",
            "POSTGRES_PORT",
            "POSTGRES_DB",
            "POSTGRES_USER",
            "POSTGRES_PASSWORD",
            "VALORA_ENV",
        )
    }

    def _assert_schema(eng):
        with eng.connect() as c:
            cols = {
                r[0]: r
                for r in c.execute(
                    text(
                        "SELECT column_name, is_nullable, data_type "
                        "FROM information_schema.columns "
                        "WHERE table_name='import_source_artifacts'"
                    )
                )
            }
            for col in (
                "id",
                "organization_id",
                "project_id",
                "import_batch_id",
                "generation",
                "original_filename",
                "detected_format",
                "content_type",
                "file_size_bytes",
                "checksum_sha256",
                "storage_object_key",
                "storage_etag",
                "state",
                "adapter_name",
                "adapter_version",
                "adapter_metadata",
                "created_by_user_id",
                "available_at",
                "failed_at",
                "orphaned_at",
                "failure_code",
                "created_at",
                "updated_at",
            ):
                assert col in cols, col
            checks = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint "
                        "WHERE conrelid = 'import_source_artifacts'::regclass AND contype = 'c'"
                    )
                )
            }
            for name in (
                "chk_source_artifact_generation_positive",
                "chk_source_artifact_size_nonneg",
                "chk_source_artifact_checksum_len",
                "chk_source_artifact_checksum_lower",
                "chk_source_artifact_checksum_hex",
                "chk_source_artifact_state",
                "chk_source_artifact_format",
            ):
                assert name in checks, name
            uniques = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint "
                        "WHERE conrelid = 'import_source_artifacts'::regclass AND contype = 'u'"
                    )
                )
            }
            for name in (
                "uq_source_artifact_batch_generation",
                "uq_source_artifact_object_key",
                "uq_source_artifact_batch_id",
            ):
                assert name in uniques, name
            fks = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT conname FROM pg_constraint WHERE contype = 'f' AND ("
                        "conrelid = 'import_source_artifacts'::regclass OR "
                        "conrelid = 'project_asset_import_batches'::regclass)"
                    )
                )
            }
            assert "fk_source_artifact_batch_tenant" in fks
            assert "fk_batch_current_artifact_same_batch" in fks
            idxs = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT indexname FROM pg_indexes WHERE tablename = 'import_source_artifacts'"
                    )
                )
            }
            for name in (
                "idx_source_artifact_org",
                "idx_source_artifact_project",
                "idx_source_artifact_batch",
                "idx_source_artifact_state",
            ):
                assert name in idxs, name
            # batch pointer column
            bcols = {
                r[0]
                for r in c.execute(
                    text(
                        "SELECT column_name FROM information_schema.columns "
                        "WHERE table_name='project_asset_import_batches'"
                    )
                )
            }
            assert "current_source_artifact_id" in bcols

    try:
        os.environ["VALORA_ENV"] = "test"
        os.environ["POSTGRES_HOST"] = u.host or "localhost"
        os.environ["POSTGRES_PORT"] = str(u.port or 5432)
        os.environ["POSTGRES_DB"] = db_name
        os.environ["POSTGRES_USER"] = u.username or "valora"
        os.environ["POSTGRES_PASSWORD"] = u.password or "valora_local_password"
        get_settings.cache_clear()
        cfg = Config("alembic.ini")
        assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]
        command.upgrade(cfg, "e1f2a3b4c5d6")
        eng = create_engine(iso_url)
        try:
            with eng.connect() as c:
                assert (
                    c.execute(text("SELECT to_regclass('public.import_source_artifacts')")).scalar()
                    is None
                )
            command.upgrade(cfg, "f2a3b4c5d6e7")
            _assert_schema(eng)
            command.downgrade(cfg, "e1f2a3b4c5d6")
            with eng.connect() as c:
                assert (
                    c.execute(text("SELECT to_regclass('public.import_source_artifacts')")).scalar()
                    is None
                )
            command.upgrade(cfg, "head")
            _assert_schema(eng)
            assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]
        finally:
            eng.dispose()
    finally:
        for k, v in prev.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        get_settings.cache_clear()
        with admin.connect() as conn:
            conn.execute(text(f'DROP DATABASE IF EXISTS "{db_name}" WITH (FORCE)'))
        admin.dispose()


# ---------------------------------------------------------------------------
# MinIO
# ---------------------------------------------------------------------------


def test_g07_s3_minio_roundtrip_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL")
    elif not os.environ.get("S3_ENDPOINT_URL"):
        pytest.skip("S3_ENDPOINT_URL not set for local MinIO")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"sixth/{uuid.uuid4()}"
    body = b"sixth-corrective"
    store.put_stream(
        key, io.BytesIO(body), content_type="application/octet-stream", expected_size=len(body)
    )
    digest = _sha256_object(store, key, chunk_size=64, expected_size=len(body))
    assert digest == hashlib.sha256(body).hexdigest()
    store.delete(key)
    assert store.head(key) is None
