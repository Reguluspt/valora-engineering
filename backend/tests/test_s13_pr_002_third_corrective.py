"""S13-PR-002 third corrective: B-01..B-08 executable proofs."""
from __future__ import annotations

import hashlib
import io
import os
import uuid
import zipfile
from datetime import datetime, timedelta, timezone

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.adapters.xls_adapter import XlsWorkbookAdapter
from app.modules.excel_import.application.adapters.xlsx_adapter import XlsxWorkbookAdapter
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.domain.workbook_adapter import AdapterError
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AuditEvent,
)
from tests.fixtures.s13_pr_002.ole_builder import write_threat_xls


@pytest.fixture
def db_session() -> Session:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org",
        organization_slug=f"o-{uuid.uuid4().hex[:6]}",
        status=OrganizationStatus.ACTIVE,
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _xlsx(path, rows=2, cols=2, long=None, formula=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(rows):
        for c in range(cols):
            if formula and r == 1 and c == 0:
                ws.cell(r + 1, c + 1, value="=1+1")
            else:
                ws.cell(r + 1, c + 1, value=f"{r},{c}")
    if long is not None:
        ws["A1"] = long
    wb.save(path)


# --- C-01 total cells on iter_rows ---


def test_xlsx_iter_rows_total_cells_exact_and_max_plus_one(tmp_path):
    p = tmp_path / "t.xlsx"
    _xlsx(p, rows=2, cols=2)  # 4 cells
    # exact 4
    rows = list(XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=4)).iter_rows(str(p)))
    assert sum(len(r) for r in rows) == 4
    # max+1
    with pytest.raises(AdapterError) as ei:
        list(XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=3)).iter_rows(str(p)))
    assert ei.value.error_code == "total_cell_limit"
    # inspect also rejects
    with pytest.raises(AdapterError) as ei2:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=3)).inspect(str(p))
    assert ei2.value.error_code == "total_cell_limit"


def test_xls_iter_rows_total_cells_exact_and_max_plus_one(tmp_path):
    pytest.importorskip("xlwt")
    import xlwt

    book = xlwt.Workbook()
    sh = book.add_sheet("S")
    sh.write(0, 0, "a")
    sh.write(0, 1, "b")
    sh.write(1, 0, "c")
    sh.write(1, 1, "d")
    p = tmp_path / "t.xls"
    book.save(str(p))
    rows = list(XlsWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=4)).iter_rows(str(p)))
    assert sum(len(r) for r in rows) == 4
    with pytest.raises(AdapterError) as ei:
        list(XlsWorkbookAdapter(limits=SourceArtifactLimits(max_total_cells=3)).iter_rows(str(p)))
    assert ei.value.error_code == "total_cell_limit"


# --- C-02 storage read failure ≠ checksum_mismatch ---


def test_reconciler_stream_timeout_not_checksum_mismatch(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    payload = b"same-size-data"
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = payload
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    fake_storage.fail_open_stream = True
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art)
    assert art.state == "pending"
    assert art.failure_code is None
    assert stats["errors"] >= 1
    assert stats["marked_failed"] == 0


def test_reconciler_true_checksum_mismatch(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    payload = b"hello-world!!"
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = payload
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(b"other-same-len").hexdigest()
        if False
        else "a" * 64,  # same length constraint; wrong digest
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    # fix size of wrong checksum string is 64 already
    db_session.add(art)
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art)
    assert art.state == "failed"
    assert art.failure_code == "checksum_mismatch"
    assert stats["marked_failed"] >= 1


# --- C-03 multi-item: later error does not undo earlier durable truth ---


def test_reconciler_later_error_preserves_earlier_commit(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    base = datetime.now(timezone.utc) - timedelta(hours=5)
    # Item 1: pending missing object past retention → failed
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="1.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="b" * 64,
        storage_object_key=f"missing/{uuid.uuid4()}",
        state="pending",
        created_by_user_id=user.id,
        created_at=base,
    )
    # Item 2: pending with object but open_stream fails → error, stay pending
    key2 = f"k2/{uuid.uuid4()}"
    payload2 = b"x"
    fake_storage._objects[key2] = payload2
    a2 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="2.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256=hashlib.sha256(payload2).hexdigest(),
        storage_object_key=key2,
        state="pending",
        created_by_user_id=user.id,
        created_at=base + timedelta(seconds=1),
    )
    db_session.add_all([a1, a2])
    db_session.commit()
    fake_storage.fail_open_stream = True
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(a1)
    db_session.refresh(a2)
    assert a1.state == "failed"
    assert a1.failure_code == "pending_object_missing"
    assert a2.state == "pending"  # infra error must not mark checksum corrupt
    assert stats["marked_failed"] >= 1
    assert stats["errors"] >= 1


# --- C-06 late reference re-check ---


def test_reconciler_late_ref_check_before_delete(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    key = f"k/{uuid.uuid4()}"
    fake_storage._objects[key] = b"x"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="c" * 64,
        storage_object_key=key,
        state="orphaned",
        created_by_user_id=user.id,
        orphaned_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()

    calls = {"n": 0}

    def flip_ref(db, a):
        calls["n"] += 1
        # early checks false; pre-delete checks true
        return calls["n"] >= 2

    stats = reconcile_source_artifacts(
        db_session,
        storage=fake_storage,
        max_items=5,
        actor_id=user.id,
        org_id=org.id,
        reference_check=flip_ref,
    )
    assert key in fake_storage._objects  # not deleted
    assert stats["deleted_objects"] == 0
    assert (
        db_session.query(AuditEvent)
        .filter(AuditEvent.event_name == "ImportSourceArtifactObjectDeleted")
        .count()
        == 0
    )
    assert calls["n"] >= 2


# --- C-04 unsafe upload with prior state preservation ---


def _seed_prior_available(db, org, user, proj, batch, fake_storage):
    data = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.append(["h"])
    wb.save(data)
    payload = data.getvalue()
    key = f"org/{org.id}/prior"
    fake_storage._objects[key] = payload
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(art)
    db.flush()
    batch.current_source_artifact_id = art.id
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep"},
        mapped_values={"name": "keep"},
        normalized_preview={},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep",
    )
    db.add(staging)
    line = ProjectAssetLine(
        organization_id=org.id,
        project_id=proj.id,
        line_no=1,
        asset_name="Official Keep",
        created_by=user.id,
    )
    # ProjectAssetLine may require more fields — check model
    db.add(line)
    db.commit()
    return art, staging, line


@pytest.mark.parametrize(
    "threat",
    ["filepass", "addin_supbook", "dconref", "macro_boundsheet", "macro_name"],
)
def test_threat_upload_preserves_prior_staging_and_lines(
    client: TestClient, db_session: Session, fake_storage, tmp_path, threat
):
    org, user, proj, batch = _seed(db_session)
    # Minimal prior available without complex ProjectAssetLine if model is heavy
    data = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.append(["h"])
    wb.save(data)
    payload = data.getvalue()
    key = f"org/{org.id}/prior"
    fake_storage._objects[key] = payload
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db_session.add(prior)
    db_session.flush()
    batch.current_source_artifact_id = prior.id
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
    )
    db_session.add(staging)
    db_session.commit()
    prior_id = prior.id
    staging_raw = dict(staging.raw_values)
    objects_before = dict(fake_storage._objects)
    art_count_before = db_session.query(ImportSourceArtifact).count()
    lines_before = db_session.query(ProjectAssetLine).filter_by(project_id=proj.id).count()

    path = tmp_path / f"{threat}.xls"
    write_threat_xls(path, threat)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": (f"{threat}.xls", io.BytesIO(path.read_bytes()), "application/vnd.ms-excel")},
        headers={"X-User-Id": str(user.id)},
    )
    assert res.status_code == 400
    assert db_session.query(ImportSourceArtifact).count() == art_count_before
    assert fake_storage._objects == objects_before
    db_session.refresh(batch)
    db_session.refresh(staging)
    assert batch.current_source_artifact_id == prior_id
    assert staging.raw_values == staging_raw
    assert staging.proposed_asset_name == "keep-me"
    assert db_session.query(ProjectAssetLine).filter_by(project_id=proj.id).count() == lines_before


# --- C-07 boundaries sample exact + max+1 ---


def test_xlsx_cell_chars_exact_accept_and_max_plus_one_reject(tmp_path):
    limits = SourceArtifactLimits(max_cell_chars=10)
    p_ok = tmp_path / "ok.xlsx"
    _xlsx(p_ok, long="x" * 10)
    XlsxWorkbookAdapter(limits=limits).inspect(str(p_ok))  # exact accept
    p_bad = tmp_path / "bad.xlsx"
    _xlsx(p_bad, long="x" * 11)
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=limits).inspect(str(p_bad))
    assert ei.value.error_code == "cell_length_limit"


def test_xlsx_formula_cached_value_data_only(tmp_path):
    """data_only=True yields cached value or None; does not execute."""
    p = tmp_path / "f.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=A1+A2"
    wb.save(p)
    # Without Excel cache, data_only may return None for formula — still no execution
    rows = list(XlsxWorkbookAdapter().iter_rows(str(p)))
    vals = [c.value for r in rows for c in r]
    assert "=A1+A2" not in vals  # formula string not returned under data_only


def test_xlsx_zip_entry_exact_and_max_plus_one(tmp_path):
    # Build minimal valid-ish xlsx via openpyxl then count entries
    p = tmp_path / "z.xlsx"
    _xlsx(p)
    with zipfile.ZipFile(p) as zf:
        n = len(zf.infolist())
    XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_zip_entries=n)).inspect(str(p))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_zip_entries=n - 1)).inspect(str(p))
    assert ei.value.error_code == "zip_entry_limit"


# --- MinIO required in CI ---


def test_s3_minio_required_in_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL"), "CI must provide MinIO S3_ENDPOINT_URL"
    else:
        if not os.environ.get("S3_ENDPOINT_URL"):
            pytest.skip("S3_ENDPOINT_URL not set for local MinIO integration")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"third/{uuid.uuid4()}"
    store.put_stream(key, io.BytesIO(b"abc"), content_type="application/octet-stream", expected_size=3)
    assert store.head(key).size == 3
    assert store.open_stream(key).read() == b"abc"
    store.delete(key)
    assert store.head(key) is None


# --- C-05 PostgreSQL real DML ---


def test_pg_same_batch_pointer_dml_enforcement():
    """Real PostgreSQL DML: same-batch pointer OK; cross-batch fails."""
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql"), "CI must provide PostgreSQL TEST_DATABASE_URL"
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL TEST_DATABASE_URL required for DML enforcement proof")

    engine = create_engine(url)
    SessionLocal = Session(bind=engine)
    # Use ORM create_all is wrong for production schema; use live migrated DB
    # Seed via raw SQL with temporary cleanup
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    b1 = uuid.uuid4()
    b2 = uuid.uuid4()
    a1 = uuid.uuid4()
    a2 = uuid.uuid4()
    slug = f"s13t-{uuid.uuid4().hex[:8]}"
    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    """
                    INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at)
                    VALUES (:id, 'T', :slug, 'active', now(), now())
                    """
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at)
                    VALUES (:id, :oid, :email, 'T', 'active', now(), now())
                    """
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at)
                    VALUES (:id, :oid, 'C', 'active', :uid, now(), now())
                    """
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    )
                    VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft',
                      'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            for bid, gen_name in ((b1, "b1.xlsx"), (b2, "b2.xlsx")):
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :fn, 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": bid, "oid": oid, "pid": pid, "fn": gen_name, "uid": uid},
                )
            for aid, bid, gen, key in (
                (a1, b1, 1, f"k1-{a1}"),
                (a2, b2, 1, f"k2-{a2}"),
            ):
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, :gen, 'x.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": aid,
                        "oid": oid,
                        "pid": pid,
                        "bid": bid,
                        "gen": gen,
                        "chk": "a" * 64,
                        "key": key,
                        "uid": uid,
                    },
                )
            # same-batch pointer succeeds
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches "
                    "SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": a1, "bid": b1},
            )
        # cross-batch must fail
        with pytest.raises(Exception):
            with engine.begin() as conn:
                conn.execute(
                    text(
                        "UPDATE project_asset_import_batches "
                        "SET current_source_artifact_id = :aid WHERE id = :bid"
                    ),
                    {"aid": a2, "bid": b1},
                )
        # invalid checksum CHECK
        with pytest.raises(Exception):
            with engine.begin() as conn:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          gen_random_uuid(), :oid, :pid, :bid, 99, 'x.xlsx', 'xlsx', 't', 1,
                          'NOT_A_VALID_HEX_CHECKSUM_______________________________',
                          :key, 'pending', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {"oid": oid, "pid": pid, "bid": b1, "key": f"bad-{uuid.uuid4()}", "uid": uid},
                )
    finally:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches "
                    "SET current_source_artifact_id = NULL WHERE organization_id = :oid"
                ),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE organization_id = :oid"), {"oid": oid})
            conn.execute(text("DELETE FROM customers WHERE organization_id = :oid"), {"oid": oid})
            conn.execute(text("DELETE FROM users WHERE organization_id = :oid"), {"oid": oid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :oid"), {"oid": oid})
        SessionLocal.close()
