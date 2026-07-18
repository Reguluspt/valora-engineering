"""S13-PR-003 deterministic workbook-structure evidence and adversarial gates."""
from __future__ import annotations

import io
import hashlib
import json
import os
import tempfile
import threading
import time
import uuid
from pathlib import Path

import openpyxl
import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event, func, text
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.db import Base, get_db
from app.main import app as fastapi_app
import app.modules.excel_import.models  # noqa: F401
import app.modules.excel_import.application.workbook_structure_service as structure_service
from app.modules.excel_import.application.adapters import detect_format_and_adapter
from app.modules.excel_import.application.workbook_structure_service import (
    analyze_source_artifact_structure,
    get_structure_snapshot,
    list_structure_snapshots,
)
import app.modules.excel_import.domain.workbook_structure as structure_domain
from app.modules.excel_import.domain.workbook_structure import (
    STRUCTURE_RULE_VERSION,
    analyze_workbook_structure,
    canonical_payload_digest,
    classify_row,
    require_review_for_drift,
)
from app.modules.excel_import.domain.workbook_adapter import (
    AdapterInspectionResult,
    CellValue,
    MergedRegion,
    SheetSummary,
    WorkbookFormat,
)
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from app.modules.excel_import.models import (
    ImportSourceArtifact,
    WorkbookStructureSnapshot,
)
from app.modules.project_master_data.models import (
    AuditEvent,
    Customer,
    CustomerStatus,
    ImportBatchStatus,
    OrganizationProfile,
    OrganizationStatus,
    Project,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ProjectAssetLine,
    ProjectWorkflowStatus,
    Role,
    User,
    UserRole,
    UserStatus,
)


@pytest.fixture
def structure_db() -> Session:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture
def structure_storage():
    storage = FakeObjectStorage()
    set_object_storage_override(storage)
    try:
        yield storage
    finally:
        set_object_storage_override(None)


@pytest.fixture
def structure_client(structure_db: Session, structure_storage) -> TestClient:
    def override_get_db():
        yield structure_db

    fastapi_app.dependency_overrides[get_db] = override_get_db
    try:
        yield TestClient(fastapi_app)
    finally:
        fastapi_app.dependency_overrides.pop(get_db, None)


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Structure Org",
        organization_slug=f"structure-{uuid.uuid4().hex[:8]}",
        status=OrganizationStatus.ACTIVE,
    )
    other_org = OrganizationProfile(
        legal_name="Other Org",
        organization_slug=f"other-{uuid.uuid4().hex[:8]}",
        status=OrganizationStatus.ACTIVE,
    )
    db.add_all([org, other_org])
    db.commit()

    role = Role(
        code=f"structure-editor-{uuid.uuid4().hex[:8]}",
        display_name="Structure editor",
        permissions=["project:read", "workbench:edit"],
    )
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"structure-{uuid.uuid4().hex[:8]}@example.com",
        full_name="Structure Editor",
        status=UserStatus.ACTIVE,
    )
    other_user = User(
        organization_id=other_org.id,
        email=f"other-{uuid.uuid4().hex[:8]}@example.com",
        full_name="Other Editor",
        status=UserStatus.ACTIVE,
    )
    db.add_all([user, other_user])
    db.commit()
    db.add_all(
        [
            UserRole(user_id=user.id, role_id=role.id, is_active=True),
            UserRole(user_id=other_user.id, role_id=role.id, is_active=True),
        ]
    )
    db.commit()

    customer = Customer(
        organization_id=org.id,
        legal_name="Structure Customer",
        status=CustomerStatus.ACTIVE,
        created_by=user.id,
    )
    db.add(customer)
    db.commit()
    project = Project(
        organization_id=org.id,
        customer_id=customer.id,
        name="Structure Project",
        code=f"SP-{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(project)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=project.id,
        source_filename="structure.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, other_user, project, batch


def _populate_pd001_sheet(ws, *, header_row: int = 5, multi_row: bool = False) -> None:
    ws.title = "PD-001"
    ws.merge_cells("A1:H1")
    ws["A1"] = "BẢNG KÊ TÀI SẢN THẨM ĐỊNH"
    ws["A2"] = "Khách hàng: ACME"
    ws["A3"] = "Đơn vị: VNĐ"
    if header_row > 5:
        ws.insert_rows(4, header_row - 5)
    if multi_row:
        ws.cell(header_row, 1, "Thông tin")
        ws.cell(header_row, 5, "Giá trị")
        ws.cell(header_row + 1, 1, "STT")
        ws.cell(header_row + 1, 2, "Tên tài sản")
        ws.cell(header_row + 1, 3, "Đặc điểm")
        ws.cell(header_row + 1, 4, "ĐVT")
        ws.cell(header_row + 1, 5, "Số lượng")
        ws.cell(header_row + 1, 6, "Đơn giá")
        ws.cell(header_row + 1, 7, "Thành tiền")
        ws.cell(header_row + 1, 8, "Ghi chú")
        data_row = header_row + 2
    else:
        headers = [
            "STT",
            "Tên tài sản",
            "Đặc điểm",
            "ĐVT",
            "Số lượng",
            "Đơn giá",
            "Thành tiền",
            "Ghi chú",
        ]
        for column, value in enumerate(headers, 1):
            ws.cell(header_row, column, value)
        data_row = header_row + 1

    rows = [
        [1, "Máy bơm", "Model X", "cái", 2, 100, 200, None],
        [2, "Máy phát", "Model Y", "bộ", 1, 300, 300, None],
        ["PHẦN ĐIỆN", None, None, None, None, None, None, None],
        [3, "Tủ điện", "IP54", "bộ", 1, 400, 400, None],
        ["Cộng phần điện", None, None, None, None, None, 400, None],
        ["PHẦN NƯỚC", None, None, None, None, None, None, None],
        [4, "Máy lọc", "Model Z", "bộ", 1, 500, 500, None],
        ["TỔNG CỘNG", None, None, None, None, None, 900, None],
        ["Ghi chú: tài sản đã qua sử dụng", None, None, None, None, None, None, None],
        ["Dòng chưa xác định", None, None, None, None, None, None, None],
    ]
    for offset, values in enumerate(rows):
        for column, value in enumerate(values, 1):
            ws.cell(data_row + offset, column, value)


def _xlsx_bytes(*, header_row: int = 5, multi_row: bool = False, two_sheets: bool = False) -> bytes:
    workbook = openpyxl.Workbook()
    _populate_pd001_sheet(workbook.active, header_row=header_row, multi_row=multi_row)
    if two_sheets:
        second = workbook.create_sheet("PD-002")
        _populate_pd001_sheet(second, header_row=header_row, multi_row=multi_row)
        second.title = "PD-002"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _xls_bytes() -> bytes:
    pytest.importorskip("xlwt")
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("PD-001")
    sheet.write_merge(0, 0, 0, 7, "BẢNG KÊ TÀI SẢN THẨM ĐỊNH")
    sheet.write(1, 0, "Khách hàng: ACME")
    sheet.write(2, 0, "Đơn vị: VNĐ")
    headers = [
        "STT",
        "Tên tài sản",
        "Đặc điểm",
        "ĐVT",
        "Số lượng",
        "Đơn giá",
        "Thành tiền",
        "Ghi chú",
    ]
    for column, value in enumerate(headers):
        sheet.write(4, column, value)
    for row, values in enumerate(
        [
            [1, "Máy bơm", "Model X", "cái", 2, 100, 200],
            [2, "Máy phát", "Model Y", "bộ", 1, 300, 300],
            ["PHẦN ĐIỆN"],
            [3, "Tủ điện", "IP54", "bộ", 1, 400, 400],
            ["Cộng phần điện", "", "", "", "", "", 400],
            ["PHẦN NƯỚC"],
            [4, "Máy lọc", "Model Z", "bộ", 1, 500, 500],
            ["TỔNG CỘNG", "", "", "", "", "", 900],
            ["Ghi chú: tài sản đã qua sử dụng"],
        ],
        5,
    ):
        for column, value in enumerate(values):
            sheet.write(row, column, value)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _upload(
    client: TestClient,
    user: User,
    project: Project,
    batch: ProjectAssetImportBatch,
    data: bytes,
    *,
    suffix: str = "xlsx",
) -> dict:
    content_type = (
        "application/vnd.ms-excel"
        if suffix == "xls"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response = client.post(
        f"/api/v1/projects/{project.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": (f"source.{suffix}", io.BytesIO(data), content_type)},
        headers={"X-User-Id": str(user.id)},
    )
    assert response.status_code == 201, response.text
    return response.json()


def _snapshot_path(project: Project, batch: ProjectAssetImportBatch, artifact_id: str) -> str:
    return (
        f"/api/v1/projects/{project.id}/asset-imports/{batch.id}/source-artifacts/"
        f"{artifact_id}/structure-snapshots"
    )


@pytest.mark.parametrize(
    ("suffix", "factory"),
    [("xlsx", _xlsx_bytes), ("xls", _xls_bytes)],
)
def test_real_adapters_propose_row_five_and_classify_markers(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
    suffix,
    factory,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, factory(), suffix=suffix)
    path = _snapshot_path(project, batch, artifact["id"])
    response = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    assert response.status_code == 201, response.text
    body = response.json()
    payload = body["structure_payload"]
    assert body["rule_version"] == STRUCTURE_RULE_VERSION
    assert body["disposition"] == "proposed"
    assert payload["candidates"][0]["sheet_name"] == "PD-001"
    assert payload["candidates"][0]["header_start_row"] == 5
    assert payload["candidates"][0]["header_end_row"] == 5
    assert payload["candidates"][0]["header_labels"][0] == "STT"
    counts = payload["row_classification"]["counts"]
    assert counts["asset"] == 4
    assert counts["section"] == 2
    assert counts["subtotal"] == 1
    assert counts["total"] == 1
    assert counts["note"] == 1
    assert body["analysis_digest_sha256"] == canonical_payload_digest(payload)


def test_multirow_header_is_detected_without_absorbing_asset_row(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(
        structure_client,
        user,
        project,
        batch,
        _xlsx_bytes(multi_row=True),
    )
    response = structure_client.post(
        _snapshot_path(project, batch, artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert response.status_code == 201, response.text
    candidate = response.json()["structure_payload"]["candidates"][0]
    assert candidate["header_start_row"] == 5
    assert candidate["header_end_row"] == 6
    assert candidate["data_start_row"] == 7
    assert "Thông tin" in candidate["header_labels"][0]
    assert "STT" in candidate["header_labels"][0]


def test_competing_sheets_fail_to_review_and_title_never_wins(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(
        structure_client,
        user,
        project,
        batch,
        _xlsx_bytes(two_sheets=True),
    )
    response = structure_client.post(
        _snapshot_path(project, batch, artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert response.status_code == 201, response.text
    payload = response.json()["structure_payload"]
    assert payload["disposition"] == "review_required"
    assert "competing_candidates" in payload["disposition_reasons"]
    assert payload["candidates"][0]["header_start_row"] == 5
    assert all(candidate["header_start_row"] != 1 for candidate in payload["candidates"])


def test_structure_drift_forces_review_without_changing_original_payload():
    current = {
        "disposition": "proposed",
        "disposition_reasons": [],
        "proposed_candidate_index": 0,
        "candidates": [
            {
                "sheet_name": "PD-001",
                "header_start_row": 6,
                "header_end_row": 6,
                "data_start_row": 7,
                "candidate_table_bounds": {"min_column": 1, "max_column": 8},
            }
        ],
    }
    previous = {
        **current,
        "candidates": [
            {
                **current["candidates"][0],
                "header_start_row": 5,
                "header_end_row": 5,
                "data_start_row": 6,
            }
        ],
    }
    result = require_review_for_drift(current, previous)
    assert result["disposition"] == "review_required"
    assert "structure_drift_from_previous_generation" in result["disposition_reasons"]
    assert current["disposition"] == "proposed"
    assert current["disposition_reasons"] == []


def test_generation_drift_is_detected_from_latest_prior_source(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    first_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    first = structure_client.post(
        _snapshot_path(project, batch, first_artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert first.status_code == 201
    assert first.json()["disposition"] == "proposed"

    second_artifact = _upload(
        structure_client,
        user,
        project,
        batch,
        _xlsx_bytes(header_row=6),
    )
    second = structure_client.post(
        _snapshot_path(project, batch, second_artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert second.status_code == 201, second.text
    payload = second.json()["structure_payload"]
    assert payload["candidates"][0]["header_start_row"] == 6
    assert payload["disposition"] == "review_required"
    assert "structure_drift_from_previous_generation" in payload["disposition_reasons"]


def test_reordered_headers_force_review_even_when_geometry_is_unchanged(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    first_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    first = structure_client.post(
        _snapshot_path(project, batch, first_artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert first.status_code == 201

    workbook = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes()))
    sheet = workbook["PD-001"]
    sheet["B5"], sheet["C5"] = sheet["C5"].value, sheet["B5"].value
    reordered = io.BytesIO()
    workbook.save(reordered)
    second_artifact = _upload(
        structure_client,
        user,
        project,
        batch,
        reordered.getvalue(),
    )
    second = structure_client.post(
        _snapshot_path(project, batch, second_artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert second.status_code == 201, second.text
    payload = second.json()["structure_payload"]
    assert payload["disposition"] == "review_required"
    assert "structure_drift_from_previous_generation" in payload["disposition_reasons"]


def test_scan_candidate_retention_and_preview_are_bounded_and_iterators_close():
    class TrackingRows:
        def __init__(self, rows):
            self._rows = iter(rows)
            self.next_calls = 0
            self.closed = False

        def __iter__(self):
            return self

        def __next__(self):
            self.next_calls += 1
            return next(self._rows)

        def close(self):
            self.closed = True

    def cells(row_number, values):
        return tuple(
            CellValue(
                row=row_number,
                column=index,
                coordinate=f"R{row_number}C{index}",
                value=value,
                cell_type="number" if isinstance(value, (int, float)) else "string",
            )
            for index, value in enumerate(values, 1)
        )

    rows = [
        cells(1, ["BẢNG KÊ TÀI SẢN THẨM ĐỊNH"]),
        cells(2, ["Khách hàng: ACME"]),
        cells(3, ["Đơn vị: VNĐ"]),
        cells(4, [None]),
        cells(5, ["STT", "Tên tài sản", "ĐVT", "Số lượng", "Thành tiền"]),
    ]
    rows.extend(cells(index, [index - 5, f"Tài sản {index}", "cái", 1, 100]) for index in range(6, 501))
    trackers = []

    def row_provider(_sheet_name):
        tracker = TrackingRows(rows)
        trackers.append(tracker)
        return tracker

    inspection = AdapterInspectionResult(
        format=WorkbookFormat.XLSX,
        adapter_name="tracking",
        adapter_version="1",
        sheet_names=("PD-001",),
        sheets=(SheetSummary(name="PD-001", max_row=500, max_column=5),),
    )
    payload = analyze_workbook_structure(inspection, row_provider)
    assert payload["candidate_count"] <= 25
    assert len(payload["row_classification"]["preview"]) == 200
    assert payload["row_classification"]["preview_truncated"] is True
    assert trackers[0].next_calls == 200
    assert all(tracker.closed for tracker in trackers)


def test_append_only_replay_digest_and_no_mutation(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact["id"])
    before_staging = structure_db.query(ProjectAssetImportStagingRow).count()
    before_lines = structure_db.query(ProjectAssetLine).count()

    first = structure_client.post(
        path,
        headers={"X-User-Id": str(user.id)},
        json={
            "rule_version": "caller-override",
            "candidate": {"header_start_row": 1},
            "disposition": "proposed",
        },
    )
    second = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    assert first.status_code == 201, first.text
    assert second.status_code == 201, second.text
    assert first.json()["snapshot_version"] == 1
    assert second.json()["snapshot_version"] == 2
    assert first.json()["rule_version"] == STRUCTURE_RULE_VERSION
    assert first.json()["structure_payload"]["candidates"][0]["header_start_row"] == 5
    assert first.json()["analysis_digest_sha256"] == second.json()["analysis_digest_sha256"]

    listed = structure_client.get(path, headers={"X-User-Id": str(user.id)})
    assert listed.status_code == 200, listed.text
    assert [item["snapshot_version"] for item in listed.json()] == [1, 2]
    snapshot_id = first.json()["id"]
    replayed = structure_client.get(
        f"{path}/{snapshot_id}",
        headers={"X-User-Id": str(user.id)},
    )
    assert replayed.status_code == 200
    assert replayed.json() == first.json()
    assert structure_db.query(ProjectAssetImportStagingRow).count() == before_staging
    assert structure_db.query(ProjectAssetLine).count() == before_lines

    events = (
        structure_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "WorkbookStructureAnalyzed")
        .all()
    )
    assert len(events) == 2
    serialized_audit = json.dumps(events[0].payload, ensure_ascii=False)
    assert "Máy bơm" not in serialized_audit
    assert "header_labels" not in serialized_audit


def test_cross_tenant_snapshot_routes_are_safe_not_found(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, other_user, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact["id"])
    owner_response = structure_client.post(
        path,
        headers={"X-User-Id": str(user.id)},
    )
    assert owner_response.status_code == 201
    response = structure_client.post(
        path,
        headers={"X-User-Id": str(other_user.id)},
    )
    assert response.status_code == 404
    listed = structure_client.get(path, headers={"X-User-Id": str(other_user.id)})
    assert listed.status_code == 404
    replayed = structure_client.get(
        f"{path}/{owner_response.json()['id']}",
        headers={"X-User-Id": str(other_user.id)},
    )
    assert replayed.status_code == 404
    assert structure_db.query(WorkbookStructureSnapshot).count() == 1


def test_non_available_source_fails_before_object_read(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage: FakeObjectStorage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact_body = _upload(structure_client, user, project, batch, _xlsx_bytes())
    artifact = structure_db.get(ImportSourceArtifact, uuid.UUID(artifact_body["id"]))
    artifact.state = "orphaned"
    structure_db.commit()
    structure_storage.fail_open_stream = True

    response = structure_client.post(
        _snapshot_path(project, batch, artifact_body["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert response.status_code == 409
    assert response.json()["detail"]["error_code"] == "source_artifact_not_available"
    assert structure_db.query(WorkbookStructureSnapshot).count() == 0


@pytest.mark.parametrize("failure", ["missing", "short_read", "checksum"])
def test_object_evidence_failure_creates_no_snapshot_or_success_audit(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage: FakeObjectStorage,
    failure: str,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact_body = _upload(structure_client, user, project, batch, _xlsx_bytes())
    artifact = structure_db.get(ImportSourceArtifact, uuid.UUID(artifact_body["id"]))
    original = structure_storage._objects[artifact.storage_object_key]
    temp_dir = Path(tempfile.gettempdir())
    temp_before = set(temp_dir.glob("valora-structure-*"))
    if failure == "missing":
        del structure_storage._objects[artifact.storage_object_key]
    elif failure == "short_read":
        structure_storage.truncate_open_to = len(original) - 1
    else:
        structure_storage._objects[artifact.storage_object_key] = b"X" + original[1:]

    response = structure_client.post(
        _snapshot_path(project, batch, artifact_body["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert response.status_code == 409
    assert set(temp_dir.glob("valora-structure-*")) == temp_before
    assert structure_db.query(WorkbookStructureSnapshot).count() == 0
    assert (
        structure_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "WorkbookStructureAnalyzed")
        .count()
        == 0
    )
    assert structure_db.query(ProjectAssetImportStagingRow).count() == 0
    assert structure_db.query(ProjectAssetLine).count() == 0


def test_tampered_snapshot_fails_closed_on_list_and_get(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact["id"])
    created = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    assert created.status_code == 201
    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    altered = dict(snapshot.structure_payload)
    altered["candidate_count"] = altered["candidate_count"] + 1
    snapshot.structure_payload = altered
    structure_db.commit()

    listed = structure_client.get(path, headers={"X-User-Id": str(user.id)})
    assert listed.status_code == 500
    replayed = structure_client.get(
        f"{path}/{snapshot.id}",
        headers={"X-User-Id": str(user.id)},
    )
    assert replayed.status_code == 500


def test_canonical_digest_is_order_independent_and_tamper_sensitive():
    left = {"b": [2, 1], "a": {"y": "đ", "x": 1}}
    right = {"a": {"x": 1, "y": "đ"}, "b": [2, 1]}
    assert canonical_payload_digest(left) == canonical_payload_digest(right)
    assert canonical_payload_digest(left) != canonical_payload_digest({**right, "b": [1, 2]})


def _cells(row_number: int, values: list[object]) -> tuple[CellValue, ...]:
    return tuple(
        CellValue(
            row=row_number,
            column=index,
            coordinate=f"R{row_number}C{index}",
            value=value,
            cell_type=(
                "boolean"
                if isinstance(value, bool)
                else "number"
                if isinstance(value, (int, float))
                else "empty"
                if value is None
                else "string"
            ),
        )
        for index, value in enumerate(values, 1)
    )


@pytest.mark.parametrize(
    ("values", "expected"),
    [
        ([99, "TỔNG CỘNG", 100], "total"),
        ([99, 100, "TỔNG CỘNG"], "total"),
        (["I", "Cộng phần điện", 100], "subtotal"),
        ([2, "Ghi chú: kiểm tra", 100], "note"),
        ([3, "PHẦN ĐIỆN", 100], "section"),
        ([1, "Tổng công ty ABC", 100], "asset"),
        (["X01", "MÁY BƠM"], "unresolved"),
        (["Máy bơm", 100], "unresolved"),
        ([1, "Máy bơm", 100], "asset"),
    ],
)
def test_v2_marker_precedence_and_asset_fail_closed(values, expected):
    assert classify_row(_cells(7, values)).row_class.value == expected


def _analyze_rows(rows: list[list[object]], *, merged: bool = False) -> tuple[dict, list]:
    cell_rows = [_cells(index, values) for index, values in enumerate(rows, 1)]
    trackers = []

    class TrackedIterator:
        def __init__(self):
            self._rows = iter(cell_rows)
            self.closed = False

        def __iter__(self):
            return self

        def __next__(self):
            return next(self._rows)

        def close(self):
            self.closed = True

    def provider(_sheet_name):
        iterator = TrackedIterator()
        trackers.append(iterator)
        return iterator

    max_column = max((len(row) for row in rows), default=0)
    inspection = AdapterInspectionResult(
        format=WorkbookFormat.XLSX,
        adapter_name="adversarial",
        adapter_version="1",
        sheet_names=("PD-001",),
        sheets=(SheetSummary(name="PD-001", max_row=len(rows), max_column=max_column),),
    )
    return analyze_workbook_structure(inspection, provider), trackers


def test_v2_header_cannot_absorb_immediate_section_and_rectangle_excludes_remote_cells():
    rows = [
        ["BẢNG KÊ"],
        [None],
        [None],
        [None],
        ["STT", "Tên tài sản", "Số lượng"],
        ["PHẦN ĐIỆN", None, None],
        [1, "Máy bơm", 2],
        [2, "Máy phát", 1],
        ["TỔNG CỘNG", None, 3],
    ]
    rows[0].extend([None] * 24 + ["ghi chú xa Z1"])
    payload, trackers = _analyze_rows(rows)
    candidate = payload["candidates"][0]
    assert candidate["header_start_row"] == 5
    assert candidate["header_end_row"] == 5
    assert candidate["candidate_table_bounds"]["max_column"] == 3
    assert payload["row_classification"]["counts"]["section"] == 1
    assert all(item["row_number"] >= 6 for item in payload["row_classification"]["preview"])
    assert len(trackers) == 2
    assert all(iterator.closed for iterator in trackers)


def test_v2_horizontal_tables_are_separate_and_blank_trailing_header_is_positional():
    rows = [
        ["STT", "Tên tài sản", None, None, "STT", "Tên tài sản"],
        [1, "Máy A", "cái", None, 1, "Máy B"],
        [2, "Máy C", None, None, 2, "Máy D"],
    ]
    payload, _trackers = _analyze_rows(rows)
    bounds = {
        (
            item["candidate_table_bounds"]["min_column"],
            item["candidate_table_bounds"]["max_column"],
        )
        for item in payload["candidates"]
        if item["header_start_row"] == 1
    }
    assert (1, 3) in bounds
    assert (5, 6) in bounds
    left = next(
        item
        for item in payload["candidates"]
        if item["header_start_row"] == 1
        and item["candidate_table_bounds"]["min_column"] == 1
    )
    assert left["header_labels"] == ["STT", "Tên tài sản", None]


def test_v2_late_adjacent_blank_header_evidence_extends_once_and_forces_review():
    rows = [["STT", "Tên tài sản"]]
    rows.extend([[index, f"Máy {index}"] for index in range(1, 26)])
    rows.append([26, "Máy 26", "cái"])
    rows.extend([[index, f"Máy {index}"] for index in range(27, 30)])
    payload, _trackers = _analyze_rows(rows)
    candidate = payload["candidates"][0]
    assert candidate["candidate_table_bounds"]["max_column"] == 3
    assert candidate["header_labels"][-1] is None
    assert "late_blank_header_column_evidence" in candidate["boundary_flags"]
    assert payload["disposition"] == "review_required"


@pytest.mark.parametrize(
    ("tail", "reason", "max_row"),
    [
        ([[None, None], [2, "Máy B"]], "sheet_end", 4),
        ([[None, None], [None, None], [2, "Máy B"]], "blank_run", 2),
        ([["không rõ"], ["vẫn không rõ"]], "ambiguous", 4),
        (
            [["TỔNG CỘNG", 100], ["Ghi chú: một"], ["note: hai"], ["* ba"], ["note: bốn"]],
            "terminal_total",
            6,
        ),
        (
            [["Ghi chú: một"], ["note: hai"], ["* ba"], ["note: bốn"]],
            "ambiguous",
            5,
        ),
    ],
)
def test_v2_vertical_boundary_state_machine(tail, reason, max_row):
    rows = [["STT", "Tên tài sản"], [1, "Máy A"], *tail]
    payload, _trackers = _analyze_rows(rows)
    candidate = payload["candidates"][0]
    assert candidate["boundary_reason"] == reason
    assert candidate["candidate_table_bounds"]["max_row"] == max_row


def test_repeated_header_ends_first_rectangle_before_second_table():
    rows = [
        ["STT", "Tên tài sản", "Số lượng"],
        [1, "Máy A", 1],
        [2, "Máy B", 1],
        [None, None, None],
        ["STT", "Tên tài sản", "Số lượng"],
        [1, "Máy C", 1],
    ]
    payload, _trackers = _analyze_rows(rows)
    first = next(item for item in payload["candidates"] if item["header_start_row"] == 1)
    assert first["boundary_reason"] == "next_header"
    assert first["candidate_table_bounds"]["max_row"] == 3
    primary_bounds = payload["candidates"][0]["candidate_table_bounds"]
    classification = payload["row_classification"]
    assert classification["candidate_table_bounds"] == primary_bounds
    assert all(
        payload["candidates"][0]["data_start_row"]
        <= item["row_number"]
        <= primary_bounds["max_row"]
        for item in classification["preview"]
    )


def test_post_total_accepts_exactly_three_empty_tail_rows():
    rows = [
        ["STT", "Tên tài sản"],
        [1, "Máy A"],
        ["TỔNG CỘNG", 100],
        [None, None],
        [None, None],
        [None, None],
        [None, None],
    ]
    payload, _trackers = _analyze_rows(rows)
    candidate = payload["candidates"][0]
    assert candidate["candidate_table_bounds"]["max_row"] == 6
    assert candidate["boundary_reason"] == "terminal_total"
    assert "post_total_tail_exceeded" in candidate["boundary_flags"]
    assert payload["row_classification"]["counts"]["empty"] == 3


def test_header_beyond_initial_scan_ends_region_and_forces_review():
    rows = [["STT", "Tên tài sản", "Số lượng"]]
    rows.extend([[index, f"Máy {index}", 1] for index in range(1, 200)])
    rows.append(["STT", "Tên tài sản", "Số lượng"])
    rows.append([1, "Máy bảng hai", 1])
    payload, _trackers = _analyze_rows(rows)
    first = next(item for item in payload["candidates"] if item["header_start_row"] == 1)
    assert first["boundary_reason"] == "next_header"
    assert first["candidate_table_bounds"]["max_row"] == 200
    assert "additional_table_beyond_header_scan" in first["boundary_flags"]
    assert payload["disposition"] == "review_required"


class _CloseFailingStream(io.BytesIO):
    def close(self):
        raise OSError("forced close failure")


class _CloseFailingStorage(FakeObjectStorage):
    def open_stream(self, key: str):
        data = self._objects[key]
        if self.truncate_open_to is not None:
            data = data[: self.truncate_open_to]
        return _CloseFailingStream(data)


@pytest.mark.parametrize(
    ("short_read", "status", "error_code"),
    [
        (False, 503, "source_stream_close_failed"),
        (True, 409, "source_size_mismatch"),
    ],
)
def test_stream_close_failure_precedence_leaks_no_temp_or_evidence(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage: FakeObjectStorage,
    short_read: bool,
    status: int,
    error_code: str,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact_body = _upload(structure_client, user, project, batch, _xlsx_bytes())
    artifact = structure_db.get(ImportSourceArtifact, uuid.UUID(artifact_body["id"]))
    failing = _CloseFailingStorage()
    failing._objects = dict(structure_storage._objects)
    failing._content_types = dict(structure_storage._content_types)
    if short_read:
        failing.truncate_open_to = artifact.file_size_bytes - 1
    set_object_storage_override(failing)
    before = set(Path(tempfile.gettempdir()).glob("valora-structure-*"))
    response = structure_client.post(
        _snapshot_path(project, batch, artifact_body["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert response.status_code == status
    assert response.json()["detail"]["error_code"] == error_code
    assert set(Path(tempfile.gettempdir()).glob("valora-structure-*")) == before
    assert structure_db.query(WorkbookStructureSnapshot).count() == 0
    assert (
        structure_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "WorkbookStructureAnalyzed")
        .count()
        == 0
    )


def test_adapter_close_failure_cannot_bypass_temp_cleanup(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage: FakeObjectStorage,
    monkeypatch,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    real_detect = structure_service.detect_format_and_adapter

    class CloseFailingAdapter:
        def __init__(self, inner):
            self.inner = inner

        def inspect(self, path):
            return self.inner.inspect(path)

        def iter_rows(self, path, sheet_name):
            return self.inner.iter_rows(path, sheet_name)

        def close(self):
            self.inner.close()
            raise OSError("forced adapter close failure")

    def failing_detect(path, filename, limits=None):
        detected, adapter = real_detect(path, filename, limits=limits)
        return detected, CloseFailingAdapter(adapter)

    monkeypatch.setattr(structure_service, "detect_format_and_adapter", failing_detect)
    before = set(Path(tempfile.gettempdir()).glob("valora-structure-*"))
    with pytest.raises(OSError, match="forced adapter close failure"):
        structure_client.post(
            _snapshot_path(project, batch, artifact["id"]),
            headers={"X-User-Id": str(user.id)},
        )
    assert set(Path(tempfile.gettempdir()).glob("valora-structure-*")) == before
    assert structure_db.query(WorkbookStructureSnapshot).count() == 0
    assert (
        structure_db.query(AuditEvent)
        .filter(AuditEvent.event_name == "WorkbookStructureAnalyzed")
        .count()
        == 0
    )


def test_immediate_predecessor_without_snapshot_is_bound_and_forces_review(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    first_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    second_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    response = structure_client.post(
        _snapshot_path(project, batch, second_artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert response.status_code == 201, response.text
    payload = response.json()["structure_payload"]
    assert payload["disposition"] == "review_required"
    assert "prior_generation_snapshot_missing" in payload["disposition_reasons"]
    assert payload["drift_reference"] == {
        "source_artifact_id": first_artifact["id"],
        "source_generation": 1,
        "snapshot_id": None,
        "snapshot_version": None,
        "rule_version": None,
        "analysis_digest_sha256": None,
    }


def test_drift_reference_is_digest_and_durable_snapshot_bound(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    first_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    first = structure_client.post(
        _snapshot_path(project, batch, first_artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert first.status_code == 201
    second_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    second_path = _snapshot_path(project, batch, second_artifact["id"])
    second = structure_client.post(second_path, headers={"X-User-Id": str(user.id)})
    assert second.status_code == 201, second.text
    payload = second.json()["structure_payload"]
    assert payload["drift_reference"]["snapshot_id"] == first.json()["id"]
    assert payload["drift_reference"]["analysis_digest_sha256"] == first.json()[
        "analysis_digest_sha256"
    ]

    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(second.json()["id"]))
    altered = json.loads(json.dumps(snapshot.structure_payload))
    altered["drift_reference"]["snapshot_version"] += 1
    snapshot.structure_payload = altered
    snapshot.analysis_digest_sha256 = canonical_payload_digest(altered)
    structure_db.commit()
    replayed = structure_client.get(
        f"{second_path}/{snapshot.id}",
        headers={"X-User-Id": str(user.id)},
    )
    assert replayed.status_code == 500
    assert replayed.json()["detail"]["error_code"] == "structure_snapshot_integrity_failure"


@pytest.mark.parametrize(
    "tamper",
    ["missing_reference", "null_reference", "missing_predecessor_artifact"],
)
def test_v2_drift_reference_shape_and_predecessor_artifact_are_integrity_bound(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
    tamper: str,
):
    _org, user, _other, project, batch = _seed(structure_db)
    first_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    second_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, second_artifact["id"])
    created = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    assert created.status_code == 201, created.text

    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    altered = json.loads(json.dumps(snapshot.structure_payload))
    if tamper == "missing_reference":
        altered.pop("drift_reference")
    elif tamper == "null_reference":
        altered["drift_reference"] = None
    else:
        altered["drift_reference"]["source_artifact_id"] = str(uuid.uuid4())
    snapshot.structure_payload = altered
    snapshot.analysis_digest_sha256 = canonical_payload_digest(altered)
    structure_db.commit()

    replayed = structure_client.get(
        f"{path}/{snapshot.id}",
        headers={"X-User-Id": str(user.id)},
    )
    assert replayed.status_code == 500
    assert replayed.json()["detail"]["error_code"] == "structure_snapshot_integrity_failure"
    assert first_artifact["id"] != second_artifact["id"]


def _postgres_engine_or_skip(*, application_name: str | None = None):
    url = os.getenv("TEST_DATABASE_URL") or os.getenv("DATABASE_URL")
    if not url or not url.startswith("postgres"):
        if os.getenv("CI") == "true":
            pytest.fail(
                "CI=true requires PostgreSQL TEST_DATABASE_URL for the S13-PR-003 "
                "serialization gate."
            )
        pytest.skip("PostgreSQL is required for the S13-PR-003 serialization gate")
    connect_args = {"connect_timeout": 5}
    if application_name:
        connect_args["application_name"] = application_name
    engine = create_engine(url, connect_args=connect_args, pool_pre_ping=True)
    try:
        with engine.connect() as connection:
            exists = connection.execute(
                text("SELECT to_regclass('workbook_structure_snapshots')")
            ).scalar_one()
        if exists is None:
            if os.getenv("CI") == "true":
                pytest.fail("CI PostgreSQL has not been migrated to the S13-PR-003 head")
            pytest.skip("PostgreSQL schema is not migrated to the S13-PR-003 head")
    except Exception:
        engine.dispose()
        raise
    return engine


def _add_available_artifact(
    db: Session,
    storage: FakeObjectStorage,
    *,
    org,
    user,
    project,
    batch,
    generation: int,
    data: bytes,
) -> ImportSourceArtifact:
    key = f"s13-pr-003/{uuid.uuid4().hex}.xlsx"
    storage._objects[key] = data
    storage._content_types[key] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    artifact = ImportSourceArtifact(
        organization_id=org.id,
        project_id=project.id,
        import_batch_id=batch.id,
        generation=generation,
        original_filename=f"generation-{generation}.xlsx",
        detected_format="xlsx",
        content_type=storage._content_types[key],
        file_size_bytes=len(data),
        checksum_sha256=hashlib.sha256(data).hexdigest(),
        storage_object_key=key,
        storage_etag=hashlib.md5(data).hexdigest(),
        state="available",
        adapter_metadata={},
        created_by_user_id=user.id,
    )
    db.add(artifact)
    db.commit()
    db.refresh(artifact)
    return artifact


def _offline_payload(data: bytes, artifact: ImportSourceArtifact) -> tuple[dict, str, str]:
    handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = handle.name
    adapter = None
    try:
        handle.write(data)
        handle.close()
        _format, adapter = detect_format_and_adapter(
            path,
            artifact.original_filename,
        )
        inspection = adapter.inspect(path)
        payload = analyze_workbook_structure(
            inspection,
            lambda sheet_name: adapter.iter_rows(path, sheet_name),
        )
        payload["source"] = {
            "source_artifact_id": str(artifact.id),
            "source_generation": artifact.generation,
            "source_checksum_sha256": artifact.checksum_sha256,
            "detected_format": artifact.detected_format,
            "adapter_name": inspection.adapter_name,
            "adapter_version": inspection.adapter_version,
        }
        payload["drift_reference"] = None
        return payload, inspection.adapter_name, inspection.adapter_version
    finally:
        if not handle.closed:
            handle.close()
        if adapter is not None:
            adapter.close()
        Path(path).unlink(missing_ok=True)


@pytest.mark.parametrize("changed_structure", [False, True])
def test_postgres_serializes_predecessor_drift_and_snapshot_versions(changed_structure: bool):
    """No-sleep proof: batch lock wait, exact predecessor binding, and unique versions."""
    observer_engine = _postgres_engine_or_skip()
    setup_session = sessionmaker(bind=observer_engine)()
    storage = FakeObjectStorage()
    try:
        org, user, _other, project, batch = _seed(setup_session)
        previous_bytes = _xlsx_bytes()
        current_bytes = _xlsx_bytes(header_row=6) if changed_structure else previous_bytes
        previous_artifact = _add_available_artifact(
            setup_session,
            storage,
            org=org,
            user=user,
            project=project,
            batch=batch,
            generation=1,
            data=previous_bytes,
        )
        current_artifact = _add_available_artifact(
            setup_session,
            storage,
            org=org,
            user=user,
            project=project,
            batch=batch,
            generation=2,
            data=current_bytes,
        )
        payload, adapter_name, adapter_version = _offline_payload(
            previous_bytes,
            previous_artifact,
        )
        previous_digest = canonical_payload_digest(payload)

        lock_session = sessionmaker(bind=observer_engine)()
        locked_batch = (
            lock_session.query(ProjectAssetImportBatch)
            .filter(ProjectAssetImportBatch.id == batch.id)
            .with_for_update()
            .one()
        )
        assert locked_batch.id == batch.id
        previous_snapshot = WorkbookStructureSnapshot(
            id=uuid.uuid4(),
            organization_id=org.id,
            project_id=project.id,
            import_batch_id=batch.id,
            source_artifact_id=previous_artifact.id,
            snapshot_version=1,
            source_checksum_sha256=previous_artifact.checksum_sha256,
            rule_version=STRUCTURE_RULE_VERSION,
            adapter_name=adapter_name,
            adapter_version=adapter_version,
            disposition=payload["disposition"],
            candidate_count=payload["candidate_count"],
            structure_payload=payload,
            analysis_digest_sha256=previous_digest,
            created_by_user_id=user.id,
        )
        lock_session.add(previous_snapshot)
        lock_session.flush()

        app_name = f"s13_pr_003_waiter_{uuid.uuid4().hex}"
        worker_engine = _postgres_engine_or_skip(application_name=app_name)
        worker_result: dict[str, object] = {}
        worker_started = threading.Event()

        def analyze_current():
            session = sessionmaker(bind=worker_engine)()
            worker_started.set()
            try:
                worker_result["snapshot"] = analyze_source_artifact_structure(
                    session,
                    org_id=org.id,
                    project_id=project.id,
                    batch_id=batch.id,
                    artifact_id=current_artifact.id,
                    current_user=user,
                    storage=storage,
                )
            except BaseException as exc:
                worker_result["error"] = exc
                session.rollback()
            finally:
                session.close()

        worker = threading.Thread(target=analyze_current, daemon=True)
        worker.start()
        assert worker_started.wait(timeout=5)
        deadline = time.monotonic() + 15
        lock_wait_observed = False
        while time.monotonic() < deadline:
            with observer_engine.connect() as connection:
                lock_wait_observed = bool(
                    connection.execute(
                        text(
                            "SELECT EXISTS (SELECT 1 FROM pg_stat_activity "
                            "WHERE application_name = :name "
                            "AND wait_event_type = 'Lock')"
                        ),
                        {"name": app_name},
                    ).scalar_one()
                )
            if lock_wait_observed:
                break
        assert lock_wait_observed, "PostgreSQL did not expose the expected batch-lock wait"
        lock_session.commit()
        worker.join(timeout=20)
        assert not worker.is_alive()
        assert "error" not in worker_result, repr(worker_result.get("error"))
        created = worker_result["snapshot"]
        reference = created.structure_payload["drift_reference"]
        assert reference["source_artifact_id"] == str(previous_artifact.id)
        assert reference["snapshot_id"] == str(previous_snapshot.id)
        assert reference["analysis_digest_sha256"] == previous_digest
        if changed_structure:
            assert created.disposition == "review_required"
            assert (
                "structure_drift_from_previous_generation"
                in created.structure_payload["disposition_reasons"]
            )
        else:
            assert created.disposition == "proposed"

        if not changed_structure:
            barrier = threading.Barrier(3)
            versions: list[int] = []
            errors: list[BaseException] = []

            def replay_same_artifact():
                session = sessionmaker(bind=observer_engine)()
                try:
                    barrier.wait(timeout=5)
                    snapshot = analyze_source_artifact_structure(
                        session,
                        org_id=org.id,
                        project_id=project.id,
                        batch_id=batch.id,
                        artifact_id=current_artifact.id,
                        current_user=user,
                        storage=storage,
                    )
                    versions.append(snapshot.snapshot_version)
                except BaseException as exc:
                    errors.append(exc)
                    session.rollback()
                finally:
                    session.close()

            threads = [threading.Thread(target=replay_same_artifact, daemon=True) for _ in range(2)]
            for thread in threads:
                thread.start()
            barrier.wait(timeout=5)
            for thread in threads:
                thread.join(timeout=20)
            assert not errors, repr(errors)
            assert sorted(versions) == [2, 3]
        worker_engine.dispose()
        lock_session.close()
    finally:
        setup_session.close()
        observer_engine.dispose()


class _BlockingCloseStream(io.BytesIO):
    def __init__(self, data: bytes, reached: threading.Event, release: threading.Event):
        super().__init__(data)
        self._reached = reached
        self._release = release

    def close(self):
        self._reached.set()
        if not self._release.wait(timeout=15):
            raise TimeoutError("identity-map probe was not released")
        super().close()


class _BlockingCloseStorage(FakeObjectStorage):
    def __init__(self, reached: threading.Event, release: threading.Event):
        super().__init__()
        self._reached = reached
        self._release = release

    def open_stream(self, key: str):
        return _BlockingCloseStream(self._objects[key], self._reached, self._release)


def test_postgres_populate_existing_detects_durable_authority_change():
    engine = _postgres_engine_or_skip()
    setup = sessionmaker(bind=engine)()
    reached = threading.Event()
    release = threading.Event()
    storage = _BlockingCloseStorage(reached, release)
    try:
        org, user, _other, project, batch = _seed(setup)
        artifact = _add_available_artifact(
            setup,
            storage,
            org=org,
            user=user,
            project=project,
            batch=batch,
            generation=1,
            data=_xlsx_bytes(),
        )
        outcome: dict[str, object] = {}

        def analyze():
            session = sessionmaker(bind=engine)()
            try:
                outcome["snapshot"] = analyze_source_artifact_structure(
                    session,
                    org_id=org.id,
                    project_id=project.id,
                    batch_id=batch.id,
                    artifact_id=artifact.id,
                    current_user=user,
                    storage=storage,
                )
            except BaseException as exc:
                outcome["error"] = exc
                session.rollback()
            finally:
                session.close()

        worker = threading.Thread(target=analyze, daemon=True)
        worker.start()
        assert reached.wait(timeout=10)
        mutation = sessionmaker(bind=engine)()
        try:
            durable = mutation.get(ImportSourceArtifact, artifact.id)
            durable.state = "orphaned"
            mutation.commit()
        finally:
            mutation.close()
        release.set()
        worker.join(timeout=20)
        assert not worker.is_alive()
        error = outcome.get("error")
        assert isinstance(error, HTTPException)
        assert error.status_code == 409
        assert error.detail["error_code"] == "source_artifact_changed"
        verify = sessionmaker(bind=engine)()
        try:
            assert (
                verify.query(WorkbookStructureSnapshot)
                .filter(WorkbookStructureSnapshot.source_artifact_id == artifact.id)
                .count()
                == 0
            )
            assert (
                verify.query(AuditEvent)
                .filter(
                    AuditEvent.entity_type == "WorkbookStructureSnapshot",
                    AuditEvent.payload["source_artifact_id"].as_string() == str(artifact.id),
                )
                .count()
                == 0
            )
        finally:
            verify.close()
    finally:
        release.set()
        setup.close()
        engine.dispose()


def test_v3_unicode_channels_and_header_roles_are_exact_and_accent_insensitive():
    assert structure_domain._surface_normalized("  TỔNG   CỘNG TỶ LỆ ") == "tổng cộng tỷ lệ"
    assert structure_domain._search_normalized("ĐVT") == "dvt"
    assert structure_domain._search_normalized("Đặc điểm") == "dac diem"
    assert structure_domain._search_normalized("Đơn giá") == "don gia"
    assert structure_domain._search_normalized("Đơn vị tính") == "don vi tinh"
    assert structure_domain._header_roles("ĐVT") == {"UNIT"}
    assert structure_domain._header_roles("Đặc điểm") == {"DESCRIPTION"}
    assert structure_domain._header_roles("Đơn giá") == {"VALUE"}
    assert structure_domain._header_roles("Đơn vị tính") == {"UNIT"}
    for separator in (":", "-", "–", "—", "/", ".", ",", ";", "(", "_"):
        assert structure_domain._header_roles(f"Đơn{separator}giá") == {"VALUE"}
    assert structure_domain._header_roles("x_Đơn giá") == {"VALUE"}
    assert structure_domain._header_roles("Đơn giá_suffix") == {"VALUE"}
    assert not structure_domain._header_roles("mã")
    assert not structure_domain._header_roles("xđơn giá")


@pytest.mark.parametrize(
    ("values", "expected", "reason"),
    [
        ([99, "TỔNG CỘNG TỶ LỆ", 100], "total", "total_marker"),
        ([99, "TỔNG CỘNG TỶ TRỌNG", 100], "total", "total_marker"),
        ([1, "Tổng công ty ABC", 100], "asset", "serial_number_and_content"),
        ([1, "TONG CONG TY ABC", 100], "unresolved", "ambiguous_folded_marker"),
        (["TONG CONG"], "total", "total_marker"),
        (["TỔNG-CỘNG: TỶ LỆ"], "total", "total_marker"),
        (["TỔNG/CỘNG"], "total", "total_marker"),
    ],
)
def test_v3_marker_collision_matrix(values, expected, reason):
    classified = classify_row(_cells(3, values))
    assert classified.row_class.value == expected
    assert reason in classified.reasons


@pytest.mark.parametrize("group_count", [2, 3])
def test_v3_adjacent_complete_header_groups_partition_without_covering_candidate(group_count):
    headers: list[object] = []
    body: list[object] = []
    for index in range(group_count):
        headers.extend(
            [
                "STT" if index == 0 else "Mã tài sản",
                "Tên tài sản",
                "Số lượng",
            ]
        )
        body.extend([1, f"Máy {index + 1}", index + 1])
    payload, _trackers = _analyze_rows([headers, body])
    row_one = [item for item in payload["candidates"] if item["header_start_row"] == 1]
    bounds = [
        (
            item["candidate_table_bounds"]["min_column"],
            item["candidate_table_bounds"]["max_column"],
        )
        for item in row_one
    ]
    assert bounds == [(1 + 3 * index, 3 + 3 * index) for index in range(group_count)]
    assert all("multiple_horizontal_table_groups" in item["boundary_flags"] for item in row_one)
    assert payload["disposition"] == "review_required"
    assert "multiple_horizontal_table_groups" in payload["disposition_reasons"]


def test_v3_single_table_with_two_known_starts_is_not_split():
    payload, _trackers = _analyze_rows(
        [["STT", "Mã tài sản", "Tên tài sản", "Số lượng"], [1, "TS-1", "Máy A", 2]]
    )
    row_one = [item for item in payload["candidates"] if item["header_start_row"] == 1]
    assert [
        (
            item["candidate_table_bounds"]["min_column"],
            item["candidate_table_bounds"]["max_column"],
        )
        for item in row_one
    ] == [(1, 4)]
    assert "multiple_horizontal_table_groups" not in row_one[0]["boundary_flags"]


def test_v3_unknown_second_start_with_repeated_core_bundles_forces_review():
    payload, _trackers = _analyze_rows(
        [
            ["STT", "Tên tài sản", "Số lượng", "Mã", "Tên tài sản", "Số lượng"],
            [1, "Máy A", 1, "X-1", "Máy B", 2],
        ]
    )
    candidate = next(item for item in payload["candidates"] if item["header_start_row"] == 1)
    assert candidate["candidate_table_bounds"]["min_column"] == 1
    assert candidate["candidate_table_bounds"]["max_column"] == 6
    assert "ambiguous_horizontal_table_boundary" in candidate["boundary_flags"]
    assert payload["disposition"] == "review_required"


def test_v3_repeated_value_labels_without_repeated_name_bundle_do_not_split():
    payload, _trackers = _analyze_rows(
        [
            ["STT", "Tên tài sản", "Số lượng", "Đơn giá", "Thành tiền"],
            [1, "Máy A", 1, 100, 100],
        ]
    )
    candidate = next(item for item in payload["candidates"] if item["header_start_row"] == 1)
    assert "multiple_horizontal_table_groups" not in candidate["boundary_flags"]
    assert "ambiguous_horizontal_table_boundary" not in candidate["boundary_flags"]


def test_v3_lower_ranked_horizontal_attack_forces_workbook_review():
    rows_by_sheet = {
        "Primary": [
            ["BẢNG TỔNG HỢP", None, None, None, None, None],
            ["STT", "Tên tài sản", "Đặc điểm", "ĐVT", "Số lượng", "Đơn giá"],
            [1, "Máy chính", "Model A", "cái", 1, 100],
        ],
        "Adjacent": [
            ["STT", "Tên tài sản", "Số lượng", "Mã tài sản", "Tên tài sản", "Số lượng"],
            [1, "Máy A", 1, "TS-2", "Máy B", 2],
        ],
    }
    cell_rows = {
        sheet: [_cells(index, values) for index, values in enumerate(rows, 1)]
        for sheet, rows in rows_by_sheet.items()
    }

    def provider(sheet_name):
        return iter(cell_rows[sheet_name])

    inspection = AdapterInspectionResult(
        format=WorkbookFormat.XLSX,
        adapter_name="adversarial",
        adapter_version="1",
        sheet_names=("Primary", "Adjacent"),
        sheets=(
            SheetSummary(
                name="Primary",
                max_row=3,
                max_column=6,
                merged_regions=(MergedRegion(1, 1, 1, 6),),
            ),
            SheetSummary(name="Adjacent", max_row=2, max_column=6),
        ),
    )
    payload = analyze_workbook_structure(inspection, provider)
    assert payload["candidates"][0]["sheet_name"] == "Primary"
    assert "multiple_horizontal_table_groups" not in payload["candidates"][0][
        "boundary_flags"
    ]
    assert any(
        "multiple_horizontal_table_groups" in candidate["boundary_flags"]
        for candidate in payload["candidates"][1:]
    )
    assert payload["disposition"] == "review_required"
    assert "multiple_horizontal_table_groups" in payload["disposition_reasons"]


def _clone_snapshot_versions(
    db: Session,
    snapshot: WorkbookStructureSnapshot,
    *,
    through_version: int,
) -> None:
    current_version = (
        db.query(func.max(WorkbookStructureSnapshot.snapshot_version))
        .filter(WorkbookStructureSnapshot.source_artifact_id == snapshot.source_artifact_id)
        .scalar()
    )
    for version in range(int(current_version or 0) + 1, through_version + 1):
        db.add(
            WorkbookStructureSnapshot(
                id=uuid.uuid4(),
                organization_id=snapshot.organization_id,
                project_id=snapshot.project_id,
                import_batch_id=snapshot.import_batch_id,
                source_artifact_id=snapshot.source_artifact_id,
                snapshot_version=version,
                source_checksum_sha256=snapshot.source_checksum_sha256,
                rule_version=snapshot.rule_version,
                adapter_name=snapshot.adapter_name,
                adapter_version=snapshot.adapter_version,
                disposition=snapshot.disposition,
                candidate_count=snapshot.candidate_count,
                structure_payload=json.loads(json.dumps(snapshot.structure_payload)),
                analysis_digest_sha256=snapshot.analysis_digest_sha256,
                created_by_user_id=snapshot.created_by_user_id,
            )
        )
    db.commit()


def test_v3_snapshot_list_paginates_55_as_array_without_loss_and_documents_headers(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact["id"])
    created = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    assert created.status_code == 201
    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    _clone_snapshot_versions(structure_db, snapshot, through_version=55)

    versions: list[int] = []
    cursor = None
    page_sizes: list[int] = []
    while True:
        params = {} if cursor is None else {"cursor": cursor}
        response = structure_client.get(path, params=params, headers={"X-User-Id": str(user.id)})
        assert response.status_code == 200, response.text
        assert isinstance(response.json(), list)
        assert response.headers["X-Valora-Page-Limit"] == "20"
        page_sizes.append(len(response.json()))
        versions.extend(item["snapshot_version"] for item in response.json())
        next_cursor = response.headers.get("X-Valora-Next-Cursor")
        if next_cursor is None:
            break
        cursor = int(next_cursor)
    assert page_sizes == [20, 20, 15]
    assert versions == list(range(1, 56))

    explicit = structure_client.get(
        path,
        params={"limit": 50},
        headers={"X-User-Id": str(user.id)},
    )
    assert len(explicit.json()) == 50
    assert explicit.headers["X-Valora-Page-Limit"] == "50"
    assert explicit.headers["X-Valora-Next-Cursor"] == "50"
    past_end = structure_client.get(
        path,
        params={"cursor": 999},
        headers={"X-User-Id": str(user.id)},
    )
    assert past_end.json() == []
    assert "X-Valora-Next-Cursor" not in past_end.headers
    for params in ({"limit": 0}, {"limit": 51}, {"cursor": -1}):
        invalid = structure_client.get(
            path, params=params, headers={"X-User-Id": str(user.id)}
        )
        assert invalid.status_code == 422

    route = (
        "/api/v1/projects/{project_id}/asset-imports/{batch_id}/source-artifacts/"
        "{artifact_id}/structure-snapshots"
    )
    response_headers = fastapi_app.openapi()["paths"][route]["get"]["responses"]["200"][
        "headers"
    ]
    assert {"X-Valora-Page-Limit", "X-Valora-Next-Cursor"}.issubset(response_headers)

    cors = structure_client.get(
        path,
        params={"limit": 1},
        headers={
            "Origin": "http://localhost:5173",
            "X-User-Id": str(user.id),
        },
    )
    exposed = {
        value.strip().casefold()
        for value in cors.headers["access-control-expose-headers"].split(",")
    }
    assert {"x-valora-page-limit", "x-valora-next-cursor"}.issubset(exposed)


@pytest.mark.parametrize(
    ("limit", "cursor", "error_code"),
    [
        (0, None, "invalid_snapshot_page_limit"),
        (51, None, "invalid_snapshot_page_limit"),
        (True, None, "invalid_snapshot_page_limit"),
        (20, -1, "invalid_snapshot_page_cursor"),
        (20, True, "invalid_snapshot_page_cursor"),
    ],
)
def test_v3_snapshot_service_enforces_page_bounds(
    structure_db: Session,
    limit,
    cursor,
    error_code: str,
):
    with pytest.raises(HTTPException) as caught:
        list_structure_snapshots(
            structure_db,
            org_id=uuid.uuid4(),
            project_id=uuid.uuid4(),
            batch_id=uuid.uuid4(),
            artifact_id=uuid.uuid4(),
            limit=limit,
            cursor=cursor,
        )
    assert caught.value.status_code == 422
    assert caught.value.detail["error_code"] == error_code


def test_v3_page_integrity_is_local_and_bulk_query_budget_is_constant(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    org, user, _other, project, batch = _seed(structure_db)
    artifact_body = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact_body["id"])
    created = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    _clone_snapshot_versions(structure_db, snapshot, through_version=50)
    org_id = org.id
    project_id = project.id
    batch_id = batch.id
    artifact_id = uuid.UUID(artifact_body["id"])

    statements: list[str] = []

    def count_sql(_conn, _cursor, statement, _parameters, _context, _executemany):
        if statement.lstrip().upper().startswith("SELECT"):
            statements.append(statement)

    event.listen(structure_db.bind, "before_cursor_execute", count_sql)
    try:
        page, next_cursor = list_structure_snapshots(
            structure_db,
            org_id=org_id,
            project_id=project_id,
            batch_id=batch_id,
            artifact_id=artifact_id,
            limit=50,
        )
    finally:
        event.remove(structure_db.bind, "before_cursor_execute", count_sql)
    assert len(page) == 50
    assert next_cursor is None
    assert len(statements) <= 5

    corrupt = next(
        item
        for item in structure_db.query(WorkbookStructureSnapshot).all()
        if item.snapshot_version == 21
    )
    corrupt.analysis_digest_sha256 = "0" * 64
    structure_db.commit()
    first_page = structure_client.get(path, headers={"X-User-Id": str(user.id)})
    assert first_page.status_code == 200
    assert len(first_page.json()) == 20
    corrupt_page = structure_client.get(
        path,
        params={"cursor": 20},
        headers={"X-User-Id": str(user.id)},
    )
    assert corrupt_page.status_code == 500
    assert corrupt_page.json()["detail"]["error_code"] == "structure_snapshot_integrity_failure"


def test_v3_referenced_pages_and_get_one_stay_within_five_sql_statements(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    org, user, _other, project, batch = _seed(structure_db)
    first_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    first_path = _snapshot_path(project, batch, first_artifact["id"])
    assert (
        structure_client.post(first_path, headers={"X-User-Id": str(user.id)}).status_code
        == 201
    )
    second_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    second_path = _snapshot_path(project, batch, second_artifact["id"])
    created = structure_client.post(second_path, headers={"X-User-Id": str(user.id)})
    assert created.status_code == 201, created.text
    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    _clone_snapshot_versions(structure_db, snapshot, through_version=50)
    scope = {
        "org_id": org.id,
        "project_id": project.id,
        "batch_id": batch.id,
        "artifact_id": uuid.UUID(second_artifact["id"]),
    }

    def select_count(callable_):
        statements: list[str] = []

        def count_sql(_conn, _cursor, statement, _parameters, _context, _executemany):
            if statement.lstrip().upper().startswith("SELECT"):
                statements.append(statement)

        event.listen(structure_db.bind, "before_cursor_execute", count_sql)
        try:
            result = callable_()
        finally:
            event.remove(structure_db.bind, "before_cursor_execute", count_sql)
        return result, len(statements)

    for limit in (1, 50):
        (page, _cursor), query_count = select_count(
            lambda limit=limit: list_structure_snapshots(
                structure_db,
                **scope,
                limit=limit,
            )
        )
        assert len(page) == limit
        assert query_count <= 5
    replayed, query_count = select_count(
        lambda: get_structure_snapshot(
            structure_db,
            **scope,
            snapshot_id=snapshot.id,
        )
    )
    assert replayed.id == snapshot.id
    assert query_count <= 5


def test_v3_cursor_is_append_safe_between_pages(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact["id"])
    created = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    _clone_snapshot_versions(structure_db, snapshot, through_version=25)
    first = structure_client.get(path, headers={"X-User-Id": str(user.id)})
    assert [item["snapshot_version"] for item in first.json()] == list(range(1, 21))
    _clone_snapshot_versions(structure_db, snapshot, through_version=26)
    second = structure_client.get(
        path,
        params={"cursor": first.headers["X-Valora-Next-Cursor"]},
        headers={"X-User-Id": str(user.id)},
    )
    assert [item["snapshot_version"] for item in second.json()] == list(range(21, 27))


@pytest.mark.parametrize("legacy_version", ["s13-pr-003-v1", "s13-pr-003-v2"])
def test_v3_read_allowlist_replays_legacy_without_rewrite(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
    legacy_version: str,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact["id"])
    created = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    payload = json.loads(json.dumps(snapshot.structure_payload))
    payload["rule_version"] = legacy_version
    if legacy_version == "s13-pr-003-v1":
        payload.pop("drift_reference")
    snapshot.rule_version = legacy_version
    snapshot.structure_payload = payload
    snapshot.analysis_digest_sha256 = canonical_payload_digest(payload)
    structure_db.commit()
    before = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    digest = snapshot.analysis_digest_sha256

    replay = structure_client.get(
        f"{path}/{snapshot.id}", headers={"X-User-Id": str(user.id)}
    )
    assert replay.status_code == 200, replay.text
    structure_db.expire_all()
    durable = structure_db.get(WorkbookStructureSnapshot, snapshot.id)
    assert json.dumps(durable.structure_payload, ensure_ascii=False, sort_keys=True) == before
    assert durable.analysis_digest_sha256 == digest


def test_v3_read_allowlist_rejects_unknown_rule_version(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    path = _snapshot_path(project, batch, artifact["id"])
    created = structure_client.post(path, headers={"X-User-Id": str(user.id)})
    snapshot = structure_db.get(WorkbookStructureSnapshot, uuid.UUID(created.json()["id"]))
    payload = json.loads(json.dumps(snapshot.structure_payload))
    payload["rule_version"] = "s13-pr-003-v999"
    snapshot.rule_version = payload["rule_version"]
    snapshot.structure_payload = payload
    snapshot.analysis_digest_sha256 = canonical_payload_digest(payload)
    structure_db.commit()

    replay = structure_client.get(
        f"{path}/{snapshot.id}", headers={"X-User-Id": str(user.id)}
    )
    assert replay.status_code == 500
    assert replay.json()["detail"]["error_code"] == "structure_snapshot_integrity_failure"


def test_v3_analysis_forces_review_when_predecessor_uses_v2_rules(
    structure_client: TestClient,
    structure_db: Session,
    structure_storage,
):
    _org, user, _other, project, batch = _seed(structure_db)
    first_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    first_path = _snapshot_path(project, batch, first_artifact["id"])
    first = structure_client.post(first_path, headers={"X-User-Id": str(user.id)})
    assert first.status_code == 201, first.text

    predecessor = structure_db.get(
        WorkbookStructureSnapshot,
        uuid.UUID(first.json()["id"]),
    )
    legacy_payload = json.loads(json.dumps(predecessor.structure_payload))
    legacy_payload["rule_version"] = "s13-pr-003-v2"
    predecessor.rule_version = "s13-pr-003-v2"
    predecessor.structure_payload = legacy_payload
    predecessor.analysis_digest_sha256 = canonical_payload_digest(legacy_payload)
    structure_db.commit()

    second_artifact = _upload(structure_client, user, project, batch, _xlsx_bytes())
    second = structure_client.post(
        _snapshot_path(project, batch, second_artifact["id"]),
        headers={"X-User-Id": str(user.id)},
    )
    assert second.status_code == 201, second.text
    payload = second.json()["structure_payload"]
    assert payload["disposition"] == "review_required"
    assert "structure_rule_version_changed" in payload["disposition_reasons"]


def test_v3_deterministic_payload_and_versioned_digest_change():
    rows = [
        ["STT", "Tên tài sản", "Số lượng"],
        [1, "Máy A", 1],
        ["TỔNG CỘNG", None, 1],
    ]
    first, _trackers = _analyze_rows(rows)
    second, _trackers = _analyze_rows(rows)
    assert first == second
    assert canonical_payload_digest(first) == canonical_payload_digest(second)
    legacy = json.loads(json.dumps(first))
    legacy["rule_version"] = "s13-pr-003-v2"
    for key in (
        "header_group_min_families",
        "header_group_min_columns",
        "header_group_start_lookback",
        "repeated_core_family_threshold",
    ):
        legacy["rule_config"].pop(key)
    assert canonical_payload_digest(first) != canonical_payload_digest(legacy)
