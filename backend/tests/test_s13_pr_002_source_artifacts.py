"""S13-PR-002 corrective: adapters + immutable source artifacts proof matrix."""
from __future__ import annotations

import io
import os
import struct
import uuid
import zipfile
from datetime import datetime, timedelta, timezone

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.adapters import detect_format_and_adapter
from app.modules.excel_import.application.adapters.xls_adapter import XlsWorkbookAdapter
from app.modules.excel_import.application.adapters.xls_safety import (
    reject_ole_vba_presence,
    scan_biff_threats,
)
from app.modules.excel_import.application.adapters.xlsx_adapter import XlsxWorkbookAdapter
from app.modules.excel_import.application.adapters.xlsx_merge import parse_merge_ref
from app.modules.excel_import.application.source_artifact_service import (
    count_staging_rows,
    reconcile_source_artifacts,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.domain.workbook_adapter import AdapterError
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    ObjectStorageError,
    S3ObjectStorage,
    set_object_storage_override,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ImportBatchStatus,
    ProjectAssetLine,
    AuditEvent,
)


@pytest.fixture
def db_session() -> Session:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()


def _make_xlsx_bytes(*, merged: bool = False, long_cell: str | None = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "A", "", "E", "F", "G", "H", ""])  # I blank header
    ws.append(["x", None, "y", None, 1, 2, 3, 4, "evidence"])
    if long_cell is not None:
        ws["A3"] = long_cell
    if merged:
        ws.merge_cells("A1:B1")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xls_path(tmp_path, *, with_merge: bool = False, trailing_blank_header: bool = True):
    pytest.importorskip("xlwt")
    import xlwt

    book = xlwt.Workbook()
    sh = book.add_sheet("S1")
    # B/C/D/H text, I blank — production-like header
    headers = ["", "B", "C", "D", "", "", "", "H", ""]
    for c, h in enumerate(headers):
        sh.write(0, c, h)
    # data: column I has evidence
    for c in range(9):
        if c == 8:
            sh.write(1, c, "evidence")
        elif c in (1, 2, 3, 7):
            sh.write(1, c, f"v{c}")
    if with_merge:
        sh.write_merge(2, 3, 0, 1, "merged")
    path = tmp_path / "t.xls"
    book.save(str(path))
    return path


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org 1", organization_slug=f"org-{uuid.uuid4().hex[:6]}", status=OrganizationStatus.ACTIVE
    )
    org_other = OrganizationProfile(
        legal_name="Org 2", organization_slug=f"org-{uuid.uuid4().hex[:6]}", status=OrganizationStatus.ACTIVE
    )
    db.add_all([org, org_other])
    db.commit()
    role_editor = Role(code="editor", display_name="Editor", permissions=["project:read", "workbench:edit"])
    db.add(role_editor)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"e{uuid.uuid4().hex[:8]}@example.com",
        full_name="Editor",
        status=UserStatus.ACTIVE,
    )
    user_other = User(
        organization_id=org_other.id,
        email=f"o{uuid.uuid4().hex[:8]}@example.com",
        full_name="Other",
        status=UserStatus.ACTIVE,
    )
    db.add_all([user, user_other])
    db.commit()
    db.add_all(
        [
            UserRole(user_id=user.id, role_id=role_editor.id, is_active=True),
            UserRole(user_id=user_other.id, role_id=role_editor.id, is_active=True),
        ]
    )
    db.commit()
    cust = Customer(
        organization_id=org.id,
        legal_name="Customer 1",
        status=CustomerStatus.ACTIVE,
        created_by=user.id,
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P1",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="seed.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, user_other, proj, batch


# --- Adapter unit tests ---


def test_xlsx_adapter_preserves_blank_and_duplicate_columns(tmp_path):
    p = tmp_path / "t.xlsx"
    p.write_bytes(_make_xlsx_bytes())
    adapter = XlsxWorkbookAdapter()
    insp = adapter.inspect(str(p))
    assert insp.format.value == "xlsx"
    rows = list(adapter.iter_rows(str(p)))
    header = rows[0]
    assert [c.value for c in header[:3]] == ["A", "B", "A"]
    assert header[0].column == 1 and header[2].column == 3
    # trailing blank header I (col 9)
    assert len(header) >= 9
    assert header[8].column == 9
    assert header[8].value is None or header[8].value == ""
    assert rows[1][8].value == "evidence"


def test_xlsx_merged_regions(tmp_path):
    p = tmp_path / "m.xlsx"
    p.write_bytes(_make_xlsx_bytes(merged=True))
    insp = XlsxWorkbookAdapter().inspect(str(p))
    merges = insp.sheets[0].merged_regions
    assert len(merges) >= 1
    m = merges[0]
    assert m.min_row == 1 and m.min_col == 1
    assert m.max_row == 1 and m.max_col == 2


def test_xlsx_merge_malformed_ref():
    with pytest.raises(AdapterError) as ei:
        parse_merge_ref("NOT_A_REF")
    assert ei.value.error_code == "invalid_xlsx"


def test_xlsx_rejects_macro_and_extension_mismatch(tmp_path):
    path = tmp_path / "macro.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("[Content_Types].xml", "<Types/>")
        zf.writestr("xl/workbook.xml", "<workbook/>")
        zf.writestr("xl/vbaProject.bin", b"vba")
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter().inspect(str(path))
    assert ei.value.error_code == "macro_not_allowed"

    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not-a-zip")
    with pytest.raises(AdapterError) as ei2:
        detect_format_and_adapter(str(bad), "bad.xlsx")
    assert ei2.value.error_code == "signature_mismatch"


def test_xlsx_cell_length_limit(tmp_path):
    limits = SourceArtifactLimits(max_cell_chars=10)
    p = tmp_path / "long.xlsx"
    p.write_bytes(_make_xlsx_bytes(long_cell="x" * 11))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=limits).inspect(str(p))
    assert ei.value.error_code == "cell_length_limit"


def test_xls_positional_blanks_and_duplicates(tmp_path):
    path = _make_xls_path(tmp_path)
    fmt, adapter = detect_format_and_adapter(str(path), "t.xls")
    assert fmt.value == "xls"
    insp = adapter.inspect(str(path))
    assert insp.sheet_names == ("S1",)
    rows = list(adapter.iter_rows(str(path)))
    header = rows[0]
    assert len(header) >= 9
    assert header[8].column == 9
    assert header[8].value in (None, "")
    assert header[1].value == "B" and header[7].value == "H"
    assert rows[1][8].value == "evidence"
    # duplicate empty headers still positional
    assert header[0].column != header[4].column


def test_xls_merged_regions(tmp_path):
    path = _make_xls_path(tmp_path, with_merge=True)
    insp = XlsWorkbookAdapter().inspect(str(path))
    merges = insp.sheets[0].merged_regions
    assert len(merges) >= 1
    m = merges[0]
    assert m.min_row <= m.max_row and m.min_col <= m.max_col


def test_xls_vba_stream_name_rejected():
    with pytest.raises(AdapterError) as ei:
        reject_ole_vba_presence(["Workbook", "_VBA_PROJECT_CUR"])
    assert ei.value.error_code == "macro_not_allowed"


def test_xls_filepass_biff_rejected():
    # Minimal BIFF-like blob with FILEPASS record
    blob = struct.pack("<HH", 0x002F, 2) + b"\x00\x00"
    with pytest.raises(AdapterError) as ei:
        scan_biff_threats(blob)
    assert ei.value.error_code == "encrypted_workbook"


def test_xls_external_supbook_rejected():
    # SUPBOOK longer than 4 bytes
    payload = b"\x01\x00" + b"http://evil.example/book.xls"
    blob = struct.pack("<HH", 0x01AE, len(payload)) + payload
    with pytest.raises(AdapterError) as ei:
        scan_biff_threats(blob)
    assert ei.value.error_code == "external_link_not_allowed"


def test_merged_region_limit_xlsx(tmp_path):
    limits = SourceArtifactLimits(max_merged_regions=0, max_merged_regions_per_sheet=0)
    p = tmp_path / "m.xlsx"
    p.write_bytes(_make_xlsx_bytes(merged=True))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=limits).inspect(str(p))
    assert ei.value.error_code == "merged_region_limit"


# --- API / artifact tests ---


def test_source_artifact_upload_available_and_no_staging(client: TestClient, db_session: Session, fake_storage):
    org, user, _u2, proj, batch = _seed(db_session)
    headers = {"X-User-Id": str(user.id)}
    data = _make_xlsx_bytes()
    before_staging = count_staging_rows(db_session, batch.id)
    before_lines = db_session.query(ProjectAssetLine).filter(ProjectAssetLine.project_id == proj.id).count()

    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "src.xlsx",
                io.BytesIO(data),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )
    assert res.status_code == 201, res.text
    body = res.json()
    assert body["state"] == "available"
    assert body["generation"] == 1
    assert body["detected_format"] == "xlsx"
    assert len(body["checksum_sha256"]) == 64
    assert body["checksum_sha256"] == body["checksum_sha256"].lower()
    assert "storage_object_key" not in body

    db_session.refresh(batch)
    assert str(batch.current_source_artifact_id) == body["id"]
    assert count_staging_rows(db_session, batch.id) == before_staging
    assert db_session.query(ProjectAssetLine).filter(ProjectAssetLine.project_id == proj.id).count() == before_lines

    art_id = uuid.UUID(body["id"])
    art = db_session.query(ImportSourceArtifact).filter(ImportSourceArtifact.id == art_id).one()
    assert fake_storage.head(art.storage_object_key) is not None
    ev = (
        db_session.query(AuditEvent)
        .filter(AuditEvent.event_name == "ImportSourceArtifactAvailable", AuditEvent.entity_id == art_id)
        .first()
    )
    assert ev is not None


def test_reject_oversized_cell_before_object_write(client: TestClient, db_session: Session, fake_storage):
    org, user, _u2, proj, batch = _seed(db_session)
    # Use default limit 10000 — craft cell of 10001
    data = _make_xlsx_bytes(long_cell="Z" * 10001)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "long.xlsx",
                io.BytesIO(data),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert res.status_code == 400
    assert db_session.query(ImportSourceArtifact).count() == 0
    assert fake_storage._objects == {}
    db_session.refresh(batch)
    assert batch.current_source_artifact_id is None


def test_s12_upload_rejects_xls(client: TestClient, db_session: Session, fake_storage):
    org, user, _u2, proj, batch = _seed(db_session)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/upload",
        files={"file": ("legacy.xls", io.BytesIO(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1xxxx"), "application/vnd.ms-excel")},
        headers={"X-User-Id": str(user.id)},
    )
    assert res.status_code == 400


def test_failure_keeps_prior_current(client: TestClient, db_session: Session, fake_storage):
    org, user, _u2, proj, batch = _seed(db_session)
    headers = {"X-User-Id": str(user.id)}
    data = _make_xlsx_bytes()
    r1 = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": ("a.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r1.status_code == 201
    first_id = r1.json()["id"]
    fake_storage.fail_put = True
    r2 = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": ("b.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r2.status_code == 500
    db_session.refresh(batch)
    assert str(batch.current_source_artifact_id) == first_id
    failed = (
        db_session.query(ImportSourceArtifact)
        .filter(ImportSourceArtifact.state == "failed")
        .all()
    )
    assert len(failed) >= 1
    assert failed[-1].failure_code


def test_cross_tenant_denied(client: TestClient, db_session: Session, fake_storage):
    org, user, user_other, proj, batch = _seed(db_session)
    data = _make_xlsx_bytes()
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": ("a.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"X-User-Id": str(user_other.id)},
    )
    assert res.status_code == 404


def test_list_and_get_metadata(client: TestClient, db_session: Session, fake_storage):
    org, user, _u2, proj, batch = _seed(db_session)
    headers = {"X-User-Id": str(user.id)}
    data = _make_xlsx_bytes()
    r1 = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": ("a.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r1.status_code == 201
    aid = r1.json()["id"]
    lst = client.get(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        headers=headers,
    )
    assert lst.status_code == 200
    assert len(lst.json()) == 1
    one = client.get(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts/{aid}",
        headers=headers,
    )
    assert one.status_code == 200
    assert "storage_object_key" not in one.json()


def test_reconciler_skips_current_requires_tenant_actor(client: TestClient, db_session: Session, fake_storage):
    org, user, _u2, proj, batch = _seed(db_session)
    headers = {"X-User-Id": str(user.id)}
    data = _make_xlsx_bytes()
    r1 = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": ("a.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r1.status_code == 201
    with pytest.raises(Exception):
        reconcile_source_artifacts(db_session, storage=fake_storage, max_items=10)
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] >= 0
    db_session.refresh(batch)
    art = db_session.get(ImportSourceArtifact, batch.current_source_artifact_id)
    assert fake_storage.head(art.storage_object_key) is not None


def test_reconciler_delete_failure_no_audit(db_session: Session, fake_storage):
    org, user, _u2, proj, batch = _seed(db_session)
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="x.xlsx",
        detected_format="xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_size_bytes=1,
        checksum_sha256="a" * 64,
        storage_object_key=f"org/{org.id}/k1",
        state="orphaned",
        created_by_user_id=user.id,
        orphaned_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    fake_storage._objects[art.storage_object_key] = b"x"
    fake_storage.fail_delete = True
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats["deleted_objects"] == 0
    assert stats["errors"] >= 1
    assert (
        db_session.query(AuditEvent)
        .filter(AuditEvent.event_name == "ImportSourceArtifactObjectDeleted")
        .count()
        == 0
    )


def test_s3_integration_when_configured():
    if os.environ.get("CI") == "true":
        endpoint = os.environ.get("S3_ENDPOINT_URL")
        assert endpoint, "CI must provide S3_ENDPOINT_URL for MinIO"
    else:
        endpoint = os.environ.get("S3_ENDPOINT_URL")
        if not endpoint:
            pytest.skip("S3_ENDPOINT_URL not set for local MinIO integration")

    store = S3ObjectStorage(
        endpoint_url=endpoint,
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"integration/{uuid.uuid4()}"
    payload = b"hello-valora-s13"
    st = store.put_stream(
        key, io.BytesIO(payload), content_type="application/octet-stream", expected_size=len(payload)
    )
    assert st.size == len(payload)
    assert store.head(key).size == len(payload)
    assert store.open_stream(key).read() == payload
    store.delete(key)
    assert store.head(key) is None


def test_s3_head_not_found_vs_error():
    """Fake storage: missing → None; forced error propagates."""
    store = FakeObjectStorage()
    assert store.head("missing") is None
    store.head_raises = ObjectStorageError("permission_denied")
    with pytest.raises(ObjectStorageError):
        store.head("x")


def test_content_identity_immutable_fields_not_exposed_for_update(
    client: TestClient, db_session: Session, fake_storage
):
    org, user, _u2, proj, batch = _seed(db_session)
    data = _make_xlsx_bytes()
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "a.xlsx",
                io.BytesIO(data),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert res.status_code == 201
    art = db_session.query(ImportSourceArtifact).one()
    original_key = art.storage_object_key
    original_checksum = art.checksum_sha256
    original_gen = art.generation
    # No public API mutates content identity; re-read stable
    got = client.get(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts/{art.id}",
        headers={"X-User-Id": str(user.id)},
    )
    assert got.status_code == 200
    db_session.refresh(art)
    assert art.storage_object_key == original_key
    assert art.checksum_sha256 == original_checksum
    assert art.generation == original_gen
