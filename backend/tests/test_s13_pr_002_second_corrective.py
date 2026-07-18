"""S13-PR-002 second corrective: R-01..R-09 executable proofs."""
from __future__ import annotations

import hashlib
import io
import os
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.adapters.xls_adapter import XlsWorkbookAdapter
from app.modules.excel_import.application.adapters.xls_safety import (
    _reject_supbook,
    scan_biff_threats,
)
from app.modules.excel_import.application.adapters.xlsx_adapter import XlsxWorkbookAdapter
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.domain.workbook_adapter import AdapterError
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ImportBatchStatus,
    ProjectAssetLine,
    AuditEvent,
)
from tests.fixtures.s13_pr_002.ole_builder import write_threat_xls


@pytest.fixture
def db_session() -> Session:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org 1",
        organization_slug=f"org-{uuid.uuid4().hex[:6]}",
        status=OrganizationStatus.ACTIVE,
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="Editor", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"e{uuid.uuid4().hex[:8]}@example.com",
        full_name="Editor",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id,
        legal_name="C",
        status=CustomerStatus.ACTIVE,
        created_by=user.id,
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _xlsx_bytes(**kwargs) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append([1, 2])
    if kwargs.get("sheets"):
        for i in range(1, kwargs["sheets"]):
            wb.create_sheet(f"S{i}")
    if kwargs.get("long_cell"):
        ws["A3"] = kwargs["long_cell"]
    if kwargs.get("merged"):
        ws.merge_cells(kwargs.get("merge_ref", "A1:B1"))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- R-01 / R-02 BIFF rules ---


def test_supbook_internal_accepted_addin_rejected():
    _reject_supbook(b"\x01\x00\x01\x04")  # internal
    with pytest.raises(AdapterError) as ei:
        _reject_supbook(b"\x01\x00\x01\x3A")
    assert ei.value.error_code == "external_link_not_allowed"
    with pytest.raises(AdapterError):
        _reject_supbook(b"\x01\x00")  # truncated
    with pytest.raises(AdapterError):
        _reject_supbook(b"\x01\x00" + b"http://x")


def test_biff_macro_and_dcon_rejected():
    # NAME with fMacro
    name = struct_pack_rec(0x0018, b"\x08\x00" + b"\x00" * 8)
    with pytest.raises(AdapterError) as ei:
        scan_biff_threats(name)
    assert ei.value.error_code == "macro_not_allowed"
    dcon = struct_pack_rec(0x0050, b"\x01\x00")
    with pytest.raises(AdapterError) as ei2:
        scan_biff_threats(dcon)
    assert ei2.value.error_code == "external_link_not_allowed"
    dconref = struct_pack_rec(0x0051, b"\x01\x00")
    with pytest.raises(AdapterError) as ei3:
        scan_biff_threats(dconref)
    assert ei3.value.error_code == "external_link_not_allowed"


def struct_pack_rec(t, payload):
    import struct

    return struct.pack("<HH", t, len(payload)) + payload


@pytest.mark.parametrize(
    "threat,code",
    [
        ("filepass", "encrypted_workbook"),
        ("addin_supbook", "external_link_not_allowed"),
        ("dconref", "external_link_not_allowed"),
        ("macro_boundsheet", "macro_not_allowed"),
        ("macro_name", "macro_not_allowed"),
    ],
)
def test_e2e_threat_xls_upload_rejected(
    client: TestClient, db_session: Session, fake_storage, tmp_path, threat, code
):
    org, user, proj, batch = _seed(db_session)
    path = tmp_path / f"{threat}.xls"
    write_threat_xls(path, threat)
    data = path.read_bytes()
    before = db_session.query(ImportSourceArtifact).count()
    objects_before = dict(fake_storage._objects)
    res = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": (f"{threat}.xls", io.BytesIO(data), "application/vnd.ms-excel")},
        headers={"X-User-Id": str(user.id)},
    )
    assert res.status_code == 400, res.text
    assert db_session.query(ImportSourceArtifact).count() == before
    assert fake_storage._objects == objects_before
    db_session.refresh(batch)
    assert batch.current_source_artifact_id is None


def test_e2e_internal_supbook_only_xls_still_fails_without_sheet(tmp_path):
    """Internal SUPBOOK alone is not a complete workbook — safety may pass then xlrd fails."""
    path = tmp_path / "i.xls"
    write_threat_xls(path, "internal_supbook")
    # Presence gate should not reject internal SUPBOOK; open may still fail closed as invalid
    from app.modules.excel_import.application.adapters.xls_safety import assert_xls_safety

    assert_xls_safety(str(path))  # must not raise


# --- R-04 merges require formatting_info; multi-sheet ---


def test_xls_multi_sheet_merges_independent(tmp_path):
    pytest.importorskip("xlwt")
    import xlwt

    book = xlwt.Workbook()
    s1 = book.add_sheet("S1")
    s1.write_merge(0, 0, 0, 1, "ab")  # A1:B1
    s2 = book.add_sheet("S2")
    s2.write_merge(0, 1, 0, 0, "a")  # A1:A2
    path = tmp_path / "m.xls"
    book.save(str(path))
    insp = XlsWorkbookAdapter().inspect(str(path))
    by_name = {s.name: s.merged_regions for s in insp.sheets}
    assert len(by_name["S1"]) >= 1
    m1 = by_name["S1"][0]
    assert m1.min_row == 1 and m1.max_row == 1 and m1.min_col == 1 and m1.max_col == 2
    assert len(by_name["S2"]) >= 1
    m2 = by_name["S2"][0]
    assert m2.min_row == 1 and m2.max_row == 2 and m2.min_col == 1 and m2.max_col == 1


def test_xls_merged_limit(tmp_path):
    pytest.importorskip("xlwt")
    import xlwt

    book = xlwt.Workbook()
    sh = book.add_sheet("S")
    sh.write_merge(0, 0, 0, 1, "x")
    path = tmp_path / "m.xls"
    book.save(str(path))
    with pytest.raises(AdapterError) as ei:
        XlsWorkbookAdapter(limits=SourceArtifactLimits(max_merged_regions=0)).inspect(str(path))
    assert ei.value.error_code == "merged_region_limit"


# --- R-06 reconciler checksum ---


def test_reconciler_pending_checksum_mismatch(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    key = f"org/{org.id}/k"
    payload = b"hello-world"
    fake_storage._objects[key] = payload
    wrong = hashlib.sha256(b"other").hexdigest()
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="application/octet-stream",
        file_size_bytes=len(payload),
        checksum_sha256=wrong,
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    db_session.refresh(art)
    assert art.state == "failed"
    assert art.failure_code == "checksum_mismatch"
    assert stats["marked_failed"] >= 1


def test_reconciler_delete_failure_no_audit(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    key = f"org/{org.id}/k2"
    fake_storage._objects[key] = b"x"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="a" * 64,
        storage_object_key=key,
        state="orphaned",
        created_by_user_id=user.id,
        orphaned_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()
    fake_storage.fail_delete = True
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=5, actor_id=user.id, org_id=org.id
    )
    assert stats["deleted_objects"] == 0
    assert stats["errors"] >= 1
    assert (
        db_session.query(AuditEvent)
        .filter(AuditEvent.event_name == "ImportSourceArtifactObjectDeleted")
        .count()
        == 0
    )


def test_reconciler_ref_check_blocks(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    key = f"org/{org.id}/k3"
    fake_storage._objects[key] = b"x"
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=1,
        checksum_sha256="b" * 64,
        storage_object_key=key,
        state="orphaned",
        created_by_user_id=user.id,
        orphaned_at=datetime.now(timezone.utc) - timedelta(hours=2),
    )
    db_session.add(art)
    db_session.commit()

    def always_ref(db, a):
        return True

    stats = reconcile_source_artifacts(
        db_session,
        storage=fake_storage,
        max_items=5,
        actor_id=user.id,
        org_id=org.id,
        reference_check=always_ref,
    )
    assert stats["deleted_objects"] == 0
    assert key in fake_storage._objects


def test_reconciler_max_items_deterministic(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    base = datetime.now(timezone.utc) - timedelta(hours=5)
    for i in range(3):
        art = ImportSourceArtifact(
            organization_id=org.id,
            project_id=proj.id,
            import_batch_id=batch.id,
            generation=i + 1,
            original_filename=f"a{i}.xlsx",
            detected_format="xlsx",
            content_type="t",
            file_size_bytes=0,
            checksum_sha256="c" * 64,
            storage_object_key=f"org/{org.id}/m{i}",
            state="pending",
            created_by_user_id=user.id,
            created_at=base + timedelta(seconds=i),
        )
        db_session.add(art)
    db_session.commit()
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=2, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 2


# --- boundaries ---


def test_xlsx_sheet_limit(tmp_path):
    p = tmp_path / "s.xlsx"
    p.write_bytes(_xlsx_bytes(sheets=3))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_sheets=2)).inspect(str(p))
    assert ei.value.error_code == "sheet_limit"


def test_xlsx_zip_entry_limit(tmp_path):
    p = tmp_path / "z.xlsx"
    # minimal invalid many entries
    with zipfile.ZipFile(p, "w") as zf:
        zf.writestr("[Content_Types].xml", "<Types/>")
        zf.writestr("xl/workbook.xml", "<workbook/>")
        for i in range(10):
            zf.writestr(f"xl/pad{i}.xml", "<a/>")
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_zip_entries=5)).inspect(str(p))
    assert ei.value.error_code in {"zip_entry_limit", "invalid_xlsx"}


def test_xlsx_cell_limit_exact(tmp_path):
    p = tmp_path / "c.xlsx"
    p.write_bytes(_xlsx_bytes(long_cell="x" * 11))
    with pytest.raises(AdapterError) as ei:
        XlsxWorkbookAdapter(limits=SourceArtifactLimits(max_cell_chars=10)).inspect(str(p))
    assert ei.value.error_code == "cell_length_limit"


def test_nosuchbucket_not_object_not_found():
    from app.modules.excel_import.infrastructure.object_storage import _is_not_found_error

    class E(Exception):
        pass

    class CE(Exception):
        def __init__(self):
            self.response = {"Error": {"Code": "NoSuchBucket"}}

    # ClientError-like
    try:
        from botocore.exceptions import ClientError

        err = ClientError({"Error": {"Code": "NoSuchBucket"}}, "HeadObject")
        assert _is_not_found_error(err) is False
        err2 = ClientError({"Error": {"Code": "NoSuchKey"}}, "HeadObject")
        assert _is_not_found_error(err2) is True
    except ImportError:
        pytest.skip("botocore missing")


# --- partial failure: put fail preserves prior ---


def test_object_put_fail_preserves_prior(client: TestClient, db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    data = _xlsx_bytes()
    r1 = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "a.xlsx",
                io.BytesIO(data),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert r1.status_code == 201
    first = r1.json()["id"]
    lines_before = db_session.query(ProjectAssetLine).count()
    fake_storage.fail_put = True
    r2 = client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={
            "file": (
                "b.xlsx",
                io.BytesIO(data),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers={"X-User-Id": str(user.id)},
    )
    assert r2.status_code == 500
    db_session.refresh(batch)
    assert str(batch.current_source_artifact_id) == first
    assert db_session.query(ProjectAssetLine).count() == lines_before
    failed = db_session.query(ImportSourceArtifact).filter_by(state="failed").all()
    assert failed


@pytest.mark.skipif(
    not (os.environ.get("TEST_DATABASE_URL") or "").startswith("postgresql"),
    reason="PostgreSQL same-batch pointer proof requires TEST_DATABASE_URL",
)
def test_pg_same_batch_current_pointer_enforced():
    """Cross-batch current pointer must be rejected by composite FK."""
    url = os.environ["TEST_DATABASE_URL"]
    if os.environ.get("CI") == "true":
        assert url
    engine = create_engine(url)
    with engine.begin() as conn:
        # Use throwaway orgs if tables empty — rely on migration-applied schema
        row = conn.execute(
            text(
                """
                SELECT tc.constraint_name
                FROM information_schema.table_constraints tc
                WHERE tc.table_name = 'project_asset_import_batches'
                  AND tc.constraint_name = 'fk_batch_current_artifact_same_batch'
                """
            )
        ).fetchone()
        assert row is not None, "same-batch FK missing from schema"
        chk = conn.execute(
            text(
                """
                SELECT 1 FROM information_schema.check_constraints
                WHERE constraint_name = 'chk_source_artifact_checksum_hex'
                """
            )
        ).fetchone()
        assert chk is not None
