"""S13-PR-002 ninth corrective: J-01…J-05 cross-dialect ownership and complete matrix."""
from __future__ import annotations

import hashlib
import io
import os
import threading
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from typing import Any

import openpyxl
import pytest
from fastapi import UploadFile
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy.orm import Session, sessionmaker
from starlette.datastructures import Headers

from app.main import app as fastapi_app
from app.db import Base, get_db
import app.modules.excel_import.models  # noqa: F401
from app.modules.excel_import.application.source_artifact_service import (
    reconcile_source_artifacts,
    set_pointer_probe_hook,
    set_reconcile_work_session_factory,
    set_source_limits_override,
    upload_source_artifact,
    _sha256_object,
)
from app.modules.excel_import.domain.source_artifact import SourceArtifactLimits
from app.modules.excel_import.infrastructure.object_storage import (
    FakeObjectStorage,
    set_object_storage_override,
)
from tests.support.s13_pr_002_http_preserve import (
    assert_http_rejection_preserve,
    snapshot_source_intake_preserve,
)
from app.modules.excel_import.models import ImportSourceArtifact
from app.modules.project_master_data.models import (
    OrganizationProfile,
    OrganizationStatus,
    User,
    UserStatus,
    Role,
    UserRole,
    Customer,
    CustomerStatus,
    Project,
    ProjectWorkflowStatus,
    ProjectAssetImportBatch,
    ProjectAssetImportStagingRow,
    ImportBatchStatus,
    ImportRowValidationStatus,
    ProjectAssetLine,
    AssetLineReviewStatus,
    AssetLineValidationStatus,
    AuditEvent,
)


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


@pytest.fixture
def db_session(tmp_path) -> Session:
    db_file = tmp_path / f"s13e9_{uuid.uuid4().hex}.db"
    engine = create_engine(f"sqlite:///{db_file}", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def fake_storage():
    store = FakeObjectStorage()
    set_object_storage_override(store)
    yield store
    set_object_storage_override(None)


@pytest.fixture
def client(db_session: Session, fake_storage) -> TestClient:
    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    fastapi_app.dependency_overrides[get_db] = override_get_db
    yield TestClient(fastapi_app)
    fastapi_app.dependency_overrides.clear()
    set_source_limits_override(None)
    set_reconcile_work_session_factory(None)


def _seed(db: Session):
    org = OrganizationProfile(
        legal_name="Org",
        organization_slug=f"o-{uuid.uuid4().hex[:6]}",
        status=OrganizationStatus.ACTIVE,
    )
    db.add(org)
    db.commit()
    role = Role(code="editor", display_name="E", permissions=["project:read", "workbench:edit"])
    db.add(role)
    db.commit()
    user = User(
        organization_id=org.id,
        email=f"u{uuid.uuid4().hex[:8]}@ex.com",
        full_name="U",
        status=UserStatus.ACTIVE,
    )
    db.add(user)
    db.commit()
    db.add(UserRole(user_id=user.id, role_id=role.id, is_active=True))
    db.commit()
    cust = Customer(
        organization_id=org.id, legal_name="C", status=CustomerStatus.ACTIVE, created_by=user.id
    )
    db.add(cust)
    db.commit()
    proj = Project(
        organization_id=org.id,
        customer_id=cust.id,
        name="P",
        code=f"P{uuid.uuid4().hex[:6]}",
        status=ProjectWorkflowStatus.DRAFT,
        created_by=user.id,
    )
    db.add(proj)
    db.commit()
    batch = ProjectAssetImportBatch(
        organization_id=org.id,
        project_id=proj.id,
        source_filename="s.xlsx",
        status=ImportBatchStatus.CREATED,
        created_by_user_id=user.id,
    )
    db.add(batch)
    db.commit()
    return org, user, proj, batch


def _xlsx_bytes(*, sheets=1, rows=1, cols=1, cell="x", merges=0) -> bytes:
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    for r in range(rows):
        for c in range(cols):
            wb.active.cell(r + 1, c + 1, cell)
    for i in range(merges):
        col = (i % max(cols, 1)) + 1
        end_col = min(col + 1, max(cols + 1, 2))
        wb.active.merge_cells(
            start_row=1, start_column=col, end_row=1, end_column=end_col
        )
    for i in range(1, sheets):
        wb.create_sheet(f"S{i + 1}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_bytes(tmp_path, *, sheets=1, rows=1, cols=1, cell="x", merges=0) -> bytes:
    pytest.importorskip("xlwt")
    import xlwt

    p = tmp_path / f"t-{uuid.uuid4().hex[:6]}.xls"
    book = xlwt.Workbook()
    for si in range(sheets):
        sh = book.add_sheet(f"S{si + 1}")
        for r in range(max(rows, 1)):
            for c in range(max(cols, 1)):
                # Leave row 0 free when merges are requested (write_merge owns those cells).
                if merges and r == 0:
                    continue
                sh.write(r, c, cell)
        for i in range(merges):
            c0 = i * 2
            c1 = c0 + 1
            sh.write_merge(0, 0, c0, c1, cell)
    book.save(str(p))
    return p.read_bytes()


def _xlsx_with_extra_zip_entries(base: bytes, extra_entries: int) -> bytes:
    """Add dummy zip members to exceed max_zip_entries (keeps valid xlsx parts)."""
    src = zipfile.ZipFile(io.BytesIO(base), "r")
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            dst.writestr(info, src.read(info.filename))
        for i in range(extra_entries):
            dst.writestr(f"xl/media/pad{i}.bin", b"x")
    src.close()
    return out_buf.getvalue()


def _xlsx_inflated_uncompressed(base: bytes, min_uncompressed: int) -> bytes:
    """Pad an existing xlsx zip so total uncompressed size exceeds a small limit."""
    src = zipfile.ZipFile(io.BytesIO(base), "r")
    current = sum(i.file_size for i in src.infolist())
    pad = max(0, min_uncompressed - current + 1)
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_STORED) as dst:
        for info in src.infolist():
            dst.writestr(info, src.read(info.filename))
        if pad:
            dst.writestr("xl/media/pad.bin", b"Z" * pad)
    src.close()
    return out_buf.getvalue()


def _seed_prior_full(db, org, user, proj, batch, fake_storage):
    payload = _xlsx_bytes()
    key = f"org/{org.id}/prior-{uuid.uuid4().hex[:8]}"
    ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    fake_storage._objects[key] = payload
    fake_storage._content_types[key] = ct
    prior = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="prior.xlsx",
        detected_format="xlsx",
        content_type=ct,
        file_size_bytes=len(payload),
        checksum_sha256=hashlib.sha256(payload).hexdigest(),
        storage_object_key=key,
        state="available",
        created_by_user_id=user.id,
        available_at=datetime.now(timezone.utc),
    )
    db.add(prior)
    db.flush()
    batch.current_source_artifact_id = prior.id
    batch.total_rows = 1
    batch.valid_rows = 1
    staging = ProjectAssetImportStagingRow(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        source_row_number=1,
        raw_values={"A": "keep-me"},
        mapped_values={"name": "keep-me"},
        normalized_preview={"n": 1},
        validation_status=ImportRowValidationStatus.VALID,
        validation_errors=[],
        validation_warnings=[],
        proposed_asset_name="keep-me",
        proposed_description="desc-keep",
        proposed_quantity="1",
        proposed_unit="cái",
    )
    db.add(staging)
    line = ProjectAssetLine(
        project_id=proj.id,
        asset_name="Official Keep",
        description="official-desc",
        quantity=2.5,
        review_status=AssetLineReviewStatus.PENDING,
        validation_status=AssetLineValidationStatus.UNVALIDATED,
    )
    db.add(line)
    db.commit()
    snap = snapshot_source_intake_preserve(
        db, fake_storage, project_id=proj.id, batch_id=batch.id
    )
    # Legacy convenience keys still used by non-reject nodes in this suite.
    snap.update(
        {
            "prior_id": prior.id,
            "prior_checksum": prior.checksum_sha256,
            "prior_key": prior.storage_object_key,
            "prior_state": prior.state,
            "prior_gen": prior.generation,
            "batch_current": batch.current_source_artifact_id,
            "batch_status": batch.status,
            "batch_total": batch.total_rows,
            "batch_valid": batch.valid_rows,
            "staging_id": staging.id,
            "staging_raw": dict(staging.raw_values),
            "staging_mapped": dict(staging.mapped_values),
            "staging_name": staging.proposed_asset_name,
            "staging_desc": staging.proposed_description,
            "staging_qty": staging.proposed_quantity,
            "staging_unit": staging.proposed_unit,
            "staging_status": staging.validation_status,
            "line_id": line.id,
            "line_name": line.asset_name,
            "line_desc": line.description,
            "line_qty": float(line.quantity),
            "line_review": line.review_status,
            "line_val": line.validation_status,
            "art_count": len(snap["artifacts"]),
            "line_count": len(snap["lines"]),
        }
    )
    return prior, staging, line, snap


def _assert_preserved(db, fake_storage, proj, batch, staging, line, snap):
    """Field-level prior/staging/line preserve for non-reject paths.

    HTTP N+1 rejections use _assert_reject_preserve → assert_http_rejection_preserve.
    """
    db.expire_all()
    prior = db.get(ImportSourceArtifact, snap["prior_id"])
    batch = db.get(ProjectAssetImportBatch, batch.id)
    staging = db.get(ProjectAssetImportStagingRow, staging.id)
    line = db.get(ProjectAssetLine, line.id)
    assert prior is not None
    assert prior.checksum_sha256 == snap["prior_checksum"]
    assert prior.storage_object_key == snap["prior_key"]
    assert prior.state == snap["prior_state"]
    assert prior.generation == snap["prior_gen"]
    assert batch.current_source_artifact_id == snap["batch_current"]
    assert batch.status == snap["batch_status"]
    assert batch.total_rows == snap["batch_total"]
    assert batch.valid_rows == snap["batch_valid"]
    assert staging.raw_values == snap["staging_raw"]
    assert staging.mapped_values == snap["staging_mapped"]
    assert staging.proposed_asset_name == snap["staging_name"]
    assert staging.proposed_description == snap["staging_desc"]
    assert staging.proposed_quantity == snap["staging_qty"]
    assert staging.proposed_unit == snap["staging_unit"]
    assert staging.validation_status == snap["staging_status"]
    assert line.asset_name == snap["line_name"]
    assert line.description == snap["line_desc"]
    assert float(line.quantity) == snap["line_qty"]
    assert line.review_status == snap["line_review"]
    assert line.validation_status == snap["line_val"]
    assert snap["prior_key"] in fake_storage._objects
    assert db.query(ProjectAssetLine).filter_by(project_id=proj.id).count() == snap["line_count"]


def _assert_error_code(res, expected_code: str):
    detail = res.json().get("detail")
    assert isinstance(detail, dict), f"detail must be mapping: {detail!r}"
    assert detail.get("error_code") == expected_code


def _audit_count(db, entity_id, event_name) -> int:
    return (
        db.query(AuditEvent)
        .filter(AuditEvent.entity_id == entity_id, AuditEvent.event_name == event_name)
        .count()
    )


def _pg_url():
    url = os.environ.get("TEST_DATABASE_URL") or ""
    if os.environ.get("CI") == "true":
        assert url.startswith("postgresql")
    elif not url.startswith("postgresql"):
        pytest.skip("PostgreSQL required")
    return url


def _cname(exc: IntegrityError) -> str:
    orig = getattr(exc, "orig", None)
    diag = getattr(orig, "diag", None) if orig is not None else None
    return getattr(diag, "constraint_name", None) or ""


def _pg_sqlstate(exc: BaseException) -> str:
    orig = getattr(exc, "orig", None)
    return getattr(orig, "sqlstate", None) or getattr(orig, "pgcode", None) or ""


# ---------------------------------------------------------------------------
# J-01 cross-dialect promotion + both overlap winners
# ---------------------------------------------------------------------------


def test_j01_sqlite_verified_pending_promotion(db_session: Session, fake_storage):
    """At 392614c this returned errors=1 (UUID bind). Must promote with errors=0."""
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    body = b"fresh-valid-pending"
    key = f"pending/{uuid.uuid4()}"
    fake_storage._objects[key] = body
    art = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=2,
        original_filename="a.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body),
        checksum_sha256=hashlib.sha256(body).hexdigest(),
        storage_object_key=key,
        state="pending",
        created_by_user_id=user.id,
        created_at=datetime.now(timezone.utc),
    )
    db_session.add(art)
    db_session.commit()
    aid = art.id
    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 1
    assert stats["errors"] == 0
    assert stats["marked_failed"] == 0
    assert stats["marked_orphan"] == 0
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, aid)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert art.state == "available"
    assert batch.current_source_artifact_id == aid
    assert _audit_count(db_session, aid, "ImportSourceArtifactAvailable") == 1
    # prior still present as available history
    prior = db_session.get(ImportSourceArtifact, snap["prior_id"])
    assert prior.state == "available"
    staging = db_session.get(ProjectAssetImportStagingRow, staging.id)
    line = db_session.get(ProjectAssetLine, line.id)
    assert staging.proposed_asset_name == snap["staging_name"]
    assert line.asset_name == snap["line_name"]


def test_j01_pg_verified_pending_promotion():
    url = _pg_url()
    engine = create_engine(url)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    aid = uuid.uuid4()
    prior_id = uuid.uuid4()
    slug = f"j01p-{uuid.uuid4().hex[:8]}"
    store = FakeObjectStorage()
    body = b"pg-pending-body"
    key = f"pg/{aid}"
    store._objects[key] = body
    prior_body = b"prior"
    prior_key = f"pg/prior/{prior_id}"
    store._objects[prior_key] = prior_body
    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'b.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid, "oid": oid, "pid": pid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at, available_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'p.xlsx', 'xlsx', 't', :sz,
                      :chk, :key, 'available', '{}'::jsonb, :uid, now(), now(), now()
                    )
                    """
                ),
                {
                    "id": prior_id,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid,
                    "sz": len(prior_body),
                    "chk": hashlib.sha256(prior_body).hexdigest(),
                    "key": prior_key,
                    "uid": uid,
                },
            )
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": prior_id, "bid": bid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 2, 'a.xlsx', 'xlsx', 't', :sz,
                      :chk, :key, 'pending', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": aid,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid,
                    "sz": len(body),
                    "chk": hashlib.sha256(body).hexdigest(),
                    "key": key,
                    "uid": uid,
                },
            )
        caller = SessionLocal()
        try:
            stats = reconcile_source_artifacts(
                caller, storage=store, max_items=10, actor_id=uid, org_id=oid
            )
            assert stats["scanned"] == 1
            assert stats["errors"] == 0
            assert stats["marked_failed"] == 0
            assert stats["marked_orphan"] == 0
            art = caller.get(ImportSourceArtifact, aid)
            batch = caller.get(ProjectAssetImportBatch, bid)
            assert art.state == "available"
            assert batch.current_source_artifact_id == aid
            assert (
                caller.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == aid,
                    AuditEvent.event_name == "ImportSourceArtifactAvailable",
                )
                .count()
                == 1
            )
        finally:
            caller.close()
    finally:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL "
                    "WHERE organization_id = :oid"
                ),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


def test_j01_overlap_reconciler_wins_uploader_idempotent(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    hooks = {"mid_put": 0, "reconcile_stats": None}
    real_put = fake_storage.put_stream

    def hooked_put(key, stream, content_type=None, expected_size=None):
        st = real_put(key, stream, content_type=content_type, expected_size=expected_size)
        hooks["mid_put"] += 1
        hooks["reconcile_stats"] = reconcile_source_artifacts(
            db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
        )
        return st

    fake_storage.put_stream = hooked_put  # type: ignore[method-assign]

    class R:
        headers = Headers({})

    art = upload_source_artifact(
        db_session,
        org_id=org.id,
        project_id=proj.id,
        batch_id=batch.id,
        file=UploadFile(filename="ov.xlsx", file=io.BytesIO(payload)),
        request=R(),
        current_user=user,
        storage=fake_storage,
    )
    assert hooks["mid_put"] == 1
    stats = hooks["reconcile_stats"]
    assert stats is not None
    assert stats["errors"] == 0
    assert stats["scanned"] == 1
    assert art.state == "available"
    db_session.expire_all()
    art = db_session.get(ImportSourceArtifact, art.id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert art.state == "available"
    assert batch.current_source_artifact_id == art.id
    avail = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .all()
    )
    assert len(avail) == 1
    assert avail[0].command_name == "ReconcileImportSourceArtifact"
    assert _audit_count(db_session, art.id, "ImportSourceArtifactReserved") == 1
    for name in (
        "ImportSourceArtifactFailed",
        "ImportSourceArtifactOrphaned",
        "ImportSourceArtifactObjectDeleted",
    ):
        assert _audit_count(db_session, art.id, name) == 0
    assert art.storage_object_key in fake_storage._objects
    assert fake_storage._objects[art.storage_object_key]  # bytes present
    # full preserve of prior snapshot fields that must survive
    staging = db_session.get(ProjectAssetImportStagingRow, staging.id)
    line = db_session.get(ProjectAssetLine, line.id)
    assert staging.raw_values == snap["staging_raw"]
    assert staging.mapped_values == snap["staging_mapped"]
    assert staging.proposed_asset_name == snap["staging_name"]
    assert staging.proposed_description == snap["staging_desc"]
    assert staging.proposed_quantity == snap["staging_qty"]
    assert staging.proposed_unit == snap["staging_unit"]
    assert staging.validation_status == snap["staging_status"]
    assert line.asset_name == snap["line_name"]
    assert line.description == snap["line_desc"]
    assert float(line.quantity) == snap["line_qty"]
    assert line.review_status == snap["line_review"]
    assert line.validation_status == snap["line_val"]
    assert snap["prior_key"] in fake_storage._objects
    assert fake_storage._objects[snap["prior_key"]] == snap["objects"][snap["prior_key"]]
    # idempotent retry
    stats2 = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats2["scanned"] == 0
    assert stats2["errors"] == 0
    assert _audit_count(db_session, art.id, "ImportSourceArtifactAvailable") == 1


def test_j01_overlap_uploader_wins_reconciler_skips(db_session: Session, fake_storage):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()

    class R:
        headers = Headers({})

    art = upload_source_artifact(
        db_session,
        org_id=org.id,
        project_id=proj.id,
        batch_id=batch.id,
        file=UploadFile(filename="up.xlsx", file=io.BytesIO(payload)),
        request=R(),
        current_user=user,
        storage=fake_storage,
    )
    assert art.state == "available"
    avail = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == art.id,
            AuditEvent.event_name == "ImportSourceArtifactAvailable",
        )
        .all()
    )
    assert len(avail) == 1
    assert avail[0].command_name == "UploadImportSourceArtifact"
    assert _audit_count(db_session, art.id, "ImportSourceArtifactReserved") == 1

    stats = reconcile_source_artifacts(
        db_session, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
    )
    assert stats["scanned"] == 0
    assert stats["errors"] == 0
    assert _audit_count(db_session, art.id, "ImportSourceArtifactAvailable") == 1
    db_session.expire_all()
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert batch.current_source_artifact_id == art.id
    staging = db_session.get(ProjectAssetImportStagingRow, staging.id)
    line = db_session.get(ProjectAssetLine, line.id)
    assert staging.proposed_asset_name == snap["staging_name"]
    assert line.asset_name == snap["line_name"]
    assert fake_storage._objects[snap["prior_key"]] == snap["objects"][snap["prior_key"]]


# ---------------------------------------------------------------------------
# J-02 within-retention generation race + fresh Session durable-then-raise
# ---------------------------------------------------------------------------


def test_j02_within_retention_stale_promotion_race(db_session: Session, fake_storage):
    """Older pending inside retention; probe barrier then newer finalize; revalidation orphans."""
    org, user, proj, batch = _seed(db_session)
    uid, oid = user.id, org.id
    body1 = b"older-valid-within-retention"
    k1 = f"old/{uuid.uuid4()}"
    fake_storage._objects[k1] = body1
    a1 = ImportSourceArtifact(
        organization_id=org.id,
        project_id=proj.id,
        import_batch_id=batch.id,
        generation=1,
        original_filename="old.xlsx",
        detected_format="xlsx",
        content_type="t",
        file_size_bytes=len(body1),
        checksum_sha256=hashlib.sha256(body1).hexdigest(),
        storage_object_key=k1,
        state="pending",
        created_by_user_id=user.id,
        # Within 1h retention → promote path, not pending_retention orphan
        created_at=datetime.now(timezone.utc) - timedelta(minutes=5),
    )
    db_session.add(a1)
    db_session.commit()
    a1_id = a1.id

    barrier_entered = threading.Event()
    release = threading.Event()
    hooks = {"probe_hits": 0}
    errors: list[BaseException] = []
    result: dict[str, Any] = {}

    def probe_hook():
        hooks["probe_hits"] += 1
        if hooks["probe_hits"] == 1:
            barrier_entered.set()
            assert release.wait(timeout=15), "release timeout"

    set_pointer_probe_hook(probe_hook)

    def run_reconcile():
        try:
            result["stats"] = reconcile_source_artifacts(
                db_session, storage=fake_storage, max_items=10, actor_id=uid, org_id=oid
            )
        except BaseException as exc:  # noqa: BLE001
            errors.append(exc)
        finally:
            set_pointer_probe_hook(None)

    t = threading.Thread(target=run_reconcile)
    t.start()
    assert barrier_entered.wait(timeout=15), "reconciler never reached pointer probe"

    body2 = _xlsx_bytes(cell="newer")
    UploadSession = sessionmaker(bind=db_session.get_bind(), autoflush=False, autocommit=False)
    up_db = UploadSession()
    try:

        class R:
            headers = Headers({})

        art2 = upload_source_artifact(
            up_db,
            org_id=org.id,
            project_id=proj.id,
            batch_id=batch.id,
            file=UploadFile(filename="new.xlsx", file=io.BytesIO(body2)),
            request=R(),
            current_user=user,
            storage=fake_storage,
        )
    finally:
        up_db.close()

    assert art2.state == "available"
    a2_id = art2.id
    release.set()
    t.join(timeout=20)
    assert not t.is_alive()
    assert not errors, errors
    assert hooks["probe_hits"] >= 1
    stats = result["stats"]
    assert stats["errors"] == 0

    db_session.expire_all()
    a1 = db_session.get(ImportSourceArtifact, a1_id)
    a2 = db_session.get(ImportSourceArtifact, a2_id)
    batch = db_session.get(ProjectAssetImportBatch, batch.id)
    assert a2.state == "available"
    assert batch.current_source_artifact_id == a2_id
    assert a1.state == "orphaned"
    assert a1.failure_code == "stale_generation"
    assert _audit_count(db_session, a2_id, "ImportSourceArtifactAvailable") == 1
    assert _audit_count(db_session, a1_id, "ImportSourceArtifactAvailable") == 0
    orphan_audits = (
        db_session.query(AuditEvent)
        .filter(
            AuditEvent.entity_id == a1_id,
            AuditEvent.event_name == "ImportSourceArtifactOrphaned",
        )
        .all()
    )
    assert len(orphan_audits) == 1
    reason = (orphan_audits[0].payload or {}).get("reason", "")
    assert "stale" in reason
    assert k1 in fake_storage._objects
    assert a2.storage_object_key in fake_storage._objects


def test_j02_pg_within_retention_stale_promotion_race():
    """PostgreSQL: decision-time newer_wins for within-retention older pending."""
    url = _pg_url()
    engine = create_engine(url)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    a1 = uuid.uuid4()
    a2 = uuid.uuid4()
    slug = f"j02p-{uuid.uuid4().hex[:8]}"
    store = FakeObjectStorage()
    b1 = b"old-within"
    b2 = b"new-available-xx"
    k1 = f"j02/{a1}"
    k2 = f"j02/{a2}"
    store._objects[k1] = b1
    store._objects[k2] = b2
    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'b.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid, "oid": oid, "pid": pid, "uid": uid},
            )
            # newer available current gen=2 already durable
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at, available_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 2, 'new.xlsx', 'xlsx', 't', :sz,
                      :chk, :key, 'available', '{}'::jsonb, :uid, now(), now(), now()
                    )
                    """
                ),
                {
                    "id": a2,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid,
                    "sz": len(b2),
                    "chk": hashlib.sha256(b2).hexdigest(),
                    "key": k2,
                    "uid": uid,
                },
            )
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": a2, "bid": bid},
            )
            # older pending within retention
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'old.xlsx', 'xlsx', 't', :sz,
                      :chk, :key, 'pending', '{}'::jsonb, :uid, now() - interval '5 minutes', now()
                    )
                    """
                ),
                {
                    "id": a1,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid,
                    "sz": len(b1),
                    "chk": hashlib.sha256(b1).hexdigest(),
                    "key": k1,
                    "uid": uid,
                },
            )
        caller = SessionLocal()
        try:
            stats = reconcile_source_artifacts(
                caller, storage=store, max_items=10, actor_id=uid, org_id=oid
            )
            assert stats["errors"] == 0
            assert stats["scanned"] == 1
            old = caller.get(ImportSourceArtifact, a1)
            new = caller.get(ImportSourceArtifact, a2)
            batch = caller.get(ProjectAssetImportBatch, bid)
            assert new.state == "available"
            assert batch.current_source_artifact_id == a2
            assert old.state == "orphaned"
            assert old.failure_code == "stale_generation"
            assert (
                caller.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == a2,
                    AuditEvent.event_name == "ImportSourceArtifactAvailable",
                )
                .count()
                == 0
            )  # pre-seeded available without audit required; pointer winner is gen2
            assert (
                caller.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == a1,
                    AuditEvent.event_name == "ImportSourceArtifactAvailable",
                )
                .count()
                == 0
            )
            assert (
                caller.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == a1,
                    AuditEvent.event_name == "ImportSourceArtifactOrphaned",
                )
                .count()
                == 1
            )
        finally:
            caller.close()
    finally:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL "
                    "WHERE organization_id = :oid"
                ),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


def test_j02_durable_then_raise_fresh_session_recovery(
    db_session: Session, fake_storage, monkeypatch, tmp_path
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    payload = _xlsx_bytes()
    p = tmp_path / "d.xlsx"
    p.write_bytes(payload)
    hooks = {"commit_n": 0, "durable_then_raise": 0}
    real = Session.commit

    def flaky(self):
        if self is not db_session:
            return real(self)
        hooks["commit_n"] += 1
        if hooks["commit_n"] == 2:
            real(self)
            hooks["durable_then_raise"] += 1
            raise RuntimeError("disconnect_after_durable_commit")
        return real(self)

    monkeypatch.setattr(Session, "commit", flaky)

    class R:
        headers = Headers({})

    with open(p, "rb") as f:
        art = upload_source_artifact(
            db_session,
            org_id=org.id,
            project_id=proj.id,
            batch_id=batch.id,
            file=UploadFile(filename="d.xlsx", file=f),
            request=R(),
            current_user=user,
            storage=fake_storage,
        )
    assert hooks["durable_then_raise"] == 1
    assert art.state == "available"
    aid = art.id

    # Genuinely new Session (not the same db_session)
    Fresh = sessionmaker(bind=db_session.get_bind(), autoflush=False, autocommit=False)
    fresh = Fresh()
    try:
        recovered = fresh.get(ImportSourceArtifact, aid)
        batch_f = fresh.get(ProjectAssetImportBatch, batch.id)
        assert recovered is not None
        assert recovered.state == "available"
        assert batch_f.current_source_artifact_id == aid
        assert (
            fresh.query(AuditEvent)
            .filter(
                AuditEvent.entity_id == aid,
                AuditEvent.event_name == "ImportSourceArtifactAvailable",
            )
            .count()
            == 1
        )
        stats = reconcile_source_artifacts(
            fresh, storage=fake_storage, max_items=10, actor_id=user.id, org_id=org.id
        )
        assert stats["scanned"] == 0
        assert stats["errors"] == 0
        assert (
            fresh.query(AuditEvent)
            .filter(
                AuditEvent.entity_id == aid,
                AuditEvent.event_name == "ImportSourceArtifactAvailable",
            )
            .count()
            == 1
        )
    finally:
        fresh.close()


# ---------------------------------------------------------------------------
# J-03 complete adapter HTTP boundaries
# ---------------------------------------------------------------------------


def _post_source(client, proj, batch, user, filename, payload, content_type):
    return client.post(
        f"/api/v1/projects/{proj.id}/asset-imports/{batch.id}/source-artifacts",
        files={"file": (filename, io.BytesIO(payload), content_type)},
        headers={"X-User-Id": str(user.id)},
    )


def _assert_reject_preserve(
    res, *, status, error_code, db, fake_storage, proj, batch, staging, line, snap
):
    # Full immutable contract: objects/content_types/artifacts/batch/staging/lines/audits.
    assert_http_rejection_preserve(
        res,
        status=status,
        error_code=error_code,
        db=db,
        fake_storage=fake_storage,
        snap=snap,
    )


@pytest.mark.parametrize(
    "limit_field,n_limit,build_ok,build_bad,error_code,bad_status",
    [
        (
            "max_zip_entries",
            None,  # computed from fixture
            None,
            None,
            "zip_entry_limit",
            413,
        ),
        (
            "max_uncompressed_zip_bytes",
            None,
            None,
            None,
            "zip_expansion_limit",
            413,
        ),
        (
            "max_row_chars",
            4,
            lambda: _xlsx_bytes(cols=2, cell="ab"),
            lambda: _xlsx_bytes(cols=3, cell="ab"),
            "row_char_limit",
            413,
        ),
        (
            "max_total_cells",
            4,
            lambda: _xlsx_bytes(rows=2, cols=2),
            lambda: _xlsx_bytes(rows=2, cols=3),
            "total_cell_limit",
            413,
        ),
        (
            "max_merged_regions",
            1,
            lambda: _xlsx_bytes(merges=1, cols=2),
            lambda: _xlsx_bytes(merges=2, cols=3),
            "merged_region_limit",
            413,
        ),
        (
            "max_merged_regions_per_sheet",
            1,
            lambda: _xlsx_bytes(merges=1, cols=2),
            lambda: _xlsx_bytes(merges=2, cols=3),
            "merged_region_limit",
            413,
        ),
    ],
)
def test_j03_endpoint_xlsx_extra_adapter_bounds(
    client: TestClient,
    db_session: Session,
    fake_storage,
    limit_field,
    n_limit,
    build_ok,
    build_bad,
    error_code,
    bad_status,
):
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    base = _xlsx_bytes()
    if limit_field == "max_zip_entries":
        with zipfile.ZipFile(io.BytesIO(base)) as zf:
            n_ok = len(zf.infolist())
        ok_payload = base
        bad_payload = _xlsx_with_extra_zip_entries(base, 1)
        limits = SourceArtifactLimits(max_zip_entries=n_ok)
    elif limit_field == "max_uncompressed_zip_bytes":
        with zipfile.ZipFile(io.BytesIO(base)) as zf:
            n_ok = sum(i.file_size for i in zf.infolist())
        ok_payload = base
        bad_payload = _xlsx_inflated_uncompressed(base, n_ok)
        limits = SourceArtifactLimits(max_uncompressed_zip_bytes=n_ok)
    else:
        ok_payload = build_ok()
        bad_payload = build_bad()
        limits = SourceArtifactLimits(**{limit_field: n_limit})

    set_source_limits_override(limits)
    try:
        res_bad = _post_source(
            client,
            proj,
            batch,
            user,
            "lim.xlsx",
            bad_payload,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _assert_reject_preserve(
            res_bad,
            status=bad_status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            proj=proj,
            batch=batch,
            staging=staging,
            line=line,
            snap=snap,
        )
        res_ok = _post_source(
            client,
            proj,
            batch,
            user,
            "ok.xlsx",
            ok_payload,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        assert res_ok.status_code == 201, res_ok.text
    finally:
        set_source_limits_override(None)


@pytest.mark.parametrize(
    "limit_field,n_limit,ok_kwargs,bad_kwargs,error_code,bad_status",
    [
        (
            "max_row_chars",
            4,
            {"cols": 2, "cell": "ab"},
            {"cols": 3, "cell": "ab"},
            "row_char_limit",
            413,
        ),
        (
            "max_total_cells",
            4,
            {"rows": 2, "cols": 2},
            {"rows": 2, "cols": 3},
            "total_cell_limit",
            413,
        ),
        (
            "max_merged_regions",
            1,
            {"merges": 1, "cols": 2},
            {"merges": 2, "cols": 4},
            "merged_region_limit",
            413,
        ),
        (
            "max_merged_regions_per_sheet",
            1,
            {"merges": 1, "cols": 2},
            {"merges": 2, "cols": 4},
            "merged_region_limit",
            413,
        ),
    ],
)
def test_j03_endpoint_xls_extra_adapter_bounds(
    client: TestClient,
    db_session: Session,
    fake_storage,
    tmp_path,
    limit_field,
    n_limit,
    ok_kwargs,
    bad_kwargs,
    error_code,
    bad_status,
):
    # max_zip_entries / max_uncompressed_zip_bytes are xlsx-only (OLE path has no zip).
    org, user, proj, batch = _seed(db_session)
    prior, staging, line, snap = _seed_prior_full(
        db_session, org, user, proj, batch, fake_storage
    )
    set_source_limits_override(SourceArtifactLimits(**{limit_field: n_limit}))
    try:
        res_bad = _post_source(
            client,
            proj,
            batch,
            user,
            "lim.xls",
            _xls_bytes(tmp_path, **bad_kwargs),
            "application/vnd.ms-excel",
        )
        _assert_reject_preserve(
            res_bad,
            status=bad_status,
            error_code=error_code,
            db=db_session,
            fake_storage=fake_storage,
            proj=proj,
            batch=batch,
            staging=staging,
            line=line,
            snap=snap,
        )
        res_ok = _post_source(
            client,
            proj,
            batch,
            user,
            "ok.xls",
            _xls_bytes(tmp_path, **ok_kwargs),
            "application/vnd.ms-excel",
        )
        assert res_ok.status_code == 201, res_ok.text
    finally:
        set_source_limits_override(None)


def test_j03_xls_zip_limits_not_externally_reachable_documented():
    """xls is OLE — max_zip_entries / max_uncompressed_zip_bytes are not on the xls path."""
    from app.modules.excel_import.application.adapters.xls_adapter import XlsWorkbookAdapter
    import inspect

    src = inspect.getsource(XlsWorkbookAdapter)
    assert "max_zip_entries" not in src
    assert "max_uncompressed_zip_bytes" not in src
    assert "zip_entry_limit" not in src
    assert "zip_expansion_limit" not in src


# ---------------------------------------------------------------------------
# J-04 literal FK map
# ---------------------------------------------------------------------------


def test_j04_throwaway_literal_fk_map_and_dml():
    url = _pg_url()
    from alembic import command
    from alembic.config import Config
    from alembic.script import ScriptDirectory
    from app.core.config import get_settings
    from sqlalchemy.engine.url import make_url

    admin = create_engine(url, isolation_level="AUTOCOMMIT")
    db_name = f"s13_j04_{uuid.uuid4().hex[:10]}"
    with admin.connect() as conn:
        conn.execute(text(f'CREATE DATABASE "{db_name}"'))
    u = make_url(url)
    iso_url = url.rsplit("/", 1)[0] + f"/{db_name}"
    prev = {
        k: os.environ.get(k)
        for k in (
            "POSTGRES_HOST",
            "POSTGRES_PORT",
            "POSTGRES_DB",
            "POSTGRES_USER",
            "POSTGRES_PASSWORD",
            "VALORA_ENV",
        )
    }

    # Literal expected map from migration f2a3b4c5d6e7
    expected_artifact_fks = {
        ("created_by_user_id",): "import_source_artifacts_created_by_user_id_fkey",
        ("import_batch_id",): "import_source_artifacts_import_batch_id_fkey",
        ("organization_id",): "import_source_artifacts_organization_id_fkey",
        ("project_id",): "import_source_artifacts_project_id_fkey",
        ("organization_id", "project_id", "import_batch_id"): "fk_source_artifact_batch_tenant",
    }
    expected_batch_pointer_fk = "fk_batch_current_artifact_same_batch"

    def _assert_fk_map(eng):
        with eng.connect() as c:
            fk_rows = c.execute(
                text(
                    """
                    SELECT con.conname,
                           array_agg(a.attname ORDER BY u.ord) AS cols
                    FROM pg_constraint con
                    JOIN LATERAL unnest(con.conkey) WITH ORDINALITY AS u(attnum, ord) ON true
                    JOIN pg_attribute a ON a.attrelid = con.conrelid AND a.attnum = u.attnum
                    WHERE con.contype = 'f'
                      AND con.conrelid = 'import_source_artifacts'::regclass
                    GROUP BY con.conname
                    """
                )
            ).all()
            fk_by_cols = {tuple(r[1]): r[0] for r in fk_rows}
            assert set(fk_by_cols.keys()) == set(expected_artifact_fks.keys())
            assert fk_by_cols == expected_artifact_fks

            ptr = c.execute(
                text(
                    """
                    SELECT con.conname,
                           array_agg(a.attname ORDER BY u.ord) AS cols
                    FROM pg_constraint con
                    JOIN LATERAL unnest(con.conkey) WITH ORDINALITY AS u(attnum, ord) ON true
                    JOIN pg_attribute a ON a.attrelid = con.conrelid AND a.attnum = u.attnum
                    WHERE con.contype = 'f'
                      AND con.conrelid = 'project_asset_import_batches'::regclass
                      AND con.conname = :name
                    GROUP BY con.conname
                    """
                ),
                {"name": expected_batch_pointer_fk},
            ).one()
            assert ptr[0] == expected_batch_pointer_fk
            assert tuple(ptr[1]) == ("id", "current_source_artifact_id")

    def _seed_and_dml(eng):
        oid = uuid.uuid4()
        oid2 = uuid.uuid4()
        uid = uuid.uuid4()
        cid = uuid.uuid4()
        pid = uuid.uuid4()
        b1 = uuid.uuid4()
        b2 = uuid.uuid4()
        a1 = uuid.uuid4()
        slug = f"j04-{uuid.uuid4().hex[:8]}"
        with eng.begin() as conn:
            for o, s in ((oid, slug), (oid2, slug + "2")):
                conn.execute(
                    text(
                        "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                        "VALUES (:id, 'T', :slug, 'active', now(), now())"
                    ),
                    {"id": o, "slug": s},
                )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            for bid, fn in ((b1, "b1.xlsx"), (b2, "b2.xlsx")):
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :fn, 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": bid, "oid": oid, "pid": pid, "fn": fn, "uid": uid},
                )
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'x.xlsx', 'xlsx', 't', 1,
                      :chk, :key, 'available', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": a1,
                    "oid": oid,
                    "pid": pid,
                    "bid": b1,
                    "chk": "a" * 64,
                    "key": f"k-{a1}",
                    "uid": uid,
                },
            )
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                ),
                {"aid": a1, "bid": b1},
            )

        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        "UPDATE project_asset_import_batches SET current_source_artifact_id = :aid WHERE id = :bid"
                    ),
                    {"aid": a1, "bid": b2},
                )
            assert _cname(ei.value) == "fk_batch_current_artifact_same_batch"

        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text(
                        """
                        INSERT INTO import_source_artifacts (
                          id, organization_id, project_id, import_batch_id, generation,
                          original_filename, detected_format, content_type, file_size_bytes,
                          checksum_sha256, storage_object_key, state, adapter_metadata,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, :bid, 2, 'y.xlsx', 'xlsx', 't', 1,
                          :chk, :key, 'pending', '{}'::jsonb, :uid, now(), now()
                        )
                        """
                    ),
                    {
                        "id": uuid.uuid4(),
                        "oid": oid2,
                        "pid": pid,
                        "bid": b1,
                        "chk": "b" * 64,
                        "key": f"k-t-{uuid.uuid4()}",
                        "uid": uid,
                    },
                )
            assert _cname(ei.value) == "fk_source_artifact_batch_tenant"

        with eng.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL WHERE id = :bid"
                ),
                {"bid": b1},
            )
        with eng.begin() as conn:
            with pytest.raises(IntegrityError) as ei:
                conn.execute(
                    text("DELETE FROM project_asset_import_batches WHERE id = :id"),
                    {"id": b1},
                )
            assert _cname(ei.value) == "import_source_artifacts_import_batch_id_fkey"

    try:
        os.environ["VALORA_ENV"] = "test"
        os.environ["POSTGRES_HOST"] = u.host or "localhost"
        os.environ["POSTGRES_PORT"] = str(u.port or 5432)
        os.environ["POSTGRES_DB"] = db_name
        os.environ["POSTGRES_USER"] = u.username or "valora"
        os.environ["POSTGRES_PASSWORD"] = u.password or "valora_local_password"
        get_settings.cache_clear()
        cfg = Config("alembic.ini")
        assert ScriptDirectory.from_config(cfg).get_heads() == ["f2a3b4c5d6e7"]
        command.upgrade(cfg, "e1f2a3b4c5d6")
        eng = create_engine(iso_url)
        try:
            command.upgrade(cfg, "f2a3b4c5d6e7")
            _assert_fk_map(eng)
            _seed_and_dml(eng)
            command.downgrade(cfg, "e1f2a3b4c5d6")
            command.upgrade(cfg, "head")
            _assert_fk_map(eng)
            _seed_and_dml(eng)
        finally:
            eng.dispose()
    finally:
        for k, v in prev.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        get_settings.cache_clear()
        with admin.connect() as conn:
            conn.execute(text(f'DROP DATABASE IF EXISTS "{db_name}" WITH (FORCE)'))
        admin.dispose()


# ---------------------------------------------------------------------------
# J-05 lock continuity after reconcile
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("case", ["insert", "update", "delete", "readonly"])
def test_j05_pg_lock_continuity_pre_and_post_reconcile(case):
    url = _pg_url()
    engine = create_engine(url)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    caller = SessionLocal()
    oid = uuid.uuid4()
    uid = uuid.uuid4()
    cid = uuid.uuid4()
    pid = uuid.uuid4()
    bid = uuid.uuid4()
    bid_rec = uuid.uuid4()
    aid_rec = uuid.uuid4()
    extra_batch_id = uuid.uuid4()
    slug = f"j05-{case}-{uuid.uuid4().hex[:8]}"
    store = FakeObjectStorage()
    rec_body = b"reconcile-pending-body"
    rec_key = f"rec/{aid_rec}"
    store._objects[rec_key] = rec_body

    def _assert_lock_unavailable(target_id: uuid.UUID):
        other = SessionLocal()
        try:
            other.execute(text("SET LOCAL lock_timeout = '200ms'"))
            with pytest.raises(OperationalError) as ei:
                other.execute(
                    text(
                        "SELECT id FROM project_asset_import_batches WHERE id = :id FOR UPDATE NOWAIT"
                    ),
                    {"id": target_id},
                )
            state = _pg_sqlstate(ei.value)
            assert state == "55P03", f"expected 55P03, got {state!r}: {ei.value}"
            try:
                other.rollback()
            except Exception:
                pass
        finally:
            other.close()

    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO organization_profiles (id, legal_name, organization_slug, status, created_at, updated_at) "
                    "VALUES (:id, 'T', :slug, 'active', now(), now())"
                ),
                {"id": oid, "slug": slug},
            )
            conn.execute(
                text(
                    "INSERT INTO users (id, organization_id, email, full_name, status, created_at, updated_at) "
                    "VALUES (:id, :oid, :email, 'T', 'active', now(), now())"
                ),
                {"id": uid, "oid": oid, "email": f"{slug}@ex.com"},
            )
            conn.execute(
                text(
                    "INSERT INTO customers (id, organization_id, legal_name, status, created_by, created_at, updated_at) "
                    "VALUES (:id, :oid, 'C', 'active', :uid, now(), now())"
                ),
                {"id": cid, "oid": oid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO projects (
                      id, organization_id, customer_id, name, code, status,
                      knowledge_status, fee_amount, created_by, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :cid, 'P', :code, 'draft', 'pending', 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": pid, "oid": oid, "cid": cid, "code": slug[:20], "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'base.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid, "oid": oid, "pid": pid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO project_asset_import_batches (
                      id, organization_id, project_id, source_filename, status,
                      total_rows, valid_rows, invalid_rows, warning_rows,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, 'rec.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                    )
                    """
                ),
                {"id": bid_rec, "oid": oid, "pid": pid, "uid": uid},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO import_source_artifacts (
                      id, organization_id, project_id, import_batch_id, generation,
                      original_filename, detected_format, content_type, file_size_bytes,
                      checksum_sha256, storage_object_key, state, adapter_metadata,
                      created_by_user_id, created_at, updated_at
                    ) VALUES (
                      :id, :oid, :pid, :bid, 1, 'p.xlsx', 'xlsx', 't', :sz,
                      :chk, :key, 'pending', '{}'::jsonb, :uid, now(), now()
                    )
                    """
                ),
                {
                    "id": aid_rec,
                    "oid": oid,
                    "pid": pid,
                    "bid": bid_rec,
                    "sz": len(rec_body),
                    "chk": hashlib.sha256(rec_body).hexdigest(),
                    "key": rec_key,
                    "uid": uid,
                },
            )

        caller.execute(text("SELECT 1"))
        txn_before = caller.execute(text("SELECT txid_current()")).scalar()
        assert caller.in_transaction()
        # same Python Session object identity across reconcile
        caller_id = id(caller)

        if case == "insert":
            nb = ProjectAssetImportBatch(
                id=extra_batch_id,
                organization_id=oid,
                project_id=pid,
                source_filename="flushed-insert.xlsx",
                status=ImportBatchStatus.CREATED,
                created_by_user_id=uid,
            )
            caller.add(nb)
            caller.flush()
            lock_target = None
        elif case == "update":
            batch = caller.get(ProjectAssetImportBatch, bid)
            batch.source_filename = "flushed-uncommitted.xlsx"
            caller.flush()
            lock_target = bid
        elif case == "delete":
            with engine.begin() as conn:
                conn.execute(
                    text(
                        """
                        INSERT INTO project_asset_import_batches (
                          id, organization_id, project_id, source_filename, status,
                          total_rows, valid_rows, invalid_rows, warning_rows,
                          created_by_user_id, created_at, updated_at
                        ) VALUES (
                          :id, :oid, :pid, 'to-delete.xlsx', 'created', 0, 0, 0, 0, :uid, now(), now()
                        )
                        """
                    ),
                    {"id": extra_batch_id, "oid": oid, "pid": pid, "uid": uid},
                )
            doomed = caller.get(ProjectAssetImportBatch, extra_batch_id)
            caller.delete(doomed)
            caller.flush()
            # Lock the deleted row itself (FOR UPDATE on deleted row in txn)
            lock_target = extra_batch_id
            # Hold lock explicitly after delete flush via SELECT FOR UPDATE on remaining?
            # For deleted row, PG still holds row lock until commit/rollback.
            # Competing SELECT FOR UPDATE NOWAIT on the deleted id should get 55P03
            # if the delete lock is held — use that as lock_target.
        else:
            caller.execute(
                text("SELECT id FROM project_asset_import_batches WHERE id = :id FOR UPDATE"),
                {"id": bid},
            )
            lock_target = bid

        # Pre-reconcile lock / invisibility
        if case == "insert":
            other = SessionLocal()
            try:
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
            finally:
                other.close()
        else:
            _assert_lock_unavailable(lock_target)

        stats = reconcile_source_artifacts(
            caller, storage=store, max_items=10, actor_id=uid, org_id=oid
        )
        assert stats["scanned"] >= 1
        assert stats["errors"] == 0
        assert id(caller) == caller_id
        assert caller.in_transaction()
        txn_after = caller.execute(text("SELECT txid_current()")).scalar()
        assert txn_after == txn_before

        # Fresh session sees durable reconciler promotion
        fresh = SessionLocal()
        try:
            rec = fresh.get(ImportSourceArtifact, aid_rec)
            assert rec.state == "available"
            brec = fresh.get(ProjectAssetImportBatch, bid_rec)
            assert brec.current_source_artifact_id == aid_rec
            assert (
                fresh.query(AuditEvent)
                .filter(
                    AuditEvent.entity_id == aid_rec,
                    AuditEvent.event_name == "ImportSourceArtifactAvailable",
                )
                .count()
                == 1
            )
        finally:
            fresh.close()

        # Post-reconcile lock continuity / invisibility
        if case == "insert":
            other = SessionLocal()
            try:
                assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
            finally:
                other.close()
        else:
            _assert_lock_unavailable(lock_target)

        if case == "update":
            caller.commit()
            other = SessionLocal()
            try:
                assert (
                    other.get(ProjectAssetImportBatch, bid).source_filename
                    == "flushed-uncommitted.xlsx"
                )
            finally:
                other.close()
        else:
            caller.rollback()
            if case == "insert":
                other = SessionLocal()
                try:
                    assert other.get(ProjectAssetImportBatch, extra_batch_id) is None
                finally:
                    other.close()
            elif case == "delete":
                other = SessionLocal()
                try:
                    assert other.get(ProjectAssetImportBatch, extra_batch_id) is not None
                finally:
                    other.close()
    finally:
        try:
            if caller.in_transaction():
                caller.rollback()
        except Exception:
            pass
        try:
            caller.close()
        except Exception:
            pass
        with engine.begin() as conn:
            conn.execute(
                text(
                    "UPDATE project_asset_import_batches SET current_source_artifact_id = NULL "
                    "WHERE organization_id = :oid"
                ),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM import_source_artifacts WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(
                text("DELETE FROM project_asset_import_batches WHERE organization_id = :oid"),
                {"oid": oid},
            )
            conn.execute(text("DELETE FROM projects WHERE id = :id"), {"id": pid})
            conn.execute(text("DELETE FROM customers WHERE id = :id"), {"id": cid})
            conn.execute(text("DELETE FROM users WHERE id = :id"), {"id": uid})
            conn.execute(text("DELETE FROM organization_profiles WHERE id = :id"), {"id": oid})
        engine.dispose()


# ---------------------------------------------------------------------------
# MinIO
# ---------------------------------------------------------------------------


def test_j06_s3_minio_roundtrip_ci():
    if os.environ.get("CI") == "true":
        assert os.environ.get("S3_ENDPOINT_URL")
    elif not os.environ.get("S3_ENDPOINT_URL"):
        pytest.skip("S3_ENDPOINT_URL not set for local MinIO")
    from app.modules.excel_import.infrastructure.object_storage import S3ObjectStorage

    store = S3ObjectStorage(
        endpoint_url=os.environ["S3_ENDPOINT_URL"],
        access_key=os.environ.get("S3_ACCESS_KEY_ID", "valora"),
        secret_key=os.environ.get("S3_SECRET_ACCESS_KEY", "valora_local_password"),
        bucket=os.environ.get("S3_BUCKET", "valora-local"),
        region=os.environ.get("S3_REGION", "us-east-1"),
    )
    store.ensure_bucket()
    key = f"ninth/{uuid.uuid4()}"
    body = b"ninth-corrective"
    store.put_stream(
        key, io.BytesIO(body), content_type="application/octet-stream", expected_size=len(body)
    )
    digest = _sha256_object(store, key, chunk_size=64, expected_size=len(body))
    assert digest == hashlib.sha256(body).hexdigest()
    store.delete(key)
    assert store.head(key) is None
